"""
TicimaxExport.xls → ürün stoklarını BARKOD bazında eşleştirir.
- Her satır bir barkod (varyant); STOKADEDI o barkodun stoğu.
- Ürün varyantlarının stoğu barkoda göre güncellenir, ürün.stock = varyant toplamı.
- Eşleşen ürünlerden stoğu 0 olanlar pasife alınır (is_active=False).
- Eşleşmeyen (export'ta olmayan) ürünlere DOKUNULMAZ (güvenlik).

Kullanım:
    DRY-RUN (yazma yok):  python scripts/sync_stock_from_xls.py /tmp/TicimaxExport.xlsx
    UYGULA:               python scripts/sync_stock_from_xls.py /tmp/TicimaxExport.xlsx --apply
"""
import asyncio
import os
import sys

import openpyxl
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient

load_dotenv(os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env"))

COL_CARD = 0      # URUNKARTIID
COL_BARCODE = 4   # BARKOD
COL_STOCK = 34    # STOKADEDI


def _to_int(v) -> int:
    if v is None:
        return 0
    try:
        return max(0, int(float(str(v).replace(",", ".").strip() or 0)))
    except Exception:
        return 0


def _norm(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):  # openpyxl bazen sayısal barkodu float okur
        s = s[:-2]
    return s


def parse_xls(path: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows)  # header
    barcode_stock = {}
    card_stock = {}
    n = 0
    for r in rows:
        if not r or len(r) <= COL_STOCK:
            continue
        bc = _norm(r[COL_BARCODE])
        card = _norm(r[COL_CARD])
        stok = _to_int(r[COL_STOCK])
        if bc:
            barcode_stock[bc] = stok
        if card:
            card_stock[card] = card_stock.get(card, 0) + stok
        n += 1
    return barcode_stock, card_stock, n


async def main():
    if len(sys.argv) < 2:
        print("Kullanım: sync_stock_from_xls.py <xls_path> [--apply]")
        return
    path = sys.argv[1]
    apply = "--apply" in sys.argv

    barcode_stock, card_stock, n_rows = parse_xls(path)
    print(f"[xls] {n_rows} satır | {len(barcode_stock)} benzersiz barkod | {len(card_stock)} ürün kartı")

    client = AsyncIOMotorClient(os.environ["MONGO_URL"])
    db = client[os.environ["DB_NAME"]]

    total_products = 0
    matched_products = 0
    updated_variants = 0
    to_deactivate = []
    in_stock = 0
    unmatched = []

    cursor = db.products.find({}, {"id": 1, "name": 1, "slug": 1, "barcode": 1, "stock": 1,
                                   "is_active": 1, "variants": 1, "urun_karti_id": 1})
    async for p in cursor:
        total_products += 1
        variants = p.get("variants") or []
        matched = False
        total_stock = 0
        new_variants = []

        if variants:
            for v in variants:
                v = dict(v)
                bc = _norm(v.get("barcode"))
                if bc and bc in barcode_stock:
                    new_stock = barcode_stock[bc]
                    if int(v.get("stock") or 0) != new_stock:
                        updated_variants += 1
                    v["stock"] = new_stock
                    matched = True
                total_stock += int(v.get("stock") or 0)
                new_variants.append(v)
        else:
            bc = _norm(p.get("barcode"))
            if bc and bc in barcode_stock:
                total_stock = barcode_stock[bc]
                matched = True
            else:
                total_stock = int(p.get("stock") or 0)

        # Varyant barkodu eşleşmediyse kart id ile dene
        if not matched:
            cid = _norm(p.get("urun_karti_id"))
            if not cid:
                # slug son segmenti (name-ticimaxid)
                seg = (p.get("slug") or "").rsplit("-", 1)
                cid = seg[1] if len(seg) == 2 and seg[1].isdigit() else ""
            if cid and cid in card_stock:
                total_stock = card_stock[cid]
                matched = True

        if matched:
            matched_products += 1
            update = {"stock": total_stock}
            if variants:
                update["variants"] = new_variants
            if total_stock <= 0:
                update["is_active"] = False
                to_deactivate.append((p.get("name"), p.get("slug")))
            else:
                in_stock += 1
            if apply:
                await db.products.update_one({"id": p["id"]}, {"$set": update})
        else:
            unmatched.append((p.get("name"), p.get("slug")))

    print("\n===== ÖZET =====")
    print(f"Toplam ürün           : {total_products}")
    print(f"Eşleşen ürün          : {matched_products}")
    print(f"  - Stoklu (aktif)    : {in_stock}")
    print(f"  - Stoksuz→PASİF     : {len(to_deactivate)}")
    print(f"Güncellenen varyant   : {updated_variants}")
    print(f"Eşleşmeyen ürün       : {len(unmatched)} (DOKUNULMADI)")
    print(f"\nMod: {'UYGULANDI ✅' if apply else 'DRY-RUN (yazma yok)'}")
    if unmatched[:10]:
        print("\nEşleşmeyen örnekler:")
        for nm, sl in unmatched[:10]:
            print("   -", nm, "|", sl)
    if to_deactivate[:10]:
        print("\nPasife alınacak örnekler:")
        for nm, sl in to_deactivate[:10]:
            print("   -", nm, "|", sl)


if __name__ == "__main__":
    asyncio.run(main())
