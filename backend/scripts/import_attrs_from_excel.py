"""Kullanıcının yüklediği Ticimax Excel export'undan ürün-özellik eşlemesini
ürünlere yazar.

Excel formatı (Sheet "Ürün Teknik Detayları"):
    UrunKartID | StokKodu | UrunAdi | Tanim | Ozellik | Deger

Her satır bir özellik atamasıdır. UrunKartID, DB'deki `ticimax_card_id` ile eşleşir.
StokKodu ile fallback (variants[].stock_code veya barcode).

Çalıştırma:
    python /app/backend/scripts/import_attrs_from_excel.py /tmp/test.xlsx
"""
import asyncio
import os
import sys
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from routes.deps import db  # noqa: E402


# Bu 3 global default'ı KORU
GLOBAL_DEFAULTS = {
    "Cinsiyet": "Kadın / Kız",
    "Yaş Grubu": "Yetişkin",
    "Menşei": "TR",
}

# Excel'deki bu key'leri tüm marketplace formlarına yaz (Trendyol/HB/Temu)
WRITE_TO_ALL_MARKETPLACES = True


def _read_excel(path: str) -> dict:
    """
    Returns: {urun_kart_id: {ozellik_adi: deger}}
    """
    wb = load_workbook(path, data_only=True)
    sheet_name = "Ürün Teknik Detayları"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    result: dict[int, dict[str, str]] = {}
    stok_to_kart: dict[str, int] = {}
    header_skipped = False

    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        kart_id, stok_kodu, urun_adi, tanim, ozellik, deger = (row + (None,) * 6)[:6]
        if not kart_id:
            continue
        try:
            kart_id = int(kart_id)
        except Exception:
            continue
        if stok_kodu:
            stok_to_kart[str(stok_kodu).strip()] = kart_id
        if not ozellik or not deger:
            continue
        ozellik = str(ozellik).strip()
        deger = str(deger).strip()
        if not ozellik or not deger:
            continue
        result.setdefault(kart_id, {})[ozellik] = deger

    return result, stok_to_kart


async def main(path: str):
    print(f"📥 Excel okunuyor: {path}")
    excel_attrs, stok_to_kart = _read_excel(path)
    print(f"   {len(excel_attrs)} ürün karti, {sum(len(v) for v in excel_attrs.values())} özellik bulundu.")

    prods = await db.products.find(
        {"source": {"$in": ["xml_feed", "ticimax", "csv_xml_merge"]}},
        {"_id": 0, "id": 1, "name": 1, "ticimax_card_id": 1,
         "barcode": 1, "sku": 1, "stock_code": 1, "variants": 1,
         "attributes": 1, "hepsiburada_attributes": 1, "temu_attributes": 1},
    ).to_list(None)

    matched = 0
    unmatched = 0
    total_added = 0
    overrides = 0

    for p in prods:
        # Eşleme: önce ticimax_card_id, sonra stok_kodu fallback
        kart_id = p.get("ticimax_card_id")
        if not kart_id:
            for cand in (p.get("stock_code"), p.get("sku"), p.get("barcode")):
                if cand and str(cand).strip() in stok_to_kart:
                    kart_id = stok_to_kart[str(cand).strip()]
                    break
            if not kart_id:
                # variants'tan dene
                for v in (p.get("variants") or []):
                    sc = v.get("stock_code") or v.get("barcode")
                    if sc and str(sc).strip() in stok_to_kart:
                        kart_id = stok_to_kart[str(sc).strip()]
                        break

        excel_data = excel_attrs.get(int(kart_id)) if kart_id else None
        if not excel_data:
            unmatched += 1
            continue

        matched += 1

        attrs = p.get("attributes") or {}
        if isinstance(attrs, list):
            new_dict = {}
            for it in attrs:
                if isinstance(it, dict) and it.get("name"):
                    new_dict[it["name"]] = str(it.get("value", ""))
            attrs = new_dict
        hb = p.get("hepsiburada_attributes") or {}
        temu = p.get("temu_attributes") or {}

        # Excel'deki her özelliği yaz (Excel = SOURCE OF TRUTH; mevcut yanlış olabilir → override)
        for ozellik, deger in excel_data.items():
            for target in (attrs, hb, temu):
                if target.get(ozellik) != deger:
                    if target.get(ozellik):
                        overrides += 1
                    target[ozellik] = deger
                    total_added += 1

        # 3 global default'ı her zaman ekle (boş ise)
        for k, v in GLOBAL_DEFAULTS.items():
            for target in (attrs, hb, temu):
                if not target.get(k):
                    target[k] = v

        await db.products.update_one(
            {"id": p["id"]},
            {"$set": {
                "attributes": attrs,
                "hepsiburada_attributes": hb,
                "temu_attributes": temu,
                "ticimax_card_id": int(kart_id),
            }},
        )

    print(f"\n✅ {matched} ürün eşleşti, {unmatched} ürün eşleşmedi.")
    print(f"   {total_added} alan yazıldı/güncellendi ({overrides} mevcut değer üzerine yazıldı).")


if __name__ == "__main__":
    p = sys.argv[1] if len(sys.argv) > 1 else "/tmp/test.xlsx"
    asyncio.run(main(p))
