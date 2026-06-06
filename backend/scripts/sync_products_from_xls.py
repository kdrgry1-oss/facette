"""TicimaxExport → ürün senkronu + Excel'de OLMAYAN ürünleri SİLER.

Kural (kullanıcı talebi): Excel'deki ürünler dışında ne panelde, ne sitede,
ne de veritabanında veri kalmasın.

Eşleştirme önceliği (ürün Excel'de "var" sayılır):
  1. Herhangi bir varyant barkodu Excel barkodlarında
  2. Ürün barkodu Excel barkodlarında
  3. urun_karti_id (veya slug sonundaki sayı) Excel URUNKARTIID'lerinde
  4. stock_code Excel STOKKODU'larında

Eşleşen ürün:
  - Stok barkod bazında güncellenir (varyant barkodu → STOKADEDI), ürün.stock = toplam.
  - Toplam stok 0 ise is_active=False (pasif), >0 ise is_active=True (aktif).
Eşleşmeyen ürün:
  - --apply ile birlikte SİLİNİR (db.products'tan kaldırılır).

Kullanım:
    DRY-RUN:  python -m scripts.sync_products_from_xls /tmp/TicimaxExport1.xlsx
    UYGULA:   python -m scripts.sync_products_from_xls /tmp/TicimaxExport1.xlsx --apply
"""
import asyncio
import os
import sys

import openpyxl
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient

load_dotenv(os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env"))

COL_CARD = 0      # URUNKARTIID
COL_STOCKCODE = 2  # STOKKODU
COL_BARCODE = 4   # BARKOD
COL_STOCK = 33    # STOKADEDI


def _to_int(v) -> int:
    if v is None:
        return 0
    try:
        return max(0, int(float(str(v).replace(",", ".").strip() or 0)))
    except Exception:
        return 0


def _norm(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def parse_xls(path: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows)  # header
    barcode_stock, card_stock = {}, {}
    stockcodes, cards = set(), set()
    n = 0
    for r in rows:
        if not r or r[COL_CARD] is None:
            continue
        bc = _norm(r[COL_BARCODE])
        card = _norm(r[COL_CARD])
        sc = _norm(r[COL_STOCKCODE])
        stok = _to_int(r[COL_STOCK])
        if bc:
            barcode_stock[bc] = stok
        if card:
            card_stock[card] = card_stock.get(card, 0) + stok
            cards.add(card)
        if sc:
            stockcodes.add(sc)
        n += 1
    return barcode_stock, card_stock, stockcodes, cards, n


async def main():
    if len(sys.argv) < 2:
        print("Kullanım: python -m scripts.sync_products_from_xls <xls_path> [--apply]")
        return
    path = sys.argv[1]
    apply = "--apply" in sys.argv

    barcode_stock, card_stock, stockcodes, cards, n_rows = parse_xls(path)
    print(f"[xls] {n_rows} satır | {len(barcode_stock)} barkod | {len(cards)} ürün kartı | {len(stockcodes)} stok kodu")

    client = AsyncIOMotorClient(os.environ["MONGO_URL"])
    db = client[os.environ["DB_NAME"]]

    total = matched = updated_variants = in_stock = 0
    to_deactivate, to_delete = [], []

    cursor = db.products.find({}, {"id": 1, "name": 1, "slug": 1, "barcode": 1, "stock": 1,
                                   "is_active": 1, "variants": 1, "urun_karti_id": 1,
                                   "csv_card_id": 1, "stock_code": 1})
    async for p in cursor:
        total += 1
        variants = p.get("variants") or []
        matched_flag = False
        total_stock = 0
        new_variants = []

        if variants:
            for v in variants:
                v = dict(v)
                bc = _norm(v.get("barcode"))
                if bc and bc in barcode_stock:
                    new_stock = barcode_stock[bc]
                    if int(v.get("stock") or 0) != new_stock:
                        updated_variants += 1
                    v["stock"] = new_stock
                    matched_flag = True
                total_stock += int(v.get("stock") or 0)
                new_variants.append(v)
        else:
            bc = _norm(p.get("barcode"))
            if bc and bc in barcode_stock:
                total_stock = barcode_stock[bc]
                matched_flag = True
            else:
                total_stock = int(p.get("stock") or 0)

        if not matched_flag:
            cid = _norm(p.get("urun_karti_id")) or _norm(p.get("csv_card_id"))
            if not cid:
                seg = (p.get("slug") or "").rsplit("-", 1)
                cid = seg[1] if len(seg) == 2 and seg[1].isdigit() else ""
            if cid and cid in card_stock:
                total_stock = card_stock[cid]
                matched_flag = True

        if not matched_flag:
            sc = _norm(p.get("stock_code"))
            if sc and sc in stockcodes:
                matched_flag = True

        if matched_flag:
            matched += 1
            update = {"stock": total_stock, "is_active": total_stock > 0}
            if variants:
                update["variants"] = new_variants
            if total_stock <= 0:
                to_deactivate.append((p.get("name"), p.get("slug")))
            else:
                in_stock += 1
            if apply:
                await db.products.update_one({"id": p["id"]}, {"$set": update})
        else:
            to_delete.append((p["id"], p.get("name"), p.get("slug")))

    if apply and to_delete:
        ids = [d[0] for d in to_delete]
        res = await db.products.delete_many({"id": {"$in": ids}})
        print(f"\n🗑️  SİLİNDİ: {res.deleted_count} ürün")

    print("\n===== ÖZET =====")
    print(f"Toplam ürün (önce)    : {total}")
    print(f"Eşleşen (kalacak)     : {matched}")
    print(f"  - Stoklu (aktif)    : {in_stock}")
    print(f"  - Stoksuz→PASİF     : {len(to_deactivate)}")
    print(f"Güncellenen varyant   : {updated_variants}")
    print(f"Eşleşmeyen→SİLİNECEK   : {len(to_delete)}")
    print(f"Kalan ürün (sonra)    : {matched}")
    print(f"\nMod: {'UYGULANDI ✅' if apply else 'DRY-RUN (yazma yok)'}")
    if to_delete[:15]:
        print("\nSilinecek örnekler:")
        for _id, nm, sl in to_delete[:15]:
            print("   -", nm, "|", sl)


if __name__ == "__main__":
    asyncio.run(main())
