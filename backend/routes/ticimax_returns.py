"""
=============================================================================
ticimax_returns.py — Ticimax iade / kısmi iade SİPARİŞLERİ (İadeler sayfası)
=============================================================================
Ticimax sipariş import'u (integrations.py /ticimax/orders/import) iade ve kısmi
iade durumundaki siparişleri `orders` koleksiyonuna `platform="ticimax"` +
status ∈ {returned, partial_refunded, refunded, return_*} olarak yazıyor.

Bu modül o siparişleri İadeler sayfasında göstermek için listeler — ödeme tipi,
durum ve tüm detayla. Durum değiştirme mevcut `PUT /api/orders/{id}/status`
endpoint'i ile yapılır (bildirim de oradan gider), burada ayrıca tanımlanmaz.

Endpoint'ler (full path):
  GET  /api/admin/ticimax/return-orders     → iade siparişleri listesi + istatistik
=============================================================================
"""
from fastapi import APIRouter, Query, Depends, HTTPException
from typing import Optional
import re

from .deps import db, logger, require_admin, generate_id

router = APIRouter(prefix="/admin/ticimax", tags=["ticimax-returns"])

# İade sürecindeki tüm sipariş durumları (order_statuses.py "İade" grubu)
RETURN_STATUSES = [
    "return_requested", "return_approved", "return_rejected",
    "return_in_transit", "returned", "refunded", "partial_refunded",
]

# Ödeme tipi kodu → okunabilir Türkçe etiket
PAYMENT_LABELS = {
    "bank_transfer": "Havale / EFT",
    "credit_card": "Kredi Kartı",
    "cash_on_delivery": "Kapıda Ödeme",
    "cod_card": "Kapıda Kredi Kartı",
    "ticimax": "Diğer (Ticimax)",
}


def _payment_label(method: str, raw: str = "") -> str:
    lbl = PAYMENT_LABELS.get(method or "", "")
    if lbl and lbl != "Diğer (Ticimax)":
        return lbl
    # Bilinmeyen ama ham metin varsa onu göster
    return (raw or "").strip() or lbl or "Bilinmiyor"


@router.get("/return-orders")
async def list_ticimax_return_orders(
    status: Optional[str] = Query(None, description="Tek durum filtresi (örn. partial_refunded). Boş = tüm iade durumları"),
    payment: Optional[str] = Query(None, description="Ödeme tipi filtresi (bank_transfer/credit_card/cash_on_delivery)"),
    search: Optional[str] = Query(None, description="Sipariş no / müşteri adı / telefon araması"),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    current_user: dict = Depends(require_admin),
):
    """Ticimax kaynaklı iade / kısmi iade siparişlerini listeler.

    Veri kaynağı: orders koleksiyonu, platform/source = ticimax ve
    status ∈ İade grubu. Her satır ödeme tipi (kod + okunabilir etiket) ve
    durumla döner; durumu değiştirmek için frontend PUT /api/orders/{id}/status
    çağırır.
    """
    # "Web Sitesi" iadeleri = pazaryeri (Trendyol/Hepsiburada) DISI tum siparisler.
    # Eski hali yalnizca platform/source=ticimax idi -> yeni site siparisleri (platform=facette)
    # iade/iptal edilince hicbir sekmede gorunmuyordu. Pazaryeri disi her kaynak (facette + ticimax + bos) dahil.
    base_filter = {
        "platform": {"$nin": ["trendyol", "hepsiburada"]},
    }

    # Durum filtresi: tek durum verilmişse onu, yoksa tüm iade grubunu kullan
    if status:
        # Tek veya virgülle ayrılmış çoklu durum (örn. "refunded,partial_refunded" = 5. İade Ödemeleri hanesi)
        _st = [s.strip() for s in status.split(",") if s.strip()]
        base_filter["status"] = {"$in": _st} if len(_st) > 1 else (_st[0] if _st else {"$in": RETURN_STATUSES})
    else:
        base_filter["status"] = {"$in": RETURN_STATUSES}

    if payment:
        _pm = [x.strip() for x in str(payment).split(",") if x.strip()]
        base_filter["payment_method"] = {"$in": _pm} if len(_pm) > 1 else (_pm[0] if _pm else payment)

    if search:
        s = re.escape(search.strip())
        rx = {"$regex": s, "$options": "i"}
        base_filter["$and"] = [{
            "$or": [
                {"order_number": rx},
                {"order_code": rx},
                {"shipping_address.first_name": rx},
                {"shipping_address.last_name": rx},
                {"shipping_address.phone": rx},
                {"shipping_address.email": rx},
            ]
        }]

    total = await db.orders.count_documents(base_filter)

    proj = {
        "_id": 0, "id": 1, "order_number": 1, "order_code": 1, "ticimax_order_id": 1,
        "status": 1, "payment_method": 1, "payment_method_raw": 1, "payment_status": 1,
        "total": 1, "paid_amount": 1, "subtotal": 1, "shipping_cost": 1, "discount": 1,
        "coupon_code": 1, "notes": 1, "shipping_address": 1, "billing_address": 1,
        "customer_name": 1, "full_name": 1, "items": 1,
        "created_at": 1, "updated_at": 1, "channel_source": 1, "invoice_number": 1,
        "return_approved_at": 1, "refund_paid_at": 1, "return_request": 1,
        "cargo_tracking_number": 1, "cargo_tracking_url": 1, "cargo_provider_name": 1,
    }

    cursor = (
        db.orders.find(base_filter, proj)
        .sort("created_at", -1)
        .skip((page - 1) * limit)
        .limit(limit)
    )

    rows = []
    async for o in cursor:
        addr = o.get("shipping_address") or {}
        bill = o.get("billing_address") or {}
        name = (" ".join([addr.get("first_name") or "", addr.get("last_name") or ""]).strip()
                or addr.get("full_name") or addr.get("name")
                or o.get("customer_name") or o.get("full_name")
                or " ".join([bill.get("first_name") or "", bill.get("last_name") or ""]).strip()
                or bill.get("name") or "—")
        items = o.get("items") or []
        # Brüt / iskonto order-seviyesinde yoksa kalemlerden türet (unit_price=brüt, price=net).
        _calc_gross = sum(round(float(i.get("unit_price") or i.get("list_price") or i.get("original_price") or i.get("price") or 0), 2) * int(i.get("quantity") or 1) for i in items)
        _calc_net = sum(round(float(i.get("price") or 0), 2) * int(i.get("quantity") or 1) for i in items)
        _calc_idisc = sum(round(float(i.get("discount_amount") or i.get("discount") or 0), 2) * int(i.get("quantity") or 1) for i in items)
        _o_sub = float(o.get("subtotal") or 0)
        _o_disc = float(o.get("discount") or 0)
        _o_total = float(o.get("total") or 0)
        _r_total = _o_total if _o_total > 0 else round(_calc_net, 2)
        _r_subtotal = _o_sub if _o_sub > 0 else round(_calc_gross, 2)
        _r_discount = _o_disc if _o_disc > 0 else (round(_calc_idisc, 2) if _calc_idisc > 0 else round(max(0.0, _r_subtotal - _r_total), 2))
        rows.append({
            "id": o.get("id"),
            "order_number": o.get("order_number"),
            "ticimax_order_id": o.get("ticimax_order_id"),
            "customer_name": name,
            "phone": addr.get("phone") or "",
            "email": addr.get("email") or "",
            "address": addr.get("address") or "",
            "city": addr.get("city") or "",
            "district": addr.get("district") or "",
            "status": o.get("status"),
            "payment_method": o.get("payment_method") or "",
            "payment_label": _payment_label(o.get("payment_method") or "", o.get("payment_method_raw") or ""),
            "payment_method_raw": o.get("payment_method_raw") or "",
            "payment_status": o.get("payment_status") or "",
            "total": _r_total,
            "paid_amount": o.get("paid_amount") or 0,
            "subtotal": _r_subtotal,
            "shipping_cost": o.get("shipping_cost") or 0,
            "discount": _r_discount,
            "reason": (o.get("return_request") or {}).get("reason") or "",
            "coupon_code": o.get("coupon_code") or "",
            "notes": o.get("notes") or "",
            "item_count": sum(int(i.get("quantity") or 1) for i in items),
            "items": [
                {
                    "name": i.get("product_name") or i.get("name") or "",
                    "qty": i.get("quantity") or 1,
                    "size": i.get("size") or "",
                    "color": i.get("color") or "",
                    "barcode": i.get("barcode") or "",
                    "price": i.get("price") or 0,
                    "unit_price": i.get("unit_price") or i.get("list_price") or i.get("price") or 0,
                    "discount": i.get("discount_amount") or i.get("discount") or 0,
                }
                for i in items
            ],
            "invoice_number": o.get("invoice_number") or "",
            "created_at": o.get("created_at") or "",
            "updated_at": o.get("updated_at") or "",
            "return_approved_at": o.get("return_approved_at") or "",
            "refund_paid_at": o.get("refund_paid_at") or "",
            "cargo_tracking_number": o.get("cargo_tracking_number") or "",
            "cargo_tracking_url": o.get("cargo_tracking_url") or "",
            "cargo_provider_name": o.get("cargo_provider_name") or "",
        })

    # İlgili customer_returns köprü kayıtları (kargo barkodu / iade kodu / reship / ödeme zamanı) — tek sorgu
    _oids = [r["id"] for r in rows if r.get("id")]
    _cr_map = {}
    if _oids:
        async for cr in db.customer_returns.find(
            {"order_id": {"$in": _oids}},
            {"_id": 0, "order_id": 1, "return_code": 1, "barcode_url": 1, "cargo_provider_name": 1,
             "iade_no": 1, "gonderi_no": 1, "mng_ref": 1, "contract_no": 1,
             "reship_code": 1, "reshipped_at": 1, "refund_payment": 1, "reason": 1},
        ):
            _cr_map[cr.get("order_id")] = cr
    for r in rows:
        cr = _cr_map.get(r["id"]) or {}
        if cr.get("reason"):
            r["reason"] = cr["reason"]
        r["return_code"] = cr.get("return_code") or ""
        r["iade_no"] = cr.get("iade_no") or cr.get("mng_ref") or ""
        r["gonderi_no"] = cr.get("gonderi_no") or ""
        r["contract_no"] = cr.get("contract_no") or "490059279"
        r["return_barcode_url"] = cr.get("barcode_url") or ""
        r["return_cargo_provider"] = cr.get("cargo_provider_name") or r.get("cargo_provider_name") or ""
        r["reship_code"] = cr.get("reship_code") or ""
        r["reshipped_at"] = cr.get("reshipped_at") or ""
        # İade ödeme tarihi: sipariş damgası > köprü ödeme zamanı > (refunded ise) updated_at
        if not r.get("refund_paid_at"):
            _rp = (cr.get("refund_payment") or {}).get("at")
            r["refund_paid_at"] = _rp or (r.get("updated_at") if r.get("status") in ("refunded", "partial_refunded") else "")

    # İstatistik: tüm iade grubunda durum + ödeme dağılımı (mevcut filtreden bağımsız,
    # pazaryeri DISI tum site siparisleri) — sekmedeki rozetler için
    stat_filter = {"platform": {"$nin": ["trendyol", "hepsiburada"]},
                   "status": {"$in": RETURN_STATUSES}}
    status_counts = {}
    payment_counts = {}
    try:
        async for grp in db.orders.aggregate([
            {"$match": stat_filter},
            {"$group": {"_id": "$status", "n": {"$sum": 1}}},
        ]):
            status_counts[grp["_id"]] = grp["n"]
        async for grp in db.orders.aggregate([
            {"$match": stat_filter},
            {"$group": {"_id": "$payment_method", "n": {"$sum": 1}}},
        ]):
            payment_counts[grp["_id"] or "bilinmiyor"] = grp["n"]
    except Exception as e:
        logger.warning(f"[ticimax-returns] stats hatası: {e}")

    return {
        "success": True,
        "orders": rows,
        "total": total,
        "page": page,
        "limit": limit,
        "status_counts": status_counts,
        "payment_counts": payment_counts,
        "total_returns": sum(status_counts.values()),
    }


@router.post("/orders/refresh-dates")
async def refresh_order_dates(
    page: int = Query(1, ge=1),
    per_pages: int = Query(5, ge=1, le=15),
    current_user: dict = Depends(require_admin),
):
    """Siparişlerin created_at'ini Ticimax'taki gerçek SiparisTarihi'ne çeker.

    Zaman aşımına takılmamak için her çağrıda `per_pages` sayfa (100'er kayıt)
    işler. Frontend, has_more=False dönene kadar page'i artırarak döngüyle çağırır.
    """
    import asyncio
    from datetime import datetime, timezone
    try:
        from ticimax_client import get_orders as tc_get_orders
    except Exception as e:
        return {"success": False, "message": f"Ticimax client yüklenemedi: {e}"}

    fixed = 0
    scanned = 0
    reached_end = False
    for i in range(per_pages):
        pno = page + i
        try:
            batch = tc_get_orders(page=pno, page_size=100,
                                  exclude_marketplace=False, only_with_phone=False)
        except Exception as e:
            logger.warning(f"[refresh-dates] sayfa {pno} çekilemedi: {e}")
            batch = []
        if not batch:
            reached_end = True
            break
        for o in batch:
            if not o:
                continue
            scanned += 1
            tid = o.get("SiparisID") or o.get("ID")
            d = o.get("SiparisTarihi") or o.get("SiparisTarih") or o.get("Tarih")
            if not tid or not d:
                continue
            try:
                iso = d.isoformat() if hasattr(d, "isoformat") else str(d)
                tid_int = int(tid)
            except Exception:
                continue
            try:
                r = await db.orders.update_one(
                    {"ticimax_order_id": tid_int},
                    {"$set": {"created_at": iso,
                              "updated_at": datetime.now(timezone.utc).isoformat()}},
                )
                if r.modified_count:
                    fixed += 1
            except Exception as ie:
                logger.warning(f"[refresh-dates] update hata {tid_int}: {ie}")
        await asyncio.sleep(0.4)

    return {
        "success": True,
        "fixed": fixed,
        "scanned": scanned,
        "next_page": page + per_pages,
        "has_more": (not reached_end),
    }


# ============================================================================
# EXPORT — İade siparişlerini Excel (.xlsx) indir (görseldeki kolon düzeni)
# ============================================================================
@router.get("/return-orders/export")
async def export_ticimax_return_orders(
    status: Optional[str] = Query(None),
    payment: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    current_user: dict = Depends(require_admin),
):
    """İade siparişlerini listelemeyle AYNI filtrelerle Excel'e aktarır.
    Kolonlar: Ürün Adı | Tutar | Sipariş Tarihi | Sipariş ID | Sipariş No |
    Ad Soyad | Durum | Kaynak | Ödeme Tipi."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    try:
        from order_statuses import ORDER_STATUS_CATALOG
        status_label = {s["key"]: (s.get("label") or s["key"]) for s in ORDER_STATUS_CATALOG}
    except Exception:
        status_label = {}

    base_filter = {"$or": [{"platform": "ticimax"}, {"source": "ticimax"}]}
    if status:
        _st = [s.strip() for s in status.split(",") if s.strip()]
        base_filter["status"] = {"$in": _st} if len(_st) > 1 else (_st[0] if _st else {"$in": RETURN_STATUSES})
    else:
        base_filter["status"] = {"$in": RETURN_STATUSES}
    if payment:
        _pm = [x.strip() for x in str(payment).split(",") if x.strip()]
        base_filter["payment_method"] = {"$in": _pm} if len(_pm) > 1 else (_pm[0] if _pm else payment)
    if search:
        s = re.escape(search.strip())
        rx = {"$regex": s, "$options": "i"}
        base_filter["$and"] = [{
            "$or": [
                {"order_number": rx}, {"order_code": rx},
                {"shipping_address.first_name": rx}, {"shipping_address.last_name": rx},
                {"shipping_address.phone": rx}, {"shipping_address.email": rx},
            ]
        }]

    proj = {
        "_id": 0, "id": 1, "order_number": 1, "ticimax_order_id": 1, "status": 1,
        "payment_method": 1, "payment_method_raw": 1, "total": 1,
        "shipping_address": 1, "billing_address": 1, "customer_name": 1, "full_name": 1,
        "items": 1, "created_at": 1, "channel_source": 1,
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "İade Siparişleri"
    headers = ["Ürün Adı", "Tutar", "Sipariş Tarihi", "Sipariş ID", "Sipariş No",
               "Ad Soyad", "Durum", "Kaynak", "Ödeme Tipi"]
    ws.append(headers)
    hfill = PatternFill("solid", fgColor="FCE4B6")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")

    cursor = db.orders.find(base_filter, proj).sort("created_at", -1)
    async for o in cursor:
        addr = o.get("shipping_address") or {}
        bill = o.get("billing_address") or {}
        name = (" ".join([addr.get("first_name") or "", addr.get("last_name") or ""]).strip()
                or addr.get("full_name") or addr.get("name")
                or o.get("customer_name") or o.get("full_name")
                or " ".join([bill.get("first_name") or "", bill.get("last_name") or ""]).strip()
                or bill.get("name") or "")
        items = o.get("items") or []
        urun = ", ".join([
            (i.get("product_name") or i.get("name") or "").strip()
            for i in items if (i.get("product_name") or i.get("name"))
        ])
        st = o.get("status") or ""
        ws.append([
            urun,
            float(o.get("total") or 0),
            str(o.get("created_at") or "")[:10],
            o.get("id") or "",
            o.get("order_number") or "",
            name,
            status_label.get(st, st),
            o.get("channel_source") or "Ticimax",
            _payment_label(o.get("payment_method") or "", o.get("payment_method_raw") or ""),
        ])

    widths = [42, 12, 14, 16, 14, 24, 18, 12, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=iade-siparisleri.xlsx"},
    )


# ============================================================================
# BRIDGE — Ticimax iade siparişini zengin iade akışına (customer_returns) bağlar
# ============================================================================
@router.post("/returns/{order_id}/open")
async def open_ticimax_return(order_id: str, current_user: dict = Depends(require_admin)):
    """Ticimax iade siparişinden, zengin iade akışı (onayla/reddet/gider/öde) için bir
    `customer_returns` köprü kaydı üretir. İDEMPOTENT: zaten varsa mevcut return_id'yi döndürür.
    PARA/İŞLEM YAPMAZ — sadece köprü kaydını oluşturur. Sonraki adımlar mevcut
    /api/orders/returns/{return_id}/{refund-preview,approve,reject,gider-pusulasi,refund-pay}
    endpoint'leriyle yürür (RBAC + bildirim oralarda)."""
    from datetime import datetime, timezone
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Sipariş bulunamadı")

    # İdempotent: bu siparişe ait köprü kaydı zaten varsa onu döndür
    existing = await db.customer_returns.find_one(
        {"order_id": order_id}, {"_id": 0, "id": 1, "status": 1})
    if existing:
        return {"success": True, "return_id": existing.get("id"),
                "status": existing.get("status"), "created": False}

    # Sipariş kalemlerini customer_returns şemasına eşle
    src = order.get("items") or []
    items = [{
        "name": it.get("product_name") or it.get("name") or "Ürün",
        "size": it.get("size", "") or "",
        "color": it.get("color", "") or "",
        "quantity": int(it.get("quantity", 1) or 1),
        "price": float(it.get("price") or it.get("unit_price") or 0),
        "unit_price": float(it.get("unit_price") or it.get("price") or 0),
        "product_id": it.get("barcode") or it.get("product_id") or it.get("sku") or "",
    } for it in src]

    # Sipariş durumu → customer_returns durumu
    _map = {
        "return_requested": "created", "return_approved": "approved",
        "return_in_transit": "in_transit", "returned": "received",
        "refunded": "refunded", "partial_refunded": "refunded",
        "return_rejected": "rejected",
    }
    cr_status = _map.get(order.get("status"), "created")

    rid = generate_id()
    rr = order.get("return_request") or {}
    rec = {
        "id": rid, "order_id": order_id, "order_number": order.get("order_number", ""),
        "user_id": order.get("user_id"), "items": items, "reason": "",
        "return_code": rr.get("return_code", "") or "", "mng_ok": False,
        "cargo_provider_name": order.get("cargo_provider_name", "") or "",
        "status": cr_status, "source": "ticimax_bridge",
        "created_at": order.get("created_at") or datetime.now(timezone.utc).isoformat(),
    }
    await db.customer_returns.insert_one({**rec})
    await db.orders.update_one({"id": order_id},
                               {"$set": {"return_request.return_id": rid}})
    return {"success": True, "return_id": rid, "status": cr_status, "created": True}
