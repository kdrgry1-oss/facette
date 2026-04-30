"""
=============================================================================
production_plan.py — İmalat Planı (18 Sütunlu Tablo) — FAZ 7
=============================================================================

Kullanıcının istediği tabloya birebir karşılık gelen şemalı CRUD.
Manufacturing modülünden ayrı tutuldu (o modül stage-based süreç takibi,
bu modül ise spreadsheet tarzı "üretim planı" listesi).

Sütunlar:
  1. seq_no              (auto)
  2. manufacturer_id     (vendors koleksiyonundan seçilir)
  3. collection          (product.collection'dan default)
  4. model_no            (manuel)
  5. product_description (product.name'den default)
  6. order_qty           (manuel)
  7. price               (product.purchase_price'dan default)
  8. payment_date        (ISO date)
  9. planned_delivery    (payment_date + 21 gün auto)
 10. color               (product.color'dan default)
 11. ok_date             (ISO date)
 12. sample_ok_date      (ISO date)
 13. cut_qty             (manuel)
 14. wash_barcode_date   (ISO date)
 15. inline_qc           ({date, result: "pass"|"fail", image_url})
 16. final_qc            ({date, result, image_url})
 17. actual_delivery     (ISO date, delay_days otomatik)
 18. delivered_qty       (manuel, +%/-% otomatik)
=============================================================================
"""
from datetime import datetime, timezone, timedelta
from io import BytesIO
from typing import Optional, List, Dict, Any
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from .deps import db, require_admin, generate_id

router = APIRouter(prefix="/production-plan", tags=["production-plan"])


class QCRecord(BaseModel):
    date: Optional[str] = None
    result: Optional[str] = None  # "pass" | "fail"
    image_url: Optional[str] = None
    note: Optional[str] = ""


class PlanRowReq(BaseModel):
    product_id: Optional[str] = None
    manufacturer_id: Optional[str] = None
    manufacturer_name: Optional[str] = ""
    collection: Optional[str] = ""
    model_no: Optional[str] = ""
    product_description: Optional[str] = ""
    order_qty: int = 0
    price: float = 0
    payment_date: Optional[str] = None      # ISO date "YYYY-MM-DD"
    planned_delivery: Optional[str] = None  # hesaplanır
    color: Optional[str] = ""
    ok_date: Optional[str] = None
    sample_ok_date: Optional[str] = None
    cut_qty: int = 0
    wash_barcode_date: Optional[str] = None
    inline_qc: Optional[QCRecord] = None
    final_qc: Optional[QCRecord] = None
    actual_delivery: Optional[str] = None
    delivered_qty: int = 0
    notes: Optional[str] = ""


def _add_days_iso(date_str: Optional[str], days: int) -> Optional[str]:
    if not date_str:
        return None
    try:
        d = datetime.fromisoformat(date_str[:10])
        return (d + timedelta(days=days)).date().isoformat()
    except Exception:
        return None


def _compute_derived(doc: Dict[str, Any]) -> Dict[str, Any]:
    # 9 — planned_delivery = payment_date + 21
    if doc.get("payment_date") and not doc.get("planned_delivery"):
        doc["planned_delivery"] = _add_days_iso(doc["payment_date"], 21)

    # 17 — delay_days
    delay_days = None
    if doc.get("planned_delivery") and doc.get("actual_delivery"):
        try:
            pd = datetime.fromisoformat(doc["planned_delivery"][:10])
            ad = datetime.fromisoformat(doc["actual_delivery"][:10])
            delay_days = (ad - pd).days
        except Exception:
            delay_days = None
    doc["delay_days"] = delay_days

    # 18 — qty_diff_pct (delivered_qty > 0 olduğunda anlamlı)
    pct = None
    try:
        oq = float(doc.get("order_qty") or 0)
        dq = float(doc.get("delivered_qty") or 0)
        if oq > 0 and dq > 0:
            pct = round(((dq - oq) / oq) * 100, 1)
    except Exception:
        pct = None
    doc["qty_diff_pct"] = pct
    return doc


@router.get("")
async def list_plan(
    search: Optional[str] = None,
    manufacturer_id: Optional[str] = None,
    collection: Optional[str] = None,
    limit: int = Query(500, le=2000),
    current_user: dict = Depends(require_admin),
):
    q: Dict[str, Any] = {}
    if manufacturer_id:
        q["manufacturer_id"] = manufacturer_id
    if collection:
        q["collection"] = collection
    if search:
        q["$or"] = [
            {"model_no": {"$regex": search, "$options": "i"}},
            {"product_description": {"$regex": search, "$options": "i"}},
            {"manufacturer_name": {"$regex": search, "$options": "i"}},
        ]
    rows = await db.production_plan.find(q, {"_id": 0}).sort("seq_no", 1).to_list(limit)
    return {"items": rows, "total": len(rows)}


@router.post("")
async def create_plan_row(req: PlanRowReq, current_user: dict = Depends(require_admin)):
    # Ürün kartından otomatik doldur (verilmemiş alanlar için)
    product = None
    if req.product_id:
        product = await db.products.find_one({"id": req.product_id}, {"_id": 0})

    data = req.model_dump()
    if product:
        if not data.get("collection"):
            data["collection"] = product.get("collection", "")
        if not data.get("product_description"):
            data["product_description"] = product.get("name", "")
        if not data.get("price"):
            data["price"] = float(product.get("purchase_price") or 0)
        if not data.get("color"):
            data["color"] = product.get("color", "")

    # Üretici ismi
    if req.manufacturer_id and not data.get("manufacturer_name"):
        vendor = await db.vendors.find_one({"id": req.manufacturer_id}, {"_id": 0, "name": 1})
        if vendor:
            data["manufacturer_name"] = vendor.get("name", "")

    # Seq no — mevcut en yüksek seq_no + 1 (silme sonrası boşluk kapatmaz, uniqueness korunur)
    last = await db.production_plan.find_one({}, {"_id": 0, "seq_no": 1}, sort=[("seq_no", -1)])
    seq_no = ((last or {}).get("seq_no") or 0) + 1

    doc = {
        "id": generate_id(),
        "seq_no": seq_no,
        **data,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.get("email", ""),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    doc = _compute_derived(doc)

    await db.production_plan.insert_one(doc)
    doc.pop("_id", None)
    return {"success": True, "row": doc}


@router.put("/{row_id}")
async def update_plan_row(row_id: str, payload: dict, current_user: dict = Depends(require_admin)):
    existing = await db.production_plan.find_one({"id": row_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")

    # Güncellenebilir alanlar
    allowed = {
        "manufacturer_id", "manufacturer_name", "collection", "model_no",
        "product_description", "order_qty", "price", "payment_date",
        "planned_delivery", "color", "ok_date", "sample_ok_date", "cut_qty",
        "wash_barcode_date", "inline_qc", "final_qc", "actual_delivery",
        "delivered_qty", "notes", "product_id",
    }
    update = {k: v for k, v in (payload or {}).items() if k in allowed}
    update["updated_at"] = datetime.now(timezone.utc).isoformat()

    # payment_date değiştiyse ve UI planned_delivery göndermediyse yeniden hesap
    if "payment_date" in update and "planned_delivery" not in update:
        update["planned_delivery"] = _add_days_iso(update["payment_date"], 21)

    # Birleştirilmiş doc üzerinden derived hesapla
    merged = {**existing, **update}
    merged = _compute_derived(merged)
    update["delay_days"] = merged.get("delay_days")
    update["qty_diff_pct"] = merged.get("qty_diff_pct")
    update["planned_delivery"] = merged.get("planned_delivery")

    await db.production_plan.update_one({"id": row_id}, {"$set": update})
    return {"success": True, "delay_days": update["delay_days"], "qty_diff_pct": update["qty_diff_pct"], "planned_delivery": update["planned_delivery"]}


@router.delete("/{row_id}")
async def delete_plan_row(row_id: str, current_user: dict = Depends(require_admin)):
    res = await db.production_plan.delete_one({"id": row_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    return {"success": True}


@router.get("/collections")
async def list_collections(current_user: dict = Depends(require_admin)):
    """Distinct koleksiyon listesi — ürün ve plan kayıtlarından."""
    prod_cols = await db.products.distinct("collection")
    plan_cols = await db.production_plan.distinct("collection")
    merged = sorted({c for c in (prod_cols + plan_cols) if c})
    return {"collections": merged}


# =============================================================================
# Excel Export / Import (FAZ 7 potansiyel iyileştirme)
# =============================================================================

_EXCEL_COLUMNS = [
    ("seq_no", "1. Sıra No"),
    ("manufacturer_name", "2. Üretici"),
    ("collection", "3. Koleksiyon"),
    ("model_no", "4. Model No"),
    ("product_description", "5. Ürün Açıklaması"),
    ("order_qty", "6. Sipariş Adedi"),
    ("price", "7. Fiyat"),
    ("payment_date", "8. Ödeme Tarihi"),
    ("planned_delivery", "9. Planlanan Teslimat"),
    ("color", "10. Renk"),
    ("ok_date", "11. Okey Tarihi"),
    ("sample_ok_date", "12. Numune Okey"),
    ("cut_qty", "13. Kesim Adedi"),
    ("wash_barcode_date", "14. Yıkama+Barkod"),
    ("inline_qc_date", "15a. İnline Tarih"),
    ("inline_qc_result", "15b. İnline Sonuç"),
    ("final_qc_date", "16a. Final Tarih"),
    ("final_qc_result", "16b. Final Sonuç"),
    ("actual_delivery", "17. Gerçek Teslim"),
    ("delay_days", "17b. Gecikme (gün)"),
    ("delivered_qty", "18. Teslim Adedi"),
    ("qty_diff_pct", "18b. +/-%"),
    ("notes", "Notlar"),
]


def _row_to_excel(row: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(row)
    out["inline_qc_date"] = (row.get("inline_qc") or {}).get("date") or ""
    out["inline_qc_result"] = (row.get("inline_qc") or {}).get("result") or ""
    out["final_qc_date"] = (row.get("final_qc") or {}).get("date") or ""
    out["final_qc_result"] = (row.get("final_qc") or {}).get("result") or ""
    return out


@router.get("/export")
async def export_excel(current_user: dict = Depends(require_admin)):
    """İmalat planını Excel (.xlsx) olarak indir."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = await db.production_plan.find({}, {"_id": 0}).sort("seq_no", 1).to_list(length=5000)

    wb = Workbook()
    ws = wb.active
    ws.title = "İmalat Planı"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for col_idx, (_, label) in enumerate(_EXCEL_COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    for r_idx, row in enumerate(rows, start=2):
        rd = _row_to_excel(row)
        for c_idx, (key, _) in enumerate(_EXCEL_COLUMNS, start=1):
            v = rd.get(key)
            if v is None:
                v = ""
            ws.cell(row=r_idx, column=c_idx, value=v)

    # Sütun genişlikleri
    for i in range(1, len(_EXCEL_COLUMNS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 22

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"imalat-plani-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/import")
async def import_excel(
    file: UploadFile = File(...),
    update_existing: bool = Query(True, description="model_no eşleşen kayıtları güncelle (false → hep yeni satır)"),
    current_user: dict = Depends(require_admin),
):
    """Excel'den toplu satır ekle/güncelle.
    Başlık sırası export ile aynıdır. model_no alanı key olarak kullanılır."""
    from openpyxl import load_workbook

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Sadece .xlsx dosya kabul edilir")
    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel okunamadı: {e}")
    ws = wb.active

    # Başlık satırı → key mapping
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        label = (cell.value or "").strip() if cell.value else ""
        for key, lbl in _EXCEL_COLUMNS:
            if label == lbl:
                headers[col_idx] = key
                break

    if not headers:
        raise HTTPException(status_code=400, detail="Başlık satırı tanınamadı. Önce export ederek şablonu alın.")

    created = 0
    updated = 0
    errors: List[str] = []

    for r_idx in range(2, ws.max_row + 1):
        data: Dict[str, Any] = {}
        for col_idx, key in headers.items():
            val = ws.cell(row=r_idx, column=col_idx).value
            if isinstance(val, datetime):
                val = val.date().isoformat()
            data[key] = val

        model_no = (data.get("model_no") or "").strip() if isinstance(data.get("model_no"), str) else data.get("model_no")
        if not model_no and not any(data.get(k) for k in ("manufacturer_name", "product_description")):
            continue  # boş satır

        # QC inline/final birleşim
        inline_qc = {"date": data.pop("inline_qc_date", "") or "", "result": data.pop("inline_qc_result", "") or ""}
        final_qc = {"date": data.pop("final_qc_date", "") or "", "result": data.pop("final_qc_result", "") or ""}
        data["inline_qc"] = inline_qc if (inline_qc["date"] or inline_qc["result"]) else None
        data["final_qc"] = final_qc if (final_qc["date"] or final_qc["result"]) else None
        # Auto-calc alanlarını dokunma
        data.pop("delay_days", None)
        data.pop("qty_diff_pct", None)

        # Numerik dönüşüm
        for nk in ("order_qty", "cut_qty", "delivered_qty"):
            try:
                data[nk] = int(data.get(nk) or 0)
            except (TypeError, ValueError):
                data[nk] = 0
        try:
            data["price"] = float(data.get("price") or 0)
        except (TypeError, ValueError):
            data["price"] = 0

        try:
            existing = None
            if update_existing and model_no:
                existing = await db.production_plan.find_one({"model_no": model_no}, {"_id": 0, "id": 1})
            if existing:
                merged = {**existing, **{k: v for k, v in data.items() if v not in (None, "")}}
                merged = _compute_derived(merged)
                merged["updated_at"] = datetime.now(timezone.utc).isoformat()
                await db.production_plan.update_one({"id": existing["id"]}, {"$set": merged})
                updated += 1
            else:
                last = await db.production_plan.find_one({}, {"_id": 0, "seq_no": 1}, sort=[("seq_no", -1)])
                seq_no = ((last or {}).get("seq_no") or 0) + 1
                doc = {
                    "id": generate_id(),
                    "seq_no": seq_no,
                    **data,
                    "model_no": model_no or "",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": current_user.get("email", ""),
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }
                doc = _compute_derived(doc)
                await db.production_plan.insert_one(doc)
                created += 1
        except Exception as e:
            errors.append(f"Satır {r_idx}: {e}")

    return {"success": True, "created": created, "updated": updated, "errors": errors[:20]}

