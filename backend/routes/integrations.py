"""
Integration routes - Trendyol, MNG Kargo, GIB, Netgsm, Ticimax, XML Feed
Not: Iyzico kısmı `integrations_iyzico.py` modülüne taşındı (2026-04-23).
"""
from fastapi import APIRouter, HTTPException, Query, Depends, Response, BackgroundTasks, Request, Body, UploadFile, File
from typing import List, Optional
from datetime import datetime, timezone, timedelta
import os
import base64
import uuid
import re
import xml.etree.ElementTree as ET
import httpx
import hashlib

from .deps import db, logger, get_current_user, require_admin, generate_id, generate_short_id

router = APIRouter(tags=["Integrations"])

# Iyzico kısmı ayrı modüle taşındı — alt router olarak include edilecek (server.py)
from .integrations_iyzico import router as iyzico_router  # noqa
from .integrations_iyzico import (  # eski import yollarını korumak için re-export
    IYZICO_MODE, IYZICO_API_KEY, IYZICO_SECRET_KEY, IYZICO_BASE_URL,
    is_iyzico_configured, _iyzico_auth_header,
)

# ==================== TRENDYOL ====================

async def get_trendyol_config():
    """Get Trendyol configuration from DB or env.
    NOTE: default_markup için Ana Ayarlar > trendyol_markup ÖNCELİKLİ — kullanıcı UI'da
    en son nereye girerse oradan okunur."""
    settings = await db.settings.find_one({"id": "trendyol"})
    # Main settings'ten markup override
    main_settings = await db.settings.find_one({"id": "main"}) or {}
    main_markup = main_settings.get("trendyol_markup")
    try:
        main_markup_f = float(main_markup) if (main_markup is not None and main_markup != "") else None
    except Exception:
        main_markup_f = None

    if settings:
        mode = settings.get("mode", "sandbox")
        local_markup = settings.get("default_markup", 0) or 0
        effective_markup = main_markup_f if main_markup_f is not None else local_markup
        return {
            "api_key": settings.get("api_key", ""),
            "api_secret": settings.get("api_secret", ""),
            "supplier_id": settings.get("supplier_id", ""),
            "is_active": settings.get("is_active", False),
            "mode": mode,
            "default_markup": effective_markup,
            "base_url": 'https://api.trendyol.com' if mode == 'live' else 'https://stageapigw.trendyol.com'
        }
    
    # Fallback to env
    mode = os.environ.get('TRENDYOL_MODE', 'sandbox')
    return {
        "api_key": os.environ.get('TRENDYOL_API_KEY', ''),
        "api_secret": os.environ.get('TRENDYOL_API_SECRET', ''),
        "supplier_id": os.environ.get('TRENDYOL_SUPPLIER_ID', ''),
        "is_active": bool(os.environ.get('TRENDYOL_API_KEY')),
        "mode": mode,
        "base_url": 'https://api.trendyol.com' if mode == 'live' else 'https://stageapigw.trendyol.com'
    }

async def get_trendyol_headers():
    config = await get_trendyol_config()
    if not config["api_key"] or not config["api_secret"]:
        return None
    credentials = f'{config["api_key"]}:{config["api_secret"]}'
    encoded = base64.b64encode(credentials.encode()).decode()
    return {
        "Authorization": f"Basic {encoded}",
        "User-Agent": f'{config["supplier_id"]} - FacetteIntegration',
        "Content-Type": "application/json"
    }

def calculate_trendyol_price(base_price: float, product_data: dict, trendyol_config: dict) -> float:
    """Calculate price with markup logic"""
    # SSOT: her zaman global default_markup uygulanir; urun bazli override yok sayilir
    # (kullanici talebi: belirlenen oran disinda fiyat guncellenmez)
    markup = float(trendyol_config.get("default_markup", 0) or 0)

    final_price = base_price * (1 + markup / 100)
    return round(final_price, 2)

@router.get("/trendyol/settings")
async def get_trendyol_settings(current_user: dict = Depends(require_admin)):
    """Get Trendyol settings"""
    config = await get_trendyol_config()

    # Single source of truth: Ana Ayarlar sayfasındaki `trendyol_markup`
    # main settings'ten yazıldıysa ÖNCELİKLİ olarak onu kullan; aksi halde
    # Trendyol Integration kartındaki default_markup'a düş.
    main_settings = await db.settings.find_one({"id": "main"})
    main_markup = (main_settings or {}).get("trendyol_markup")
    if main_markup is not None and main_markup != "":
        try:
            default_markup = float(main_markup)
        except Exception:
            default_markup = config.get("default_markup", 0) or 0
    else:
        default_markup = config.get("default_markup", 0) or 0

    # Mask secrets
    return {
        "supplier_id": config.get("supplier_id", ""),
        "api_key": config.get("api_key", ""),
        "api_secret": "********" if config.get("api_secret") else "",
        "mode": config.get("mode", "sandbox"),
        "is_active": config.get("is_active", False),
        "default_markup": default_markup
    }

@router.post("/trendyol/settings")
async def save_trendyol_settings(
    settings: dict,
    current_user: dict = Depends(require_admin)
):
    """Save Trendyol settings"""
    from datetime import datetime, timezone

    # Required alan validasyonu — is_active=True ise supplier_id/api_key/api_secret zorunlu
    if settings.get("is_active"):
        existing = await db.settings.find_one({"id": "trendyol"}, {"_id": 0}) or {}
        supplier_id = settings.get("supplier_id") or existing.get("supplier_id")
        api_key = settings.get("api_key") or existing.get("api_key")
        api_secret = settings.get("api_secret")
        if api_secret in (None, "", "********"):
            api_secret = existing.get("api_secret")
        missing = [k for k, v in {"supplier_id": supplier_id, "api_key": api_key, "api_secret": api_secret}.items() if not v]
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"Trendyol aktifleştirmek için zorunlu alanlar eksik: {', '.join(missing)}"
            )

    update_data = {
        "supplier_id": settings.get("supplier_id", ""),
        "api_key": settings.get("api_key", ""),
        "mode": settings.get("mode", "sandbox"),
        "is_active": settings.get("is_active", False),
        "default_markup": settings.get("default_markup", 0),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }

    if settings.get("api_secret") and settings.get("api_secret") != "********":
        update_data["api_secret"] = settings.get("api_secret")

    await db.settings.update_one(
        {"id": "trendyol"},
        {"$set": update_data},
        upsert=True
    )

    # SSOT: trendyol_markup'i main settings'e de yansıt (UI'da ana ayarlar sayfası bu key'i okur)
    try:
        await db.settings.update_one(
            {"id": "main"},
            {"$set": {"trendyol_markup": update_data["default_markup"]}},
            upsert=True,
        )
    except Exception:
        pass

    return {"success": True, "message": "Trendyol ayarları kaydedildi"}


@router.post("/trendyol/test-connection")
async def test_trendyol_connection(current_user: dict = Depends(require_admin)):
    """Trendyol gerçek bağlantı testi — brands endpoint'i üzerinden."""
    cfg = await get_trendyol_config()
    if not cfg.get("api_key") or not cfg.get("api_secret") or not cfg.get("supplier_id"):
        return {"success": False, "message": "Trendyol API bilgileri eksik (api_key / api_secret / supplier_id zorunlu)"}
    try:
        from trendyol_client import TrendyolClient
        client = TrendyolClient(
            supplier_id=cfg["supplier_id"],
            api_key=cfg["api_key"],
            api_secret=cfg["api_secret"],
            mode=cfg["mode"],
        )
        # Hafif bir probe — ilk 1 marka yeter
        data = await client.get_brands(size=1, page=0)
        if isinstance(data, dict) and ("brands" in data or "content" in data or "totalElements" in data):
            return {"success": True, "message": "Trendyol bağlantısı başarılı", "mode": cfg["mode"]}
        return {"success": False, "message": f"Beklenmeyen yanıt: {str(data)[:200]}"}
    except Exception as e:
        msg = str(e)[:300]
        status_hint = "401/403 kimlik hatası" if ("401" in msg or "403" in msg or "Unauthorized" in msg) else "HTTP hatası"
        return {"success": False, "message": f"{status_hint}: {msg}"}

@router.get("/trendyol/status")
async def get_trendyol_status():
    """Get Trendyol integration status"""
    config = await get_trendyol_config()
    return {
        "configured": config["is_active"],
        "mode": config["mode"],
        "supplier_id": config["supplier_id"] if config["is_active"] else None
    }

@router.get("/trendyol/debug")
async def debug_trendyol_orders():
    config = await get_trendyol_config()
    import sys
    import os
    import time
    from datetime import datetime, timezone
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )
    
    now = datetime.now()
    start = now - timedelta(days=14)
    end_date_ms = int(now.timestamp() * 1000)
    start_date_ms = int(start.timestamp() * 1000)
    
    try:
        resp = await client.get_orders(start_date_ms=start_date_ms, end_date_ms=end_date_ms, size=50)
        return resp
    except Exception as e:
        return {"error": str(e)}

@router.post("/trendyol/categories/sync")
async def sync_trendyol_categories(current_user: dict = Depends(require_admin)):
    """Sync and save category tree from Trendyol API to local DB"""
    config = await get_trendyol_config()
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    
    try:
        client = TrendyolClient(
            supplier_id=config["supplier_id"],
            api_key=config["api_key"],
            api_secret=config["api_secret"],
            mode=config["mode"]
        )
        categories = await client.get_categories()
        
        # Save to DB (Drop and re-insert for clean sync)
        if categories:
            await db.trendyol_categories.delete_many({})
            await db.trendyol_categories.insert_many(categories)
            
        return {"success": True, "message": f"{len(categories)} kategori senkronize edildi."}
    except Exception as e:
        logger.error(f"Error fetching trendyol categories: {str(e)}")
        raise HTTPException(status_code=500, detail="Trendyol kategorileri alınamadı")



@router.get("/trendyol/categories/{category_id}/attributes")
async def get_trendyol_category_attributes(category_id: int, refresh: bool = Query(False, description="True ise cache atlanır, Trendyol'dan taze ve eksiksiz çekilir"), current_user: dict = Depends(require_admin)):
    """Get attributes for a specific category (From DB or Trendyol API directly)"""
    config = await get_trendyol_config()
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    
    # Check if already in local DB (refresh=true ise cache atla — eksik/eski cache'i yenile)
    if not refresh:
        existing = await db.trendyol_attributes.find_one({"category_id": category_id})
        if existing and existing.get("attributes"):
            return {"success": True, "attributes": existing.get("attributes", []), "cached": True}
        
    try:
        client = TrendyolClient(
            supplier_id=config["supplier_id"],
            api_key=config["api_key"],
            api_secret=config["api_secret"],
            mode=config["mode"]
        )
        attributes = await client.get_category_attributes(category_id)
        
        # Save to local DB for future use
        if attributes:
            await db.trendyol_attributes.update_one(
                {"category_id": category_id},
                {"$set": {"category_id": category_id, "attributes": attributes}},
                upsert=True
            )
            
        return {"success": True, "attributes": attributes}
    except Exception as e:
        logger.error(f"Error fetching trendyol attributes for category {category_id}: {str(e)}")
        raise HTTPException(status_code=500, detail="Özellikler (attributes) alınamadı")


@router.get("/hepsiburada/categories/{category_id}/attributes")
async def get_hepsiburada_category_attributes(category_id: str, current_user: dict = Depends(require_admin)):
    """Bir HB kategorisinin (canli) ozelliklerini doner — urun editorunun HB bolumu icin.
    Once cache, yoksa canli cekip cache'ler."""
    key = int(category_id) if str(category_id).isdigit() else str(category_id)
    cached = await db.hepsiburada_category_attributes.find_one({"category_id": key}, {"_id": 0})
    if cached and cached.get("_v") == 8 and cached.get("attributes") is not None:
        return {"success": True, "attributes": cached.get("attributes", []),
                "media_attributes": cached.get("media_attributes", []),
                "base_attributes": cached.get("base_attributes", []),
                "raw_structure": cached.get("raw_structure", {})}
    from .category_mapping import _fetch_hb_category_attributes
    attrs, err = await _fetch_hb_category_attributes(category_id)
    if err:
        return {"success": False, "attributes": [], "message": err}
    fresh = await db.hepsiburada_category_attributes.find_one({"category_id": key}, {"_id": 0}) or {}
    return {"success": True, "attributes": attrs,
            "media_attributes": fresh.get("media_attributes", []),
            "base_attributes": fresh.get("base_attributes", []),
            "raw_structure": fresh.get("raw_structure", {})}


HB_COMMON_ATTRS = [
    {"key": "cinsiyet", "label": "Cinsiyet"},
]


async def _hb_common_attr_values(norm_name: str):
    """Önbellekteki HB kategorilerinden, adı norm_name olan özelliğin TÜM değerlerini (uniq) toplar.
    Örn 'cinsiyet' -> ['Erkek','Kadın','Unisex',...] (kategoriler arası birleşik liste)."""
    out, seen = [], set()
    try:
        cursor = db.hepsiburada_category_attributes.find({}, {"_id": 0, "attributes": 1})
        async for doc in cursor:
            for a in (doc.get("attributes") or []):
                if _hb_norm(a.get("name")) == norm_name:
                    for v in (a.get("attributeValues") or []):
                        nm = v.get("name")
                        k = _hb_norm(nm)
                        if nm and k not in seen:
                            seen.add(k)
                            out.append(nm)
    except Exception:
        pass
    return sorted(out)


@router.get("/hepsiburada/base-field-mappings")
async def hb_get_base_field_mappings(current_user: dict = Depends(require_admin)):
    """HB temel/sistem alanlarının (Satıcı Stok Kodu / Ürün Adı / Barkod / Marka / KDV / Desi /
    Görsel ...) ürün-kartı kaynağına ya da sabit değere GLOBAL eşleştirmesi. Üst panel için
    alan listesi + kaynak seçenekleri + kayıtlı config döner."""
    from .category_mapping import HB_BASE_FIELDS, HB_PRODUCT_SOURCES
    s = await db.settings.find_one({"id": "hepsiburada"}, {"_id": 0}) or {}
    saved = s.get("base_field_mappings") or {}
    fields = []
    for f in HB_BASE_FIELDS:
        cfg = saved.get(f["key"]) or {}
        fields.append({
            "key": f["key"], "label": f["label"],
            "default_source": f.get("default_source"),
            "source": cfg.get("source") or f.get("default_source"),
            "default": cfg.get("default", f.get("default_value", "")),
        })
    try:
        markup = float(s.get("default_markup", 0) or 0)
    except Exception:
        markup = 0.0
    price_source = s.get("price_source") or "price"
    if price_source not in _HB_PRICE_SOURCE_KEYS:
        price_source = "price"
    gad = s.get("global_attr_defaults") or {}
    common_attrs = []
    for c in HB_COMMON_ATTRS:
        nk = _hb_norm(c["key"])
        vals = await _hb_common_attr_values(nk)
        common_attrs.append({"key": nk, "label": c["label"], "values": vals,
                             "selected": gad.get(nk) or gad.get(c["key"]) or ""})
    return {"success": True, "fields": fields, "sources": HB_PRODUCT_SOURCES, "saved": saved,
            "markup": markup, "price_source": price_source, "price_sources": HB_PRICE_SOURCES,
            "common_attrs": common_attrs, "global_attr_defaults": gad}


@router.post("/hepsiburada/base-field-mappings")
async def hb_save_base_field_mappings(request: Request, current_user: dict = Depends(require_admin)):
    """{mappings: {key: {source, default}}} kaydeder (db.settings id=hepsiburada)."""
    from .category_mapping import _HB_BASE_BY_KEY
    payload = await request.json()
    mappings = payload.get("mappings") if isinstance(payload, dict) and "mappings" in payload else payload
    clean = {}
    for k, v in (mappings or {}).items():
        if k not in _HB_BASE_BY_KEY or not isinstance(v, dict):
            continue
        clean[k] = {"source": v.get("source") or "", "default": str(v.get("default") or "")}
    set_doc = {"id": "hepsiburada", "base_field_mappings": clean}
    if isinstance(payload, dict) and "markup" in payload:
        try:
            set_doc["default_markup"] = max(0.0, float(payload.get("markup") or 0))
        except Exception:
            pass
    if isinstance(payload, dict) and "price_source" in payload:
        ps = str(payload.get("price_source") or "price")
        set_doc["price_source"] = ps if ps in _HB_PRICE_SOURCE_KEYS else "price"
    if isinstance(payload, dict) and "global_attr_defaults" in payload:
        gin = payload.get("global_attr_defaults") or {}
        valid = {_hb_norm(c["key"]) for c in HB_COMMON_ATTRS}
        gclean = {}
        if isinstance(gin, dict):
            for k, v in gin.items():
                nk = _hb_norm(k)
                if nk in valid and v not in (None, ""):
                    gclean[nk] = str(v)
        set_doc["global_attr_defaults"] = gclean
    await db.settings.update_one({"id": "hepsiburada"}, {"$set": set_doc}, upsert=True)
    return {"success": True, "saved": clean, "count": len(clean),
            "markup": set_doc.get("default_markup"), "price_source": set_doc.get("price_source"),
            "global_attr_defaults": set_doc.get("global_attr_defaults")}


@router.post("/trendyol/brands/sync")
async def sync_trendyol_brands(current_user: dict = Depends(require_admin)):
    """Sync brands from Trendyol API to local DB"""
    config = await get_trendyol_config()
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    
    try:
        client = TrendyolClient(
            supplier_id=config["supplier_id"],
            api_key=config["api_key"],
            api_secret=config["api_secret"],
            mode=config["mode"]
        )
        # Assuming maximum 500 per page, just fetching first page for demonstration. In prod, paginate.
        brands_data = await client.get_brands(size=5000)
        brands = brands_data.get("brands", [])
        
        if brands:
            await db.trendyol_brands.delete_many({})
            await db.trendyol_brands.insert_many(brands)
            
        return {"success": True, "message": f"{len(brands)} marka senkronize edildi."}
    except Exception as e:
        logger.error(f"Error fetching trendyol brands: {str(e)}")
        raise HTTPException(status_code=500, detail="Trendyol markaları alınamadı")

def _dedupe_products_by_stock_code(products: list) -> list:
    """Aynı stock_code/barcode ile birden fazla ürün dokümanı varsa
    (örn. csv_xml_merge ve xml_feed kaynaklarından gelen dublikatlar),
    her grup için EN İYİ dokümanı seçer.

    "İyi" doküman skoru:
      +1000  görseli varsa (images non-empty)
      +100   thumbnail varsa
      +len(images)
      +50    source != csv_xml_merge   (xml_feed > ticimax > csv_xml_merge)
      +10    aktif ürünse

    Aynı stock_code'a sahip diğer dokümanlar elenir. Anahtar olarak
    stock_code/sku/barcode sırayla denenir; hiçbiri yoksa id ile (tek başına).
    """
    def _score(p):
        s = 0
        imgs = p.get("images") or []
        if imgs:
            s += 1000
            s += min(len(imgs), 20)
        if p.get("thumbnail"):
            s += 100
        src = (p.get("source") or "").lower()
        if src and src != "csv_xml_merge":
            s += 50
        if p.get("is_active"):
            s += 10
        return s

    groups: dict = {}
    out_no_key: list = []
    for p in products:
        # Aynı stock_code'a farklı renkler/varyantlar tek bir SKU paylaşıyorsa
        # name (ürün adı) ile ayrıştır → "Bordo" ve "Siyah" ayrı kalır.
        stock = (p.get("stock_code") or p.get("sku") or p.get("barcode") or "").strip()
        name = (p.get("name") or "").strip().lower()
        key = f"{stock}|{name}" if stock else ""
        if not key:
            out_no_key.append(p)
            continue
        groups.setdefault(key, []).append(p)

    deduped = []
    for key, plist in groups.items():
        if len(plist) == 1:
            deduped.append(plist[0])
        else:
            best = max(plist, key=_score)
            deduped.append(best)
    deduped.extend(out_no_key)
    return deduped


def _normalize_attr_key(s: str) -> str:
    """Türkçe duyarsız normalize (İ/ı/ş/ğ/ü/ö/ç + birleşik nokta)."""
    s = (s or "").casefold()
    for a, b in (("ı", "i"), ("İ", "i"), ("ş", "s"), ("ğ", "g"),
                 ("ü", "u"), ("ö", "o"), ("ç", "c"), ("\u0307", "")):
        s = s.replace(a, b)
    return s.strip()


def _norm_val(s: str) -> str:
    """Listeli özellik değerlerini eşlerken kullanılan agresif normalize:
    Türkçe duyarsız + boşluk/eğik çizgi/noktalama tamamen kaldırılır.
    Örn. 'Kısa / Mini' ≈ 'Kısa/Mini' ≈ 'kisamini'."""
    return re.sub(r"[^a-z0-9]", "", _normalize_attr_key(s))


# Açık ve evrensel (kategoriden bağımsız) değer eşanlamlıları: lokal değer (norm) →
# kabul edilebilir Trendyol değer adları (norm). Sadece anlamı net olanlar; belirsiz
# olanlar (Sezon, Materyal vb.) kullanıcının kategori bazında eşleştirmesine bırakılır.
_VALUE_SYNONYMS = {
    "yakasiz": ["sifiryaka", "yakayok"],
    "sifiryaka": ["yakasiz"],
}


def _resolve_value_id(name_map: dict, local_val: str):
    """local_val'i Trendyol value_id'ye çöz: önce birebir (norm), sonra eşanlamlı."""
    if not name_map or local_val in (None, ""):
        return None
    nv = _norm_val(str(local_val))
    if nv in name_map:
        return name_map[nv]
    for syn in _VALUE_SYNONYMS.get(nv, []):
        if syn in name_map:
            return name_map[syn]
    return None

# Trendyol özellik adı -> bu değeri besleyebilecek lokal özellik kaynakları (normalize edilmiş).
# Trendyol "Materyal Bileşeni" (serbest metin, allowCustom) bizdeki "Ürün İçerik Bilgisi"ne karşılık gelir.
_TRENDYOL_ATTR_SYNONYMS = {
    "materyal bileşeni": ["urun icerik bilgisi", "kumas bilgisi", "kumas icerigi", "urun icerigi", "icerik bilgisi"],
}
# Materyal Bileşeni'ne köprülenirken atlanacak açıkça hatalı (yaş/cinsiyet) değerler.
_BAD_COMPOSITION_VALUES = {"yetişkin", "yetiskin", "genç", "genc", "çocuk", "cocuk", "bebek", "kadın", "kadin", "erkek", "unisex"}


def _bridge_trendyol_attr_synonyms(local_vals: dict) -> dict:
    """Trendyol özellik adlarıyla (ör. 'Materyal Bileşeni') lokal özellik adları
    (ör. 'Ürün İçerik Bilgisi') farklı olabilir. Eksik Trendyol anahtarını uygun
    lokal kaynaktan köprüler. local_vals: {lower(label): value}."""
    if not local_vals:
        return local_vals
    norm_index: dict = {}
    for k, v in local_vals.items():
        norm_index.setdefault(_normalize_attr_key(k), v)
    for target, sources in _TRENDYOL_ATTR_SYNONYMS.items():
        if target in local_vals and local_vals.get(target):
            continue
        for src in sources:
            val = norm_index.get(src)
            if val and str(val).strip().casefold() not in _BAD_COMPOSITION_VALUES:
                local_vals[target] = val
                break
    return local_vals



async def _build_product_query_from_payload(payload: dict) -> dict:
    """Trendyol sync/validate payload'undan products koleksiyon sorgusu üretir."""
    product_ids = payload.get("product_ids", [])
    category_filters = payload.get("category_filters", [])
    barcodes_raw = payload.get("barcodes", [])
    stock_codes_raw = payload.get("stock_codes", [])
    barcodes = [str(b).strip() for b in (barcodes_raw or []) if str(b).strip()]
    stock_codes = [str(s).strip() for s in (stock_codes_raw or []) if str(s).strip()]
    date_from = payload.get("date_from")
    date_to = payload.get("date_to")

    query: dict = {}
    if product_ids:
        query = {"id": {"$in": product_ids}}
    elif barcodes or stock_codes:
        or_conditions = []
        if barcodes:
            or_conditions.append({"barcode": {"$in": barcodes}})
            or_conditions.append({"variants.barcode": {"$in": barcodes}})
        if stock_codes:
            or_conditions.append({"stock_code": {"$in": stock_codes}})
            or_conditions.append({"sku": {"$in": stock_codes}})
            or_conditions.append({"variants.stock_code": {"$in": stock_codes}})
        query = {"$or": or_conditions}
    elif category_filters:
        or_conditions = []
        for cf in category_filters:
            cat_id = cf.get("category_id")
            filters = cf.get("filters", {})
            cat = None
            try:
                from bson.objectid import ObjectId
                cat = await db.categories.find_one({"_id": ObjectId(cat_id)})
            except Exception:
                cat = await db.categories.find_one({"id": cat_id})
            if not cat:
                continue
            cat_name = cat.get("name")
            # Hem category_id hem category_name ile match (ürünlerin çoğu category_id=None olabiliyor)
            inner_or = []
            if cat_id:
                inner_or.append({"category_id": cat_id})
            if cat_name:
                inner_or.append({"category_name": cat_name})
            if not inner_or:
                continue
            cat_q = {"$or": inner_or} if len(inner_or) > 1 else inner_or[0]
            extra_q = {}
            if filters.get("stock_code"):
                extra_q["stock_code"] = {"$regex": filters["stock_code"], "$options": "i"}
            if filters.get("date_range"):
                try:
                    date_obj = datetime.strptime(filters["date_range"], "%Y-%m-%d").replace(tzinfo=timezone.utc)
                    extra_q["created_at"] = {"$gte": date_obj.isoformat()}
                except Exception:
                    pass
            if extra_q:
                cat_q = {"$and": [cat_q, extra_q]}
            or_conditions.append(cat_q)
        if or_conditions:
            query = {"$or": or_conditions}
        else:
            query = {"is_active": True}
    else:
        query = {"is_active": True}

    if date_from or date_to:
        date_q: dict = {}
        if date_from:
            date_q["$gte"] = date_from
        if date_to:
            date_q["$lte"] = date_to + "T23:59:59" if "T" not in str(date_to) else date_to
        query = {"$and": [query, {"created_at": date_q}]} if query else {"created_at": date_q}
    return query


@router.post("/trendyol/products/validate")
async def validate_products_for_trendyol(
    request: Request,
    current_user: dict = Depends(require_admin),
):
    """
    Aktarım öncesi DOĞRULAMA paneli — ürün(ler)i Trendyol'a göndermeden,
    eksik zorunlu alanları (kategori mapping, barkod, görsel, zorunlu attribute)
    listeleyip raporlar. Body sync ile aynı.
    """
    payload = await request.json()
    query = await _build_product_query_from_payload(payload)
    products = await db.products.find(query, {"_id": 0}).to_list(length=None)
    # 🛡️ Aynı stock_code ile dublike doküman varsa, görseli/source'u en iyi olanı seç
    products = _dedupe_products_by_stock_code(products)

    # Category mappings (category_id -> mapping doc) — Trendyol için
    cm_list = await db.category_mappings.find(
        {"marketplace": "trendyol"}, {"_id": 0}
    ).to_list(length=3000)
    cm_by_local = {str(c.get("category_id")): c for c in cm_list}
    # Fallback: kategori adına göre
    all_cats = await db.categories.find({}, {"_id": 0}).to_list(length=5000)
    cat_by_id = {str(c.get("id")): c for c in all_cats}
    cat_by_name = {(c.get("name") or "").strip(): c for c in all_cats}
    cm_by_name = {}
    for c in all_cats:
        nm = (c.get("name") or "").strip()
        if nm and str(c.get("id")) in cm_by_local:
            cm_by_name[nm] = cm_by_local[str(c.get("id"))]

    # Trendyol mp_cat -> required attribute listesini cache'le
    attr_cache: dict = {}

    async def get_required_attrs(mp_cat_id):
        if not mp_cat_id:
            return []
        key = str(mp_cat_id)
        if key in attr_cache:
            return attr_cache[key]
        try:
            cached = await db.trendyol_category_attributes.find_one(
                {"category_id": int(mp_cat_id) if str(mp_cat_id).isdigit() else str(mp_cat_id)},
                {"_id": 0},
            )
            attrs = (cached or {}).get("attributes", []) or []
        except Exception:
            attrs = []
        required = [a for a in attrs if a.get("required")]
        attr_cache[key] = required
        return required

    results = []
    valid_count = 0
    invalid_count = 0

    for p in products:
        errors: list[str] = []
        warnings: list[str] = []
        missing_required_attrs: list[dict] = []
        unmatched_values: list[dict] = []

        cat_id = p.get("category_id")
        cat_name = p.get("category_name") or ""
        # Sync ile aynı sıra: category_id → category_name → categories.trendyol_category_id
        cm = cm_by_local.get(str(cat_id)) if cat_id else None
        if not cm and cat_name:
            cm = cm_by_name.get(cat_name.strip())
        cat_doc = cat_by_id.get(str(cat_id)) or cat_by_name.get(cat_name.strip())

        mp_cat_id = None
        if cm and cm.get("marketplace_category_id"):
            mp_cat_id = cm.get("marketplace_category_id")
        elif cat_doc and cat_doc.get("trendyol_category_id"):
            mp_cat_id = cat_doc.get("trendyol_category_id")

        if not mp_cat_id:
            errors.append("Trendyol kategori eşleştirmesi yok")

        # Görsel kontrolü
        if not (p.get("images") or []):
            errors.append("En az 1 ürün görseli yok")

        # Barkod kontrolü (varyantlı ürünlerde tüm varyantlar)
        # ⚠️ barcode_uncertain=True ise barkod yok kabul edilir (stock_code'tan kopyalanmış olabilir)
        variants = p.get("variants") or []
        if not variants:
            if not p.get("barcode") or p.get("barcode_uncertain"):
                errors.append("Barkod yok / belirsiz (Ticimax'tan doğrulayın)")
            if not p.get("stock_code"):
                warnings.append("Stok kodu yok")
        else:
            missing_v_barcode = sum(
                1 for v in variants if not v.get("barcode") or v.get("barcode_uncertain")
            )
            if missing_v_barcode:
                errors.append(f"{missing_v_barcode} varyantın barkodu eksik/belirsiz")

        # Fiyat
        try:
            if float(p.get("price", 0) or 0) <= 0:
                errors.append("Fiyat 0 veya boş")
        except Exception:
            errors.append("Fiyat geçersiz")

        # Stok
        total_stock = int(p.get("stock", 0) or 0) + sum(int(v.get("stock", 0) or 0) for v in variants)
        if total_stock <= 0:
            warnings.append("Toplam stok 0")

        # Açıklama
        if not (p.get("description") or p.get("short_description")):
            warnings.append("Açıklama boş")

        # Zorunlu attribute kontrolü (category_mappings → attribute_mappings + default_mappings)
        if mp_cat_id:
            req_attrs = await get_required_attrs(mp_cat_id)
            attr_mappings = (cm or {}).get("attribute_mappings", []) or []
            default_mappings = (cm or {}).get("default_mappings", {}) or {}
            # local_attr → mp_attr_id map'ını da kur
            mp_id_to_local = {}
            for am in attr_mappings:
                mid = str(am.get("mp_attr_id") or am.get("trendyol_attr_id") or "")
                if mid:
                    mp_id_to_local[mid] = am.get("local_attr")

            # Ürünün attribute'larından lokal isim → değer (DICT veya LIST formatını destekler)
            local_vals: dict = {}

            def _add_lv(nm, vv):
                if not nm or vv in (None, ""):
                    return
                local_vals.setdefault(str(nm).lower().strip(), str(vv))

            def _walk(attrs):
                if isinstance(attrs, dict):
                    items = sorted(
                        attrs.items(),
                        key=lambda kv: 1 if str(kv[0]).lower().startswith("ticimax_") else 0,
                    )
                    for k, v in items:
                        if isinstance(v, dict):
                            nm = v.get("label") or v.get("name") or k
                            vv = v.get("value") or v.get("attribute_value")
                            _add_lv(nm, vv)
                        elif v is not None:
                            _add_lv(k, v)
                elif isinstance(attrs, list):
                    for a in attrs:
                        if isinstance(a, dict):
                            nm = a.get("label") or a.get("name") or a.get("type") or a.get("attribute_name")
                            vv = a.get("value") or a.get("attribute_value")
                            _add_lv(nm, vv)

            _walk(p.get("attributes"))
            for v in variants:
                _walk(v.get("attributes"))
                if v.get("color"):
                    _add_lv("Renk", v["color"])
                    _add_lv("Web Color", v["color"])
                if v.get("size"):
                    _add_lv("Beden", v["size"])
            _bridge_trendyol_attr_synonyms(local_vals)

            for ra in req_attrs:
                ra_id = str(ra.get("id") or ra.get("attribute", {}).get("id") or "")
                ra_name = ra.get("name") or ra.get("attribute", {}).get("name") or "(?)"
                if not ra_id:
                    continue
                # default mapping var mı?
                default_val = default_mappings.get(ra_id) or default_mappings.get(str(ra_id))
                if default_val:
                    continue
                # local attribute mapping var mı + üründe değer var mı?
                local_attr = mp_id_to_local.get(ra_id)
                has_val = False
                if local_attr:
                    has_val = bool(local_vals.get(local_attr.lower()))
                if not has_val:
                    missing_required_attrs.append({
                        "id": ra_id,
                        "name": ra_name,
                        "mapped_local": local_attr,
                    })
            if missing_required_attrs:
                errors.append(f"{len(missing_required_attrs)} zorunlu özellik eksik")

            # 🔎 Listeli (enum) değerlerin Trendyol karşılığı var mı? Yoksa aktarım engellenir
            # (allowCustom alanlar serbest metindir, her zaman gönderilir → kontrol edilmez).
            val_mappings_v = (cm or {}).get("value_mappings", {}) or {}
            try:
                _cache_attrs = await db.trendyol_category_attributes.find_one(
                    {"category_id": int(mp_cat_id) if str(mp_cat_id).isdigit() else str(mp_cat_id)},
                    {"_id": 0},
                )
            except Exception:
                _cache_attrs = None
            local_norm_index_v = {}
            for lk, lv in local_vals.items():
                local_norm_index_v.setdefault(_normalize_attr_key(lk), lv)
            for a in (_cache_attrs or {}).get("attributes", []) or []:
                aid = a.get("id") or a.get("attribute", {}).get("id")
                if aid is None:
                    continue
                if bool(a.get("allowCustom") or a.get("attribute", {}).get("allowCustom")):
                    continue
                aname = a.get("name") or a.get("attribute", {}).get("name") or ""
                lval = local_norm_index_v.get(_normalize_attr_key(aname))
                if not lval:
                    la = mp_id_to_local.get(str(aid))
                    if la:
                        lval = local_vals.get(la.lower())
                if not lval:
                    continue
                if val_mappings_v.get(f"{aid}|{lval}"):
                    continue
                vname_map = {
                    _norm_val(v.get("name")): str(v.get("id"))
                    for v in (a.get("attributeValues") or [])
                    if v.get("id") is not None and v.get("name")
                }
                if _resolve_value_id(vname_map, lval):
                    continue
                unmatched_values.append({
                    "mp_attr_id": int(aid),
                    "attr_name": aname,
                    "local_value": lval,
                    "required": bool(a.get("required")),
                })
            if unmatched_values:
                errors.append(f"{len(unmatched_values)} değerin Trendyol karşılığı yok (eşleştirme gerekli)")

        is_valid = len(errors) == 0
        if is_valid:
            valid_count += 1
        else:
            invalid_count += 1

        results.append({
            "id": p.get("id"),
            "name": p.get("name"),
            "stock_code": p.get("stock_code"),
            "barcode": p.get("barcode"),
            "category_name": cat_name,
            "marketplace_category_id": mp_cat_id,
            "is_valid": is_valid,
            "errors": errors,
            "warnings": warnings,
            "missing_required_attrs": missing_required_attrs,
            "unmatched_values": unmatched_values,
        })

    # Eksik attribute istatistikleri (en sık eksik olanları üstte göster)
    attr_freq: dict = {}
    for r in results:
        for m in r["missing_required_attrs"]:
            k = m["name"]
            attr_freq[k] = attr_freq.get(k, 0) + 1
    top_missing = sorted(attr_freq.items(), key=lambda x: -x[1])[:10]

    return {
        "success": True,
        "total": len(products),
        "valid_count": valid_count,
        "invalid_count": invalid_count,
        "results": results,
        "top_missing_attrs": [{"name": k, "count": v} for k, v in top_missing],
    }


@router.get("/trendyol/batch/{batch_id}")
async def get_trendyol_batch_status(
    batch_id: str,
    current_user: dict = Depends(require_admin),
):
    """Trendyol batch işleminin (ürün oluşturma vb.) gerçek durumunu döndürür.
    UI'da kullanıcı 'Detayları Gör' butonuna basınca her item'ın SUCCESS/FAILED
    durumu ve failureReasons listesini görür."""
    config = await get_trendyol_config()
    if not config or not config.get("is_active"):
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu aktif değil")
    from trendyol_client import TrendyolClient
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config.get("mode", "live"),
    )
    try:
        data = await client.get_batch_request_result(batch_id)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Trendyol batch detayı alınamadı: {e}")

    items = data.get("items") or []
    success_count = sum(1 for it in items if str(it.get("status")).upper() == "SUCCESS")
    failed_count = sum(1 for it in items if str(it.get("status")).upper() == "FAILED")
    # Hata özetlerini topla
    fail_freq: dict = {}
    for it in items:
        for fr in (it.get("failureReasons") or []):
            key = str(fr).split(".")[0][:120]
            fail_freq[key] = fail_freq.get(key, 0) + 1
    return {
        "batch_id": batch_id,
        "status": data.get("status"),
        "source_type": data.get("sourceType"),
        "item_count": data.get("itemCount") or len(items),
        "success_count": success_count,
        "failed_count": failed_count,
        "top_failures": [{"reason": k, "count": v} for k, v in
                         sorted(fail_freq.items(), key=lambda x: -x[1])[:10]],
        "items": items,
        "raw": data,
    }


# ============== GHOST PRODUCT SCANNER & DB DUPLICATE DETECTOR ==============

@router.get("/trendyol/barcode-duplicates")
async def trendyol_barcode_duplicates(current_user: dict = Depends(require_admin)):
    """DB içinde aynı barkoda atanmış birden fazla varyantı tespit eder.
    Bu Trendyol'a aktarımı bloklayan ana sebep oluyor. Manuel düzeltme için liste döner.
    """
    pipeline = [
        {"$match": {"variants.barcode": {"$nin": [None, ""]}}},
        {"$unwind": "$variants"},
        {"$match": {"variants.barcode": {"$nin": [None, ""]}}},
        {"$group": {
            "_id": "$variants.barcode",
            "count": {"$sum": 1},
            "products": {"$push": {
                "product_id": "$id",
                "stock_code": "$stock_code",
                "name": "$name",
                "is_active": "$is_active",
                "variant_size": "$variants.size",
                "variant_color": "$variants.color",
                "variant_stock": "$variants.stock",
            }}
        }},
        {"$match": {"count": {"$gt": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 500},
    ]
    rows = await db.products.aggregate(pipeline).to_list(length=500)
    out = []
    for r in rows:
        out.append({
            "barcode": r["_id"],
            "count": r["count"],
            "assignments": r["products"],
        })
    return {"total": len(out), "duplicates": out}


@router.post("/trendyol/ghost-scanner")
async def trendyol_ghost_scanner(
    payload: dict = Body(default={}),
    current_user: dict = Depends(require_admin),
):
    """Trendyol panelindeki tüm ürünleri (max 5000) sayfalı çekip,
    DB'de KARŞILIK BULAMAYAN ya da DB'de archive edilmiş olanları "hayalet" olarak listeler.
    Bu hayaletler genelde eski yanlış barkod kayıtlarıdır ve duplicate çakışmalara sebep olur.
    """
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    import sys
    import os as _os
    sys.path.insert(0, _os.path.dirname(_os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient

    cli = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    only_unmatched = bool(payload.get("only_unmatched", True))
    include_archived = bool(payload.get("include_archived", False))
    page_limit = int(payload.get("page_limit", 50))  # 50 sayfa x 200 = 10K ürün max

    # DB barkodlarını VE stock_code'larını topla
    db_barcodes = set()
    db_stock_codes = set()
    async for p in db.products.find({}, {"_id": 0, "barcode": 1, "stock_code": 1, "variants.barcode": 1, "variants.stock_code": 1}):
        if p.get("barcode"):
            db_barcodes.add(str(p["barcode"]))
        if p.get("stock_code"):
            db_stock_codes.add(str(p["stock_code"]))
        for v in (p.get("variants") or []):
            if v.get("barcode"):
                db_barcodes.add(str(v["barcode"]))
            if v.get("stock_code"):
                db_stock_codes.add(str(v["stock_code"]))

    ghosts = []
    matched = 0
    total_scanned = 0
    page = 0
    while page < page_limit:
        try:
            res = await cli.get_filtered_products(
                page=page, size=200,
                archived=None if include_archived else False,
            )
        except Exception as e:
            return {"error": str(e), "scanned": total_scanned, "ghosts": ghosts}
        content = res.get("content") or []
        total_pages = res.get("totalPages") or 0
        if not content:
            break
        for row in content:
            total_scanned += 1
            bc = str(row.get("barcode") or "")
            sc = str(row.get("stockCode") or "")
            pmi = str(row.get("productMainId") or "")
            if not bc:
                continue
            # Match: barkod DB'de varsa VEYA stockCode/productMainId DB'de varsa "matched"
            if bc in db_barcodes or sc in db_barcodes or sc in db_stock_codes or pmi in db_stock_codes:
                matched += 1
                if only_unmatched:
                    continue
            else:
                ghosts.append({
                    "barcode": bc,
                    "stockCode": row.get("stockCode"),
                    "productMainId": row.get("productMainId"),
                    "title": row.get("title"),
                    "brand": row.get("brand"),
                    "approved": row.get("approved"),
                    "archived": row.get("archived"),
                    "onSale": row.get("onSale"),
                    "salePrice": row.get("salePrice"),
                    "quantity": row.get("quantity"),
                    "rejectReasonDetails": row.get("rejectReasonDetails"),
                })
        page += 1
        if page >= total_pages:
            break

    return {
        "scanned": total_scanned,
        "matched_in_db": matched,
        "ghosts_count": len(ghosts),
        "ghosts": ghosts,
    }


@router.post("/trendyol/archive-barcodes")
async def trendyol_archive_barcodes(
    payload: dict = Body(...),
    current_user: dict = Depends(require_admin),
):
    """Verilen barkodları Trendyol'da arşivler (panelden gizler, slot iadesi sağlar).
    Hayalet ürünlerin temizliği için kullanılır.
    """
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    barcodes = [str(b).strip() for b in (payload.get("barcodes") or []) if str(b).strip()]
    if not barcodes:
        raise HTTPException(status_code=400, detail="Arşivlenecek barkod listesi boş.")
    if len(barcodes) > 1000:
        raise HTTPException(status_code=400, detail="Tek seferde max 1000 barkod arşivlenebilir.")

    import sys
    import os as _os
    sys.path.insert(0, _os.path.dirname(_os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient

    cli = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    try:
        resp = await cli.archive_products(barcodes)
        batch_id = (resp or {}).get("batchRequestId")
        from datetime import datetime, timezone
        log_doc = {
            "id": generate_id(),
            "started_at": datetime.now(timezone.utc).isoformat(),
            "status": "archive",
            "products_attempted": len(barcodes),
            "batch_request_id": batch_id,
            "archived_barcodes": barcodes,
            "trendyol_response": resp,
            "message": f"{len(barcodes)} barkod için arşivleme batch'i Trendyol'a gönderildi (batch: {batch_id}).",
        }
        await db.trendyol_sync_logs.insert_one(log_doc)
        log_doc.pop("_id", None)
        return {
            "success": bool(batch_id),
            "batchRequestId": batch_id,
            "barcodes_count": len(barcodes),
            "response": resp,
        }
    except Exception as e:
        logger.error(f"Archive barcodes error: {e}")
        raise HTTPException(status_code=500, detail=f"Trendyol arşiv hatası: {str(e)}")





@router.post("/trendyol/products/sync")
async def sync_products_to_trendyol(
    request: Request,
    current_user: dict = Depends(require_admin)
):
    """Sync products to Trendyol via Batch Request"""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")
    
    payload = await request.json()
    product_ids = payload.get("product_ids", [])
    category_filters = payload.get("category_filters", [])
    # Yeni: Barkod/stok kodu ile filtre (kullanıcı UI'da yazıp aktarabilsin)
    barcodes_raw = payload.get("barcodes", [])
    stock_codes_raw = payload.get("stock_codes", [])
    barcodes = [str(b).strip() for b in (barcodes_raw or []) if str(b).strip()]
    stock_codes = [str(s).strip() for s in (stock_codes_raw or []) if str(s).strip()]
    # Yeni: Tarih aralığı (created_at — ürün eklenme tarihi)
    date_from = payload.get("date_from")  # ISO format "2026-01-01"
    date_to = payload.get("date_to")
    
    query = {}
    if product_ids:
        query = {"id": {"$in": product_ids}}
    elif barcodes or stock_codes:
        # Hem ürünün kendi barcode/stock_code'una hem de variants[] içine bak
        or_conditions = []
        if barcodes:
            or_conditions.append({"barcode": {"$in": barcodes}})
            or_conditions.append({"variants.barcode": {"$in": barcodes}})
        if stock_codes:
            or_conditions.append({"stock_code": {"$in": stock_codes}})
            or_conditions.append({"sku": {"$in": stock_codes}})
            or_conditions.append({"variants.stock_code": {"$in": stock_codes}})
        query = {"$or": or_conditions}
    elif category_filters:
        # Build an $or query for each category + its filters
        or_conditions = []
        for cf in category_filters:
            cat_id = cf.get("category_id")
            filters = cf.get("filters", {})
            try:
                from bson.objectid import ObjectId
                cat = await db.categories.find_one({"_id": ObjectId(cat_id)})
            except Exception:
                cat = await db.categories.find_one({"id": cat_id})

            if not cat:
                continue
            
            cat_q = {"category_name": cat.get("name")}
            if filters.get("stock_code"):
                cat_q["stock_code"] = {"$regex": filters["stock_code"], "$options": "i"}
            if filters.get("date_range"):
                try:
                    from datetime import datetime, timezone
                    # Format: YYYY-MM-DD
                    date_obj = datetime.strptime(filters["date_range"], "%Y-%m-%d").replace(tzinfo=timezone.utc)
                    cat_q["created_at"] = {"$gte": date_obj.isoformat()}
                except Exception:
                    pass
            or_conditions.append(cat_q)
            
        if or_conditions:
            query = {"$or": or_conditions}
        else:
            query = {"is_active": True}
    else:
        # Sync only active products if no specific IDs are provided
        query = {"is_active": True}

    # Tarih aralığı filtresi (created_at) — diğer filtrelerle AND ile birleşir
    if date_from or date_to:
        date_q: dict = {}
        if date_from:
            date_q["$gte"] = date_from
        if date_to:
            date_q["$lte"] = date_to + "T23:59:59" if "T" not in str(date_to) else date_to
        query = {"$and": [query, {"created_at": date_q}]} if query else {"created_at": date_q}

    products = await db.products.find(query).to_list(length=None)
    # 🛡️ Aynı stock_code ile dublike doküman varsa, görseli/source'u en iyi olanı seç
    products = _dedupe_products_by_stock_code(products)

    # 🔍 Hangi stok kodları / barkodlar DB'de bulundu? Bulunmayanları topla.
    not_found_codes = []
    if barcodes or stock_codes:
        found_barcodes = set()
        found_stock_codes = set()
        for p in products:
            if p.get("barcode"):
                found_barcodes.add(str(p["barcode"]))
            if p.get("stock_code"):
                found_stock_codes.add(str(p["stock_code"]))
            if p.get("sku"):
                found_stock_codes.add(str(p["sku"]))
            for v in (p.get("variants") or []):
                if v.get("barcode"):
                    found_barcodes.add(str(v["barcode"]))
                if v.get("stock_code"):
                    found_stock_codes.add(str(v["stock_code"]))
        requested = set([str(c) for c in (barcodes or [])] + [str(c) for c in (stock_codes or [])])
        not_found_codes = sorted([c for c in requested if c not in found_barcodes and c not in found_stock_codes])

    # ❗ Hiç ürün bulunamadıysa anında net mesajla dön — kullanıcı "DB'de yok mu, validasyon mu hatalı?" diye anlamayabiliyor
    if not products:
        from datetime import datetime, timezone
        msg = (
            f"DB'de bu kod(lar)la ürün bulunamadı: {', '.join(not_found_codes[:10])}"
            + (f" (+{len(not_found_codes)-10} daha)" if len(not_found_codes) > 10 else "")
        ) if not_found_codes else "Sorgu kriterlerine uyan ürün bulunamadı."
        await db.trendyol_sync_logs.insert_one({
            "id": generate_id(),
            "started_at": datetime.now(timezone.utc).isoformat(),
            "status": "failed",
            "products_attempted": 0,
            "products_sent": 0,
            "batch_request_id": None,
            "errors": [msg],
            "not_found_codes": not_found_codes,
            "message": msg,
        })
        return {
            "success": False,
            "message": msg,
            "total": 0,
            "successful": 0,
            "failed": 0,
            "not_found_codes": not_found_codes,
            "errors": [msg],
        }
    
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )
    
    items_to_send = []
    errors = []

    # Trendyol kategori özellik cache'lerini bu sync için yükle (attribute meta + geçerli value_id'ler)
    _attr_meta_cache: dict = {}

    async def _get_attr_meta(mp_cat_id):
        if mp_cat_id in _attr_meta_cache:
            return _attr_meta_cache[mp_cat_id]
        meta: dict = {}
        try:
            cache = await db.trendyol_category_attributes.find_one(
                {"category_id": int(mp_cat_id) if str(mp_cat_id).isdigit() else str(mp_cat_id)},
                {"_id": 0},
            )
        except Exception:
            cache = None
        for a in (cache or {}).get("attributes", []) or []:
            aid = a.get("id") or a.get("attribute", {}).get("id")
            if aid is None:
                continue
            valid_value_ids = {str(v.get("id")) for v in (a.get("attributeValues") or []) if v.get("id") is not None}
            # Normalize edilmiş "değer adı → value_id" haritası (isimle otomatik eşleştirme için)
            value_name_to_id = {}
            for v in (a.get("attributeValues") or []):
                if v.get("id") is None or not v.get("name"):
                    continue
                value_name_to_id[_norm_val(v["name"])] = str(v["id"])
            meta[int(aid)] = {
                "allow_custom": bool(a.get("allowCustom") or a.get("attribute", {}).get("allowCustom")),
                "required": bool(a.get("required")),
                "valid_value_ids": valid_value_ids,
                "value_name_to_id": value_name_to_id,
                "name": a.get("name") or a.get("attribute", {}).get("name") or "",
            }
        _attr_meta_cache[mp_cat_id] = meta
        return meta

    def _collect_local_values(product, variant):
        """Ürünün ve varyantın attribute değerlerini lokalName(lower) → value şeklinde toplar.
        attributes hem LIST hem DICT formatını destekler."""
        out = {}
        def _put(nm, vv):
            if not nm or vv in (None, ""):
                return
            out.setdefault(str(nm).lower().strip(), str(vv))

        def _walk(attrs):
            if isinstance(attrs, dict):
                # Çakışan mükerrer özelliklerde TEMİZ (güncel, elle/teknik) anahtarlar
                # önce işlenir; `ticimax_*` (eski/ham 113-kolon) anahtarları yalnızca
                # fallback olur. _put setdefault olduğundan ilk gelen kazanır.
                items = sorted(
                    attrs.items(),
                    key=lambda kv: 1 if str(kv[0]).lower().startswith("ticimax_") else 0,
                )
                for k, v in items:
                    if isinstance(v, dict):
                        nm = v.get("label") or v.get("name") or k
                        vv = v.get("value") or v.get("attribute_value")
                        _put(nm, vv)
                    elif v is not None:
                        _put(k, v)
            elif isinstance(attrs, list):
                for a in attrs:
                    if isinstance(a, dict):
                        nm = a.get("label") or a.get("name") or a.get("type") or a.get("attribute_name")
                        vv = a.get("value") or a.get("attribute_value")
                        _put(nm, vv)
        _walk(product.get("attributes"))
        _walk((product.get("trendyol_attributes_labels") or {}))  # opsiyonel
        if variant:
            _walk(variant.get("attributes"))
            if variant.get("color"):
                _put("Renk", variant["color"])
                _put("Web Color", variant["color"])
            if variant.get("size"):
                _put("Beden", variant["size"])
        _bridge_trendyol_attr_synonyms(out)
        return out

    def resolve_attributes(base_attrs, product, variant, category, meta):
        """meta = _get_attr_meta sonucu: {attr_id: {allow_custom, required, valid_value_ids, name}}
        Geçersiz value_id ya da allow_custom=False olan custom değerleri SESSİZCE atlar."""
        item_attrs = list(base_attrs)
        processed = {int(a["attributeId"]) for a in item_attrs if "attributeId" in a}
        
        attr_mappings = category.get("attribute_mappings", []) or []
        val_mappings = category.get("value_mappings", {}) or {}
        default_mappings = category.get("default_mappings", {}) or {}

        local_vals = _collect_local_values(product, variant)

        def _push(ty_id: int, value_id=None, custom=None):
            """Cache'e karşı doğrulayarak append."""
            am = meta.get(ty_id) or {}
            # Dosya linki / sertifika gerektiren attribute'ları skip (custom text kabul etmez)
            am_name = (am.get("name") or "").lower()
            if any(p in am_name for p in ["analiz testi", "test raporu", "sertifika dosya", "dosya linki"]):
                return False
            if value_id is not None:
                vid = str(value_id)
                # Cache'de yoksa custom'a düşür
                if am.get("valid_value_ids") and vid not in am["valid_value_ids"]:
                    if am.get("allow_custom") and custom:
                        item_attrs.append({"attributeId": ty_id, "customAttributeValue": str(custom)})
                        processed.add(ty_id)
                        return True
                    return False
                item_attrs.append({"attributeId": ty_id, "attributeValueId": int(vid)})
                processed.add(ty_id)
                return True
            if custom is not None:
                if not am.get("allow_custom"):
                    return False
                item_attrs.append({"attributeId": ty_id, "customAttributeValue": str(custom)})
                processed.add(ty_id)
                return True
            return False
        
        for mapping in attr_mappings:
            # Yeni format: mp_attr_id, eski format: trendyol_attr_id
            ty_id = mapping.get("mp_attr_id") or mapping.get("trendyol_attr_id")
            if not ty_id:
                continue
            try:
                ty_id = int(ty_id)
            except (ValueError, TypeError):
                continue
            if ty_id in processed:
                continue
                
            local_attr_name = str(mapping.get("local_attr") or "").strip()
            local_key = local_attr_name.lower()
            local_val = None

            if variant:
                if local_key in ["renk", "color", "web color"]:
                    local_val = variant.get("color")
                elif local_key in ["beden", "size"]:
                    local_val = variant.get("size") or local_vals.get(local_key)
                    
            if not local_val:
                local_val = local_vals.get(local_key)
            if not local_val and local_attr_name:
                local_val = local_vals.get(local_attr_name.lower())
                        
            if local_val:
                str_ty_id = str(ty_id)
                # value_mapping format: "343|Erkek" → "value_id"
                mapped_val = val_mappings.get(f"{str_ty_id}|{local_val}")
                # eski yapı: {"343": {"Erkek": "value_id"}}
                if not mapped_val and str_ty_id in val_mappings and isinstance(val_mappings[str_ty_id], dict):
                    mapped_val = val_mappings[str_ty_id].get(str(local_val))
                if mapped_val:
                    if str(mapped_val).isdigit():
                        if _push(ty_id, value_id=mapped_val, custom=local_val):
                            continue
                    else:
                        if _push(ty_id, custom=mapped_val):
                            continue
                # Kaydedilmiş eşleştirme yok → listeli (enum) değeri ADIYLA otomatik eşle.
                # Örn. local "Dokuma"/"Örme"/"Kısa" → Trendyol value_id (Türkçe/boşluk duyarsız).
                am_meta = meta.get(ty_id) or {}
                name_map = am_meta.get("value_name_to_id") or {}
                if name_map:
                    auto_vid = _resolve_value_id(name_map, local_val)
                    if auto_vid and _push(ty_id, value_id=auto_vid, custom=local_val):
                        continue
                # Mapping yok ama allow_custom varsa local_val'i custom olarak yolla
                if _push(ty_id, custom=local_val):
                    continue
                
            # Default mapping
            str_ty_id = str(ty_id)
            if str_ty_id in default_mappings and default_mappings[str_ty_id]:
                def_val = default_mappings[str_ty_id]
                if str(def_val).isdigit():
                    _push(ty_id, value_id=def_val, custom=local_val)
                else:
                    _push(ty_id, custom=def_val)

        # Default mapping'de olup attribute_mappings'de olmayanları da ekle
        for ty_str, def_val in default_mappings.items():
            if not def_val:
                continue
            try:
                ty_id = int(ty_str)
            except (ValueError, TypeError):
                continue
            if ty_id not in processed:
                if str(def_val).isdigit():
                    _push(ty_id, value_id=def_val)
                else:
                    _push(ty_id, custom=def_val)

        # 🎯 GARANTİ: Trendyol "Materyal Bileşeni" (serbest metin / allowCustom) alanı,
        # mapping yapılmamış olsa bile HER kategoride ürünün "Ürün İçerik Bilgisi"
        # değerinden otomatik gönderilir. local_vals["materyal bileşeni"] köprü ile dolar.
        icerik_val = local_vals.get("materyal bileşeni")
        if icerik_val:
            for m_ty_id, m_meta in meta.items():
                if m_ty_id in processed:
                    continue
                mname = (m_meta.get("name") or "").lower()
                if "materyal bileşeni" in mname and m_meta.get("allow_custom"):
                    _push(m_ty_id, custom=icerik_val)

        # 🎯 GENEL GARANTİ: Trendyol özellik adı ile lokal özellik adı eşleşen ama
        # attribute_mappings'te tanımlanmamış alanları (Boy, Desen, Kumaş Tipi vb.)
        # otomatik gönder. Listeli alanlar değeri ADIYLA value_id'ye eşlenir,
        # serbest alanlar custom olarak yollanır.
        local_norm_index = {}
        for lk, lv in local_vals.items():
            local_norm_index.setdefault(_normalize_attr_key(lk), lv)
        for m_ty_id, m_meta in meta.items():
            if m_ty_id in processed:
                continue
            lval = local_norm_index.get(_normalize_attr_key(m_meta.get("name") or ""))
            if not lval:
                continue
            name_map = m_meta.get("value_name_to_id") or {}
            auto_vid = _resolve_value_id(name_map, lval)
            if auto_vid:
                _push(m_ty_id, value_id=auto_vid, custom=lval)
            else:
                _push(m_ty_id, custom=lval)

        return item_attrs
    
    for product in products:
        try:
            # 1. Category Mapping check — önce category_id, sonra category_name üzerinden category_mappings
            trendyol_cat_id = None
            category = None
            cat_id = product.get("category_id")
            cat_name = (product.get("category_name") or "").strip()
            cm = None

            # 1a. Doğrudan category_id ile mapping arayalım
            if cat_id:
                cm = await db.category_mappings.find_one(
                    {"category_id": str(cat_id), "marketplace": "trendyol"}, {"_id": 0}
                )

            # 1b. Bulunamadıysa: kategori adından sistem kategorisini, oradan mapping'i bul
            if not cm and cat_name:
                sys_cat = await db.categories.find_one({"name": cat_name}, {"_id": 0, "id": 1, "trendyol_category_id": 1})
                if sys_cat and sys_cat.get("id"):
                    cm = await db.category_mappings.find_one(
                        {"category_id": str(sys_cat["id"]), "marketplace": "trendyol"}, {"_id": 0}
                    )

            if cm and cm.get("marketplace_category_id"):
                trendyol_cat_id = cm["marketplace_category_id"]
                category = cm  # category_mappings doc — attribute_mappings, default_mappings içerir

            # 1c. Eski legacy fallback: db.categories.trendyol_category_id
            if not trendyol_cat_id and cat_name:
                cat_doc = await db.categories.find_one({"name": cat_name})
                if cat_doc and cat_doc.get("trendyol_category_id"):
                    trendyol_cat_id = cat_doc["trendyol_category_id"]
                    category = cat_doc

            if not trendyol_cat_id:
                errors.append(f"{product.get('name')} - Trendyol kategori eşleştirmesi (Mapping) yok.")
                continue
                
            # 2. Attributes check and formatting
            raw_attrs = product.get("trendyol_attributes", {})
            attributes = []
            for attr_id, val_id in raw_attrs.items():
                if val_id:
                    # Trendyol expects attributeId and attributeValueId, or customAttributeValue
                    if str(val_id).isdigit():
                        attributes.append({
                            "attributeId": int(attr_id),
                            "attributeValueId": int(val_id)
                        })
                    else:
                        attributes.append({
                            "attributeId": int(attr_id),
                            "customAttributeValue": str(val_id)
                        })
            
            # Description (HTML → düz metin: br/li/p satır sonuna, entity decode, min 30 karakter)
            import re as _re
            import html as _html
            raw_desc = (product.get("description") or product.get("short_description")
                        or product.get("long_description") or "").strip()
            # 1) Blok/satır kıran tagları newline'a çevir (görsel paragraf yapısı korunur).
            _desc = raw_desc
            _desc = _re.sub(r"<\s*br\s*/?\s*>", "\n", _desc, flags=_re.IGNORECASE)
            _desc = _re.sub(r"</\s*(p|div|li|h[1-6]|tr)\s*>", "\n", _desc, flags=_re.IGNORECASE)
            _desc = _re.sub(r"<\s*li\b[^>]*>", "• ", _desc, flags=_re.IGNORECASE)
            # 2) Tüm kalan HTML tag'lerini at.
            _desc = _re.sub(r"<[^>]+>", " ", _desc)
            # 3) HTML entity'leri decode et (&nbsp;, &amp;, &uuml; vb.).
            _desc = _html.unescape(_desc)
            # 4) Yatay boşlukları sıkıştır ama satır sonlarını koru.
            _desc = _re.sub(r"[ \t]+", " ", _desc)
            _desc = _re.sub(r"\n[ \t]+", "\n", _desc)
            _desc = _re.sub(r"[ \t]+\n", "\n", _desc)
            _desc = _re.sub(r"\n{3,}", "\n\n", _desc)
            clean_desc = _desc.strip()
            if not clean_desc or len(clean_desc) < 30:
                # Description yoksa veya çok kısaysa ürün adından bir minimum açıklama üret
                fallback = (product.get("name") or "").strip()
                if fallback and len(fallback) >= 10:
                    clean_desc = f"{fallback}. Kaliteli kumaş, modern kesim, şık tasarım. Günlük ve özel kullanım için ideal."
                else:
                    errors.append(f"{product.get('name')} - Açıklama eksik (Trendyol min 30 karakter zorunlu).")
                    continue

            # 3. Base Product Details
            base_item = {
                "title": product.get("name"),
                "productMainId": product.get("stock_code") or product.get("id"),
                "brandId": int(product.get("trendyol_brand_id") or 975755),
                "categoryId": int(trendyol_cat_id),
                "description": clean_desc,
                "currencyType": product.get("currency", "TRY"),
                "listPrice": calculate_trendyol_price(float(product.get("price", 0)), product, config),
                "salePrice": calculate_trendyol_price(float(product.get("price", 0)), product, config),
                "vatRate": int(product.get("vat_rate", 20)),
                "cargoCompanyId": 10, # Assuming 10 is MNG Kargo (Needs specific Cargo Provider ID)
                "dimensionalWeight": float(product.get("cargo_weight", 1)),
                "images": [{"url": img} for img in product.get("images", [])[:8]]
            }
            
            if not base_item["images"]:
                errors.append(f"{product.get('name')} - En az 1 görsel gerekli.")
                continue

            # 3b. Trendyol attribute meta (cache) — value_id ve allowCustom validasyonu için
            meta = await _get_attr_meta(trendyol_cat_id)
            
            # 4. Handle Variants or No-Variants
            variants = product.get("variants", [])
            if not variants:
                if not product.get("barcode") or product.get("barcode_uncertain"):
                    errors.append(f"{product.get('name')} - Barkod yok ya da belirsiz (Ticimax'tan doğrulayın).")
                    continue
                item = base_item.copy()
                item["barcode"] = product.get("barcode")
                item["stockCode"] = product.get("stock_code") or product.get("barcode")
                item["quantity"] = int(product.get("stock", 0))
                item["attributes"] = resolve_attributes(attributes, product, None, category, meta)
                items_to_send.append(item)
            else:
                # 🎯 Eğer kullanıcı barcodes parametresi ile spesifik barkodlar istediyse,
                # sadece o barkodları batch'e ekle. Aksi halde parent'ın tüm varyantları
                # gönderilir → Trendyol'da kardeş varyantlar zaten varsa duplicate hatası
                # oluşur ve yeni eklenmek istenen barkod da reject olur.
                # AMA: Kullanıcı stok kodu girip parent ürünü hedeflediyse, tüm varyantlar
                # gönderilmelidir (stockCode == "FCSSxxx" hiçbir varyant barkoduyla
                # eşleşmediği için aksi halde sıfır varyant gönderilir).
                product_match_by_stock = False
                if stock_codes:
                    sc_set = set(str(s) for s in stock_codes)
                    if (
                        str(product.get("stock_code") or "") in sc_set
                        or str(product.get("sku") or "") in sc_set
                        or any(str(v.get("stock_code") or "") in sc_set for v in variants)
                    ):
                        product_match_by_stock = True
                if product_match_by_stock:
                    requested_set = None  # tüm varyantları gönder
                else:
                    requested_set = set(barcodes) if barcodes else None
                for v in variants:
                    if not v.get("barcode") or v.get("barcode_uncertain"):
                        errors.append(f"{product.get('name')} - Varyant ({v.get('size') or v.get('color') or '?'}) barkodu yok / belirsiz.")
                        continue
                    if requested_set and str(v.get("barcode")) not in requested_set:
                        continue  # kullanıcı bu barkodu istemedi, atla
                    item = base_item.copy()
                    item["barcode"] = v.get("barcode")
                    # stockCode = parent stok kodu (FCSS.../FCFW...). Trendyol unique
                    # check'i barcode üzerinden yapar, stockCode aynı olabilir.
                    parent_stock = (product.get("stock_code") or product.get("sku") or "").strip()
                    v_stock = (v.get("stock_code") or v.get("sku") or "").strip()
                    # Varyantın kendi unique stock_code'u parent'tan farklıysa onu kullan,
                    # değilse parent stock_code'u gönder.
                    if v_stock and v_stock != parent_stock:
                        item["stockCode"] = v_stock
                    elif parent_stock:
                        item["stockCode"] = parent_stock
                    else:
                        item["stockCode"] = v.get("barcode")
                    item["quantity"] = int(v.get("stock", 0))
                    item["attributes"] = resolve_attributes(attributes, product, v, category, meta)
                    items_to_send.append(item)
                    
        except Exception as e:
            errors.append(f"{product.get('name')} - Hazırlama Hatası: {str(e)}")
            
    if not items_to_send:
        # Save failure log — products bulundu ama validasyondan geçemedi
        from datetime import datetime, timezone
        first_reason = errors[0] if errors else "Bilinmeyen validasyon hatası."
        msg = (
            f"{len(products)} ürün DB'de bulundu ama Trendyol'a gönderilemedi. "
            f"İlk hata: {first_reason}"
        )
        log_doc = {
            "id": generate_id(),
            "started_at": datetime.now(timezone.utc).isoformat(),
            "status": "failed",
            "products_attempted": len(products),
            "products_sent": 0,
            "batch_request_id": None,
            "errors": errors,
            "not_found_codes": not_found_codes,
            "message": msg,
        }
        await db.trendyol_sync_logs.insert_one(log_doc)
        return {
            "success": False,
            "message": msg,
            "total": len(products),
            "successful": 0,
            "failed": len(errors),
            "not_found_codes": not_found_codes,
            "errors": errors
        }
        
    try:
        from datetime import datetime, timezone
        import asyncio as _asyncio
        started_at = datetime.now(timezone.utc).isoformat()
        response = await client.create_products(items_to_send)
        batch_id = response.get("batchRequestId")
        # Trendyol returned an error response (no batchRequestId)?
        trendyol_error = None
        if not batch_id:
            # response may contain {"errors":[...]} or {"message":...}
            trendyol_error = (
                response.get("message")
                or (response.get("errors") if isinstance(response.get("errors"), list) else None)
                or str(response)[:500]
            )
            if isinstance(trendyol_error, list) and trendyol_error:
                trendyol_error = "; ".join([
                    (e.get("message") if isinstance(e, dict) else str(e)) for e in trendyol_error
                ])[:1000]

            # 🛟 FALLBACK: "Tekrarlı ürün oluşturma isteği atılamaz" (Trendyol anti-spam)
            # geldiyse, ürün zaten Trendyol'da yaşıyor demektir → create yerine
            # price-and-inventory POST (stok/fiyat) ile güncellemeyi dene. Bu endpoint
            # farklı bir kapı (inventory/...) ve recurring throttling'e takılmaz.
            # Hala başarısızsa update_products (PUT) ile son bir deneme yap.
            if trendyol_error and (
                "tekrarl" in trendyol_error.lower()
                or "duplicate request" in trendyol_error.lower()
                or "recurring" in trendyol_error.lower()
            ):
                logger.info("Trendyol create reddetti (tekrarlı). price-and-inventory fallback deneniyor.")
                try:
                    pi_items = []
                    for it in items_to_send:
                        if not it.get("barcode"):
                            continue
                        entry = {
                            "barcode": str(it.get("barcode")),
                            "quantity": int(it.get("quantity", 0)),
                        }
                        if it.get("salePrice") is not None:
                            try:
                                entry["salePrice"] = float(it["salePrice"])
                            except Exception:
                                pass
                        if it.get("listPrice") is not None:
                            try:
                                entry["listPrice"] = float(it["listPrice"])
                            except Exception:
                                pass
                        pi_items.append(entry)
                    pi_response = await client.update_price_and_inventory(pi_items) if pi_items else {}
                    pi_batch = (pi_response or {}).get("batchRequestId")
                    if pi_batch:
                        batch_id = pi_batch
                        response = pi_response
                        trendyol_error = None
                        logger.info(f"Trendyol price-and-inventory fallback başarılı: batch={pi_batch}")
                    else:
                        # Son çare: update_products PUT
                        logger.info("price-and-inventory de batch_id dönmedi. update_products fallback.")
                        upd_response = await client.update_products(items_to_send)
                        upd_batch = (upd_response or {}).get("batchRequestId")
                        if upd_batch:
                            batch_id = upd_batch
                            response = upd_response
                            trendyol_error = None
                            logger.info(f"Trendyol update_products fallback başarılı: batch={upd_batch}")
                except Exception as upd_err:
                    logger.error(f"Update fallback exception: {upd_err}")

        # 🔍 Gerçek batch sonucunu sorgula: Trendyol asenkron işliyor; 5-15sn'de tamamlanır.
        # "aktarıldı diyor ama Trendyol'da yok" tuzağını önler.
        # Trendyol BUG: bazen status=COMPLETED dönüp items[]'in işlenmiş olduğunu söyler,
        # ama failedItemCount=0 olur halbuki birkaç sn sonra items[].status FAILED olur.
        # Bu yüzden polling birkaç tur daha sürdürüyoruz: items[].status sayısı itemCount'a ulaşmalı
        # ve "tüm item'lar terminal status'ta" (SUCCESS|FAILED|COMPLETED) olmalı.
        batch_failed_items = []
        batch_success_count = 0
        batch_final_status = "INPROGRESS"

        def _norm_status(s: str) -> str:
            """Trendyol bazen 'IN_PROGRESS' bazen 'INPROGRESS' döndürüyor — normalize."""
            return (s or "").upper().replace("_", "").replace(" ", "").replace("-", "")

        if batch_id:
            for attempt in range(16):  # 16 deneme x 2.5sn = max ~40sn (ingress 60sn limitine güvenli)
                await _asyncio.sleep(2.5)
                try:
                    br = await client.get_batch_request_result(batch_id)
                    batch_final_status = (br or {}).get("status") or "INPROGRESS"
                    items = (br or {}).get("items", []) or []
                    item_count = (br or {}).get("itemCount", 0) or len(items)
                    # Tüm item'ların terminal status'ta olmasını bekle
                    terminal_set = {"SUCCESS", "FAILED", "COMPLETED"}
                    statuses_present = [(it.get("status") or "").upper() for it in items]
                    all_terminal = (
                        len(items) >= item_count
                        and all(s in terminal_set for s in statuses_present)
                    )
                    if _norm_status(batch_final_status) in ("COMPLETED", "FAILED") and all_terminal:
                        batch_failed_items = []
                        for it in items:
                            if (it.get("status") or "").upper() == "FAILED":
                                req = it.get("requestItem", {}) or {}
                                prod = req.get("product", {}) if isinstance(req, dict) else {}
                                batch_failed_items.append({
                                    "stock_code": prod.get("stockCode") or req.get("barcode") or "",
                                    "barcode": prod.get("barcode") or req.get("barcode") or "",
                                    "title": prod.get("title") or "",
                                    "reasons": it.get("failureReasons") or [],
                                })
                        # ✔ Gerçek başarı = items listesinden say (failedItemCount değil!)
                        batch_success_count = sum(1 for s in statuses_present if s == "SUCCESS")
                        break
                except Exception as poll_err:
                    logger.warning(f"Batch poll {attempt+1}/12 failed: {poll_err}")

        # 🛟 Polling timeout durumunda da elde olan son durumu işle — fallback'ler
        # böylece her senaryoda tetiklenir (kritik: aksi halde duplicate'lar sessizce kalır).
        if batch_id and not batch_failed_items and batch_success_count == 0:
            try:
                br = await client.get_batch_request_result(batch_id)
                items = (br or {}).get("items", []) or []
                if items:
                    batch_final_status = (br or {}).get("status") or batch_final_status
                    for it in items:
                        if (it.get("status") or "").upper() == "FAILED":
                            req = it.get("requestItem", {}) or {}
                            prod = req.get("product", {}) if isinstance(req, dict) else {}
                            batch_failed_items.append({
                                "stock_code": prod.get("stockCode") or req.get("barcode") or "",
                                "barcode": prod.get("barcode") or req.get("barcode") or "",
                                "title": prod.get("title") or "",
                                "reasons": it.get("failureReasons") or [],
                            })
                    batch_success_count = sum(
                        1 for it in items if (it.get("status") or "").upper() == "SUCCESS"
                    )
            except Exception as final_poll_err:
                logger.warning(f"Final batch poll failed: {final_poll_err}")

        # 🔁 SMART CONFLICT RESOLUTION:
        # - "Self duplicate" (sent_bc == conflict_bc) → PUT (update_products)
        # - "Cross duplicate" (sent_bc != conflict_bc) → eski barkodu ARCHIVE + yeni barkodu CREATE
        upsert_attempted = 0
        upsert_succeeded = 0
        upsert_failed_items = []
        upsert_batch_id = None
        upsert_final_status = None
        archived_barcodes: list = []
        archive_batch_id = None
        retry_create_batch_id = None
        retry_create_succeeded = 0

        def _is_duplicate_error(reasons):
            """Trendyol'un duplicate (çakışma) hatalarını tespit eder."""
            if not reasons:
                return False
            text = " ".join([
                (r.get("message") if isinstance(r, dict) else str(r)) for r in reasons
            ]).lower()
            patterns = [
                "aynı barkodlu",
                "ayni barkodlu",
                "aynı barkod",
                "barkod zaten",
                "stockcode zaten",
                "stok kodu zaten",
                "duplicate barcode",
                "duplicate stockcode",
                "already exist",
                "zaten mevcut",
                "zaten kayıtlı",
                "zaten kayitli",
                "productmainid",
                "bulunduğundan",
                "bulundugundan",
            ]
            return any(p in text for p in patterns)

        def _parse_conflict_barcode(reasons):
            """Trendyol hata mesajından çakışan barkodu çıkar."""
            import re
            text = " ".join([
                (r.get("message") if isinstance(r, dict) else str(r)) for r in (reasons or [])
            ])
            m = re.search(r"Barkod[:\s]+(\d{6,})", text)
            return m.group(1) if m else None

        def _is_not_found_error(reasons):
            """Trendyol'un 'ürün bulunamadı' (price-and-inventory'de Trendyol'da olmayan
            barkod) hatasını tespit eder."""
            if not reasons:
                return False
            text = " ".join([
                (r.get("message") if isinstance(r, dict) else str(r)) for r in reasons
            ]).lower()
            patterns = [
                "ürün bulunamadı",
                "urun bulunamadi",
                "product not found",
                "tedarikçi id si",
                "tedarikci id si",
            ]
            return any(p in text for p in patterns)

        # Hangi item'lar duplicate yüzünden patladı?
        duplicate_failed = [f for f in batch_failed_items if _is_duplicate_error(f.get("reasons"))]
        if batch_id and duplicate_failed:
            # Cross conflict: eski barkod ARCHIVE + yeni barkod CREATE
            # Self conflict: PUT update_products
            cross_conflicts = []   # (sent_barcode, conflict_old_barcode, item_payload)
            self_conflicts = []    # item_payload

            # items_to_send → barcode lookup
            items_by_barcode = {str(it.get("barcode") or ""): it for it in items_to_send}

            for f in duplicate_failed:
                sent_bc = str(f.get("barcode") or "")
                conflict_bc = _parse_conflict_barcode(f.get("reasons")) or ""
                payload = items_by_barcode.get(sent_bc)
                if not payload:
                    continue
                if conflict_bc and conflict_bc != sent_bc:
                    cross_conflicts.append((sent_bc, conflict_bc, payload))
                else:
                    self_conflicts.append(payload)

            # 1) CROSS CONFLICTS: önce ARCHIVE + RETRY CREATE.
            # (Önceki versiyonlarda PUT update_products deneniyordu ama Trendyol
            # stockCode altında çoklu kayıt olduğunda hayalet SUCCESS dönüyor → kaldırıldı.)
            if cross_conflicts:
                put_cross_succeeded_barcodes = set()  # şimdilik boş (PUT yok)

                # Hâlâ başarısız olan cross item'lar için archive + retry create
                remaining_cross = [
                    c for c in cross_conflicts
                    if str(c[2].get("barcode") or "") not in put_cross_succeeded_barcodes
                ]
                if remaining_cross:
                    # 🧹 AGRESİF TEMİZLİK: cross-conflict varsa, conflict'in tek barkodunu
                    # arşivlemek yetmez. Trendyol cache'inde aynı stockCode altında onlarca
                    # eski varyant olabiliyor. DB'de OLMAYAN tüm Trendyol barkodlarını
                    # bul ve arşivle. Sonra retry create yapılabilir.
                    old_barcodes_to_archive: set = set()
                    db_barcodes_per_sc: dict = {}
                    # 1) cross_conflicts'tan stockCode'ları topla, her stockCode için DB barkodlarını çek
                    stock_codes_in_play: set = set()
                    for c in remaining_cross:
                        old_barcodes_to_archive.add(c[1])
                        sc = (c[2] or {}).get("productMainId") or (c[2] or {}).get("stockCode")
                        if sc:
                            stock_codes_in_play.add(str(sc))
                    for sc in stock_codes_in_play:
                        # DB'deki barkodlar (bu stockCode için)
                        if sc not in db_barcodes_per_sc:
                            db_prod = await db.products.find_one(
                                {"$or": [{"stock_code": sc}, {"sku": sc}, {"variants.stock_code": sc}]},
                                {"_id": 0, "barcode": 1, "variants.barcode": 1, "stock_code": 1}
                            )
                            db_bcs = set()
                            if db_prod:
                                if db_prod.get("barcode"):
                                    db_bcs.add(str(db_prod.get("barcode")))
                                for v in (db_prod.get("variants") or []):
                                    if v.get("barcode"):
                                        db_bcs.add(str(v.get("barcode")))
                            db_barcodes_per_sc[sc] = db_bcs
                        # 2) Trendyol'da bu stockCode altındaki tüm barkodları çek
                        try:
                            tr_resp = await client.get_filtered_products(stock_code=sc, archived=False, size=100)
                            for p in (tr_resp or {}).get("content", []):
                                tr_bc = p.get("barcode")
                                if tr_bc and str(tr_bc) not in db_barcodes_per_sc[sc]:
                                    old_barcodes_to_archive.add(str(tr_bc))
                        except Exception as fe:
                            logger.warning(f"get_filtered_products(stock_code={sc}) failed: {fe}")
                    old_barcodes_to_archive = list(old_barcodes_to_archive)
                    logger.info(f"Trendyol CROSS-CONFLICT (deep): {len(old_barcodes_to_archive)} eski barkod arşivlenecek: {old_barcodes_to_archive[:10]}…")
                    try:
                        arch_resp = await client.archive_products(old_barcodes_to_archive) if old_barcodes_to_archive else {}
                        archive_batch_id = (arch_resp or {}).get("batchRequestId")
                        archived_barcodes = old_barcodes_to_archive
                        # Archive batch poll
                        if archive_batch_id:
                            for attempt in range(5):
                                await _asyncio.sleep(1.5)
                                try:
                                    abr = await client.get_batch_request_result(archive_batch_id)
                                    if _norm_status((abr or {}).get("status", "")) in ("COMPLETED", "FAILED"):
                                        break
                                except Exception:
                                    pass
                    except Exception as ae:
                        logger.error(f"Archive products error: {ae}")

                    # Retry CREATE with the new barcodes
                    retry_items = [c[2] for c in remaining_cross]
                    if retry_items:
                        try:
                            retry_resp = await client.create_products(retry_items)
                            retry_create_batch_id = (retry_resp or {}).get("batchRequestId")
                            if retry_create_batch_id:
                                for attempt in range(5):
                                    await _asyncio.sleep(2.0)
                                    try:
                                        rbr = await client.get_batch_request_result(retry_create_batch_id)
                                        rstatus = (rbr or {}).get("status") or "INPROGRESS"
                                        if _norm_status(rstatus) in ("COMPLETED", "FAILED"):
                                            for rit in (rbr or {}).get("items", []) or []:
                                                if (rit.get("status") or "").upper() == "FAILED":
                                                    req = rit.get("requestItem", {}) or {}
                                                    prod = req.get("product", {}) if isinstance(req, dict) else {}
                                                    upsert_failed_items.append({
                                                        "stock_code": prod.get("stockCode") or req.get("barcode") or "",
                                                        "barcode": prod.get("barcode") or req.get("barcode") or "",
                                                        "title": prod.get("title") or "",
                                                        "reasons": rit.get("failureReasons") or [],
                                                        "phase": "retry_create",
                                                    })
                                            retry_create_succeeded = max(
                                                0,
                                                (rbr or {}).get("itemCount", 0) - (rbr or {}).get("failedItemCount", 0),
                                            )
                                            upsert_succeeded += retry_create_succeeded
                                            break
                                    except Exception as rp:
                                        logger.warning(f"Retry-create poll {attempt+1}/5 failed: {rp}")
                        except Exception as re_err:
                            logger.error(f"Retry create after archive error: {re_err}")

            # 2) SELF CONFLICTS: PUT update_products + price-and-inventory POST
            # PUT mevcut ürünün kategori/attribute/image alanlarını günceller; ancak
            # stok/fiyat değerleri için Trendyol'un ayrı bir endpoint'i daha güvenilir
            # (`/inventory/.../price-and-inventory`). Stok güncellemesinin Trendyol
            # panelinde garantili yansıması için ikisini birden çalıştırıyoruz.
            if self_conflicts:
                # 2a) Önce price-and-inventory POST (stok/fiyat senkronu)
                try:
                    pi_items = []
                    for it in self_conflicts:
                        if not it.get("barcode"):
                            continue
                        entry = {
                            "barcode": str(it.get("barcode")),
                            "quantity": int(it.get("quantity", 0)),
                        }
                        if it.get("salePrice") is not None:
                            try:
                                entry["salePrice"] = float(it["salePrice"])
                            except Exception:
                                pass
                        if it.get("listPrice") is not None:
                            try:
                                entry["listPrice"] = float(it["listPrice"])
                            except Exception:
                                pass
                        pi_items.append(entry)
                    if pi_items:
                        pi_resp = await client.update_price_and_inventory(pi_items)
                        pi_batch = (pi_resp or {}).get("batchRequestId")
                        if pi_batch:
                            logger.info(f"Self-conflict price-and-inventory batch: {pi_batch}")
                            # Polling — kısa süre içinde tamamlanır
                            for attempt in range(5):
                                await _asyncio.sleep(1.5)
                                try:
                                    pbr = await client.get_batch_request_result(pi_batch)
                                    if _norm_status((pbr or {}).get("status", "")) in ("COMPLETED", "FAILED"):
                                        break
                                except Exception:
                                    pass
                except Exception as pi_err:
                    logger.warning(f"Self-conflict price-and-inventory exception: {pi_err}")

                # 2b) Sonra update_products PUT (kategori/attribute/image senkronu)
                upsert_attempted += len(self_conflicts)
                try:
                    upd_resp = await client.update_products(self_conflicts)
                    upsert_batch_id = (upd_resp or {}).get("batchRequestId")
                    if upsert_batch_id:
                        for attempt in range(8):
                            await _asyncio.sleep(2.5)
                            try:
                                ubr = await client.get_batch_request_result(upsert_batch_id)
                                upsert_final_status = (ubr or {}).get("status") or "INPROGRESS"
                                if _norm_status(upsert_final_status) in ("COMPLETED", "FAILED"):
                                    put_succeeded = max(
                                        0,
                                        (ubr or {}).get("itemCount", 0) - (ubr or {}).get("failedItemCount", 0),
                                    )
                                    for uit in (ubr or {}).get("items", []) or []:
                                        if (uit.get("status") or "").upper() == "FAILED":
                                            req = uit.get("requestItem", {}) or {}
                                            prod = req.get("product", {}) if isinstance(req, dict) else {}
                                            upsert_failed_items.append({
                                                "stock_code": prod.get("stockCode") or req.get("barcode") or "",
                                                "barcode": prod.get("barcode") or req.get("barcode") or "",
                                                "title": prod.get("title") or "",
                                                "reasons": uit.get("failureReasons") or [],
                                                "phase": "put_update",
                                            })
                                    upsert_succeeded += put_succeeded
                                    break
                            except Exception as up_poll_err:
                                logger.warning(f"Upsert PUT poll {attempt+1}/8 failed: {up_poll_err}")
                    else:
                        upd_err = (upd_resp or {}).get("message") or str(upd_resp)[:500]
                        logger.warning(f"Trendyol PUT update_products reddetti: {upd_err}")
                except Exception as up_err:
                    logger.error(f"Trendyol PUT update_products exception: {up_err}")

        # 🔁 3. FALLBACK: Price-and-inventory "ürün bulunamadı" hatası verirse
        # → Bu varyant Trendyol'da hiç yok. Asıl payload ile create_products'a gönder.
        # (Recurring fallback'te tüm itemlar price-and-inventory'ye yönlendirilince,
        # Trendyol'da olmayan varyantlar "ürün bulunamadı" der; onları create'e geri gönderelim.)
        not_found_failed = [f for f in batch_failed_items if _is_not_found_error(f.get("reasons"))]
        if not_found_failed:
            items_by_barcode = {str(it.get("barcode") or ""): it for it in items_to_send}
            nf_create_items = []
            for f in not_found_failed:
                bc = str(f.get("barcode") or "")
                payload = items_by_barcode.get(bc)
                if payload:
                    nf_create_items.append(payload)
            if nf_create_items:
                upsert_attempted += len(nf_create_items)
                try:
                    nf_resp = await client.create_products(nf_create_items)
                    nf_batch = (nf_resp or {}).get("batchRequestId")
                    if nf_batch:
                        logger.info(f"Not-found → create_products fallback batch: {nf_batch}")
                        if not retry_create_batch_id:
                            retry_create_batch_id = nf_batch
                        for attempt in range(5):
                            await _asyncio.sleep(2.0)
                            try:
                                nbr = await client.get_batch_request_result(nf_batch)
                                if _norm_status((nbr or {}).get("status", "")) in ("COMPLETED", "FAILED"):
                                    nf_succeeded = 0
                                    nf_failed_barcodes = set()
                                    for nit in (nbr or {}).get("items", []) or []:
                                        if (nit.get("status") or "").upper() == "SUCCESS":
                                            nf_succeeded += 1
                                        else:
                                            req = nit.get("requestItem", {}) or {}
                                            prod = req.get("product", {}) if isinstance(req, dict) else {}
                                            bc_n = prod.get("barcode") or req.get("barcode") or ""
                                            nf_failed_barcodes.add(str(bc_n))
                                            upsert_failed_items.append({
                                                "stock_code": prod.get("stockCode") or bc_n,
                                                "barcode": bc_n,
                                                "title": prod.get("title") or "",
                                                "reasons": nit.get("failureReasons") or [],
                                                "phase": "not_found_create",
                                            })
                                    upsert_succeeded += nf_succeeded
                                    # not-found-success olanları batch_failed_items'tan düş
                                    succeeded_barcodes = {
                                        str(f.get("barcode"))
                                        for f in not_found_failed
                                        if str(f.get("barcode")) not in nf_failed_barcodes
                                    }
                                    if succeeded_barcodes:
                                        batch_failed_items = [
                                            f for f in batch_failed_items
                                            if str(f.get("barcode")) not in succeeded_barcodes
                                        ]
                                        batch_success_count += nf_succeeded
                                    break
                            except Exception as nf_poll_err:
                                logger.warning(f"Not-found-create poll {attempt+1}/5 failed: {nf_poll_err}")
                except Exception as nf_err:
                    logger.error(f"Not-found → create_products exception: {nf_err}")

        # Upsert ile düzelen item'ları batch_failed_items'tan düş
        if upsert_succeeded > 0 and duplicate_failed:
            still_failed_keys = set()
            for f in upsert_failed_items:
                if f.get("barcode"):
                    still_failed_keys.add(str(f["barcode"]))

            def _still_failed(f):
                if not _is_duplicate_error(f.get("reasons")):
                    return True
                return str(f.get("barcode") or "") in still_failed_keys

            batch_failed_items = [f for f in batch_failed_items if _still_failed(f)]
            batch_success_count = batch_success_count + upsert_succeeded

        # Local "errors" + Trendyol API hataları + Batch failure'ları birleştir
        all_errors = list(errors)
        if trendyol_error:
            all_errors.append(f"Trendyol API: {trendyol_error}")
        for f in batch_failed_items:
            reason = "; ".join([
                (r.get("message") if isinstance(r, dict) else str(r)) for r in (f.get("reasons") or [])
            ])
            all_errors.append(f"{f.get('title') or f.get('stock_code')} [{f.get('stock_code')}]: {reason}")
        for f in upsert_failed_items:
            reason = "; ".join([
                (r.get("message") if isinstance(r, dict) else str(r)) for r in (f.get("reasons") or [])
            ])
            all_errors.append(f"[UPDATE] {f.get('title') or f.get('stock_code')} [{f.get('stock_code')}]: {reason}")

        # Final status hesapla
        if not batch_id:
            final_status = "failed"
        elif batch_failed_items and batch_success_count == 0:
            final_status = "failed"
        elif batch_failed_items:
            final_status = "partial"
        elif _norm_status(batch_final_status) == "INPROGRESS":
            final_status = "pending"  # 60sn'de tamamlanmadı, kullanıcıya logdan takibi öneriliyor
        else:
            final_status = "success" if not errors else "partial"

        log_doc = {
            "id": generate_id(),
            "started_at": started_at,
            "finished_at": datetime.now(timezone.utc).isoformat(),
            "status": final_status,
            "products_attempted": len(products),
            "products_sent": batch_success_count if batch_id else 0,
            "products_failed": len(batch_failed_items),
            "batch_request_id": batch_id,
            "batch_final_status": batch_final_status,
            "upsert_attempted": upsert_attempted,
            "upsert_succeeded": upsert_succeeded,
            "upsert_batch_id": upsert_batch_id,
            "upsert_final_status": upsert_final_status,
            "upsert_failed_items": upsert_failed_items,
            "archived_barcodes": archived_barcodes,
            "archive_batch_id": archive_batch_id,
            "retry_create_batch_id": retry_create_batch_id,
            "retry_create_succeeded": retry_create_succeeded,
            "errors": all_errors,
            "failed_items": batch_failed_items,
            "trendyol_response": response,
            "message": (
                f"{batch_success_count} ürün başarıyla aktarıldı." + (f" ({upsert_succeeded} adet UPDATE ile)." if upsert_succeeded else "") if batch_id and _norm_status(batch_final_status) == "COMPLETED" and not batch_failed_items
                else f"{batch_success_count} başarılı, {len(batch_failed_items)} HATA — detaylar loglarda." if batch_failed_items
                else f"Batch alındı, Trendyol işliyor (durum: {batch_final_status}). Loglardan takip edin." if batch_id
                else f"Trendyol kabul etmedi: {trendyol_error}"
            ),
        }
        await db.trendyol_sync_logs.insert_one(log_doc)
        # _id ekleniyor MongoDB tarafından — response'a koyma
        log_doc.pop("_id", None)

        # 🤖 OTOMATİK RETRY KUYRUĞU: "Aynı barkod var" / "ürün bulunamadı" / "Trendyol cache"
        # tipi sıkışmalarda barkodları otomatik kuyruğa ekle (saatte bir retry edilir).
        # Header'da X-Internal-Retry varsa atla (loop önlemi).
        try:
            from fastapi import Request as _Req  # noqa
            # request header'ına direkt erişim yok; bunun yerine Request inject etmiyoruz.
            # Bunun yerine sadece "stuck" tipi hataları kuyruğa al.
            stuck_barcodes = []
            for f in batch_failed_items:
                reasons_text = " ".join([
                    (r.get("message") if isinstance(r, dict) else str(r))
                    for r in (f.get("reasons") or [])
                ]).lower()
                if any(p in reasons_text for p in ["aynı barkodlu", "ayni barkodlu", "ürün bulunamadı", "urun bulunamadi", "tedarikçi id si", "tedarikci id si"]):
                    if f.get("barcode"):
                        stuck_barcodes.append(str(f["barcode"]))
            for f in upsert_failed_items:
                reasons_text = " ".join([
                    (r.get("message") if isinstance(r, dict) else str(r))
                    for r in (f.get("reasons") or [])
                ]).lower()
                if any(p in reasons_text for p in ["aynı barkodlu", "ayni barkodlu", "ürün bulunamadı", "urun bulunamadi"]):
                    if f.get("barcode"):
                        stuck_barcodes.append(str(f["barcode"]))
            if stuck_barcodes:
                from routes.trendyol_retry_queue import add_to_queue as _add_to_queue
                added = await _add_to_queue(db, stuck_barcodes, reason="Trendyol cache conflict — auto-queued")
                logger.info(f"Trendyol auto-queue: {added} barkod retry kuyruğuna eklendi")
        except Exception as q_err:
            logger.warning(f"Auto-queue failed: {q_err}")

        return {
            "success": bool(batch_id) and not batch_failed_items and _norm_status(batch_final_status) != "FAILED",
            "message": log_doc["message"],
            "total": len(products),
            "successful": batch_success_count if batch_id else 0,
            "failed": len(batch_failed_items) + (0 if batch_id else len(items_to_send)),
            "batchRequestId": batch_id,
            "batch_final_status": batch_final_status,
            "failed_items": batch_failed_items,
            "upsert_attempted": upsert_attempted,
            "upsert_succeeded": upsert_succeeded,
            "upsert_batch_id": upsert_batch_id,
            "upsert_failed_items": upsert_failed_items,
            "archived_barcodes": archived_barcodes,
            "archive_batch_id": archive_batch_id,
            "retry_create_batch_id": retry_create_batch_id,
            "errors": all_errors,
            "trendyol_response": response,
        }
    except Exception as e:
        logger.error(f"Error syncing products to Trendyol: {str(e)}")
        from datetime import datetime, timezone
        log_doc = {
            "id": generate_id(),
            "started_at": datetime.now(timezone.utc).isoformat(),
            "status": "error",
            "products_attempted": len(products),
            "products_sent": 0,
            "batch_request_id": None,
            "errors": errors + [f"API Hatası: {str(e)}"],
            "message": "Trendyol API hatası oluştu."
        }
        await db.trendyol_sync_logs.insert_one(log_doc)
        if hasattr(e, "response") and hasattr(e.response, "json"):
            raise HTTPException(status_code=500, detail=e.response.json())
        raise HTTPException(status_code=500, detail="Trendyol API ürün gönderimi başarısız oldu.")

@router.get("/trendyol/sync-logs")
async def get_trendyol_sync_logs(
    page: int = 1,
    limit: int = 20,
    current_user: dict = Depends(require_admin)
):
    """Get paginated Trendyol sync logs"""
    skip = (page - 1) * limit
    logs = await db.trendyol_sync_logs.find({}, {"_id": 0}).sort("started_at", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.trendyol_sync_logs.count_documents({})
    return {"logs": logs, "total": total, "page": page}

@router.post("/trendyol/products/inventory-sync")
async def sync_trendyol_inventory(current_user: dict = Depends(require_admin)):
    """Bulk sync stock and prices to Trendyol"""
    products = await db.products.find({"is_active": True}).to_list(length=None)
    return await _sync_inventory_to_trendyol(products)

@router.post("/trendyol/products/{product_id}/sync-inventory")
async def sync_single_product_inventory(
    product_id: str,
    current_user: dict = Depends(require_admin)
):
    """Sync stock and prices for a single product to Trendyol"""
    product = await db.products.find_one({"id": product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    return await _sync_inventory_to_trendyol([product])

async def _sync_inventory_to_trendyol(products: list):
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")
        
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )
    
    items_to_send = []
    # Global markup (Ana Ayarlar/Trendyol default_markup) uygula — cron ile manuel push
    # AYNI mantıkta çalışsın (aksi halde otomatik senkron zammı geri alıyordu).
    default_markup = float(config.get("default_markup", 0) or 0)

    for product in products:
        _mult = product.get("trendyol_multiplier")
        markup = float(_mult) if (_mult is not None and float(_mult) > 0) else default_markup
        factor = 1 + markup / 100.0
        base_price = float(product.get("price", 0) or 0) * factor
        sale_price = float(product.get("price", 0) or 0) * factor  # Trendyol: indirimsiz satis fiyati
        variants = product.get("variants", [])
        if not variants:
            if product.get("barcode"):
                items_to_send.append({
                    "barcode": product["barcode"],
                    "quantity": int(product.get("stock", 0)),
                    "salePrice": round(sale_price, 2),
                    "listPrice": round(base_price, 2)
                })
        else:
            for v in variants:
                if v.get("barcode"):
                    diff = float(v.get("price_diff") or 0)
                    items_to_send.append({
                        "barcode": v["barcode"],
                        "quantity": int(v.get("stock", 0)),
                        "salePrice": round(sale_price + diff, 2),
                        "listPrice": round(base_price + diff, 2)
                    })
    
    from datetime import datetime, timezone
    started_at = datetime.now(timezone.utc).isoformat()
    
    if not items_to_send:
        # Log failure to sync screen
        log_doc = {
            "id": generate_id(),
            "started_at": started_at,
            "status": "failed",
            "products_attempted": len(products),
            "products_sent": 0,
            "batch_request_id": None,
            "errors": ["Gönderilecek stok/fiyat bilgisi bulunamadı (barkodlar eksik olabilir)."],
            "message": "Envanter güncellemesi başarısız (geçerli barkod yok)."
        }
        await db.trendyol_sync_logs.insert_one(log_doc)
        return {"success": False, "message": "Gönderilecek stok/fiyat bilgisi bulunamadı (barkod eksik?)"}
        
    try:
        res = await client.update_price_and_inventory(items_to_send)
        batch_id = res.get("batchRequestId", "")
        
        # Log to the new sync logs screen
        log_doc = {
            "id": generate_id(),
            "started_at": started_at,
            "finished_at": datetime.now(timezone.utc).isoformat(),
            "status": "success",
            "products_attempted": len(products),
            "products_sent": len(items_to_send),
            "batch_request_id": batch_id,
            "errors": [],
            "message": "Stok ve fiyat güncellemesi başarıyla gönderildi."
        }
        await db.trendyol_sync_logs.insert_one(log_doc)
        
        return {"success": True, "message": f"{len(items_to_send)} kalem ürünün stok/fiyat bilgisi Trendyol'a gönderildi.", "batch_id": batch_id}
    except Exception as e:
        logger.error(f"Trendyol inventory sync error: {str(e)}")
        log_doc = {
            "id": generate_id(),
            "started_at": started_at,
            "status": "error",
            "products_attempted": len(products),
            "products_sent": 0,
            "batch_request_id": None,
            "errors": [f"API Hatası: {str(e)}"],
            "message": "Stok/fiyat güncellemesi sırasında hata oluştu."
        }
        await db.trendyol_sync_logs.insert_one(log_doc)
        raise HTTPException(status_code=500, detail=f"Trendyol stok/fiyat güncelleme hatası: {str(e)}")


@router.get("/products/barcode-issues")
async def list_products_with_barcode_issues(
    limit: int = 1000,
    current_user: dict = Depends(require_admin),
):
    """Barkodu eksik veya belirsiz (barcode_uncertain=True) ürünleri listeler.
    
    Bu ürünler Trendyol vb. pazaryerlerine ÖNCESİ aktarılırken sistem yanlış
    barkod (stock_code'tan kopyalanmış) gönderiyordu. Şimdi bunlar uncertain
    olarak işaretlendi ve push'tan engelleniyor — kullanıcı doğru barkodu
    Ticimax'tan alıp manuel düzeltmeli.
    """
    or_q = [
        {"barcode_uncertain": True},
        {"barcode": {"$in": [None, ""]}},
        {"variants.barcode_uncertain": True},
        {"variants.barcode": {"$in": [None, ""]}},
    ]
    items = []
    async for p in db.products.find(
        {"$or": or_q},
        {"_id": 0, "id": 1, "name": 1, "stock_code": 1, "barcode": 1, "sku": 1,
         "category_name": 1, "barcode_uncertain": 1, "variants": 1, "source": 1,
         "ticimax_id": 1, "is_active": 1},
    ).limit(int(limit) or 1000):
        bad_variants = []
        for v in p.get("variants") or []:
            if v.get("barcode_uncertain") or not v.get("barcode"):
                bad_variants.append({
                    "id": v.get("id"),
                    "stock_code": v.get("stock_code") or v.get("sku"),
                    "color": v.get("color"),
                    "size": v.get("size"),
                    "barcode": v.get("barcode"),
                    "uncertain": v.get("barcode_uncertain", False),
                })
        items.append({
            "id": p.get("id"),
            "name": p.get("name"),
            "stock_code": p.get("stock_code") or p.get("sku"),
            "main_barcode": p.get("barcode"),
            "main_barcode_uncertain": p.get("barcode_uncertain", False),
            "category_name": p.get("category_name"),
            "source": p.get("source"),
            "ticimax_id": p.get("ticimax_id"),
            "is_active": p.get("is_active"),
            "bad_variants": bad_variants,
            "bad_variant_count": len(bad_variants),
        })
    items.sort(key=lambda x: (
        not x["main_barcode_uncertain"], -(x["bad_variant_count"] or 0), x.get("name") or "",
    ))
    return {
        "success": True,
        "count": len(items),
        "items": items,
        "message": f"{len(items)} ürünün barkodu eksik/belirsiz. Manuel düzeltin.",
    }


@router.post("/products/barcode-fix")
async def fix_product_barcode(
    payload: dict,
    current_user: dict = Depends(require_admin),
):
    """Belirsiz/eksik bir ürün ya da varyant barkodunu manuel düzeltir.
    
    Payload: {
        "product_id": "...",
        "main_barcode": "...",              # (opsiyonel) ana ürün barkodu
        "variants": [                          # (opsiyonel) varyant bazlı düzeltme
            {"variant_id": "...", "barcode": "..."}
        ]
    }
    """
    pid = payload.get("product_id")
    if not pid:
        raise HTTPException(status_code=400, detail="product_id zorunlu")
    p = await db.products.find_one({"id": pid}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    set_fields = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if payload.get("main_barcode"):
        set_fields["barcode"] = str(payload["main_barcode"]).strip()
        set_fields["barcode_uncertain"] = False
        set_fields["barcode_note"] = "Manuel düzeltildi."
    if payload.get("variants"):
        existing = p.get("variants") or []
        fix_map = {str(v.get("variant_id")): str(v.get("barcode", "")).strip()
                   for v in (payload.get("variants") or []) if v.get("variant_id")}
        for v in existing:
            vid = str(v.get("id"))
            if vid in fix_map and fix_map[vid]:
                v["barcode"] = fix_map[vid]
                v["barcode_uncertain"] = False
        set_fields["variants"] = existing
    await db.products.update_one({"id": pid}, {"$set": set_fields})
    return {"success": True, "message": "Barkod güncellendi."}



@router.get("/trendyol/products/batch-status/{batch_id}")
async def get_trendyol_batch_status_v2(batch_id: str, current_user: dict = Depends(require_admin)):
    """Check the status of a batch request"""
    config = await get_trendyol_config()
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    
    try:
        client = TrendyolClient(
            supplier_id=config["supplier_id"],
            api_key=config["api_key"],
            api_secret=config["api_secret"],
            mode=config["mode"]
        )
        status = await client.get_batch_request_result(batch_id)
        return {"success": True, "status": status}
    except Exception as e:
        logger.error(f"Error fetching batch status: {str(e)}")
        raise HTTPException(status_code=500, detail="Batch durumu alınamadı.")

from pydantic import BaseModel

class TrendyolOrderPreviewReq(BaseModel):
    order_number: Optional[str] = None
    start_date_ms: Optional[int] = None
    end_date_ms: Optional[int] = None

class TrendyolOrderImportReq(BaseModel):
    orders: List[dict]

async def log_integration_event(platform: str, action: str, entity_type: str, entity_id: str, status: str, message: str, details: dict = None):
    try:
        from datetime import datetime, timezone
        await db.integration_logs.insert_one({
            "platform": platform,
            "action": action,
            "entity_type": entity_type,
            "entity_id": entity_id,
            "status": status,
            "message": message,
            "details": details or {},
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    except Exception as e:
        logger.error(f"Failed to log integration event: {str(e)}")

async def _sync_trendyol_status_passes(client, start_date_ms, end_date_ms, widen_cancel=True, statuses=("Cancelled", "Returned", "UnDelivered")):
    """İptal/iade/teslim edilemedi gibi durum değişikliklerini ayrı status
    sorgularıyla yakalar; yalnızca MEVCUT siparişlerin status alanını günceller
    (kaydı baştan ezmez). Trendyol orders ucu ~2 haftalık pencereyle sınırlıdır;
    daha eski iadeler İadeler (claims) modülünden takip edilir."""
    from datetime import datetime, timezone, timedelta
    updated = 0
    _logged_sample = False
    for st in statuses:
        # İptaller daha GENİŞ pencerede taranır: Trendyol startDate/endDate sipariş
        # TARİHİNE göre filtreler → 14 günden eski bir sipariş sonradan iptal olursa dar
        # pencereye girmez ve İptaller'e hiç düşmezdi ("bazıları düşmemiş" kök nedeni).
        # Cancelled için pencereyi 45 güne kadar geriye çek.
        _start = start_date_ms
        if st == "Cancelled" and widen_cancel:
            _wide = int((datetime.now(timezone.utc) - timedelta(days=45)).timestamp() * 1000)
            _start = min(start_date_ms, _wide) if start_date_ms else _wide
        try:
            page = 0
            while page < 25:
                resp = await client.get_orders(
                    start_date_ms=_start, end_date_ms=end_date_ms,
                    status=st, size=200, page=page,
                )
                chunk = resp.get("content", []) or []
                for t_order in chunk:
                    onum = str(t_order.get("orderNumber"))
                    mapped = map_trendyol_order(t_order)
                    _raw_status = t_order.get("status")
                    # GERÇEK SEBEP ARAYIŞI — ground truth için: iptal kaydının HAM yapısını
                    # her taramada BİR KEZ logla; Trendyol'un sebebi hangi alanda verdiğini
                    # (varsa) buradan kesin görürüz.
                    if st == "Cancelled" and not _logged_sample:
                        try:
                            import json as _json
                            _lines0 = (t_order.get("lines") or [{}])[0]
                            logger.info(
                                "[trendyol cancel sample %s] top_keys=%s line_keys=%s "
                                "cancellationReason=%r cancelReason=%r packageHistories=%r "
                                "line.cancellationReason=%r line.orderLineItemStatusName=%r",
                                onum, sorted([str(k) for k in t_order.keys()]),
                                sorted([str(k) for k in _lines0.keys()]),
                                t_order.get("cancellationReason"), t_order.get("cancelReason"),
                                str(t_order.get("packageHistories"))[:300],
                                _lines0.get("cancellationReason"), _lines0.get("orderLineItemStatusName"),
                            )
                            _logged_sample = True
                        except Exception:
                            _logged_sample = True
                    # Sebebi orders ucundaki TÜM olası alanlardan dene (sipariş + line seviyesi).
                    _l0 = (t_order.get("lines") or [{}])[0] if t_order.get("lines") else {}
                    _reason = (
                        t_order.get("cancellationReason") or t_order.get("cancelReason")
                        or t_order.get("cancellationReasonText") or t_order.get("customerCancellationReason")
                        or _l0.get("cancellationReason") or _l0.get("cancelReason") or ""
                    )
                    if isinstance(_reason, dict):
                        _reason = _reason.get("name") or _reason.get("text") or _reason.get("reason") or ""
                    _reason = str(_reason or "").strip()
                    if not _reason and st == "Cancelled":
                        # Yedek: müşteri iade/iptal claim'i varsa oradaki sebep.
                        try:
                            _cl = await db.trendyol_claims.find_one(
                                {"order_number": onum, "claim_reason": {"$nin": ["", None]}},
                                {"_id": 0, "claim_reason": 1},
                                sort=[("updated_at", -1)],
                            )
                            if _cl and _cl.get("claim_reason"):
                                _reason = _cl["claim_reason"]
                        except Exception:
                            pass
                    if not _reason:
                        _reason = ("Trendyol iptali" if st == "Cancelled"
                                   else ("Teslim edilemedi (Trendyol)" if st == "UnDelivered"
                                         else "Trendyol iadesi"))
                    _set = {
                        "status": mapped.get("status"),
                        "trendyol_status_raw": _raw_status,
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    }
                    if st == "Cancelled":
                        _set["cancel_reason"] = _reason
                        _set["cancel_source"] = "trendyol"
                    res = await db.orders.update_one(
                        {"order_number": onum, "platform": "trendyol"},
                        {"$set": _set},
                    )
                    if res.modified_count:
                        updated += 1
                    # ÖNEMLİ ("bi düşüyor bi düşmüyor" kök nedeni): sipariş bizde YOKSA
                    # (eski / hiç senkronlanmamış) eskiden update_one sessizce düşüyordu.
                    # Artık tam kaydı iptal/iade durumuyla EKLERİZ → İptaller/İadeler
                    # sayfasına garanti düşer. created_at = gerçek sipariş tarihi (orderDate)
                    # ki tarih sıralaması doğru olsun.
                    _was_new = (res.matched_count == 0)
                    if _was_new:
                        try:
                            mapped["id"] = generate_id()
                            mapped["created_at"] = _ms_to_iso(t_order.get("orderDate")) or datetime.now(timezone.utc).isoformat()
                            mapped["trendyol_status_raw"] = _raw_status
                            if st == "Cancelled":
                                mapped["cancel_reason"] = _reason
                                mapped["cancel_source"] = "trendyol"
                            await db.orders.insert_one(mapped)
                            updated += 1
                        except Exception as _ie:
                            logger.error(f"[trendyol status upsert {onum}] {_ie}")
                    # İptal senkronu → stoğu BİR KEZ geri ekle (idempotent; manuel iptalle aynı
                    # order_cancelled guard). YENİ eklenen kayıtta stok geri EKLENMEZ: bu sipariş
                    # bizde hiç olmadığı için stok daha önce DÜŞÜLMEDİ → +1 yanlış olurdu.
                    if st == "Cancelled" and not _was_new:
                        try:
                            _o = await db.orders.find_one({"order_number": onum, "platform": "trendyol"}, {"_id": 0, "id": 1, "items": 1})
                            if _o:
                                _already = await db.stock_movements.find_one({"order_id": _o.get("id"), "type": "order_cancelled"}, {"_id": 1})
                                if not _already:
                                    from routes.orders import _stock_delta_for_order
                                    _moves = await _stock_delta_for_order(_o, +1)
                                    await db.stock_movements.insert_one({
                                        "id": str(uuid.uuid4()), "type": "order_cancelled",
                                        "order_id": _o.get("id"), "order_number": onum,
                                        "items": _moves, "source": "trendyol_cancel_sync",
                                        "created_at": datetime.now(timezone.utc).isoformat(),
                                    })
                        except Exception as _re:
                            logger.error(f"[trendyol cancel restock {onum}] {_re}")
                total_pages = resp.get("totalPages") or 0
                page += 1
                if not chunk or page >= total_pages:
                    break
        except Exception as e:
            logger.error(f"[trendyol status pass {st}] {e}")
    return updated


def _ms_to_iso(v):
    """Trendyol ms-epoch (int) → ISO string; sayı değilse olduğu gibi/boş döner."""
    if v is None or v == "":
        return ""
    try:
        from datetime import datetime, timezone
        if isinstance(v, (int, float)) or (isinstance(v, str) and str(v).isdigit()):
            return datetime.fromtimestamp(int(v) / 1000, tz=timezone.utc).isoformat()
    except Exception:
        pass
    return str(v)


def map_trendyol_order(t_order: dict) -> dict:
    from datetime import datetime, timezone
    order_number = t_order.get("orderNumber")
    
    items = []
    total_price = t_order.get("totalPrice", 0)
    gross_amount = t_order.get("grossAmount", 0)
    total_discount = t_order.get("totalDiscount", 0)
    
    for line in t_order.get("lines", []):
        qty = max(line.get("quantity", 1), 1)
        line_gross = line.get("lineGrossAmount", line.get("amount", 0))
        unit_price = line_gross / qty # İndirimsiz birim fiyat
        
        # Trendyol 'price' field is actually the net discounted price
        net_price = line.get("price", line.get("lineUnitPrice", 0))
        
        discount = line.get("discount", 0)
        discount_per_item = discount / qty if discount else 0

        items.append({
            "product_id": line.get("productCode"),
            "product_name": line.get("productName"),
            "quantity": qty,
            "unit_price": unit_price, # İndirimsiz birim fiyat
            "discount_amount": discount_per_item, # Birim başına indirim
            "price": net_price, # Net ödenen birim fiyat (Faturalandırılan)
            "size": line.get("productSize", ""),
            "color": line.get("productColor", ""),
            "barcode": line.get("barcode", ""),
            "currency": line.get("currencyCode", "TRY")
        })

    shipment_address = t_order.get("shipmentAddress", {})
    invoice_address = t_order.get("invoiceAddress", {})

    # --- Kurumsal fatura alanları (VKN / vergi dairesi / ünvan) ---
    # Trendyol bu alanları invoiceAddress İÇİNDE de, sipariş ÜST SEVİYESİNDE de
    # (taxNumber / taxOffice) verebilir. Eskiden yalnızca invoiceAddress.* okunduğu
    # için müşteri kurumsal fatura talebinde VKN girse bile (üst seviye taxNumber)
    # billing'e düşmüyor → e-Arşiv VKN'siz/vergi-dairesiz kalıp Doğan reddediyordu.
    # Artık ikisi de okunur, dolu olan kullanılır.
    _cust_tax_number = (str(invoice_address.get("taxNumber") or "").strip()
                        or str(t_order.get("taxNumber") or "").strip())
    _cust_tax_office = (str(invoice_address.get("taxOffice") or "").strip()
                        or str(t_order.get("taxOffice") or "").strip())
    _cust_company = (str(invoice_address.get("company") or "").strip()
                     or str(invoice_address.get("companyName") or "").strip()
                     or str(invoice_address.get("companyTitle") or "").strip())
    _is_corporate = (len(_cust_tax_number) == 10) or bool(_cust_company)

    # --- Mikro İhracat Tespiti ---
    # KESİN gösterge: Trendyol paketindeki `micro` bayrağı (ve ETGB no/tarihi).
    # Adres ülkesine güvenilmez: mikro ihracatta kargo Türkiye içi aktarım
    # noktasına teslim edildiği için shipmentAddress.country = "Türkiye" görünür;
    # yabancı ülke invoiceAddress (alıcı) tarafındadır. Bu yüzden yedek ülke
    # kontrolü invoiceAddress üzerinden yapılır.
    delivery_type = (t_order.get("deliveryType") or "").lower()
    _buyer_country = (invoice_address.get("country") or "").strip().lower()
    _is_intl_buyer = bool(_buyer_country) and _buyer_country not in ("turkey", "türkiye", "tr")
    is_micro_export = bool(
        t_order.get("micro")                 # birincil/kesin bayrak
        or t_order.get("etgbNo")             # ETGB numarası => mikro ihracat
        or t_order.get("etgbDate")
        or "micro" in delivery_type
        or "international" in delivery_type
        or _is_intl_buyer                    # yedek: ALICI ülkesi TR dışı
    )
    
    # Pazaryeri siparisi panele DAIMA "confirmed" (Onaylandi) duser; Trendyol'un is
    # akisi durumu (Picking/Invoiced/Shipped/Delivered) bizim operasyonel durumumuzu
    # EZMEZ. Yalnizca iptal/iade/teslim-edilemedi terminal durumlari yansir.
    status_map = {
        "Cancelled": "cancelled",
        "Returned": "returned",
        "UnDelivered": "returned",
    }
    
    order_doc = {
        "order_number": str(order_number),
        "platform": "trendyol",
        "trendyol_package_id": t_order.get("id"),
        "user_id": None,
        "items": items,
        "shipping_address": {
            "first_name": shipment_address.get("firstName", "Trendyol"),
            "last_name": shipment_address.get("lastName", "Müşterisi"),
            "phone": shipment_address.get("phone", ""),
            "email": t_order.get("customerEmail", ""),
            "address": shipment_address.get("fullAddress", ""),
            "city": shipment_address.get("city", ""),
            "district": shipment_address.get("district", ""),
            "country": shipment_address.get("country", "")
        },
        "billing_address": {
            "first_name": invoice_address.get("firstName", "Trendyol"),
            "last_name": invoice_address.get("lastName", "Müşterisi"),
            "phone": invoice_address.get("phone", ""),
            "address": invoice_address.get("fullAddress", ""),
            "city": invoice_address.get("city", ""),
            "district": invoice_address.get("district", ""),
            "country": invoice_address.get("country", ""),
            "company_name": _cust_company,
            "tax_number": _cust_tax_number,
            "tax_office": _cust_tax_office,
            "is_corporate": _is_corporate
        },
        "billing_info": {
            "is_corporate": _is_corporate,
            "company_name": _cust_company,
            "tax_number": _cust_tax_number,
            "tax_office": _cust_tax_office,
            # Müşteri self-deklarasyonu burada yok; e-Fatura mükellefiyeti kesimde
            # Doğan CheckUser ile sorgulanır → mükellefse e-Fatura, değilse e-Arşiv.
            "e_invoice_user": False,
        },
        "subtotal": gross_amount if gross_amount else total_price,
        "shipping_cost": 0,
        "discount_amount": total_discount,
        "total": total_price,
        "payment_method": "marketplace",
        "payment_status": "paid",
        "status": status_map.get(t_order.get("status"), "confirmed"),
        "cargo_tracking_number": t_order.get("cargoTrackingNumber", ""),
        "cargo_tracking_link": t_order.get("cargoTrackingLink", ""),
        "cargo_provider_name": t_order.get("cargoProviderName", ""),
        "invoice_link": t_order.get("invoiceLink", ""),
        # Mikro ihracat bilgileri — ayrı faturalama/ETGB akışı için
        "is_micro_export": is_micro_export,
        "shipment_country": shipment_address.get("country", ""),
        "delivery_type": t_order.get("deliveryType", ""),
        "trendyol_customer_id": str(t_order.get("customerId") or ""),
        "trendyol_identity_number": str(t_order.get("identityNumber") or ""),
        # --- FAZ 1b: Pazaryeri / SLA bilgileri (ham .get; alan yoksa boş) ---
        "marketplace_status": t_order.get("status", ""),
        "marketplace_agreed_delivery_date": _ms_to_iso(t_order.get("agreedDeliveryDate")),
        "marketplace_estimated_delivery_start": _ms_to_iso(t_order.get("estimatedDeliveryStartDate")),
        "marketplace_estimated_delivery_end": _ms_to_iso(t_order.get("estimatedDeliveryEndDate")),
        "marketplace_last_modified": _ms_to_iso(t_order.get("lastModifiedDate")),
        "cargo_sender_number": str(t_order.get("cargoSenderNumber") or ""),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    return order_doc

@router.post("/trendyol/orders/preview")
async def preview_trendyol_orders(req: TrendyolOrderPreviewReq, current_user: dict = Depends(require_admin)):
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    client = TrendyolClient(
        supplier_id=config["supplier_id"], api_key=config["api_key"],
        api_secret=config["api_secret"], mode=config["mode"]
    )
    try:
        resp = await client.get_orders(
            start_date_ms=req.start_date_ms, 
            end_date_ms=req.end_date_ms, 
            order_number=req.order_number, 
            size=100
        )
        content = resp.get("content", [])
        return {"success": True, "orders": content}
    except Exception as e:
        logger.error(f"Error previewing Trendyol orders: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

async def _decrement_stock_for_imported_order(order_data: dict, source: str):
    """Facette stok master — yeni içe aktarılan pazaryeri/Ticimax siparişi için
    idempotent stok DÜŞÜMÜ. Aynı sipariş tekrar senkronlanırsa stock_movements
    guard'ı çift düşümü engeller. Kalemde barkod yoksa o kalem atlanır."""
    try:
        oid = order_data.get("id") or order_data.get("order_number")
        if not oid:
            return
        already = await db.stock_movements.find_one(
            {"order_id": oid, "type": "order_imported"}, {"_id": 1}
        )
        if already:
            return
        from .orders import _stock_delta_for_order  # local import — döngü önleme
        moves = await _stock_delta_for_order(order_data, -1)
        await db.stock_movements.insert_one({
            "order_id": oid,
            "order_number": order_data.get("order_number"),
            "type": "order_imported",
            "source": source,
            "delta": -1,
            "moves": moves,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    except Exception as _e:
        logger.error(f"[stock] import decrement failed ({source}): {_e}")


@router.post("/trendyol/orders/import-selected")
async def import_selected_trendyol_orders(req: TrendyolOrderImportReq, current_user: dict = Depends(require_admin)):
    try:
        from datetime import datetime, timezone
        from .deps import generate_id
        imported_count = 0
        updated_count = 0
        errors = []
        for t_order in req.orders:
            order_number = str(t_order.get("orderNumber"))
            try:
                order_data = map_trendyol_order(t_order)
                existing = await db.orders.find_one({"order_number": order_number, "platform": "trendyol"})
                if existing:
                    await db.orders.update_one({"_id": existing["_id"]}, {"$set": {k: v for k, v in order_data.items() if k != "status"}})
                    updated_count += 1
                else:
                    order_data["id"] = generate_id()
                    order_data["created_at"] = _ms_to_iso(t_order.get("orderDate")) or datetime.now(timezone.utc).isoformat()
                    await db.orders.insert_one(order_data)
                    imported_count += 1
                    await _decrement_stock_for_imported_order(order_data, "trendyol")
                await log_integration_event("trendyol", "import_order", "order", order_number, "success", "Sipariş başarıyla aktarıldı.")
            except Exception as e:
                err_msg = str(e)
                errors.append({"orderNumber": order_number, "error": err_msg})
                await log_integration_event("trendyol", "import_order", "order", order_number, "error", f"Aktarım hatası: {err_msg}", {"raw": t_order})
                
        return {"success": True, "imported": imported_count, "updated": updated_count, "errors": errors}
    except Exception as e:
        logger.error(f"Error importing selected orders: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================ HEPSİBURADA SİPARİŞLERİ (OMS) ============================
class HbOrderPreviewReq(BaseModel):
    begin_date: Optional[str] = None   # ISO: 2026-06-01T00:00:00
    end_date: Optional[str] = None
    order_number: Optional[str] = None

class HbOrderImportReq(BaseModel):
    orders: List[dict]

def _hb_g(d: dict, *keys, default=""):
    for k in keys:
        v = (d or {}).get(k)
        if v not in (None, ""):
            return v
    return default

def _hb_normalize_lines(resp):
    """OMS yanıtını kalem listesine indirger (list | {items|data|orders|content}...)."""
    if isinstance(resp, list):
        return resp
    if isinstance(resp, dict):
        for k in ("items", "data", "orders", "content", "lineItems", "details"):
            v = resp.get(k)
            if isinstance(v, list):
                return v
        if resp.get("orderNumber") or resp.get("orderId"):
            return [resp]
    return []

def _hb_group_orders(lines):
    """Kalemleri sipariş numarasına göre gruplar -> [{orderNumber, ...ortak..., lines:[...]}]."""
    groups = {}
    for ln in lines:
        no = str(_hb_g(ln, "orderNumber", "orderId", "id", default="?"))
        g = groups.get(no)
        if not g:
            g = {k: ln.get(k) for k in ("orderNumber", "orderId", "orderDate", "status",
                 "customerName", "email", "customerEmail", "totalPrice", "shippingAddress",
                 "shippingAddressDetail", "invoiceAddress", "customer") if k in ln}
            g["orderNumber"] = no
            g["lines"] = []
            groups[no] = g
        g["lines"].append(ln)
    return list(groups.values())

def _hb_money(v, default=0.0):
    """HB OMS fiyatı düz sayı VEYA {amount/value/grossAmount/...} dict olabilir → güvenle float'a indirger."""
    if isinstance(v, dict):
        for k in ("amount", "value", "grossAmount", "totalPrice", "unitPrice", "price", "paidPrice"):
            iv = v.get(k)
            if iv is not None and not isinstance(iv, dict):
                try:
                    return float(iv)
                except Exception:
                    pass
        return float(default)
    try:
        return float(v if v not in (None, "") else default)
    except Exception:
        return float(default)


def map_hepsiburada_order(o: dict) -> dict:
    from datetime import datetime, timezone
    raw_no = str(_hb_g(o, "orderNumber", "orderId", "id"))
    order_number = raw_no if raw_no.upper().startswith("HB") else f"HB{raw_no}"
    lines = o.get("lines") or o.get("items") or []
    items, subtotal = [], 0.0
    for ln in lines:
        try:
            qty = int(_hb_money(_hb_g(ln, "quantity", "qty", default=1), 1)) or 1
        except Exception:
            qty = 1
        unit = _hb_money(_hb_g(ln, "price", "unitPrice", "totalPrice", "amount", default=0))
        items.append({
            "product_id": _hb_g(ln, "merchantSku", "sku", "hbSku", "productBarcode"),
            "product_name": _hb_g(ln, "productName", "name", "lineItemName"),
            "quantity": qty, "unit_price": unit, "price": unit,
            "barcode": _hb_g(ln, "barcode", "productBarcode"),
            "size": _hb_g(ln, "size", "variantValue"), "color": _hb_g(ln, "color"),
            "currency": "TRY",
        })
        subtotal += unit * qty
    total = _hb_money(_hb_g(o, "totalPrice", "totalAmount", default=subtotal), subtotal)
    ship = o.get("shippingAddress") or o.get("shippingAddressDetail") or {}
    inv = o.get("invoiceAddress") or {}
    cust = o.get("customer") or {}
    cust_name = (_hb_g(o, "customerName") or _hb_g(cust, "name")
                 or _hb_g(ship, "name", "firstName") or "Hepsiburada Müşterisi")
    parts = str(cust_name).split(" ", 1)
    return {
        "order_number": order_number, "platform": "hepsiburada", "marketplace": "hepsiburada",
        "hepsiburada_order_number": raw_no, "user_id": None, "items": items,
        "shipping_address": {
            "first_name": parts[0] if parts else "Hepsiburada",
            "last_name": parts[1] if len(parts) > 1 else "Müşterisi",
            "phone": _hb_g(ship, "phoneNumber", "phone", "gsm"),
            "email": _hb_g(o, "email", "customerEmail"),
            "address": _hb_g(ship, "address", "addressDetail", "fullAddress", "detail"),
            "city": _hb_g(ship, "city"), "district": _hb_g(ship, "district", "town"),
            "country": _hb_g(ship, "countryCode", "country", default="TR"),
        },
        "billing_address": {
            "first_name": _hb_g(inv, "name", "firstName") or (parts[0] if parts else ""),
            "last_name": _hb_g(inv, "lastName") or (parts[1] if len(parts) > 1 else ""),
            "phone": _hb_g(inv, "phoneNumber", "phone"),
            "address": _hb_g(inv, "address", "addressDetail", "fullAddress"),
            "city": _hb_g(inv, "city"), "district": _hb_g(inv, "district", "town"),
            "country": _hb_g(inv, "countryCode", "country", default="TR"),
            "company_name": _hb_g(inv, "companyName", "company"),
            "tax_number": _hb_g(inv, "taxNumber", "vkn"), "tax_office": _hb_g(inv, "taxOffice"),
        },
        "subtotal": subtotal, "shipping_cost": 0, "discount_amount": 0, "total": total,
        "payment_method": "marketplace", "payment_status": "paid", "status": "confirmed",
        "marketplace_status": _hb_g(o, "status"), "hb_order_date": _hb_g(o, "orderDate"),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

def _hb_created_at(o):
    from datetime import datetime, timezone
    d = o.get("hb_order_date")
    if isinstance(d, (int, float)) and d > 0:
        try:
            return datetime.fromtimestamp(d / 1000, tz=timezone.utc).isoformat()
        except Exception:
            pass
    if isinstance(d, str) and len(d) >= 10 and "-" in d:
        return d
    return datetime.now(timezone.utc).isoformat()

@router.post("/hepsiburada/orders/preview")
async def preview_hepsiburada_orders(req: HbOrderPreviewReq, current_user: dict = Depends(require_admin)):
    """Hepsiburada OMS'ten geçmiş siparişleri tarih aralığı veya sipariş no ile listeler (içe aktarmadan).
    Hata durumunda 200 + {success:False, error, attempted_url} döner ki proxy/timeout mesajı maskelemesin."""
    import asyncio
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        return {"success": False, "error": err}
    try:
        oms_base = client._oms_base()
        mid = client.merchant_id
    except Exception:
        oms_base, mid = "?", "?"
    attempted = f"{oms_base}/orders/merchantid/{mid}"
    try:
        if req.order_number and req.order_number.strip():
            on = req.order_number.strip()
            attempted = f"{oms_base}/orders/merchantid/{mid}/ordernumber/{on}"
            resp = await asyncio.wait_for(asyncio.to_thread(client.get_order_by_number, on), timeout=15)
        else:
            if req.begin_date or req.end_date:
                attempted += f"?beginDate={req.begin_date}&endDate={req.end_date}&offset=0&limit=50"
            else:
                attempted += "?offset=0&limit=50"  # tarihsiz: ödemesi tamamlanmış (Open) listesi
            resp = await asyncio.wait_for(
                asyncio.to_thread(client.get_orders, req.begin_date, req.end_date, 0, 50), timeout=15)
    except asyncio.TimeoutError:
        return {"success": False,
                "error": "HB OMS 15 sn icinde yanit vermedi (oms-external-sit Railway agindan yavas/erisilemez "
                         "gorunuyor). Bu HB SIT tarafi kaynakli; 1-2 dk sonra tekrar deneyin, surekli olursa "
                         "ticket acariz.",
                "attempted_url": attempted}
    except Exception as e:
        return {"success": False, "error": str(e), "attempted_url": attempted}
    lines = _hb_normalize_lines(resp)
    grouped = _hb_group_orders(lines)
    preview = [map_hepsiburada_order(g) for g in grouped]
    nums = [p["order_number"] for p in preview]
    existing = set()
    if nums:
        async for o in db.orders.find({"order_number": {"$in": nums}, "platform": "hepsiburada"},
                                       {"_id": 0, "order_number": 1}):
            existing.add(o["order_number"])
    for p in preview:
        p["_already_imported"] = p["order_number"] in existing
    return {"success": True, "count": len(preview), "orders": grouped, "preview": preview,
            "attempted_url": attempted,
            "raw_sample": (lines[:2] if isinstance(lines, list) else lines)}

@router.get("/hepsiburada/oms-diag")
async def hepsiburada_oms_diag(on: str = "", key: str = ""):
    """GEÇİCİ TEŞHİS UCU: Railway backend'inden HB OMS-SIT'e bağlantıyı ölçer. HER ZAMAN 200 döner.
    Tarayıcı adres çubuğundan aç:  /api/integrations/hepsiburada/oms-diag?key=facette_oms_diag&on=<sipariş_no>
    Yorum: list_1.ok=true+düşük ms => OMS erişilebilir (sorun frontend/cache);
           error 'timeout' => oms-external-sit Railway ağından erişilemez/yavaş (HB SIT tarafı, ticket);
           '401/403' => OMS auth; '400' => erişim var, istek formatı."""
    if key != "facette_oms_diag":
        return {"ok": False, "error": "?key=facette_oms_diag gerekli (gecici teshis ucu)"}
    import asyncio, time
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        return {"ok": False, "stage": "client", "error": err}
    info = {"test_mode": getattr(client, "test", None)}
    try:
        info["base"] = client._oms_base(); info["mid"] = client.merchant_id
    except Exception as e:
        info["base_err"] = str(e)

    async def timed(fn, timeout):
        t0 = time.time()
        try:
            r = await asyncio.wait_for(asyncio.to_thread(fn), timeout=timeout)
            return {"ok": True, "ms": int((time.time() - t0) * 1000), "sample": str(r)[:300]}
        except asyncio.TimeoutError:
            return {"ok": False, "ms": int((time.time() - t0) * 1000),
                    "error": f"timeout>{timeout}s (OMS yanit vermedi — Railway agindan erisilemez/yavas)"}
        except Exception as e:
            return {"ok": False, "ms": int((time.time() - t0) * 1000), "error": str(e)[:300]}

    results = {"list_1": await timed(lambda: client.get_orders(None, None, 0, 1, read_timeout=10), 11)}
    on = (on or "").strip()
    if on:
        results["by_number"] = await timed(lambda: client.get_order_by_number(on), 10)
    return {"ok": True, "info": info, "results": results}

@router.get("/hepsiburada/orders/import-by-number")
async def hepsiburada_import_by_number(on: str = "", key: str = ""):
    """GEÇİCİ: Siparişi numarayla OMS'ten çekip doğrudan panele (db.orders) aktarır — frontend'e bağlı değil.
    Tarayıcıdan aç: /api/integrations/hepsiburada/orders/import-by-number?key=facette_oms_diag&on=<sipariş_no>"""
    if key != "facette_oms_diag":
        return {"ok": False, "error": "?key=facette_oms_diag gerekli"}
    on = (on or "").strip()
    if not on:
        return {"ok": False, "error": "on (siparis no) gerekli"}
    import asyncio, traceback
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        return {"ok": False, "error": err}
    try:
        resp = await asyncio.wait_for(asyncio.to_thread(client.get_order_by_number, on), timeout=12)
        lines = _hb_normalize_lines(resp)
        grouped = _hb_group_orders(lines)
        if not grouped:
            return {"ok": False, "error": "siparis bulunamadi/bos", "raw": str(resp)[:400]}
        imported = updated = 0
        out = []
        for g in grouped:
            order_data = map_hepsiburada_order(g)
            onum = order_data["order_number"]
            existing = await db.orders.find_one({"order_number": onum, "platform": "hepsiburada"})
            if existing:
                await db.orders.update_one({"_id": existing["_id"]},
                                           {"$set": {k: v for k, v in order_data.items() if k != "status"}})
                updated += 1
            else:
                order_data["id"] = generate_id()
                order_data["created_at"] = _hb_created_at(order_data)
                await db.orders.insert_one(order_data)
                imported += 1
                try:
                    await _decrement_stock_for_imported_order(order_data, "hepsiburada")
                except Exception:
                    pass  # ürün FACETTE'de yoksa stok düşmez — sipariş yine kaydedildi
            out.append({"order_number": onum, "items": len(order_data.get("items", []))})
        return {"ok": True, "imported": imported, "updated": updated, "orders": out,
                "mesaj": "Siparis(ler) panele aktarildi — Siparisler sayfasinda gorunur."}
    except Exception as e:
        return {"ok": False, "stage": "import", "error": str(e)[:400],
                "trace": traceback.format_exc()[-900:]}

@router.post("/hepsiburada/orders/create-test")
async def create_hepsiburada_test_order(payload: dict = None, current_user: dict = Depends(require_admin)):
    """SADECE TEST (SIT): oms-stub üzerinden bir test siparişi oluşturur; sonra 'Çek' ile panele alınır.
    SKU olarak senin gerçek SIT listing'lerinden bir HBSKU otomatik kullanılır (stub listing arar).
    Override: {sku:"HBV..."} / {skus:[...]} / {body:{...tam gövde...}}."""
    import asyncio
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        return {"success": False, "error": err}
    if not getattr(client, "test", True):
        return {"success": False,
                "error": "Test siparişi yalnızca SANDBOX/TEST modunda oluşturulur. Hepsiburada modunu 'Sandbox' yapın."}
    payload = payload or {}
    body = payload.get("body")
    skus = payload.get("skus") or ([payload["sku"]] if payload.get("sku") else [])

    # Tam gövde de SKU da verilmediyse: gerçek listing'lerden HBSKU çek
    if not body and not skus:
        try:
            lst = await asyncio.to_thread(client.get_listings, 0, 50)
        except Exception as e:
            return {"success": False, "error": f"Listing çekilemedi: {e}"}
        rows = lst.get("listings") if isinstance(lst, dict) else (lst if isinstance(lst, list) else [])

        def _hbsku(r):
            for k in ("hepsiburadaSku", "HepsiburadaSku", "hbSku", "hepsiburada_sku", "sku", "Sku"):
                v = (r or {}).get(k)
                if v:
                    return str(v)
            return None
        for r in (rows or []):
            s = _hbsku(r)
            if s and s not in skus:
                skus.append(s)
            if len(skus) >= 2:
                break
        if not skus:
            return {"success": False,
                    "error": "SIT kataloğunda HBSKU'lu listing bulunamadı. Önce ürünü gönderip listing oluştur "
                             "(gerekirse satışa aç) ya da gövdeye gerçek HBSKU gir.",
                    "listings_sample": (rows or [])[:1]}

    if not body:
        def _li(sku):
            return {"Sku": sku, "Quantity": 1, "Price": {"Amount": 301.4, "Currency": "TRY"},
                    "Vat": 0, "TotalPrice": {"Amount": 301.4, "Currency": "TRY"},
                    "CargoCompanyId": 1, "DeliveryOptionId": 1}
        body = {
            "Customer": {"CustomerId": "dfc8a27f-faae-4cb2-859c-8a7d50ee77be", "Name": "Test User"},
            "DeliveryAddress": {
                "AddressId": "e66765b3-d37d-488c-ae15-47051245dc9b", "Name": "Hepsiburada Office",
                "AddressDetail": "Trump Towers", "Email": "customer@hepsiburada.com.tr",
                "CountryCode": "TR", "PhoneNumber": "902822613231", "AlternatePhoneNumber": "045321538212",
                "Town": "Sisli", "District": "Kustepe", "City": "İstanbul"},
            "LineItems": [_li(s) for s in skus[:2]] or [_li(skus[0])],
        }
    attempted = f"{client.OMS_STUB_SANDBOX}/orders/merchantId/{client.merchant_id}"
    try:
        resp = await asyncio.to_thread(client.create_test_order, body)
    except Exception as e:
        return {"success": False, "error": str(e), "attempted_url": attempted, "used_skus": skus}
    return {"success": True, "order_number": (resp or {}).get("_orderNumber"),
            "used_skus": skus, "attempted_url": attempted, "response": resp}


@router.post("/hepsiburada/orders/import-selected")
async def import_selected_hepsiburada_orders(req: HbOrderImportReq, current_user: dict = Depends(require_admin)):
    """Önizlemeden seçilen Hepsiburada siparişlerini sisteme aktarır (Trendyol akışıyla aynı: stok düşümü + log)."""
    imported = updated = 0
    errors = []
    for raw in req.orders:
        order_data = map_hepsiburada_order(raw)
        on = order_data["order_number"]
        try:
            existing = await db.orders.find_one({"order_number": on, "platform": "hepsiburada"})
            if existing:
                await db.orders.update_one({"_id": existing["_id"]},
                                           {"$set": {k: v for k, v in order_data.items() if k != "status"}})
                updated += 1
            else:
                order_data["id"] = generate_id()
                order_data["created_at"] = _hb_created_at(order_data)
                await db.orders.insert_one(order_data)
                imported += 1
                await _decrement_stock_for_imported_order(order_data, "hepsiburada")
            await log_integration_event("hepsiburada", "import_order", "order", on, "success", "Sipariş aktarıldı.")
        except Exception as e:
            errors.append({"orderNumber": on, "error": str(e)})
            await log_integration_event("hepsiburada", "import_order", "order", on, "error", f"Aktarım hatası: {e}", {"raw": raw})
    return {"success": True, "imported": imported, "updated": updated, "errors": errors}


# ============================ HEPSİBURADA — LISTING / ÜRÜN / PAKET / İADE ============================
# Resmi MPOP/Listing/OMS API'sine göre tam entegrasyon. Listing (fiyat/stok) standart MPOP
# kimliğiyle; sipariş & iade OMS kimliğiyle (ayrı Basic auth gerekebilir) çalışır.

class HbBulkListingReq(BaseModel):
    items: List[dict]                 # [{merchantSku?, hepsiburadaSku?, price?, availableStock?}]
    update_stock: bool = True
    update_price: bool = True

class HbPackageReq(BaseModel):
    line_items: List[dict]            # [{id, quantity}]
    parcel_quantity: int = 1
    deci: Optional[int] = None

class HbInvoiceReq(BaseModel):
    invoice_link: str

class HbCargoReq(BaseModel):
    cargo_company_short_name: str

class HbCancelReq(BaseModel):
    reason_id: str = "83"

class HbClaimRejectReq(BaseModel):
    reason: int
    merchant_statement: Optional[str] = ""

class HbProductSyncReq(BaseModel):
    product_ids: Optional[List[str]] = None
    category_id: Optional[str] = None


async def _hb_markup() -> float:
    s = await db.settings.find_one({"id": "hepsiburada"}, {"_id": 0}) or {}
    try:
        return float(s.get("default_markup", 0) or 0)
    except Exception:
        return 0.0


# Fiyat kaynağı seçenekleri (UI + çözümleme). Varsayılan: price (mevcut davranış).
HB_PRICE_SOURCES = [
    {"value": "price", "label": "Satış Fiyatı (price) — varsayılan"},
    {"value": "auto", "label": "Otomatik (İndirimli varsa onu kullan, yoksa Satış Fiyatı)"},
    {"value": "sale_price", "label": "İndirimli Fiyat (sale_price)"},
]
_HB_PRICE_SOURCE_KEYS = {x["value"] for x in HB_PRICE_SOURCES}


async def _hb_price_source() -> str:
    s = await db.settings.find_one({"id": "hepsiburada"}, {"_id": 0}) or {}
    ps = (s.get("price_source") or "price")
    return ps if ps in _HB_PRICE_SOURCE_KEYS else "price"


def _hb_pick_base_price(obj: dict, price_source: str = "auto") -> float:
    """Ürün/varyanttan, seçili kaynağa göre temel fiyatı çöz."""
    try:
        price = float(obj.get("price", 0) or 0)
    except Exception:
        price = 0.0
    sale = obj.get("sale_price")
    try:
        sale = float(sale) if sale not in (None, "") else 0.0
    except Exception:
        sale = 0.0
    if price_source == "price":
        return price
    if price_source == "sale_price":
        return sale if sale > 0 else price
    # auto: indirimli (>0) varsa onu kullan, yoksa satış fiyatı
    return sale if sale > 0 else price


def _hb_merchant_sku(obj: dict) -> str:
    return str(obj.get("stock_code") or obj.get("barcode") or obj.get("sku") or "").strip()


def _hb_listing_items_from_product(product: dict, markup: float = 0.0, price_source: str = "auto"):
    """Yerel ürün -> HB listing kalemleri [{merchantSku, price, availableStock}].
    Fiyat, price_source'a göre (auto/price/sale_price) seçilir; markup uygulanır."""
    items = []
    base_price = _hb_pick_base_price(product, price_source)
    if markup > 0:
        base_price = base_price * (1 + markup / 100)
    variants = product.get("variants", []) or []
    if variants:
        for v in variants:
            sku = str(v.get("stock_code") or v.get("barcode") or "").strip()
            if not sku:
                continue
            # Varyantın kendi fiyatı varsa onu (kaynağa göre) baz al, yoksa ürün fiyatı + price_diff
            v_price = _hb_pick_base_price(v, price_source) if (v.get("price") or v.get("sale_price")) else 0.0
            if v_price > 0:
                p = v_price * (1 + markup / 100) if markup > 0 else v_price
            else:
                p = base_price + float(v.get("price_diff", 0) or 0)
            items.append({"merchantSku": sku, "price": round(p, 2),
                          "availableStock": int(v.get("stock", 0) or 0)})
    else:
        sku = _hb_merchant_sku(product)
        if sku:
            items.append({"merchantSku": sku, "price": round(base_price, 2),
                          "availableStock": int(product.get("stock", 0) or 0)})
    return items


async def _hb_push_stock_price(client, items, do_price=True, do_stock=True):
    """price-uploads / stock-uploads çağrılarını yapar, upload id'lerini döner."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    out = {"price_upload_id": None, "stock_upload_id": None, "errors": []}
    price_items = [{k: it[k] for k in ("merchantSku", "hepsiburadaSku", "price") if k in it and it.get(k) not in (None, "")}
                   for it in items if it.get("price") is not None]
    stock_items = [{k: it[k] for k in ("merchantSku", "hepsiburadaSku", "availableStock") if k in it and it.get(k) not in (None, "")}
                   for it in items if it.get("availableStock") is not None]
    if do_price and price_items:
        try:
            r = await asyncio.to_thread(client.update_prices, price_items)
            out["price_upload_id"] = (r or {}).get("id") if isinstance(r, dict) else None
        except HepsiburadaError as e:
            out["errors"].append(f"Fiyat: {e}")
    if do_stock and stock_items:
        try:
            r = await asyncio.to_thread(client.update_stocks, stock_items)
            out["stock_upload_id"] = (r or {}).get("id") if isinstance(r, dict) else None
        except HepsiburadaError as e:
            out["errors"].append(f"Stok: {e}")
    return out


@router.post("/hepsiburada/products/{product_id}/update-stock-price")
async def hb_update_product_stock_price(product_id: str, body: HbBulkListingReq = None,
                                        current_user: dict = Depends(require_admin)):
    """Tek ürünün stok ve fiyatını Hepsiburada listing'ine gönderir (price/stock-uploads)."""
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    markup = await _hb_markup()
    price_source = await _hb_price_source()
    items = _hb_listing_items_from_product(product, markup, price_source)
    if not items:
        raise HTTPException(status_code=400, detail="Ürünün stok kodu/barkodu bulunamadı")
    do_price = True if (body is None) else body.update_price
    do_stock = True if (body is None) else body.update_stock
    res = await _hb_push_stock_price(client, items, do_price, do_stock)
    await db.products.update_one({"id": product_id}, {"$set": {
        "hb_listing_updated": datetime.now(timezone.utc).isoformat(),
        "hb_price_upload_id": res.get("price_upload_id"),
        "hb_stock_upload_id": res.get("stock_upload_id"),
    }})
    status = "success" if not res["errors"] else "error"
    await log_integration_event("hepsiburada", "update_stock_price", "product", product_id, status,
                                f"{len(items)} kalem gönderildi" + (f" — {'; '.join(res['errors'])}" if res["errors"] else ""))
    return {"success": not res["errors"], "items_count": len(items), **res}


@router.post("/hepsiburada/categories/{category_id}/update-stock-price")
async def hb_update_category_stock_price(category_id: str, body: HbBulkListingReq = None,
                                         current_user: dict = Depends(require_admin)):
    """Bir kategorideki tüm ürünlerin stok/fiyatını Hepsiburada'ya gönderir."""
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    category = await db.categories.find_one({"id": category_id}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Kategori bulunamadı")
    products = await db.products.find({"category_name": category.get("name"), "is_active": True}, {"_id": 0}).to_list(1000)
    if not products:
        products = await db.products.find({"category_id": category_id, "is_active": True}, {"_id": 0}).to_list(1000)
    if not products:
        raise HTTPException(status_code=404, detail="Bu kategoride ürün bulunamadı")
    markup = await _hb_markup()
    price_source = await _hb_price_source()
    items = []
    for p in products:
        items.extend(_hb_listing_items_from_product(p, markup, price_source))
    if not items:
        raise HTTPException(status_code=400, detail="Ürünlerin stok kodu/barkodu bulunamadı")
    do_price = True if (body is None) else body.update_price
    do_stock = True if (body is None) else body.update_stock
    # HB tek istekte max 4000 sku — parça parça gönder
    all_res = {"price_upload_ids": [], "stock_upload_ids": [], "errors": []}
    for i in range(0, len(items), 4000):
        chunk = items[i:i + 4000]
        r = await _hb_push_stock_price(client, chunk, do_price, do_stock)
        if r.get("price_upload_id"):
            all_res["price_upload_ids"].append(r["price_upload_id"])
        if r.get("stock_upload_id"):
            all_res["stock_upload_ids"].append(r["stock_upload_id"])
        all_res["errors"].extend(r.get("errors", []))
    await log_integration_event("hepsiburada", "update_stock_price", "category", category_id,
                                "success" if not all_res["errors"] else "error",
                                f"{category.get('name')}: {len(items)} kalem")
    return {"success": not all_res["errors"], "items_count": len(items), **all_res}


@router.post("/hepsiburada/listings/update")
async def hb_update_listings_bulk(req: HbBulkListingReq, current_user: dict = Depends(require_admin)):
    """Serbest kalem listesiyle toplu fiyat/stok güncelleme.
    items: [{merchantSku?, hepsiburadaSku?, price?, availableStock?}]"""
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    if not req.items:
        raise HTTPException(status_code=400, detail="items boş")
    res = await _hb_push_stock_price(client, req.items, req.update_price, req.update_stock)
    await log_integration_event("hepsiburada", "update_listings", "bulk", str(len(req.items)),
                                "success" if not res["errors"] else "error", f"{len(req.items)} kalem")
    return {"success": not res["errors"], "items_count": len(req.items), **res}


@router.get("/hepsiburada/listings/status/{kind}/{upload_id}")
async def hb_listing_upload_status(kind: str, upload_id: str, current_user: dict = Depends(require_admin)):
    """Fiyat/stok güncelleme işlem kontrolü. kind: price | stock | inventory."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    try:
        data = await asyncio.to_thread(client.get_upload_status, kind, upload_id)
        return {"success": True, "data": data}
    except HepsiburadaError as e:
        raise HTTPException(status_code=502, detail=str(e))


@router.get("/hepsiburada/listings")
async def hb_get_listings(offset: int = 0, limit: int = 100, merchant_sku: Optional[str] = None,
                          current_user: dict = Depends(require_admin)):
    """Satıcı listing bilgilerini çeker."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    try:
        skus = [merchant_sku] if merchant_sku else None
        data = await asyncio.to_thread(client.get_listings, offset, limit, skus)
        return {"success": True, "data": data}
    except HepsiburadaError as e:
        raise HTTPException(status_code=502, detail=str(e))


@router.post("/hepsiburada/listings/{hbsku}/activate")
async def hb_activate_listing(hbsku: str, current_user: dict = Depends(require_admin)):
    """Listingi satışa açar (stok ve fiyat > 0 olmalı)."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    try:
        data = await asyncio.to_thread(client.activate_listing, hbsku)
        await log_integration_event("hepsiburada", "activate_listing", "listing", hbsku, "success", "Satışa açıldı")
        return {"success": True, "data": data}
    except HepsiburadaError as e:
        raise HTTPException(status_code=502, detail=str(e))


@router.post("/hepsiburada/listings/{hbsku}/deactivate")
async def hb_deactivate_listing(hbsku: str, current_user: dict = Depends(require_admin)):
    """Listingi satışa kapatır."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    try:
        data = await asyncio.to_thread(client.deactivate_listing, hbsku)
        await log_integration_event("hepsiburada", "deactivate_listing", "listing", hbsku, "success", "Satışa kapatıldı")
        return {"success": True, "data": data}
    except HepsiburadaError as e:
        raise HTTPException(status_code=502, detail=str(e))


# ---------------------------- ÜRÜN OLUŞTURMA (import) ----------------------------
def _hb_norm(s) -> str:
    """Türkçe-duyarsız normalize (eşleştirme için). HB'ye bağımsız."""
    import unicodedata as _u
    s = _u.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not _u.combining(c))
    s = (s.lower().replace("ı", "i").replace("ş", "s").replace("ç", "c")
         .replace("ğ", "g").replace("ü", "u").replace("ö", "o"))
    return " ".join(s.split())


def _hb_collect_local(product: dict, variant: dict | None = None) -> dict:
    """Ürün (+varsa varyant) verisinden {normalize(özellik_adı): değer} toplar.
    Tamamen yerel — hiçbir pazaryerine bağlı değil."""
    out: dict = {}

    def put(nm, vv):
        if not nm or vv in (None, ""):
            return
        k = _hb_norm(nm)
        if k and k not in out:
            out[k] = str(vv).strip()

    def walk(attrs):
        if isinstance(attrs, dict):
            for k, v in attrs.items():
                if isinstance(v, dict):
                    put(v.get("label") or v.get("name") or k, v.get("value") or v.get("attribute_value"))
                else:
                    put(k, v)
        elif isinstance(attrs, list):
            for a in attrs:
                if isinstance(a, dict):
                    put(a.get("label") or a.get("name") or a.get("type") or a.get("attribute_name"),
                        a.get("value") or a.get("attribute_value"))

    walk(product.get("attributes"))
    walk(product.get("hepsiburada_attributes"))
    if product.get("brand") or product.get("brand_name"):
        put("marka", product.get("brand") or product.get("brand_name"))
    if product.get("gender"):
        put("cinsiyet", product.get("gender"))
    if variant:
        walk(variant.get("attributes"))
        if variant.get("color"):
            put("renk", variant["color"]); put("color", variant["color"])
        if variant.get("size"):
            put("beden", variant["size"]); put("numara", variant["size"])
    return out


def _hb_local_for_attr(attr_name: str, local: dict) -> str | None:
    """HB özellik adına göre yerel değeri semantik olarak bulur (Renk/Beden/Cinsiyet/Materyal/Marka)."""
    n = _hb_norm(attr_name)
    if local.get(n):
        return local[n]
    if any(w in n for w in ("renk", "color")):
        return local.get("renk") or local.get("color") or local.get("web color")
    if any(w in n for w in ("beden", "size", "numara", "olcu", "olcusu")):
        return local.get("beden") or local.get("size") or local.get("numara")
    if "cinsiyet" in n or "gender" in n:
        return local.get("cinsiyet") or local.get("gender")
    if any(w in n for w in ("materyal", "kumas", "icerik", "material", "kumas tipi", "kumas bilgisi")):
        return (local.get("materyal") or local.get("materyal bilesimi") or local.get("kumas bilgisi")
                or local.get("kumas icerigi") or local.get("urun icerik bilgisi") or local.get("urun icerigi")
                or local.get("kumas tipi"))
    if "marka" in n or "brand" in n:
        return local.get("marka")
    return None


def _hb_value_from_name(product_name, attr: dict):
    """Ürün adındaki bir KELİME, HB özelliğinin enum değerlerinden biriyle tam eşleşirse onu döner.
    Örn: 'Mira Dantelli Mini Etek Ekru' + Renk(enum: ...,Ekru,Siyah) -> 'Ekru'. Yapısal renk
    alanı olmayan, rengi yalnız adında geçen ürünler için. Tam-kelime eşleşme (parça değil) → güvenli."""
    vals = attr.get("attributeValues") or []
    if not vals or not product_name:
        return None
    words = {_hb_norm(w) for w in re.split(r"[\s/,\.\-_()\[\]]+", str(product_name)) if w}
    words.discard("")
    if not words:
        return None
    for v in vals:
        vn = _hb_norm(v.get("name"))
        if vn and len(vn) >= 2 and vn in words:
            return v.get("name")
    return None


def _hb_resolve_value(attr: dict, raw):
    """raw değeri HB özelliğinin izin verdiği değerlerden (enum) birine çözer.
    raw bir değer ADı veya değer ID'si olabilir (Özel Değer dropdown'u id kaydediyor).
    enum değilse serbest metin döner. Uyuşmayan ve allowCustom kapalıysa None."""
    vals = attr.get("attributeValues") or []
    if not vals:
        return str(raw)  # serbest metin / varchar
    nr = _hb_norm(raw)
    if not nr:
        return None
    # 1) Değer ID eşleşmesi (Özel Değer/Listeden Seçin dropdown'u value.id kaydeder)
    for v in vals:
        vid = v.get("id")
        if vid is not None and _hb_norm(str(vid)) == nr:
            return v.get("name")
    # 2) Tam ad eşleşmesi
    for v in vals:
        if _hb_norm(v.get("name")) == nr:
            return v.get("name")
    # 3) En iyi parçalı/kelime eşleşmesi: tam-kelime içeren ve EN KISA değer tercih edilir.
    #    Böylece "Ekru" daima "Altın - Ekru"ya tercih edilir (yanlış birebir-olmayan eşleşmeyi önler).
    nr_words = set(nr.split())
    nr2 = " ".join(nr.replace("-", " ").split())
    best = None  # (skor, uzunluk, ad) — küçük skor + kısa ad daha iyi
    for v in vals:
        nm = v.get("name")
        vn = _hb_norm(nm)
        if not vn or len(vn) < 2:
            continue
        vn_words = set(vn.split())
        vn2 = " ".join(vn.replace("-", " ").split())
        score = None
        if nr in vn_words:               # raw, değerin bir kelimesi: "ekru" ∈ "Altın - Ekru"
            score = 10
        elif vn in nr_words:             # değer, raw'ın bir kelimesi: "Ekru" ∈ "Altın Ekru"
            score = 20
        elif len(nr2) >= 3 and (nr2 in vn2 or vn2 in nr2):  # gevşek substring (son çare)
            score = 100
        if score is not None:
            cand = (score, len(vn), nm)
            if best is None or cand < best:
                best = cand
    if best:
        return best[2]
    if attr.get("allowCustom"):
        return str(raw)
    return None


async def _hb_category_attributes_for(hb_cat):
    """HB kategori özelliklerini (cache → yoksa canlı) getirir. (attrs_list, error)."""
    key = int(hb_cat) if str(hb_cat).isdigit() else str(hb_cat)
    cad = await db.hepsiburada_category_attributes.find_one({"category_id": key}, {"_id": 0})
    if cad and cad.get("_v") == 8 and cad.get("attributes"):
        return cad.get("attributes") or [], None
    from .category_mapping import _fetch_hb_category_attributes
    attrs, ferr = await _fetch_hb_category_attributes(hb_cat)
    if not attrs and ferr:
        return [], ferr
    return attrs or [], None


async def _build_hb_product_item(product: dict, merchant_id: str):
    """Yerel ürün -> HB import kalem(ler)i. Liste döner (varyant başına bir kalem).

    HB'ye TAMAMEN BAĞIMSIZ: kategorinin HB API özelliklerini (zorunlu/opsiyonel + izin
    verilen değerler) alır ve her özelliği ÜRÜN VERİSİNDEN otomatik türetir; enum değerleri
    HB'nin kabul ettiği değere çözer. Kaydedilmiş attribute_mappings/value_mappings/
    default_mappings override olarak kullanılır. Çözülemeyen ZORUNLU özellik varsa kalem
    atlanır ve sebebi (hangi özellikler eksik) raporlanır → kullanıcı yalnız onları doldurur.

    Döner: (items_list, error). items_list boşsa error doludur.
    """
    cm = await db.category_mappings.find_one(
        {"marketplace": "hepsiburada", "category_id": product.get("category_id")}, {"_id": 0})
    if not cm:
        cm = await db.category_mappings.find_one(
            {"marketplace": "hepsiburada", "category_name": product.get("category_name")}, {"_id": 0})
    hb_cat = (cm or {}).get("marketplace_category_id")
    if not hb_cat:
        return [], "HB kategori eşleşmesi yok (Kategori Eşleştirme ekranından eşleyin)"

    hb_attrs_list, ferr = await _hb_category_attributes_for(hb_cat)
    if not hb_attrs_list:
        return [], f"HB kategori özellikleri çekilemedi: {ferr or 'boş'}"

    saved_maps = (cm or {}).get("attribute_mappings") or []
    map_by_attr_id = {str(m.get("mp_attr_id") or m.get("trendyol_attr_id")): m
                      for m in saved_maps if (m.get("mp_attr_id") or m.get("trendyol_attr_id"))}
    vmaps = (cm or {}).get("value_mappings") or {}
    defaults = (cm or {}).get("default_mappings") or {}
    # Global ortak-özellik default'ları (panel: "Ortak Özellikler", ör. Cinsiyet=Kadın).
    # Her kategoride geçerli; o kategorinin enum'una ada göre çözülür.
    _hbset = await db.settings.find_one({"id": "hepsiburada"}, {"_id": 0}) or {}
    gad = {_hb_norm(k): v for k, v in (_hbset.get("global_attr_defaults") or {}).items()
           if v not in (None, "")}

    brand = product.get("brand") or product.get("brand_name")
    desc = re.sub(r"<[^>]+>", " ", product.get("description") or "").strip()
    imgs = []
    for img in (product.get("images") or [])[:5]:
        u = img.get("url") if isinstance(img, dict) else img
        if u:
            imgs.append(u)
    if not imgs and product.get("image"):
        imgs.append(product["image"])

    variants = product.get("variants") or []
    targets = variants if variants else [None]
    vgroup = _hb_merchant_sku(product) or str(product.get("id") or "")
    cat_val = int(hb_cat) if str(hb_cat).isdigit() else hb_cat

    # Global "Varsayılan Alan Eşleştirme" — temel HB alanlarının ürün-kartı kaynağı / sabit değeri
    from .category_mapping import _HB_BASE_BY_KEY
    bfm = (await db.settings.find_one({"id": "hepsiburada"}, {"_id": 0}) or {}).get("base_field_mappings") or {}

    def _src_val(src, variant):
        if src == "name":
            return product.get("name")
        if src == "description":
            return desc
        if src == "stock_code":
            return ((variant or {}).get("stock_code") or (variant or {}).get("barcode")
                    or product.get("stock_code") or _hb_merchant_sku(product))
        if src == "barcode":
            return (variant or {}).get("barcode") or product.get("barcode")
        if src == "brand":
            return brand
        if src == "category_name":
            return product.get("category_name")
        if src == "price":
            return product.get("price")
        if src == "weight":
            return product.get("weight") or product.get("kg")
        if src == "images":
            return imgs
        return None

    def _base_val(key, variant):
        meta = _HB_BASE_BY_KEY.get(key, {})
        cfg = bfm.get(key) or {}
        src = cfg.get("source") or meta.get("default_source") or "__default"
        dflt = cfg.get("default") if cfg.get("default") not in (None, "") else meta.get("default_value", "")
        if src in ("__default", "__auto"):
            return dflt
        v = _src_val(src, variant)
        if v in (None, "", []):
            v = dflt
        return v

    items, errors = [], set()
    used_skus: set = set()
    hb_attrs_for_product = dict(product.get("hepsiburada_attributes") or {})

    for vi, variant in enumerate(targets):
        local = _hb_collect_local(product, variant)
        v_hb = dict((variant or {}).get("hepsiburada_attributes") or {})
        attrs: dict = {}
        missing_req = []

        for a in hb_attrs_list:
            aid = str(a.get("id"))
            aname = a.get("name")
            if not aname:
                continue
            # 1) Manuel girilmiş HB değeri (varyant > ürün)
            raw = v_hb.get(aname) or hb_attrs_for_product.get(aname)
            # 2) Kaydedilmiş attribute_mapping (local_attr → bu HB özelliği)
            if not raw:
                m = map_by_attr_id.get(aid)
                if m and m.get("local_attr"):
                    raw = local.get(_hb_norm(m["local_attr"]))
            # 3) Otomatik: HB özellik adından ürün verisini türet
            if not raw:
                raw = _hb_local_for_attr(aname, local)
            # 3b) Hâlâ yoksa, ürün adındaki bir kelime HB enum değeriyle TAM eşleşiyorsa kullan
            #     (Renk yapısal alanda değil, yalnız adında geçiyorsa: "...Etek Ekru" -> Ekru)
            if not raw:
                raw = _hb_value_from_name(product.get("name"), a)
            # 4) value_mapping çevirisi (Kırmızı↔Red gibi)
            if raw:
                mapped = vmaps.get(f"{aid}|{raw}")
                if not mapped and isinstance(vmaps.get(aid), dict):
                    mapped = vmaps[aid].get(str(raw))
                if mapped:
                    raw = mapped
            # 5) enum'a/serbest metne çöz
            if raw not in (None, ""):
                rv = _hb_resolve_value(a, raw)
                if rv not in (None, ""):
                    attrs[aname] = rv
            # 6) şirket default'u (per-kategori)
            if aname not in attrs:
                dv = defaults.get(aname) or defaults.get(aid)
                if dv not in (None, ""):
                    rv = _hb_resolve_value(a, dv)
                    if rv not in (None, ""):
                        attrs[aname] = rv
            # 6b) global ortak-özellik default'u (panel, ör. Cinsiyet=Kadın — tüm kategoriler)
            if aname not in attrs:
                gdv = gad.get(_hb_norm(aname))
                if gdv not in (None, ""):
                    rv = _hb_resolve_value(a, gdv)
                    if rv not in (None, ""):
                        attrs[aname] = rv
            if a.get("required") and aname not in attrs:
                missing_req.append(aname)

        # Taban alanlar — global "Varsayılan Alan Eşleştirme" panelinden çözülür
        # merchantSku VARYANT BAŞINA BENZERSIZ olmalı; yoksa HB tüm bedenleri tek ürüne indirger.
        sku = (str((variant or {}).get("stock_code") or "").strip()
               or str((variant or {}).get("barcode") or "").strip())
        if not sku:
            base_sku = str(_hb_merchant_sku(product) or product.get("stock_code")
                           or product.get("id") or "").strip()
            if variant is not None:
                suffix = (local.get("beden") or local.get("size") or local.get("numara")
                          or local.get("renk") or local.get("color") or "")
                suffix = _hb_norm(suffix).replace(" ", "").upper() if suffix else f"V{vi + 1}"
                sku = f"{base_sku}-{suffix}" if base_sku else (suffix or f"V{vi + 1}")
            else:
                sku = base_sku
        if not sku:
            errors.add("stok kodu/barkod yok")
            continue
        # Aynı ürün içinde çakışan merchantSku'yu benzersizleştir (HB tekilleştirmesin)
        if sku in used_skus:
            sku = f"{sku}-V{vi + 1}"
        used_skus.add(sku)
        bc = (str(_base_val("Barcode", variant) or "").strip()
              or (variant or {}).get("barcode") or product.get("barcode") or sku)
        attrs.setdefault("merchantSku", sku)
        attrs.setdefault("Barcode", bc)
        if variants:
            attrs.setdefault("VaryantGroupID", str(vgroup))
        attrs.setdefault("UrunAdi", str(_base_val("UrunAdi", variant) or product.get("name") or ""))
        attrs.setdefault("UrunAciklamasi",
                         str(_base_val("UrunAciklamasi", variant) or desc or product.get("name") or ""))
        mk = _base_val("Marka", variant)
        if mk not in (None, ""):
            attrs.setdefault("Marka", str(mk))
        gar = _base_val("GarantiSuresi", variant)
        gar_s = str(gar or "").strip()
        # HB tam sayı (ay) ister, en fazla 2 hane, >0. Geçerli sayı -> onu; "Yok"/boş/geçersiz
        # -> "24" (HB bazı kategorilerde zorunlu kılıyor; istemeyen kategoride zararsızca yok sayılır).
        attrs.setdefault("GarantiSuresi",
                         str(int(gar_s)) if (gar_s.isdigit() and 1 <= int(gar_s) <= 99) else "24")
        # KDV (zorunlu) -> HB anahtarı "tax", tam sayı. Panel/default'tan gelir (varsayılan 10).
        kdv_raw = _base_val("kdv", variant)
        kdv_s = str(kdv_raw or "").strip().replace("%", "").replace(",", ".")
        if kdv_s:
            try:
                kdv_i = int(round(float(kdv_s)))
            except Exception:
                kdv_i = None
            if kdv_i and kdv_i > 0:
                attrs.setdefault("tax", str(kdv_i))
        img_src = (bfm.get("Image") or {}).get("source") or "images"
        if img_src == "images":
            for i, u in enumerate(imgs, 1):
                attrs.setdefault(f"Image{i}", u)
        attrs.setdefault("kg", str(_base_val("kg", variant) or "1"))
        # Kategori-özelliği olmayan şirket default'larını da ekle
        for k, v in defaults.items():
            if v not in (None, "") and not str(k).isdigit():
                attrs.setdefault(k, v)

        if missing_req:
            errors.add("zorunlu HB özellikleri eksik: " + ", ".join(sorted(set(missing_req))))
            continue
        items.append({"categoryId": cat_val, "merchant": merchant_id, "attributes": attrs})

    if not items:
        return [], ("; ".join(sorted(errors)) if errors else "Gönderilebilir varyant yok")
    return items, None


@router.post("/hepsiburada/products/sync")
async def hb_sync_products(request: Request, current_user: dict = Depends(require_admin)):
    """Seçili ürünleri Hepsiburada kataloğuna gönderir (import). FilteredPushPanel sözleşmesi:
    body {stock_codes, barcodes, date_from, date_to, category_filters} → {successful, failed, ...}.
    NOT: Kategori-özellik eşleşmesi eksikse ilgili ürün atlanır; HB sandbox ile doğrulanmalıdır."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    payload = await request.json()
    query = await _build_product_query_from_payload(payload)
    products = await db.products.find(query, {"_id": 0}).to_list(length=None)
    products = _dedupe_products_by_stock_code(products)
    if not products:
        return {"successful": 0, "failed": 0, "message": "Filtreye uyan ürün bulunamadı"}
    items, skipped = [], []
    for p in products:
        built, e = await _build_hb_product_item(p, client.merchant_id)
        if e:
            skipped.append({"product_id": p.get("id"), "name": p.get("name"), "reason": e})
        else:
            items.extend(built)
    if not items:
        return {"successful": 0, "failed": len(skipped),
                "message": "Hiçbir ürün gönderilemedi (kategori/özellik eşleşmesi eksik).",
                "skipped": skipped}
    try:
        res = await asyncio.to_thread(client.create_products, items)
    except HepsiburadaError as e:
        await log_integration_event("hepsiburada", "product_import", "bulk", str(len(items)), "error", str(e))
        raise HTTPException(status_code=502, detail=str(e))
    tracking_id = (res or {}).get("trackingId") or (res or {}).get("tracking_id") or (res or {}).get("id")
    await log_integration_event("hepsiburada", "product_import", "bulk", str(tracking_id or len(items)),
                                "success", f"{len(items)} ürün gönderildi, {len(skipped)} atlandı")
    return {"successful": len(items), "failed": len(skipped), "tracking_id": tracking_id,
            "message": f"{len(items)} ürün gönderildi"
                       + (f", {len(skipped)} atlandı (eşleşme eksik)" if skipped else "")
                       + (f" · Takip: {tracking_id}" if tracking_id else ""),
            "skipped": skipped, "raw": res}


@router.post("/hepsiburada/products/validate")
async def hb_validate_products(request: Request, current_user: dict = Depends(require_admin)):
    """Aktarım öncesi DOĞRULAMA — ürünlerin HB kategori eşleşmesi + stok kodunu kontrol eder.
    Body sync ile aynı. Dönüş: {valid_count, invalid_count, results}."""
    payload = await request.json()
    query = await _build_product_query_from_payload(payload)
    products = await db.products.find(query, {"_id": 0}).to_list(length=None)
    products = _dedupe_products_by_stock_code(products)
    results = []
    valid_count = invalid_count = 0
    for p in products:
        # Gerçek gönderim mantığıyla doğrula: motor zorunlu HB özelliklerini ürün
        # verisinden türetir; türetemediği zorunluları sebep olarak raporlar.
        built, e = await _build_hb_product_item(p, "")
        if e:
            invalid_count += 1
            results.append({"product_id": p.get("id"), "name": p.get("name"),
                            "errors": [e], "variant_count": 0})
        else:
            valid_count += 1
            results.append({"product_id": p.get("id"), "name": p.get("name"),
                            "errors": [], "variant_count": len(built)})
    return {"valid_count": valid_count, "invalid_count": invalid_count, "results": results}


@router.post("/hepsiburada/products/inventory-sync")
async def hb_inventory_sync(current_user: dict = Depends(require_admin)):
    """Tüm aktif ürünlerin güncel stok+fiyatını Hepsiburada listing'ine gönderir (StockPriceUpdatePanel)."""
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    products = await db.products.find({"is_active": True}, {"_id": 0}).to_list(length=None)
    markup = await _hb_markup()
    price_source = await _hb_price_source()
    items = []
    for p in products:
        items.extend(_hb_listing_items_from_product(p, markup, price_source))
    if not items:
        return {"message": "Gönderilecek stok/fiyat kalemi bulunamadı", "items_count": 0}
    res = await _hb_push_stock_price(client, items, True, True)
    status = "success" if not res["errors"] else "error"
    await log_integration_event("hepsiburada", "inventory_sync", "bulk", str(len(items)), status,
                                f"{len(items)} kalem" + (f" — {'; '.join(res['errors'])}" if res["errors"] else ""))
    return {"message": f"{len(items)} kalem stok/fiyat gönderildi"
                       + (f" — uyarı: {'; '.join(res['errors'])}" if res["errors"] else ""),
            "items_count": len(items), **res}


@router.post("/hepsiburada/products/autofill-attributes")
async def hb_autofill_attributes(request: Request, current_user: dict = Depends(require_admin)):
    """Filtreye uyan (boş = tüm HB-eşleşmiş kategori) ürünlerin `hepsiburada_attributes`
    alanını, HB kategori özelliklerinden + ürün verisinden OTOMATİK türetip kalıcı doldurur.
    Renk/Beden varyant-bazlıdır → ürün-alanına yazılmaz (gönderimde varyanttan türetilir);
    ürün-seviyesi özellikler (Cinsiyet, Materyal, Marka, Kalıp vb.) doldurulur.
    Mevcut (manuel) değerler KORUNUR — yalnız boş alanlar doldurulur."""
    payload = await request.json()
    query = await _build_product_query_from_payload(payload)
    products = await db.products.find(query, {"_id": 0}).to_list(length=None)
    cm_list = await db.category_mappings.find(
        {"marketplace": "hepsiburada", "marketplace_category_id": {"$nin": [None, ""]}}, {"_id": 0}
    ).to_list(length=3000)
    cm_by_id = {str(c.get("category_id")): c for c in cm_list}
    cm_by_name = {(c.get("category_name") or "").strip(): c for c in cm_list}
    attr_cache: dict = {}
    skip_norm = ("renk", "color", "beden", "size", "numara")
    updated = filled = scanned = 0
    for p in products:
        cm = (cm_by_id.get(str(p.get("category_id")))
              or cm_by_name.get((p.get("category_name") or "").strip()))
        if not cm or not cm.get("marketplace_category_id"):
            continue
        scanned += 1
        hb_cat = cm["marketplace_category_id"]
        ck = str(hb_cat)
        if ck not in attr_cache:
            alist, _e = await _hb_category_attributes_for(hb_cat)
            attr_cache[ck] = alist or []
        attrs_list = attr_cache[ck]
        if not attrs_list:
            continue
        local = _hb_collect_local(p, None)
        defaults = cm.get("default_mappings") or {}
        vmaps = cm.get("value_mappings") or {}
        cur = dict(p.get("hepsiburada_attributes") or {})
        changed = False
        for a in attrs_list:
            aname = a.get("name")
            if not aname or cur.get(aname):
                continue
            if any(w in _hb_norm(aname) for w in skip_norm):
                continue  # varyant-bazlı → gönderimde türetilir
            raw = _hb_local_for_attr(aname, local)
            if raw:
                aid = str(a.get("id"))
                mapped = vmaps.get(f"{aid}|{raw}")
                if not mapped and isinstance(vmaps.get(aid), dict):
                    mapped = vmaps[aid].get(str(raw))
                if mapped:
                    raw = mapped
            if not raw:
                raw = defaults.get(aname) or defaults.get(str(a.get("id")))
            if not raw:
                continue
            rv = _hb_resolve_value(a, raw)
            if rv not in (None, ""):
                cur[aname] = rv
                changed = True
                filled += 1
        if changed:
            await db.products.update_one({"id": p["id"]}, {"$set": {"hepsiburada_attributes": cur}})
            updated += 1
    return {"success": True, "scanned": scanned, "updated_products": updated, "filled_values": filled,
            "message": f"{updated} üründe {filled} HB özelliği otomatik dolduruldu"
                       + (" · Renk/Beden gönderimde varyanttan gelir" if updated else "")}


@router.get("/hepsiburada/products/tracking/{tracking_id}")
async def hb_product_tracking(tracking_id: str, current_user: dict = Depends(require_admin)):
    """Ürün import (tracking) durumunu döner."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    try:
        data = await asyncio.to_thread(client.get_product_tracking, tracking_id)
        return {"success": True, "data": data}
    except HepsiburadaError as e:
        raise HTTPException(status_code=502, detail=str(e))


@router.get("/hepsiburada/products/by-status")
async def hb_products_by_status(product_status: str = "WAITING", task_status: bool = False,
                                page: int = 0, size: int = 100,
                                current_user: dict = Depends(require_admin)):
    """Statü bazlı ürün listesi (WAITING, MATCHED, REJECTED, CREATED ...)."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    try:
        data = await asyncio.to_thread(client.get_products_by_status, product_status, task_status, page, size)
        return {"success": True, "data": data}
    except HepsiburadaError as e:
        raise HTTPException(status_code=502, detail=str(e))


# ---------------------------- SİPARİŞ / PAKET (OMS) ----------------------------
@router.get("/hepsiburada/orders/{order_number}")
async def hb_order_detail(order_number: str, current_user: dict = Depends(require_admin)):
    """Sipariş detayını OMS'ten getirir."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    raw_no = order_number[2:] if order_number.upper().startswith("HB") else order_number
    try:
        data = await asyncio.to_thread(client.get_order_detail, raw_no)
        return {"success": True, "data": data}
    except HepsiburadaError as e:
        raise HTTPException(status_code=502, detail=str(e))


@router.get("/hepsiburada/packages")
async def hb_packages(offset: int = 0, limit: int = 100, current_user: dict = Depends(require_admin)):
    """Paket listesini döner."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    try:
        data = await asyncio.to_thread(client.get_packages, offset, limit)
        return {"success": True, "data": data}
    except HepsiburadaError as e:
        raise HTTPException(status_code=502, detail=str(e))


@router.post("/hepsiburada/packages")
async def hb_create_package(req: HbPackageReq, current_user: dict = Depends(require_admin)):
    """Kalemleri paketler (kargoya hazırlar)."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    if not req.line_items:
        raise HTTPException(status_code=400, detail="line_items boş")
    try:
        data = await asyncio.to_thread(client.package_items, req.line_items, req.parcel_quantity, req.deci)
        await log_integration_event("hepsiburada", "package", "order", str(len(req.line_items)), "success", "Paketlendi")
        return {"success": True, "data": data}
    except HepsiburadaError as e:
        raise HTTPException(status_code=502, detail=str(e))


@router.put("/hepsiburada/packages/{package_number}/invoice")
async def hb_send_invoice(package_number: str, req: HbInvoiceReq, current_user: dict = Depends(require_admin)):
    """Pakete fatura linki iletir."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    try:
        data = await asyncio.to_thread(client.send_invoice, package_number, req.invoice_link)
        await log_integration_event("hepsiburada", "send_invoice", "package", package_number, "success", "Fatura iletildi")
        return {"success": True, "data": data}
    except HepsiburadaError as e:
        raise HTTPException(status_code=502, detail=str(e))


@router.get("/hepsiburada/packages/{package_number}/label")
async def hb_cargo_label(package_number: str, fmt: str = "base64zpl", current_user: dict = Depends(require_admin)):
    """Hepsiburada kargo etiketini döner (zpl | base64zpl | png)."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    try:
        data = await asyncio.to_thread(client.get_cargo_label, package_number, fmt)
        return {"success": True, "data": data}
    except HepsiburadaError as e:
        raise HTTPException(status_code=502, detail=str(e))


@router.put("/hepsiburada/packages/{package_number}/cargo")
async def hb_change_cargo(package_number: str, req: HbCargoReq, current_user: dict = Depends(require_admin)):
    """Paketin kargo firmasını değiştirir."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    try:
        data = await asyncio.to_thread(client.change_package_cargo, package_number, req.cargo_company_short_name)
        return {"success": True, "data": data}
    except HepsiburadaError as e:
        raise HTTPException(status_code=502, detail=str(e))


@router.post("/hepsiburada/lineitems/{line_item_id}/cancel")
async def hb_cancel_line(line_item_id: str, req: HbCancelReq = None, current_user: dict = Depends(require_admin)):
    """Sipariş kalemini iptal eder (para cezasına tabidir)."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    reason = (req.reason_id if req else "83")
    try:
        data = await asyncio.to_thread(client.cancel_line_item, line_item_id, reason)
        await log_integration_event("hepsiburada", "cancel_line", "lineitem", line_item_id, "success", f"İptal (sebep {reason})")
        return {"success": True, "data": data}
    except HepsiburadaError as e:
        raise HTTPException(status_code=502, detail=str(e))


@router.post("/hepsiburada/packages/{package_number}/deliver")
async def hb_mark_delivered(package_number: str, body: dict = Body(default={}),
                            current_user: dict = Depends(require_admin)):
    """Teslim edildi bilgisi gönderir."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    try:
        data = await asyncio.to_thread(client.send_delivered, package_number,
                                       body.get("received_by"), body.get("received_date"),
                                       body.get("digital_codes"))
        return {"success": True, "data": data}
    except HepsiburadaError as e:
        raise HTTPException(status_code=502, detail=str(e))


# ---------------------------- İADE / TALEP (OMS claim) ----------------------------
@router.get("/hepsiburada/claims")
async def hb_claims(status: Optional[str] = None, offset: int = 0, limit: int = 100,
                    current_user: dict = Depends(require_admin)):
    """Talep (iade) listesini döner. status verilirse statü bazlı."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    try:
        if status:
            data = await asyncio.to_thread(client.get_claims_by_status, status, offset, limit)
        else:
            data = await asyncio.to_thread(client.get_claims, offset, limit)
        return {"success": True, "data": data}
    except HepsiburadaError as e:
        raise HTTPException(status_code=502, detail=str(e))


@router.post("/hepsiburada/claims/{claim_number}/accept")
async def hb_accept_claim(claim_number: str, current_user: dict = Depends(require_admin)):
    """Talebi (iadeyi) kabul eder."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    try:
        data = await asyncio.to_thread(client.accept_claim, claim_number)
        await log_integration_event("hepsiburada", "accept_claim", "claim", claim_number, "success", "İade kabul")
        return {"success": True, "data": data}
    except HepsiburadaError as e:
        raise HTTPException(status_code=502, detail=str(e))


@router.post("/hepsiburada/claims/{claim_number}/reject")
async def hb_reject_claim(claim_number: str, req: HbClaimRejectReq, current_user: dict = Depends(require_admin)):
    """Talebi (iadeyi) reddeder. reason: HB ret sebep kodu (int)."""
    import asyncio
    from hepsiburada_client import HepsiburadaError
    from .category_mapping import _get_hb_client
    client, err = await _get_hb_client()
    if err:
        raise HTTPException(status_code=400, detail=err)
    try:
        data = await asyncio.to_thread(client.reject_claim, claim_number, req.reason, req.merchant_statement)
        await log_integration_event("hepsiburada", "reject_claim", "claim", claim_number, "success", f"İade ret (sebep {req.reason})")
        return {"success": True, "data": data}
    except HepsiburadaError as e:
        raise HTTPException(status_code=502, detail=str(e))



@router.post("/trendyol/orders/import")
async def import_trendyol_orders(current_user: dict = Depends(require_admin)):
    """Import orders from Trendyol (Last 15 days) auto job"""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")
    
    import sys
    import os
    import time
    from datetime import datetime, timezone
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    from .deps import generate_id
    
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )
    
    # Trendyol siparis ucu en fazla ~2 hafta (14 gun) araliga izin verir.
    now = datetime.now()
    start = now - timedelta(days=14)
    end_date_ms = int(now.timestamp() * 1000)
    start_date_ms = int(start.timestamp() * 1000)
    
    imported_count = 0
    updated_count = 0
    
    try:
        # TUM sayfalari dolas - daha once tek sayfa (200) cekilip fazlasi atlaniyordu.
        content = []
        page = 0
        MAX_PAGES = 50
        while page < MAX_PAGES:
            resp = await client.get_orders(
                start_date_ms=start_date_ms, end_date_ms=end_date_ms, size=200, page=page
            )
            chunk = resp.get("content", []) or []
            content.extend(chunk)
            total_pages = resp.get("totalPages") or 0
            page += 1
            if not chunk or page >= total_pages:
                break
        
        for t_order in content:
            order_number = t_order.get("orderNumber")
            existing_order = await db.orders.find_one({"order_number": str(order_number), "platform": "trendyol"})
            try:
                order_data = map_trendyol_order(t_order)
                if existing_order:
                    await db.orders.update_one(
                        {"_id": existing_order["_id"]},
                        {"$set": {k: v for k, v in order_data.items() if k != "status"}}
                    )
                    updated_count += 1
                else:
                    order_data["id"] = generate_id()
                    order_data["created_at"] = _ms_to_iso(t_order.get("orderDate")) or datetime.now(timezone.utc).isoformat()
                    await db.orders.insert_one(order_data)
                    imported_count += 1
                    await _decrement_stock_for_imported_order(order_data, "trendyol")
            except Exception as e:
                logger.error(f"Error mapping/saving order {order_number}: {e}")
                await log_integration_event("trendyol", "auto_import", "order", str(order_number), "error", f"Otomatik aktarım hatası: {str(e)}", {"raw": t_order})
        
        try:
            await _sync_trendyol_status_passes(client, start_date_ms, end_date_ms)
        except Exception as _e:
            logger.error(f"[trendyol status passes] {_e}")

        return {
            "success": True, 
            "message": f"Trendyol'dan {imported_count} yeni sipariş aktarıldı, {updated_count} sipariş güncellendi.",
            "imported": imported_count,
            "updated": updated_count
        }
    except Exception as e:
        logger.error(f"Error importing Trendyol orders: {str(e)}")
        await log_integration_event("trendyol", "auto_import_job", "system", "-", "error", f"Toplu aktarım hatası: {str(e)}")
        raise HTTPException(status_code=500, detail="Sipariş aktarımı sırasında bir hata oluştu.")

class CategoryMappingReq(BaseModel):
    local_category_id: str
    local_name: str
    trendyol_category_id: int
    trendyol_category_name: str

@router.get("/trendyol/category-mappings")
async def get_trendyol_category_mappings(current_user: dict = Depends(require_admin)):
    """Get all local categories with their Trendyol mappings, excluding hidden ones"""
    try:
        categories = await db.categories.find({"trendyol_hidden": {"$ne": True}}).to_list(1000)
        mappings = []
        for c in categories:
            mappings.append({
                "id": str(c["_id"]) if "_id" in c else c.get("id"),
                "local_name": c.get("name"),
                "trendyol_category_id": c.get("trendyol_category_id"),
                "trendyol_category_name": c.get("trendyol_category_name"),
                "attribute_mappings": c.get("attribute_mappings", []),
                "value_mappings": c.get("value_mappings", {}),
                "default_mappings": c.get("default_mappings", {}),
                "has_children": c.get("has_children", c.get("children_count", 0) > 0),
                "is_matched": bool(c.get("trendyol_category_id"))
            })
        return {"success": True, "mappings": mappings}
    except Exception as e:
        logger.error(f"Error fetching category mappings: {e}")
        raise HTTPException(status_code=500, detail="Kategori eşleştirmeleri alınamadı.")

@router.post("/trendyol/category-mappings")
async def save_trendyol_category_mapping(req: CategoryMappingReq, current_user: dict = Depends(require_admin)):
    """Save a Trendyol category mapping to a local category"""
    from bson.objectid import ObjectId
    try:
        # Support both string UUID and ObjectId
        filter_q = {"id": req.local_category_id} if len(req.local_category_id) > 24 else {"$or": [{"id": req.local_category_id}, {"_id": ObjectId(req.local_category_id)}]}
        
        await db.categories.update_one(
            filter_q,
            {"$set": {
                "trendyol_category_id": req.trendyol_category_id,
                "trendyol_category_name": req.trendyol_category_name
            }}
        )
        return {"success": True, "message": "Eşleştirme kaydedildi"}
    except Exception as e:
        logger.error(f"Error saving category mapping: {e}")
        raise HTTPException(status_code=500, detail="Eşleştirme kaydedilemedi.")



@router.post("/trendyol/category-mappings/{local_category_id}/value-mappings")
async def save_trendyol_category_value_mappings(local_category_id: str, req: Request, current_user: dict = Depends(require_admin)):
    payload = await req.json()
    value_mappings = payload.get("value_mappings", {})
    from bson.objectid import ObjectId
    filter_q = {"id": local_category_id} if len(local_category_id) > 24 else {"$or": [{"id": local_category_id}, {"_id": ObjectId(local_category_id)}]}
    
    await db.categories.update_one(
        filter_q,
        {"$set": {"value_mappings": value_mappings}}
    )
    return {"success": True}

@router.get("/trendyol/category-values/{local_category_id}")
async def get_local_category_values(local_category_id: str, current_user: dict = Depends(require_admin)):
    from bson.objectid import ObjectId
    import re
    
    filter_q = {"id": local_category_id} if len(local_category_id) > 24 else {"$or": [{"id": local_category_id}, {"_id": ObjectId(local_category_id)}]}
    category = await db.categories.find_one(filter_q)
    
    if not category:
         raise HTTPException(status_code=404, detail="Kategori bulunamadı")
         
    # Case-insensitive category name search
    name = category.get("name")
    cat_id = category.get("id") or str(category.get("_id"))
    
    query = {
        "$or": [
            {"category_name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}},
            {"category_id": cat_id}
        ]
    }
    
    products = await db.products.find(query).to_list(None)
    
    val_map = {
        "Renk": set(),
        "Beden": set(),
        "Boy": set()
    }
    
    for p in products:
        # Pull from variants (standard for clothing)
        for v in p.get("variants", []):
            if v.get("color"):
                c = str(v["color"]).strip()
                if c and c.lower() != "none":
                    val_map["Renk"].add(c)
            if v.get("size"): 
                s = str(v["size"]).strip()
                if s and s.lower() != "none":
                    val_map["Beden"].add(s)
        
        # Pull from attributes array (from CSV imports or manual entry)
        for a in p.get("attributes", []):
            t = str(a.get("type", "")).strip()
            val = str(a.get("value", "")).strip()
            if t and val and val.lower() != "none":
                if t not in val_map:
                    val_map[t] = set()
                val_map[t].add(val)
                
    result = []
    for k, v in val_map.items():
        if v:
            # Sort values naturally if possible
            sorted_vals = sorted(list(v))
            result.append({"attribute_name": k, "values": sorted_vals})
            
    return {"success": True, "local_values": result}

@router.delete("/trendyol/category-mappings/{local_category_id}")
async def delete_trendyol_category_mapping(local_category_id: str, current_user: dict = Depends(require_admin)):
    """Hide a category from the Trendyol mappings list and clear its mapping"""
    from bson.objectid import ObjectId
    try:
        filter_q = {"id": local_category_id} if len(local_category_id) > 24 else {"$or": [{"id": local_category_id}, {"_id": ObjectId(local_category_id)}]}
        
        await db.categories.update_one(
            filter_q,
            {
                "$unset": {
                    "trendyol_category_id": "",
                    "trendyol_category_name": "",
                    "attribute_mappings": ""
                },
                "$set": {
                    "trendyol_hidden": True
                }
            }
        )
        return {"success": True, "message": "Kategori listeden kaldırıldı ve eşleştirmesi silindi"}
    except Exception as e:
        logger.error(f"Error hiding category mapping: {e}")
        raise HTTPException(status_code=500, detail="Eşleştirme silinemedi.")

@router.post("/trendyol/category-mappings/bulk-delete")
async def bulk_delete_trendyol_category_mappings(req: Request, current_user: dict = Depends(require_admin)):
    payload = await req.json()
    category_ids = payload.get("category_ids", [])
    if not category_ids:
        return {"success": True}
    from bson.objectid import ObjectId
    try:
        flat_filters = []
        for cid in category_ids:
            cid_str = str(cid)
            flat_filters.append({"id": cid_str})
            if len(cid_str) <= 24:
                try:
                    flat_filters.append({"_id": ObjectId(cid_str)})
                except Exception:
                    pass

        if not flat_filters:
            return {"success": True}

        await db.categories.update_many(
            {"$or": flat_filters},
            {
                "$unset": {
                    "trendyol_category_id": "",
                    "trendyol_category_name": "",
                    "attribute_mappings": ""
                },
                "$set": {
                    "trendyol_hidden": True
                }
            }
        )
        return {"success": True, "message": "Seçili kategoriler kaldırıldı"}
    except Exception as e:
        logger.error(f"Error bulk hiding categories: {e}")
        raise HTTPException(status_code=500, detail="Kategoriler silinemedi.")

class AttributeMapping(BaseModel):
    local_attr: str
    trendyol_attr_id: int

class AttributeMappingReq(BaseModel):
    attribute_mappings: List[AttributeMapping]
    default_mappings: Optional[dict] = {}

@router.post("/trendyol/category-mappings/{local_category_id}/attributes")
async def save_trendyol_attribute_mapping(local_category_id: str, req: AttributeMappingReq, current_user: dict = Depends(require_admin)):
    from bson.objectid import ObjectId
    try:
        filter_q = {"id": local_category_id} if len(local_category_id) > 24 else {"$or": [{"id": local_category_id}, {"_id": ObjectId(local_category_id)}]}
        
        mappings = [{"local_attr": m.local_attr, "trendyol_attr_id": m.trendyol_attr_id} for m in req.attribute_mappings]
        
        await db.categories.update_one(
            filter_q,
            {"$set": {
                "attribute_mappings": mappings,
                "default_mappings": req.default_mappings
            }}
        )
        return {"success": True, "message": "Özellik eşleştirmeleri kaydedildi"}
    except Exception as e:
        logger.error(f"Error saving attribute mapping: {e}")
        raise HTTPException(status_code=500, detail="Özellik eşleştirmeleri kaydedilemedi.")

@router.get("/trendyol/categories")
async def get_local_trendyol_categories(current_user: dict = Depends(require_admin)):
    """Fetch previously downloaded Trendyol categories for UI lists"""
    try:
        categories = await db.trendyol_categories.find({}, {"_id": 0, "id": 1, "name": 1, "subCategories": 1}).to_list(1000)
        # Flatten simple list for datalist mapping
        def flatten(cats, parent_name=""):
            result = []
            for c in cats:
                full_name = f"{parent_name} > {c['name']}" if parent_name else c["name"]
                result.append({"id": c["id"], "name": full_name})
                if c.get("subCategories"):
                    result.extend(flatten(c["subCategories"], full_name))
            return result
        flat_list = flatten(categories)
        return {"success": True, "categories": flat_list}
    except Exception as e:
        logger.error(f"Error fetching trendyol categories from db: {e}")
        return {"success": False, "categories": []}

@router.get("/integration-logs")
async def get_integration_logs(
    platform: str = Query(None),
    status: str = Query(None),
    limit: int = 50,
    current_user: dict = Depends(require_admin)
):
    """Fetch integration logs for UI"""
    try:
        query = {}
        if platform:
            query["platform"] = platform
        if status:
            query["status"] = status
            
        logs = await db.integration_logs.find(query).sort("created_at", -1).limit(limit).to_list(1000)
        # remove mongo _id
        for log in logs:
            if "_id" in log:
                log["_id"] = str(log["_id"])
        return {"success": True, "logs": logs}
    except Exception as e:
        logger.error(f"Error fetching integration logs: {e}")
        raise HTTPException(status_code=500, detail="Loglar alınamadı.")


@router.get("/trendyol/orders/label/{cargo_tracking_number}")
async def get_trendyol_cargo_label(cargo_tracking_number: str, current_user: dict = Depends(require_admin)):
    """Fetch PDF Cargo label from Trendyol"""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")
        
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    from fastapi.responses import Response
    
    try:
        client = TrendyolClient(
            supplier_id=config["supplier_id"],
            api_key=config["api_key"],
            api_secret=config["api_secret"],
            mode=config["mode"]
        )
        pdf_bytes = await client.get_cargo_label(cargo_tracking_number)
        return Response(content=pdf_bytes, media_type="application/pdf")
    except Exception as e:
        logger.error(f"Error fetching Trendyol cargo label: {str(e)}")
        raise HTTPException(status_code=500, detail="Kargo etiketi alınırken hata oluştu.")

# ==================== GIB E-FATURA ====================
GIB_MODE = os.environ.get('GIB_MODE', 'test')
GIB_USERNAME = os.environ.get('GIB_USERNAME', '')
GIB_PASSWORD = os.environ.get('GIB_PASSWORD', '')
GIB_VKN = os.environ.get('GIB_VKN', '')
GIB_COMPANY_NAME = os.environ.get('GIB_COMPANY_NAME', 'FACETTE')

def is_gib_configured():
    return bool(GIB_USERNAME and GIB_PASSWORD and GIB_VKN and len(GIB_VKN) == 10)

@router.get("/gib/status")
async def get_gib_status():
    """Get GIB integration status"""
    return {
        "configured": is_gib_configured(),
        "mode": GIB_MODE,
        "vkn": GIB_VKN[:4] + "******" if GIB_VKN else None,
        "company_name": GIB_COMPANY_NAME
    }

# ==================== TİCİMAX ====================

def _generate_slug(name: str) -> str:
    slug = name.lower()
    tr_map = {'ı':'i','ğ':'g','ü':'u','ş':'s','ö':'o','ç':'c',
              'İ':'i','Ğ':'g','Ü':'u','Ş':'s','Ö':'o','Ç':'c'}
    for tr, en in tr_map.items():
        slug = slug.replace(tr, en)
    slug = re.sub(r'[^a-z0-9\s-]', '', slug)
    slug = re.sub(r'[\s_]+', '-', slug).strip('-')
    return slug or str(uuid.uuid4())[:8]





@router.post("/site/categories/sync-missing-from-products")
async def sync_missing_categories_from_products(current_user: dict = Depends(require_admin)):
    """
    Ticimax kategori senkronizasyonunda kaçırılmış (örn. çok derin alt-kategori veya
    silinmiş ama ürünleri kalmış) kategorileri ürünlerin `category_name` alanından
    bulup yerel `categories` koleksiyonuna ekler ve ilgili ürünleri category_id ile
    günceller.

    Tipik kullanım: "Tulum kategorisi gelmemiş" gibi durumlarda; Ticimax API'sini
    tekrar çağırmadan, mevcut veriden eksiklikleri tamamlar.
    """
    # 1) Mevcut yerel kategoriler (isim → id)
    existing = {}
    async for c in db.categories.find({}, {"_id": 0, "id": 1, "name": 1}):
        nm = (c.get("name") or "").strip()
        if nm:
            existing[nm.lower()] = c.get("id")

    # 2) Ürünlerdeki tüm kategori isimleri (boş olmayanlar)
    pipeline = [
        {"$match": {"category_name": {"$nin": [None, ""]}}},
        {"$group": {"_id": "$category_name", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    product_cats = []
    async for row in db.products.aggregate(pipeline):
        product_cats.append({"name": (row["_id"] or "").strip(), "count": row["count"]})

    # 3) Eksik olanları bul → oluştur
    created = []
    relinked = 0
    for pc in product_cats:
        nm = pc["name"]
        if not nm or nm.lower() in existing:
            continue
        # Yeni kategori oluştur
        new_id = await generate_short_id("categories")
        slug = _generate_slug(nm)
        doc = {
            "id": new_id,
            "ticimax_id": None,  # Ticimax'tan gelmediği için None
            "name": nm,
            "slug": slug,
            "parent_id": None,
            "is_active": True,
            "source": "products_backfill",
            "ticimax_sub_count": 0,
            "ticimax_sira": 999,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "note": "Bu kategori Ticimax sync sırasında kaçırılmıştı; ürünlerden geri-yüklendi.",
        }
        await db.categories.insert_one(doc)
        existing[nm.lower()] = new_id
        created.append({"id": new_id, "name": nm, "product_count": pc["count"]})

    # 4) Ürünlerde category_id boş olup category_name dolu olanları bağla
    for pc in product_cats:
        cat_id = existing.get(pc["name"].lower())
        if not cat_id:
            continue
        res = await db.products.update_many(
            {"category_name": pc["name"], "$or": [{"category_id": {"$exists": False}}, {"category_id": None}, {"category_id": ""}]},
            {"$set": {"category_id": cat_id}}
        )
        relinked += res.modified_count

    return {
        "success": True,
        "created_categories": created,
        "created_count": len(created),
        "relinked_products": relinked,
        "message": f"{len(created)} kategori oluşturuldu, {relinked} ürün bağlandı.",
    }





@router.post("/site/teknik-detay/sync")
async def sync_ticimax_teknik_detay(
    use_cache: bool = Query(True, description="Cache (DB) varsa kullan, yoksa SOAP'tan çek"),
    current_user: dict = Depends(require_admin),
):
    """
    Ticimax 'Teknik Detay Özellik + Değer' master listesini çekip her ürünün
    name+description text'inde değerleri arayarak attributes alanına otomatik
    eşler. Trendyol/HB/Temu özellik formlarındaki Boy, Cep, Astar Durumu, Bel,
    Web Color, Materyal, Kalıp vs. alanları DOLDURULUR.

    use_cache=True: DB'deki master cache'i kullan (~3 sn, anında).
    use_cache=False: Ticimax SOAP'a sorgu at, master'ı yenile (~30 sn, rate limit).
    """
    import sys as _sys
    import os as _os
    _sys.path.insert(0, _os.path.dirname(_os.path.dirname(__file__)))
    from scripts.enrich_attrs_from_ticimax_master import (
        fetch_master, enrich_products, _build_value_pattern,
    )

    if use_cache:
        cached = await db.ticimax_attribute_master.find({}, {"_id": 0}).to_list(None)
        if cached:
            ozellik_map = {c["ozellik_id"]: c["ozellik_tanim"] for c in cached}
            deger_by_ozellik: dict = {}
            for c in cached:
                ozid = c["ozellik_id"]
                deger_by_ozellik[ozid] = []
                for d in c.get("degerler", []):
                    pat = _build_value_pattern(d["tanim"])
                    if pat:
                        deger_by_ozellik[ozid].append({
                            "id": d["id"], "tanim": d["tanim"], "pattern": pat,
                        })
        else:
            ozellik_map, deger_by_ozellik = fetch_master()
    else:
        ozellik_map, deger_by_ozellik = fetch_master()

    # Run enrichment (görüntülemek için stdout yakalamadan; loglar zaten)
    # `enrich_products` is async-safe; just await it directly via patched runner:
    total = 0
    enriched = 0
    added_keys: dict = {}

    prods = await db.products.find(
        {"source": {"$in": ["xml_feed", "ticimax", "csv_xml_merge"]}},
        {"_id": 0, "id": 1, "name": 1, "description": 1, "attributes": 1,
         "hepsiburada_attributes": 1, "temu_attributes": 1},
    ).to_list(None)

    import re as _re
    import unicodedata as _u

    def _norm(s: str) -> str:
        s = _u.normalize("NFKD", s or "")
        s = "".join(c for c in s if not _u.combining(c))
        s = (s.lower()
               .replace("ı", "i").replace("ş", "s").replace("ç", "c")
               .replace("ğ", "g").replace("ü", "u").replace("ö", "o"))
        return _re.sub(r"\s+", " ", s).strip()

    for p in prods:
        total += 1
        name = p.get("name") or ""
        desc = _re.sub(r"<[^>]+>", " ", p.get("description") or "")
        text = _norm(name + " " + desc)

        existing = p.get("attributes") or {}
        if isinstance(existing, list):
            new_attrs = {}
            for item in existing:
                if isinstance(item, dict) and item.get("name"):
                    new_attrs[item["name"]] = str(item.get("value", ""))
            existing = new_attrs

        hb = p.get("hepsiburada_attributes") or {}
        temu = p.get("temu_attributes") or {}

        added = False
        for ozid, tanim in ozellik_map.items():
            matched = None
            for d in deger_by_ozellik.get(ozid, []):
                if d["pattern"].search(text):
                    matched = d
                    break
            if not matched:
                continue
            if existing.get(tanim):
                continue
            value = matched["tanim"]
            existing[tanim] = value
            if not hb.get(tanim):
                hb[tanim] = value
            if not temu.get(tanim):
                temu[tanim] = value
            added_keys[tanim] = added_keys.get(tanim, 0) + 1
            added = True

        await db.products.update_one(
            {"id": p["id"]},
            {"$set": {
                "attributes": existing,
                "hepsiburada_attributes": hb,
                "temu_attributes": temu,
            }},
        )
        if added:
            enriched += 1

    return {
        "success": True,
        "total_products": total,
        "enriched_products": enriched,
        "added_by_attribute": added_keys,
        "ozellik_count": len(ozellik_map),
        "message": f"{enriched}/{total} ürüne otomatik teknik detay eşlendi.",
    }






@router.post("/rooftr/orders/import")
async def import_ticimax_orders(
    limit: int = Query(200, ge=1, le=2000),
    days: int = Query(365, ge=1, le=3650, description="Son kaç günün siparişleri çekilsin"),
    exclude_marketplace: bool = Query(False, description="True ise Trendyol/HB/N11 vb. pazaryeri siparişleri hariç tutulur"),
    only_with_phone: bool = Query(False, description="True ise telefon numarası olmayan siparişler atlanır"),
    pages: int = Query(20, ge=1, le=100, description="Kaç sayfa çekilecek"),
    current_user: dict = Depends(require_admin)
):
    # [ticimax-off 2026-06-22] Ticimax SOAP entegrasyonu kapatildi; bu uc devre disi.
    return {"success": False, "message": "Ticimax siparis cekme kapatildi. Site siparisleri React/iyzico checkout'tan gelir."}


@router.post("/ticimax/orders/backfill")
async def backfill_broken_ticimax_orders(
    limit: int = Query(1000, ge=1, le=5000, description="En fazla kaç bozuk sipariş düzeltilsin"),
    days: int = Query(365, ge=1, le=3650, description="Son kaç günü tara"),
    pages: int = Query(20, ge=1, le=100, description="Kaç sayfa Ticimax'tan çekilsin"),
    page_size: int = Query(100, ge=50, le=200),
    items_chunk: int = Query(40, ge=0, le=300, description="Her çağrıda kaç eski sipariş için ürün listesi çekilsin (0=atla)"),
    current_user: dict = Depends(require_admin)
):
    # [ticimax-off 2026-06-22] Ticimax SOAP entegrasyonu kapatildi; bu uc devre disi.
    return {"success": False, "message": "Ticimax backfill kapatildi (entegrasyon sonlandirildi)."}







@router.post("/rooftr/products/upload-excel")
async def upload_rooftr_products_excel(
    file: UploadFile = File(..., description="TicimaxExport .xls/.xlsx dosyası"),
    default_stock: int = Query(5, description="Excel'de stok yoksa varsayılan stok adedi"),
    current_user: dict = Depends(require_admin),
):
    """Ticimax ürün Excel'ini (drag-drop UI'dan) yükleyip TAM resync yapar.

    `scripts/ticimax_full_resync.py` ile aynı mantık: URUNKARTIID'e göre gruplar,
    DB'de eşleştirir (urun_karti_id > stock_code+color > name+color), eşleşeni günceller,
    eşleşmeyeni yeni ürün olarak ekler.

    Beklenen sütunlar: URUNKARTIID, URUNID, STOKKODU, BARKOD, URUNADI, ACIKLAMA,
    BREADCRUMBKAT, TEDARIKCI, ALISFIYATI, SATISFIYATI, INDIRIMLIFIYAT, UYETIPIFIYAT1,
    KDVORANI, RENK, BEDEN
    """
    import tempfile
    import unicodedata
    from uuid import uuid4
    from fastapi.concurrency import run_in_threadpool

    filename = (file.filename or "").lower()
    if not (filename.endswith(".xls") or filename.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="Sadece .xls veya .xlsx dosyası yükleyin")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Dosya boş")

    suffix = ".xlsx" if filename.endswith(".xlsx") else ".xls"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(raw)
    tmp.close()
    tmp_path = tmp.name

    def _slugify(text):
        t = unicodedata.normalize("NFKD", str(text or ""))
        t = "".join(c for c in t if not unicodedata.combining(c))
        t = re.sub(r"[^a-zA-Z0-9]+", "-", t).strip("-").lower()
        return t[:200] or "urun"

    def _read_df():
        import pandas as pd
        engine = "openpyxl" if suffix == ".xlsx" else None
        try:
            return pd.read_excel(tmp_path, engine=engine)
        except Exception:
            # .xls aslında xlsx içeriği olabilir → openpyxl dene, yoksa xlrd
            try:
                return pd.read_excel(tmp_path, engine="openpyxl")
            except Exception:
                return pd.read_excel(tmp_path, engine="xlrd")

    try:
        df = await run_in_threadpool(_read_df)
    except Exception as e:
        os.unlink(tmp_path)
        raise HTTPException(status_code=400, detail=f"Excel okunamadı: {e}")

    required_cols = {"URUNKARTIID", "URUNADI", "BARKOD"}
    missing = required_cols - set(df.columns)
    if missing:
        os.unlink(tmp_path)
        raise HTTPException(status_code=400, detail=f"Eksik sütunlar: {', '.join(missing)}")

    import pandas as pd

    def _f(x, default=0.0):
        try:
            if pd.isna(x):
                return default
            return float(x)
        except Exception:
            return default

    def _s(x, default=""):
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return default
        return str(x).strip()

    def _int(x, default=0):
        try:
            if pd.isna(x):
                return default
            return int(x)
        except Exception:
            return default

    def _has(col):
        return col in df.columns

    def _cell(row, col, default=""):
        return _s(row[col]) if _has(col) else default

    def _category_from_breadcrumb(crumb):
        if not crumb:
            return ""
        parts = [p.strip() for p in crumb.split(">") if p.strip()]
        return parts[-1] if parts else ""

    sys_cats = await db.categories.find({}, {"_id": 0}).to_list(None)
    cat_by_name = {(c.get("name") or "").strip().lower(): c for c in sys_cats}

    stats = {
        "parents_in_excel": 0,
        "parents_updated_db": 0,
        "parents_created_new": 0,
        "variants_total": 0,
        "errors": [],
    }

    by_kart = df.groupby("URUNKARTIID", sort=False)
    for kart_id_raw, grp in by_kart:
        try:
            stats["parents_in_excel"] += 1
            # URUNKARTIID numerikse int-string normalize; değilse ham string koru
            try:
                if pd.isna(kart_id_raw):
                    kart_id = ""
                else:
                    kart_id = str(int(float(kart_id_raw)))
            except (ValueError, TypeError):
                kart_id = _s(kart_id_raw)
            first = grp.iloc[0]

            urun_adi = _cell(first, "URUNADI")
            renk = _cell(first, "RENK")
            base_name = urun_adi
            if renk and base_name.lower().endswith(" " + renk.lower()):
                base_name = base_name[: -(len(renk) + 1)].strip()

            list_price = _f(first["SATISFIYATI"]) if _has("SATISFIYATI") else 0.0
            sale_price = (_f(first["INDIRIMLIFIYAT"]) if _has("INDIRIMLIFIYAT") else 0.0) or list_price
            member_price_1 = (_f(first["UYETIPIFIYAT1"]) if _has("UYETIPIFIYAT1") else 0.0) or list_price
            cost_price = _f(first["ALISFIYATI"]) if _has("ALISFIYATI") else 0.0
            vat_rate = _f(first["KDVORANI"], 10) if _has("KDVORANI") else 10
            vendor = _cell(first, "TEDARIKCI", "FACETTE")
            description = _cell(first, "ACIKLAMA")
            breadcrumb = _cell(first, "BREADCRUMBKAT")
            category_leaf = _category_from_breadcrumb(breadcrumb)
            cat_doc = cat_by_name.get(category_leaf.lower()) if category_leaf else None
            parent_stock_code = _cell(first, "STOKKODU")

            variants = []
            for _, row in grp.iterrows():
                v = {
                    "size": (_cell(row, "BEDEN").upper() or "STD"),
                    "color": (_cell(row, "RENK").title() or renk.title()),
                    "barcode": _cell(row, "BARKOD"),
                    "stock_code": _cell(row, "STOKKODU"),
                    "urun_id": _cell(row, "URUNID"),
                    "stock": default_stock,
                    "price": _f(row["SATISFIYATI"]) if _has("SATISFIYATI") else list_price,
                    "sale_price": (_f(row["INDIRIMLIFIYAT"]) if _has("INDIRIMLIFIYAT") else 0.0) or (_f(row["SATISFIYATI"]) if _has("SATISFIYATI") else list_price),
                }
                variants.append(v)
                stats["variants_total"] += 1

            existing = await db.products.find_one({"urun_karti_id": kart_id})
            if not existing and parent_stock_code:
                existing = await db.products.find_one({
                    "$or": [
                        {"stock_code": parent_stock_code, "color": renk.title()},
                        {"variants.stock_code": parent_stock_code, "color": renk.title()},
                    ]
                })
            if not existing and renk:
                existing = await db.products.find_one({
                    "name": {"$regex": f"^{re.escape(base_name)}.*{re.escape(renk)}$", "$options": "i"}
                })

            update_doc = {
                "name": urun_adi,
                "color": renk.title(),
                "stock_code": parent_stock_code,
                "sku": parent_stock_code,
                "urun_karti_id": kart_id,
                "price": list_price,
                "sale_price": sale_price,
                "member_price_1": member_price_1,
                "cost_price": cost_price,
                "vat_rate": vat_rate,
                "vendor": vendor,
                "vendor_name": vendor,
                "description": description,
                "category_name": category_leaf,
                "breadcrumb": breadcrumb,
                "variants": variants,
            }
            if cat_doc:
                update_doc["category_id"] = cat_doc.get("id")
                update_doc["category_name"] = cat_doc.get("name")

            if existing:
                await db.products.update_one({"id": existing["id"]}, {"$set": update_doc})
                stats["parents_updated_db"] += 1
            else:
                # Slug çakışmasını önle: temiz slug kullan, ancak başka bir ürün
                # aynı slug'ı kullanıyorsa benzersizlik için kart_id ekle.
                base_slug = _slugify(urun_adi)
                slug = base_slug
                if await db.products.find_one({"slug": slug}):
                    slug = f"{base_slug}-{kart_id}" if kart_id else f"{base_slug}-{str(uuid4())[:6]}"
                new_doc = {
                    "id": str(uuid4()),
                    "slug": slug,
                    "is_active": True,
                    "is_published": True,
                    "images": [],
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    **update_doc,
                }
                await db.products.insert_one(new_doc)
                stats["parents_created_new"] += 1
        except Exception as e:
            stats["errors"].append(f"KART {kart_id_raw}: {e}")

    try:
        os.unlink(tmp_path)
    except Exception:
        pass

    return {
        "success": True,
        "message": f"{stats['parents_updated_db']} güncellendi, {stats['parents_created_new']} yeni eklendi, {stats['variants_total']} varyant işlendi",
        "filename": file.filename,
        "stats": stats,
    }







# ==================== XML FEED IMPORT ====================

XML_FEED_URL = "https://www.facette.com.tr/XMLExport/7BECCB0A782647BFAB843E68AD11E468"
_NS = {"g": "http://base.google.com/ns/1.0"}

def _xml_text(item: ET.Element, tag: str, ns: dict = _NS) -> str:
    el = item.find(tag, ns)
    return (el.text or "").strip() if el is not None else ""

def _xml_all(item: ET.Element, tag: str, ns: dict = _NS) -> list:
    return [(el.text or "").strip() for el in item.findall(tag, ns) if el is not None and el.text]

@router.post("/xml/products/import")
async def import_xml_products(
    xml_url: str = Query(XML_FEED_URL, description="Google Shopping XML URL"),
    deactivate_missing: bool = Query(
        True,
        description="Feed'de bulunmayan (Ticimax'ta pasif/silinmiş) ürünleri pasif yap"
    ),
    current_user: dict = Depends(require_admin)
):
    """
    Google Shopping XML feed'inden ürünleri çekip MongoDB'ye upsert eder.

    - Açıklamadan TÜM `<strong>Etiket:</strong>Değer` özellikleri dinamik olarak
      çıkarılır (Boy, Cep, Astar, Web Color, Kumaş, Kalıp, Model Ölçüleri vb.).
    - deactivate_missing=True (varsayılan): Feed'de olmayan tüm `source=xml_feed`
      ürünleri otomatik olarak `is_active=False` yapılır (Ticimax'ta pasif olanlar).
    """
    import html
    from utils.attr_parser import parse_description_attributes

    try:
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            resp = await client.get(xml_url)
            resp.raise_for_status()
            xml_bytes = resp.content
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"XML feed çekilemedi: {str(e)}")

    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError as e:
        raise HTTPException(status_code=422, detail=f"XML parse hatası: {str(e)}")

    channel = root.find("channel")
    items = channel.findall("item") if channel is not None else root.findall(".//item")

    if not items:
        raise HTTPException(status_code=422, detail="XML'de hiç ürün (item) bulunamadı")

    imported = 0
    updated = 0
    errors = 0
    seen_xml_ids: set[str] = set()

    for item in items:
        try:
            xml_id = _xml_text(item, "g:id")
            title  = _xml_text(item, "g:title")
            if not xml_id or not title:
                continue
            seen_xml_ids.add(xml_id)

            desc = html.unescape(_xml_text(item, "g:description"))
            # Teknik detayları parse et — admin'de "Özellikler" sekmesinde gösterilecek
            tech_attrs, _clean_text = parse_description_attributes(desc)

            # Bedenleri description metninden yakala — "S M L XL bedenlerine uyumlu" vb.
            import re as _re_sizes
            sizes_found: list[str] = []
            # 1) "S M L XL" gibi boşluklu/peş peşe harf bedenleri
            for m in _re_sizes.findall(r"\b((?:XXS|XS|S|M|L|XL|XXL|XXXL)(?:[ ,/\-]+(?:XXS|XS|S|M|L|XL|XXL|XXXL))+)\b", _clean_text):
                tokens = _re_sizes.split(r"[ ,/\-]+", m.strip())
                for t in tokens:
                    t = t.strip().upper()
                    if t and t not in sizes_found:
                        sizes_found.append(t)
            # 2) Numerik bedenler "34 36 38 40 42 44"
            num_match = _re_sizes.search(r"\b(\d{2}(?:[ ,/\-]+\d{2}){2,})\b", _clean_text)
            if num_match:
                for t in _re_sizes.split(r"[ ,/\-]+", num_match.group(1)):
                    if t and t not in sizes_found:
                        sizes_found.append(t)

            def parse_price(s: str) -> Optional[float]:
                if not s:
                    return None
                try:
                    return float(s.split()[0])
                except (ValueError, IndexError):
                    return None

            price      = parse_price(_xml_text(item, "g:price")) or 0.0
            sale_price = parse_price(_xml_text(item, "g:sale_price"))

            availability  = _xml_text(item, "g:availability")
            in_stock      = availability.lower() == "in stock"
            product_type  = _xml_text(item, "g:product_type")
            goog_cat      = _xml_text(item, "g:google_product_category")
            category_name = product_type or goog_cat
            brand         = _xml_text(item, "g:brand") or "FACETTE"
            product_url   = _xml_text(item, "g:link")
            mpn           = _xml_text(item, "g:mpn")
            label_0       = _xml_text(item, "g:custom_label_0")
            label_1       = _xml_text(item, "g:custom_label_1")

            main_image   = _xml_text(item, "g:image_link")
            extra_images = _xml_all(item, "g:additional_image_link")
            all_images: list = []
            seen_imgs: set = set()
            for img in [main_image] + extra_images:
                if img and img not in seen_imgs:
                    all_images.append(img)
                    seen_imgs.add(img)

            slug = product_url.rstrip("/").split("/")[-1] if product_url else None
            if not slug:
                slug = _generate_slug(title) + f"-{xml_id}"

            doc = {
                "xml_id":        xml_id,
                "name":          title,
                "slug":          slug,
                "description":   desc,
                "attributes":    tech_attrs,  # parsed teknik detaylar: urun_bilgisi, kumas, kalip, beden_olculeri, model_olculeri, ...
                "sizes":         sizes_found,  # description'dan parse edilen beden listesi (S/M/L/XL veya 36/38/40)
                "stock_code":    label_0 or "",  # Ticimax'ın "Ürün Kodu" — varyantlar arası ortak
                "sku":           mpn or "",     # MPN/barkod, varyanta özel
                "price":         price,
                "sale_price":    sale_price,
                "brand":         brand,
                "category_name": category_name,
                "stock":         1 if in_stock else 0,
                "is_active":     True,
                "is_featured":   False,
                "is_new":        False,
                "images":        all_images,
                "thumbnail":     all_images[0] if all_images else "",
                "barcode":       mpn,
                "source":        "xml_feed",
                "product_url":   product_url,
                "availability":  availability,
                "xml_label_0":   label_0,
                "xml_label_1":   label_1,
                "updated_at":    datetime.now(timezone.utc).isoformat(),
            }

            existing = await db.products.find_one({"xml_id": xml_id})
            if existing:
                await db.products.update_one(
                    {"xml_id": xml_id},
                    {"$set": doc, "$unset": {"deactivated_reason": ""}},
                )
                updated += 1
            else:
                doc["id"]         = generate_id()
                doc["created_at"] = datetime.now(timezone.utc).isoformat()
                doc["variants"]   = []
                await db.products.insert_one(doc)
                imported += 1

        except Exception as e:
            logger.error(f"XML item parse hatası (id={_xml_text(item, 'g:id')}): {e}")
            errors += 1
            continue

    await db.settings.update_one(
        {"id": "xml_feed"},
        {"$set": {"xml_url": xml_url, "last_sync": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )

    # Feed'de OLMAYAN xml_feed ürünleri pasif yap (Ticimax'ta pasif/silinmiş)
    deactivated = 0
    if deactivate_missing and seen_xml_ids:
        result = await db.products.update_many(
            {
                "source": "xml_feed",
                "xml_id": {"$nin": list(seen_xml_ids)},
                "is_active": {"$ne": False},
            },
            {
                "$set": {
                    "is_active": False,
                    "deactivated_reason": "ticimax_xml_missing",
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }
            },
        )
        deactivated = result.modified_count

    return {
        "success":     True,
        "imported":    imported,
        "updated":     updated,
        "total":       imported + updated,
        "errors":      errors,
        "deactivated": deactivated,
        "message":     (
            f"{imported} yeni ürün eklendi, {updated} ürün güncellendi"
            + (f", {deactivated} ürün pasife alındı" if deactivated else "")
            + (f", {errors} hata" if errors else "")
        ),
    }

@router.get("/xml/status")
async def get_xml_feed_status():
    """XML feed son senkronizasyon bilgisi"""
    settings = await db.settings.find_one({"id": "xml_feed"}) or {}
    return {
        "xml_url":    settings.get("xml_url", XML_FEED_URL),
        "last_sync":  settings.get("last_sync"),
        "configured": True,
    }

# ==================== TRENDYOL CLAIMS (İADE/İPTAL) ====================

# Trendyol claim objesinde claim-level "status" alanı YOKTUR. Statü item bazındadır:
# items[].claimItems[].claimItemStatus.name. Genel statü item'lardan türetilir.
_CLAIM_STATUS_PRIORITY = ["WaitingInAction", "InAnalysis", "Created", "Unresolved", "Rejected", "Cancelled", "Accepted"]


def _derive_claim_status(claim: dict) -> str:
    """Trendyol claim objesinden genel statüyü item statülerinden türetir.

    Bir item bile aksiyon bekliyorsa (WaitingInAction/InAnalysis) claim aksiyon bekler;
    yoksa açık talep (Created) baskındır; hepsi terminal ise Accepted/Rejected/Cancelled.
    Hiç statü bulunamazsa "" döner.
    """
    statuses = []
    for item in (claim.get("items") or []):
        for ci in (item.get("claimItems") or []):
            nm = ((ci.get("claimItemStatus") or {}).get("name") or "").strip()
            if nm:
                statuses.append(nm)
    if not statuses:
        return ""
    sset = set(statuses)
    for p in _CLAIM_STATUS_PRIORITY:
        if p in sset:
            return p
    return statuses[0]


def _claim_bucket(c: dict) -> str:
    """Bir Trendyol claim'ini sekme kovasına eşler (status + kargo takip durumuna göre).

    Created + takip no BOŞ  -> talep_olusturulan (müşteri henüz kargoya vermedi)
    Created + takip no DOLU  -> kargoya_verilen  (iade kargoda)
    WaitingInAction/InAnalysis -> aksiyon_bekleyen
    Accepted                  -> onaylanan
    Rejected/Unresolved       -> reddedilen
    Cancelled                 -> iptal (iptal edilmiş iade talebi; ayrı sekme)

    NOT: Bu eşleme canlı 34/54/16/3583/49 sayılarıyla kalibre edilecek tek noktadır;
    backfill sonrası gerçek dağılım ile bire bir tutmazsa yalnız burası ayarlanır.
    """
    st = (c.get("claim_status") or "").strip()
    has_cargo = bool(str(c.get("cargo_tracking_number") or "").strip())
    if st == "Accepted":
        return "onaylanan"
    if st == "Cancelled":
        return "iptal"
    if st in ("Rejected", "Unresolved"):
        return "reddedilen"
    if st in ("WaitingInAction", "InAnalysis"):
        return "aksiyon_bekleyen"
    if st == "Created":
        return "kargoya_verilen" if has_cargo else "talep_olusturulan"
    return "talep_olusturulan"


def _first_seen_stamps(claim: dict) -> dict:
    """Yeni bir claim ilk kez yazılırken status'una göre onay/ret tarih damgası üretir.

    lastModifiedDate (ms epoch) varsa onu, yoksa şimdiyi kullanır. İdempotent katkı:
    yalnız ilgili terminal durumda alan döner; aksi halde boş dict.
    """
    st = _derive_claim_status(claim)
    lm = claim.get("lastModifiedDate")
    iso = ""
    if lm:
        try:
            iso = datetime.fromtimestamp(lm / 1000, tz=timezone.utc).isoformat()
        except Exception:
            iso = ""
    if not iso:
        iso = datetime.now(timezone.utc).isoformat()
    if st == "Accepted":
        return {"return_approved_at": iso}
    if st in ("Rejected", "Cancelled"):
        return {"return_rejected_at": iso}
    return {}


async def _sync_trendyol_claims_core(days_back: int = 1095):
    """Trendyol claims çekirdek senkron — endpoint + scheduler ortak kullanır.

    days_back: geçmiş tarama penceresi (ilk backfill 1095=3yıl; scheduler kısa pencere).
    """
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    from trendyol_client import TrendyolClient

    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    end_date = datetime.now(timezone.utc)
    total_synced = 0
    order_cache = {}  # order_number -> order_data cache

    # Trendyol max 15 günlük aralık destekliyor, parçalıyoruz
    chunk_days = 15
    current_end = end_date

    while True:
        current_start = current_end - timedelta(days=chunk_days)
        days_elapsed = (end_date - current_start).days

        if days_elapsed > days_back:
            current_start = end_date - timedelta(days=days_back)

        start_ts = int(current_start.timestamp() * 1000)
        end_ts = int(current_end.timestamp() * 1000)

        current_page = 0
        page_size = 200

        while True:
            try:
                url = f"{client.base_url}/order/sellers/{client.supplier_id}/claims"
                params = {
                    "page": current_page,
                    "size": page_size,
                    "startDate": start_ts,
                    "endDate": end_ts,
                }
                async with httpx.AsyncClient(timeout=30.0) as http_client:
                    headers = client._get_headers()
                    response = await http_client.get(url, headers=headers, params=params)
                    response.raise_for_status()
                    result = response.json()
            except Exception as e:
                logger.error(f"Claims sync error: {str(e)}")
                break

            data = result if isinstance(result, dict) else {}
            content = data.get("content", [])
            total_pages = data.get("totalPages", 0)

            if not content:
                break

            for claim in content:
                claim_id = str(claim.get("claimId", claim.get("id", "")))
                if not claim_id:
                    continue

                # Mevcut claim: pahalı order API'sini TEKRAR ÇEKME (iskonto zaten kayıtlı),
                # ama DURUM/kargo/tarih alanlarını canlı tazele — durum geçişleri (Created→
                # Accepted/Rejected) yansısın. Trendyol kayıtları lastModifiedDate sırasıyla gelir.
                existing_claim = await db.trendyol_claims.find_one(
                    {"claim_id": claim_id},
                    {"_id": 1, "return_approved_at": 1, "return_rejected_at": 1, "manual_locked": 1}
                )
                if existing_claim:
                    _new_status = _derive_claim_status(claim)
                    _lm = claim.get("lastModifiedDate")
                    _lm_iso = ""
                    if _lm:
                        try:
                            _lm_iso = datetime.fromtimestamp(_lm / 1000, tz=timezone.utc).isoformat()
                        except Exception:
                            _lm_iso = ""
                    _now_iso = datetime.now(timezone.utc).isoformat()
                    _set = {
                        "cargo_tracking_number": str(claim.get("cargoTrackingNumber", "")),
                        "cargo_provider_name": claim.get("cargoProviderName", ""),
                        "raw_data": claim,
                        "updated_at": _now_iso,
                    }
                    # MANUEL KİLİT: admin durumu elle ilerlettiyse (manual_locked) Trendyol senkronu
                    # claim_status'u EZMEZ — yalnız kargo/raw bilgisi tazelenir, durum manuel kalır.
                    if not existing_claim.get("manual_locked"):
                        _set["claim_status"] = _new_status
                        # Durum geçiş tarih damgaları (idempotent — yalnız ilk geçişte yazılır)
                        if _new_status == "Accepted" and not existing_claim.get("return_approved_at"):
                            _set["return_approved_at"] = _lm_iso or _now_iso
                        if _new_status in ("Rejected", "Cancelled") and not existing_claim.get("return_rejected_at"):
                            _set["return_rejected_at"] = _lm_iso or _now_iso
                    await db.trendyol_claims.update_one({"claim_id": claim_id}, {"$set": _set})
                    total_synced += 1
                    continue

                # Claim items'dan tip ve sebep çıkar
                claim_items = []
                claim_type = ""
                claim_reason = ""
                refund_amount = 0
                
                # Claims API'sinde iskonto bilgisi yok. Sipariş API'sinden çek.
                order_number = str(claim.get("orderNumber", ""))
                order_discount_map = {}  # barcode -> {discount, gross_price, net_price}
                if order_number:
                    # Cache kontrolü: aynı sipariş numarasını tekrar çekme
                    cache_key = f"order_{order_number}"
                    if cache_key not in order_cache:
                        try:
                            order_data = await client.get_orders(order_number=order_number)
                            order_cache[cache_key] = order_data
                        except Exception as e:
                            logger.warning(f"Could not fetch order {order_number} for discount: {e}")
                            order_cache[cache_key] = {}
                    
                    cached = order_cache.get(cache_key, {})
                    for pkg in cached.get("content", []):
                        for line in pkg.get("lines", []):
                            bc = line.get("barcode", "")
                            line_gross = line.get("lineGrossAmount", line.get("amount", 0))
                            line_net = line.get("price", 0)
                            line_disc = line.get("discount", 0)
                            qty = max(line.get("quantity", 1), 1)
                            if bc:
                                order_discount_map[bc] = {
                                    "gross": line_gross / qty if line_gross else 0,
                                    "net": line_net / qty if line_net else 0,
                                    "discount": line_disc / qty if line_disc else 0,
                                }

                for item in claim.get("items", []):
                    order_line = item.get("orderLine", {})
                    for ci in item.get("claimItems", []):
                        reason_info = ci.get("customerClaimItemReason", {})
                        if not claim_type:
                            code = reason_info.get("code", "").upper()
                            if code in ["ABANDON", "UNDELIVERED", "NOTDELIVERED"]:
                                claim_type = "CANCEL"
                            else:
                                claim_type = "RETURN"
                        if not claim_reason:
                            claim_reason = reason_info.get("name", "")

                        barcode = order_line.get("barcode", "")
                        claim_price = order_line.get("price", 0)
                        
                        # İskontoyu sipariş verisinden al
                        order_info = order_discount_map.get(barcode, {})
                        if order_info:
                            gross_price = order_info.get("gross", claim_price)
                            net_price = order_info.get("net", claim_price)
                            discount = order_info.get("discount", 0)
                        else:
                            # Fallback: Claims API verisini kullan (iskonto yok)
                            gross_price = claim_price
                            net_price = claim_price
                            discount = 0
                        
                        claim_items.append({
                            "claim_item_id": str(ci.get("id", "")),
                            "productName": order_line.get("productName", ""),
                            "barcode": barcode,
                            "unit_price": gross_price,
                            "discount_amount": discount,
                            "price": net_price,
                            "quantity": 1,
                            "reason": reason_info.get("name", "")
                        })
                        refund_amount += net_price

                # Tarih formatı
                claim_date = claim.get("claimDate")
                created_date_str = ""
                if claim_date:
                    try:
                        created_date_str = datetime.fromtimestamp(claim_date / 1000, tz=timezone.utc).isoformat()
                    except Exception:
                        created_date_str = str(claim_date)

                # Fatura numarasını çıkar: sipariş verisinden veya claim'den
                invoice_number = ""
                for item in claim.get("items", []):
                    ol = item.get("orderLine", {})
                    inv = ol.get("invoiceNumber", "") or item.get("invoiceNumber", "")
                    if inv:
                        invoice_number = str(inv)
                        break
                if not invoice_number:
                    invoice_number = str(claim.get("invoiceNumber", "") or "")
                # Sipariş verisinden fatura no çek
                if not invoice_number and order_discount_map:
                    try:
                        _order_data = await client.get_orders(order_number=order_number)
                        for pkg in _order_data.get("content", []):
                            inv_no = pkg.get("invoiceNumber", "")
                            if inv_no:
                                invoice_number = str(inv_no)
                                break
                    except Exception:
                        pass

                claim_doc = {
                    "claim_id": claim_id,
                    "order_number": order_number,
                    "claim_type": claim_type,
                    "claim_reason": claim_reason,
                    "claim_status": _derive_claim_status(claim),
                    **_first_seen_stamps(claim),
                    "customer_name": f"{claim.get('customerFirstName', '')} {claim.get('customerLastName', '')}".strip(),
                    "created_date": created_date_str,
                    "items": claim_items,
                    "refund_amount": refund_amount,
                    "invoice_number": invoice_number,
                    "invoice_link": claim.get("invoiceLink", ""), # Yeni eklendi
                    "cargo_tracking_number": str(claim.get("cargoTrackingNumber", "")),
                    "cargo_provider_name": claim.get("cargoProviderName", ""),
                    "raw_data": claim,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }

                await db.trendyol_claims.update_one(
                    {"claim_id": claim_id},
                    {"$set": claim_doc, "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": datetime.now(timezone.utc).isoformat()}},
                    upsert=True
                )
                total_synced += 1

                # Claim sebebini eşleşen İPTAL siparişine bağla. claim_type sınıflandırması
                # (CANCEL/RETURN) Trendyol sebep koduna göre dar kalabildiğinden ONA GÜVENMEYİZ:
                # sipariş zaten status=cancelled ise (iade post-teslimat olduğundan iptalle
                # karışmaz) o claim'in sebebi = iptal sebebidir. Böylece "Stok tükendi",
                # "Adreste bulunmayacağım" gibi gerçek sebepler İptaller'e yazılır.
                if claim_reason and order_number:
                    try:
                        await db.orders.update_one(
                            {"order_number": str(order_number), "platform": "trendyol", "status": "cancelled"},
                            {"$set": {"cancel_reason": claim_reason, "cancel_source": "trendyol",
                                      "updated_at": datetime.now(timezone.utc).isoformat()}},
                        )
                    except Exception as _ce:
                        logger.error(f"[trendyol claim->order reason {order_number}] {_ce}")

            current_page += 1
            if current_page >= total_pages:
                break

        current_end = current_start
        if days_elapsed >= days_back:
            break

    return {
        "message": f"Son {days_back} gündeki toplam {total_synced} iade/iptal kaydı senkronize edildi",
        "total_synced": total_synced,
        "days_back": days_back
    }


@router.get("/trendyol/claims/sync")
async def sync_trendyol_claims(
    days_back: int = 1095,
    current_user: dict = Depends(require_admin)
):
    """Trendyol'dan iade/iptal (claim) kayıtlarını çeker ve MongoDB'ye kaydeder.

    days_back varsayılan 1095 (3 yıl) — geçmiş backfill için. UI bu ucu tetikler;
    durum geçişleri her çağrıda canlı tazelenir.
    """
    return await _sync_trendyol_claims_core(days_back)


@router.post("/trendyol/claims/fix-discounts")
async def fix_claim_discounts(current_user: dict = Depends(require_admin)):
    """Fix discount data for existing claims by fetching from order API"""
    settings = await db.settings.find_one({"id": "trendyol"}, {"_id": 0})
    if not settings or not settings.get("api_key"):
        raise HTTPException(status_code=400, detail="Trendyol API ayarları eksik")
    
    from trendyol_client import TrendyolClient
    client = TrendyolClient(settings["supplier_id"], settings["api_key"], settings["api_secret"])
    
    # Get claims that need discount fix (where items have 0 discount)
    claims = await db.trendyol_claims.find({}, {"_id": 0, "claim_id": 1, "order_number": 1, "items": 1}).to_list(None)
    
    order_cache = {}
    fixed = 0
    
    for claim in claims:
        order_number = claim.get("order_number", "")
        if not order_number:
            continue
        
        items = claim.get("items", [])
        needs_fix = any(item.get("discount_amount", 0) == 0 and item.get("unit_price", 0) == item.get("price", 0) for item in items)
        if not needs_fix:
            continue
        
        # Get order data (cached)
        if order_number not in order_cache:
            try:
                order_cache[order_number] = await client.get_orders(order_number=order_number)
            except Exception:
                order_cache[order_number] = {}
        
        cached = order_cache.get(order_number, {})
        discount_map = {}
        invoice_number = ""
        for pkg in cached.get("content", []):
            if not invoice_number:
                invoice_number = pkg.get("invoiceNumber", "")
            for line in pkg.get("lines", []):
                bc = line.get("barcode", "")
                qty = max(line.get("quantity", 1), 1)
                if bc:
                    discount_map[bc] = {
                        "gross": (line.get("lineGrossAmount", line.get("amount", 0)) or 0) / qty,
                        "net": (line.get("price", 0) or 0) / qty,
                        "discount": (line.get("discount", 0) or 0) / qty,
                    }
        
        updated_items = []
        refund_amount = 0
        for item in items:
            bc = item.get("barcode", "")
            if bc in discount_map:
                item["unit_price"] = discount_map[bc]["gross"]
                item["discount_amount"] = discount_map[bc]["discount"]
                item["price"] = discount_map[bc]["net"]
            refund_amount += item.get("price", 0)
            updated_items.append(item)
        
        update_set = {"items": updated_items, "refund_amount": refund_amount}
        if invoice_number:
            update_set["invoice_number"] = invoice_number
        
        await db.trendyol_claims.update_one(
            {"claim_id": claim["claim_id"]},
            {"$set": update_set}
        )
        fixed += 1
    
    return {"success": True, "fixed": fixed, "message": f"{fixed} iadenin iskonto bilgisi güncellendi"}


def _claim_is_site_order(claim: dict) -> bool:
    """Bir iade kaydının SİTE (web) siparişi mi yoksa PAZARYERİ mi olduğunu belirler.

    Kural (öncelik sırası — Kadir'in onayladığı kesin sinyal):
      1. trendyol_package_id dolu  -> Trendyol (kargo DHL/MNG olsa bile)
      2. cargo_provider_name içinde "marketplace" geçiyor -> pazaryeri
      3. hepsiburada paket izi -> pazaryeri
      4. platform alanı bilinen bir pazaryeri -> pazaryeri
      5. Hiçbiri değil -> SİTE siparişi
    Not: 'platform'/'source' alanı tek başına güvenilir değil (eski Ticimax-çekme
    döneminde Trendyol siparişleri de 'ticimax' damgası almış olabilir), bu yüzden
    önce sağlam paket-kimliği sinyallerine bakılır.
    """
    if claim.get("trendyol_package_id"):
        return False
    cpn = str(claim.get("cargo_provider_name") or "").lower()
    if "marketplace" in cpn:
        return False
    if claim.get("hepsiburada_package_id") or claim.get("hb_package_id"):
        return False
    plt = str(claim.get("platform") or "").lower()
    _MARKETPLACES = {"trendyol", "hepsiburada", "n11", "amazon_tr", "amazon_de",
                     "temu", "aliexpress", "etsy", "ciceksepeti", "emag", "fruugo",
                     "hepsiglobal", "trendyol_export"}
    if plt in _MARKETPLACES:
        return False
    return True


@router.get("/trendyol/claims")
async def get_trendyol_claims(
    page: int = 1,
    limit: int = 20,
    claim_type: str = "",
    search: str = "",
    status: str = "",
    platform: str = "trendyol",
    current_user: dict = Depends(require_admin)
):
    """Yerel veritabanındaki iade kayıtlarını listele.

    Tüm claim'ler İADE'dir (İptal ayrı menüde, sipariş durumu üzerinden).
    `status` = durum sekmesi anahtarı: all / talep_olusturulan / kargoya_verilen /
    aksiyon_bekleyen / onaylanan / reddedilen.
    `platform` = 'facette' (web/site iadeleri) veya bir pazaryeri anahtarı
    (trendyol, hepsiburada, ...). Ayrım `_claim_is_site_order` kuralıyla yapılır.
    """
    # Sekme kovası eşlemesi artık _claim_bucket(c) helper'ında (status + kargo takip durumu).
    _VALID_TABS = {"talep_olusturulan", "kargoya_verilen", "acik_iade", "aksiyon_bekleyen", "onaylanan", "reddedilen"}

    base_query = {}
    if claim_type:
        base_query["claim_type"] = claim_type
    if search:
        base_query["$or"] = [
            {"order_number": {"$regex": search, "$options": "i"}},
            {"customer_name": {"$regex": search, "$options": "i"}},
            {"claim_id": {"$regex": search, "$options": "i"}},
            {"invoice_number": {"$regex": search, "$options": "i"}},
            {"cargo_tracking_number": {"$regex": search, "$options": "i"}},
        ]

    want_tab = status if (status and status != "all" and status in _VALID_TABS) else None

    # base_query (arama/tip filtreli, AMA status filtresiz) ile TÜM kayıtları çek.
    # Status filtresi bellekte uygulanır ki tab_counts tüm kovaları doğru sayabilsin.
    raw = await db.trendyol_claims.find(
        base_query, {"_id": 0, "raw_data": 0}
    ).sort("created_date", -1).to_list(None)

    # (a) claim_id'ye göre tekilleştir — aynı claim birden fazla belge olarak yazılmışsa
    # en güncel (created_date'e göre zaten sıralı) ilk görüleni tut.
    seen = set()
    deduped = []
    for c in raw:
        cid = c.get("claim_id") or c.get("order_number")
        if cid in seen:
            continue
        seen.add(cid)
        deduped.append(c)

    # (b) trendyol_claims koleksiyonu TAMAMEN pazaryeri (Trendyol) kayitlaridir.
    # Site iadeleri ayri koleksiyonda (customer_returns) ve ayri sekmede (Web Sitesi)
    # gosterilir. Bu yuzden burada site/pazaryeri ayrimi YAPILMAZ; mikro ihracat dahil
    # TUM claim'ler dondurulur. (`platform` parametresi geriye-donuk uyumluluk icin
    # kabul edilir ama bu endpoint'te artik filtreleme yapmaz.)
    platform_scoped = deduped

    # İptal (Cancelled iade statüsü) bu iade ekranından TAMAMEN dışlanır; iptaller
    # ayrı bir alandan yönetilir. Böylece "Tüm İadeler" sekmesi ve "Toplam İade"
    # kartı aynı evreni (iptal-hariç tekil iade) sayar.
    iade_scoped = [c for c in platform_scoped if _claim_bucket(c) != "iptal"]

    # (c) status sekmesi filtresi (bellekte) — _claim_bucket ile
    if want_tab == "acik_iade":
        filtered = [c for c in iade_scoped if _claim_bucket(c) in ("talep_olusturulan", "kargoya_verilen")]
    elif want_tab is not None:
        filtered = [c for c in iade_scoped if _claim_bucket(c) == want_tab]
    else:
        filtered = iade_scoped

    total = len(filtered)
    skip = (page - 1) * limit
    claims = filtered[skip: skip + limit]

    # Her claim'e kova + Türkçe durum etiketi ekle (frontend durum rozeti için).
    # talep/kargoda tek "Açık İade" altında birleşir; geçmiş kovalar korunur.
    _BUCKET_LABEL = {
        "talep_olusturulan": "Açık İade",
        "kargoya_verilen": "Açık İade",
        "aksiyon_bekleyen": "Aksiyon Bekleyen",
        "onaylanan": "Onaylandı",
        "reddedilen": "Reddedildi",
        "iptal": "İptal",
    }
    for c in claims:
        _b = _claim_bucket(c)
        c["bucket"] = _b
        c["bucket_label"] = _BUCKET_LABEL.get(_b, "—")

    # Sekme adetleri — iade_scoped (iptal hariç) üzerinden, _claim_bucket ile.
    _bcount = {"talep_olusturulan": 0, "kargoya_verilen": 0, "aksiyon_bekleyen": 0, "onaylanan": 0, "reddedilen": 0}
    for c in iade_scoped:
        _b = _claim_bucket(c)
        if _b in _bcount:
            _bcount[_b] += 1
    tab_counts = {"all": len(iade_scoped), **_bcount, "acik_iade": _bcount["talep_olusturulan"] + _bcount["kargoya_verilen"]}

    # İstatistikler — "Toplam İade" kartı = "Tüm İadeler" sekmesi (iptal hariç tekil iade).
    total_returns = len(iade_scoped)
    total_cancels = await db.trendyol_claims.count_documents({"claim_type": "CANCEL"})
    total_refund = sum((c.get("refund_amount") or 0) for c in iade_scoped)

    return {
        "claims": claims,
        "total": total,
        "page": page,
        "limit": limit,
        "tab_counts": tab_counts,
        "stats": {
            "total_returns": total_returns,
            "total_cancels": total_cancels,
            "total_refund": total_refund
        }
    }


@router.get("/trendyol/claims/export")
async def export_trendyol_claims(
    status: str = "",
    search: str = "",
    current_user: dict = Depends(require_admin),
):
    """Trendyol iadelerini bulunulan sekme (status) + arama filtresiyle Excel'e aktarır.
    Liste endpoint'iyle (get_trendyol_claims) AYNI dedup + kova mantığını kullanır;
    böylece hangi sekmedeyse o sekmenin kayıtları döner.
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    _VALID_TABS = {"talep_olusturulan", "kargoya_verilen", "acik_iade", "aksiyon_bekleyen", "onaylanan", "reddedilen"}
    base_query = {}
    if search:
        base_query["$or"] = [
            {"order_number": {"$regex": search, "$options": "i"}},
            {"customer_name": {"$regex": search, "$options": "i"}},
            {"claim_id": {"$regex": search, "$options": "i"}},
            {"cargo_tracking_number": {"$regex": search, "$options": "i"}},
        ]
    want_tab = status if (status and status != "all" and status in _VALID_TABS) else None

    raw = await db.trendyol_claims.find(base_query, {"_id": 0, "raw_data": 0}).sort("created_date", -1).to_list(None)
    seen = set(); deduped = []
    for c in raw:
        cid = c.get("claim_id") or c.get("order_number")
        if cid in seen:
            continue
        seen.add(cid); deduped.append(c)
    iade_scoped = [c for c in deduped if _claim_bucket(c) != "iptal"]
    if want_tab == "acik_iade":
        rows = [c for c in iade_scoped if _claim_bucket(c) in ("talep_olusturulan", "kargoya_verilen")]
    elif want_tab is not None:
        rows = [c for c in iade_scoped if _claim_bucket(c) == want_tab]
    else:
        rows = iade_scoped

    _BUCKET_LABEL = {
        "talep_olusturulan": "Açık İade", "kargoya_verilen": "Açık İade",
        "aksiyon_bekleyen": "Aksiyon Bekleyen", "onaylanan": "Onaylandı",
        "reddedilen": "Reddedildi", "iptal": "İptal",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Trendyol İadeleri"
    headers = ["Sipariş No", "Müşteri", "Ürün", "Tutar", "Tarih", "Durum", "Gider Pusulası No"]
    ws.append(headers)
    hfill = PatternFill("solid", fgColor="FCE4B6")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for c in rows:
        items = c.get("items") or []
        urun = ", ".join([
            (i.get("product_name") or i.get("name") or "").strip()
            for i in items if (i.get("product_name") or i.get("name"))
        ])
        b = _claim_bucket(c)
        ws.append([
            c.get("order_number") or "",
            c.get("customer_name") or "",
            urun,
            float(c.get("refund_amount") or 0),
            str(c.get("created_date") or "")[:10],
            _BUCKET_LABEL.get(b, "—"),
            c.get("gider_pusulasi_no") or "",
        ])

    for col in ws.columns:
        ml = max((len(str(cc.value)) for cc in col if cc.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 2, 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"trendyol-iadeleri-{status or 'tum'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/trendyol/claims/diagnostics")
async def trendyol_claims_diagnostics(current_user: dict = Depends(require_admin)):
    """Kova kalibrasyonu teşhisi: ham status dağılımı + kova eşlemesi + kargo kırılımı.

    Panel'deki Trendyol sekme sayılarını (talep/kargoda/aksiyon/onay/ret) gerçek veriyle
    karşılaştırıp _claim_bucket eşlemesini doğrulamak/ayarlamak için kullanılır.
    Backfill bitince çağır; by_bucket panel sayılarıyla tutmazsa by_status_raw'a bakıp
    _claim_bucket tek noktadan düzeltilir.
    """
    rows = await db.trendyol_claims.find(
        {}, {"_id": 0, "claim_status": 1, "cargo_tracking_number": 1, "claim_id": 1}
    ).to_list(None)
    seen = set()
    uniq = []
    for c in rows:
        cid = c.get("claim_id")
        if cid in seen:
            continue
        seen.add(cid)
        uniq.append(c)

    by_status = {}
    by_bucket = {"talep_olusturulan": 0, "kargoya_verilen": 0, "aksiyon_bekleyen": 0, "onaylanan": 0, "reddedilen": 0, "iptal": 0}
    created_with_cargo = 0
    created_without_cargo = 0
    for c in uniq:
        st = (c.get("claim_status") or "").strip() or "(boş)"
        by_status[st] = by_status.get(st, 0) + 1
        b = _claim_bucket(c)
        if b in by_bucket:
            by_bucket[b] += 1
        if (c.get("claim_status") or "").strip() == "Created":
            if str(c.get("cargo_tracking_number") or "").strip():
                created_with_cargo += 1
            else:
                created_without_cargo += 1

    return {
        "total_unique_claims": len(uniq),
        "by_status_raw": by_status,
        "by_bucket": by_bucket,
        "created_with_cargo": created_with_cargo,
        "created_without_cargo": created_without_cargo,
        "expected_from_panel": {
            "talep_olusturulan": 34, "kargoya_verilen": 54,
            "aksiyon_bekleyen": 16, "onaylanan": 3583, "reddedilen": 49,
        },
        "note": "by_bucket panel ile tutmazsa _claim_bucket eşlemesi by_status_raw'a göre ayarlanır.",
    }


@router.post("/trendyol/claims/repair-status")
async def repair_trendyol_claim_status(current_user: dict = Depends(require_admin)):
    """Mevcut TÜM claim kayıtlarının statüsünü KAYITLI raw_data'dan yeniden türetir.

    Eski sync, claim-level olmayan `status` alanını okuduğu için kayıtlar statüsüz
    yazılmıştı. Bu uç Trendyol API'sine YENİ istek atmadan raw_data'daki gerçek item
    statülerinden (claimItemStatus.name) claim_status'ü, kargo alanlarını ve onay/ret
    tarih damgalarını yeniden hesaplar. İdempotent — tekrar çağrılabilir.
    """
    cursor = db.trendyol_claims.find(
        {}, {"_id": 0, "claim_id": 1, "raw_data": 1, "return_approved_at": 1, "return_rejected_at": 1}
    )
    scanned = 0
    fixed = 0
    bucket_after = {"talep_olusturulan": 0, "kargoya_verilen": 0, "aksiyon_bekleyen": 0, "onaylanan": 0, "reddedilen": 0, "iptal": 0}
    async for c in cursor:
        scanned += 1
        raw = c.get("raw_data") or {}
        if not raw:
            continue
        new_status = _derive_claim_status(raw)
        cargo_no = str(raw.get("cargoTrackingNumber", "") or "")
        _set = {
            "claim_status": new_status,
            "cargo_tracking_number": cargo_no,
            "cargo_provider_name": raw.get("cargoProviderName", "") or "",
        }
        stamps = _first_seen_stamps(raw)
        if stamps.get("return_approved_at") and not c.get("return_approved_at"):
            _set["return_approved_at"] = stamps["return_approved_at"]
        if stamps.get("return_rejected_at") and not c.get("return_rejected_at"):
            _set["return_rejected_at"] = stamps["return_rejected_at"]
        await db.trendyol_claims.update_one({"claim_id": c["claim_id"]}, {"$set": _set})
        fixed += 1
        b = _claim_bucket({"claim_status": new_status, "cargo_tracking_number": cargo_no})
        if b in bucket_after:
            bucket_after[b] += 1
    return {"scanned": scanned, "fixed": fixed, "bucket_after": bucket_after}


@router.post("/trendyol/claims/dedupe")
async def dedupe_trendyol_claims(current_user: dict = Depends(require_admin)):
    """Aynı claim_id'ye sahip MÜKERRER belgeleri temizler (her claim_id için en güncel
    updated_at olanı tutar), sonra claim_id üzerinde unique index kurarak gelecekte
    mükerrerlenmeyi engeller. İdempotent — tekrar çağrılabilir.
    """
    pipeline = [
        {"$group": {"_id": "$claim_id", "count": {"$sum": 1}}},
        {"$match": {"count": {"$gt": 1}}},
    ]
    groups = await db.trendyol_claims.aggregate(pipeline).to_list(None)
    removed = 0
    affected = 0
    for g in groups:
        cid = g["_id"]
        docs = await db.trendyol_claims.find(
            {"claim_id": cid}, {"_id": 1, "updated_at": 1, "created_date": 1, "created_at": 1}
        ).to_list(None)
        if len(docs) <= 1:
            continue
        docs.sort(key=lambda d: (str(d.get("updated_at") or ""), str(d.get("created_date") or ""), str(d.get("created_at") or "")), reverse=True)
        to_delete = [d["_id"] for d in docs[1:]]
        if to_delete:
            res = await db.trendyol_claims.delete_many({"_id": {"$in": to_delete}})
            removed += res.deleted_count
            affected += 1
    index_created = False
    index_error = ""
    try:
        await db.trendyol_claims.create_index("claim_id", unique=True, name="uniq_claim_id")
        index_created = True
    except Exception as e:
        index_error = str(e)[:200]
    remaining = await db.trendyol_claims.count_documents({})
    return {
        "duplicate_groups": len(groups),
        "claims_affected": affected,
        "removed": removed,
        "index_created": index_created,
        "index_error": index_error,
        "remaining_docs": remaining,
    }


@router.post("/trendyol/claims/{claim_id}/set-status")
async def set_trendyol_claim_status(claim_id: str, payload: dict, current_user: dict = Depends(require_admin)):
    """Trendyol iade durumunu MANUEL ilerletir ve KİLİTLER.

    manual_locked=True olunca sonraki Trendyol senkronları bu claim'in claim_status'unu
    EZMEZ (yalnız kargo/raw bilgisi tazelenir). Kilidi kaldırmak için /unlock çağrılır.
    """
    new_status = (payload.get("status") or "").strip()
    valid = {"Created", "WaitingInAction", "InAnalysis", "Accepted", "Rejected", "Unresolved", "Cancelled"}
    if new_status not in valid:
        raise HTTPException(status_code=400, detail=f"Geçersiz durum. Geçerli: {', '.join(sorted(valid))}")
    now = datetime.now(timezone.utc).isoformat()
    _set = {"claim_status": new_status, "manual_locked": True, "manual_status_at": now, "updated_at": now}
    if new_status == "Accepted":
        _set["return_approved_at"] = now
    if new_status in ("Rejected", "Cancelled"):
        _set["return_rejected_at"] = now
    res = await db.trendyol_claims.update_one({"claim_id": claim_id}, {"$set": _set})
    if not res.matched_count:
        raise HTTPException(status_code=404, detail="İade (claim) bulunamadı")
    return {"success": True, "claim_id": claim_id, "status": new_status, "manual_locked": True}


@router.post("/trendyol/claims/{claim_id}/unlock")
async def unlock_trendyol_claim(claim_id: str, current_user: dict = Depends(require_admin)):
    """Manuel kilidi kaldırır → durum tekrar Trendyol senkronundan güncellenmeye başlar."""
    res = await db.trendyol_claims.update_one(
        {"claim_id": claim_id},
        {"$set": {"manual_locked": False, "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    if not res.matched_count:
        raise HTTPException(status_code=404, detail="İade (claim) bulunamadı")
    return {"success": True, "claim_id": claim_id, "manual_locked": False}


@router.get("/trendyol/claims/shipment-probe")
async def trendyol_shipment_probe(order_number: str = "", current_user: dict = Depends(require_admin)):
    """GEÇİCİ ARAŞTIRMA: Trendyol sipariş paketi servisini (getShipmentPackages) bir iadenin
    orderNumber'ı ile sorgular ve paket durum/satır alanlarını döndürür. Amaç: talep vs
    kargoya-verilen ayrımı için kullanılabilir bir 'shipped/returned' sinyali var mı görmek.
    """
    headers = await get_trendyol_headers()
    if not headers:
        raise HTTPException(status_code=400, detail="Trendyol kimliği yapılandırılmamış")
    config = await get_trendyol_config()
    base = config["base_url"]
    sid = config["supplier_id"]
    probed = []
    # order_number verilmezse birkaç açık (Created) claim üzerinde dene
    targets = []
    if order_number:
        targets = [order_number]
    else:
        async for c in db.trendyol_claims.find({"claim_status": "Created"}, {"_id": 0, "order_number": 1}).limit(3):
            on = c.get("order_number")
            if on:
                targets.append(on)
    for on in targets:
        url = f"{base}/sapigw/suppliers/{sid}/orders?orderNumber={on}"
        try:
            async with httpx.AsyncClient(timeout=30) as cx:
                r = await cx.get(url, headers=headers)
            try:
                data = r.json()
            except Exception:
                data = {"_text": r.text[:500]}
            pkgs = []
            for pkg in (data.get("content") or [])[:5]:
                pkgs.append({
                    "id": pkg.get("id"),
                    "status": pkg.get("status"),
                    "cargoTrackingNumber": pkg.get("cargoTrackingNumber"),
                    "lines_status": [l.get("orderLineItemStatusName") for l in (pkg.get("lines") or [])[:4]],
                    "top_keys": list(pkg.keys()),
                })
            probed.append({"order_number": on, "http": r.status_code, "package_count": len(data.get("content") or []), "packages": pkgs})
        except Exception as e:
            probed.append({"order_number": on, "error": str(e)[:200]})
    return {"base": base, "supplier_id": sid, "probed": probed}


@router.get("/trendyol/claims/issue-reasons")
async def get_trendyol_issue_reasons(current_user: dict = Depends(require_admin)):
    """Fetch claim issue reasons from Trendyol"""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    from trendyol_client import TrendyolClient
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    try:
        url = f"{client.base_url}/order/sellers/{client.supplier_id}/claim-issue-reasons"
        async with httpx.AsyncClient(timeout=30.0) as http_client:
            headers = client._get_headers()
            response = await http_client.get(url, headers=headers)
            response.raise_for_status()
            return response.json()
    except Exception as e:
        logger.error(f"Failed to fetch issue reasons: {str(e)}")
        # Return common fallback reasons
        return [
            {"id": 1, "name": "Kullanım Hatası / Tüketici Kaynaklı Hasar"},
            {"id": 2, "name": "Ürün Orijinal Kutusunda / Ambalajında Değil"},
            {"id": 4, "name": "Eksik Aksesuar / Parça"},
            {"id": 6, "name": "İade Süresi Geçmiş"},
            {"id": 21, "name": "Ürün Kullanılmış / Etiketi Koparılmış"}
        ]


@router.get("/trendyol/claims/{claim_id}")
async def get_trendyol_claim_detail(claim_id: str, current_user: dict = Depends(require_admin)):
    """Tek bir iade/iptal kaydının detayını getir."""
    claim = await db.trendyol_claims.find_one({"claim_id": claim_id}, {"_id": 0})
    if not claim:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    return claim


async def restock_claim_once(claim_id: str, source: str, by_email: str = "") -> list:
    """İade kalemlerini ürün stoğuna BİR KEZ geri ekler (idempotent: claim.stock_restored).
    Eşleşme: variant.barcode -> variant.stock; yoksa product.barcode -> product.stock."""
    claim = await db.trendyol_claims.find_one(
        {"claim_id": claim_id}, {"_id": 0, "items": 1, "order_number": 1, "stock_restored": 1}
    )
    if not claim or claim.get("stock_restored"):
        return []
    restocked = []
    for item in (claim.get("items") or []):
        barcode = str(item.get("barcode", "") or "").strip()
        qty = int(item.get("quantity", 1) or 1)
        if not barcode or qty <= 0:
            continue
        prod = await db.products.find_one({"variants.barcode": barcode}, {"_id": 0, "id": 1, "variants": 1, "stock": 1})
        if prod:
            for v in (prod.get("variants") or []):
                if v.get("barcode") == barcode:
                    v["stock"] = int(v.get("stock", 0) or 0) + qty
                    break
            new_total = sum(int(v.get("stock", 0) or 0) for v in prod.get("variants", []))
            await db.products.update_one(
                {"id": prod["id"]},
                {"$set": {"variants": prod["variants"], "stock": new_total, "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
            restocked.append({"barcode": barcode, "qty": qty, "product_id": prod["id"]})
        else:
            p2 = await db.products.find_one({"barcode": barcode}, {"_id": 0, "id": 1, "stock": 1})
            if p2:
                await db.products.update_one(
                    {"id": p2["id"]},
                    {"$inc": {"stock": qty}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
                )
                restocked.append({"barcode": barcode, "qty": qty, "product_id": p2["id"]})
    # Tekrarı önlemek için her durumda işaretle
    await db.trendyol_claims.update_one(
        {"claim_id": claim_id},
        {"$set": {"stock_restored": True, "stock_restored_at": datetime.now(timezone.utc).isoformat(), "stock_restored_source": source}}
    )
    if restocked:
        await db.stock_movements.insert_one({
            "id": str(uuid.uuid4()), "type": "return_restock", "source": source,
            "claim_id": claim_id, "order_number": claim.get("order_number", ""),
            "items": restocked, "created_by": by_email,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    return restocked


@router.post("/trendyol/claims/{claim_id}/gider-pusulasi")
async def generate_gider_pusulasi(claim_id: str, payload: Optional[dict] = Body(default=None), current_user: dict = Depends(require_admin)):
    """Generate expense receipt (gider pusulası) data for a return claim"""
    claim = await db.trendyol_claims.find_one({"claim_id": claim_id}, {"_id": 0})
    if not claim:
        raise HTTPException(status_code=404, detail="İade kaydı bulunamadı")

    # Kurumsal/e-Fatura siparişinde gider pusulası DÜZENLENEMEZ — iade faturası gerekir.
    _onum = claim.get("order_number", "")
    if _onum:
        _ord = await db.orders.find_one({"order_number": _onum}, {"_id": 0, "invoice_type": 1, "billing_info": 1})
        if _ord and ((_ord.get("invoice_type") == "e-fatura") or bool((_ord.get("billing_info") or {}).get("is_corporate"))):
            raise HTTPException(status_code=400, detail="Bu sipariş kurumsal/e-Fatura siparişi — gider pusulası düzenlenemez. Müşteriden iade faturası gerekir (Doğan'dan panele düşecek, onaylayınca stok +1).")

    # İade onayımız = gider pusulası oluşturmak. Stoğu BİR KEZ geri ekle (idempotent).
    await restock_claim_once(claim_id, "gider_pusulasi", current_user.get("email", ""))

    settings = await db.settings.find_one({"id": "main"}, {"_id": 0})
    company = settings.get("company_info", {}) if settings else {}
    # Alıcı adresi claim'de tutulmuyor; sipariş kaydından çek
    order = await db.orders.find_one({"order_number": claim.get("order_number", "")}, {"_id": 0}) or {}
    _ship = order.get("shipping_address", {}) or {}
    _cust_name = claim.get("customer_name", "") or (f"{_ship.get('first_name','')} {_ship.get('last_name','')}".strip())
    _cust_addr = _ship.get("address", "") or (claim.get("shipping_address", "") if isinstance(claim.get("shipping_address"), str) else "")
    _cust_district = _ship.get("district", "")
    _cust_city = _ship.get("city", "") or claim.get("shipping_city", "")
    _cust_country = _ship.get("country", "") or "Türkiye"

    items = claim.get("items", [])
    # Kısmi gider pusulası: yalnızca seçili kalemler (item_indexes verilirse SADECE onlar hesaplanır)
    _sel_idx = (payload or {}).get("item_indexes")
    if isinstance(_sel_idx, list) and _sel_idx:
        _filtered = []
        for _i in _sel_idx:
            try:
                _filtered.append(items[int(_i)])
            except Exception:
                continue
        if _filtered:
            items = _filtered
    total_net = sum(item.get("price", 0) * item.get("quantity", 1) for item in items)
    total_discount = sum(item.get("discount_amount", 0) * item.get("quantity", 1) for item in items)
    total_gross = sum(item.get("unit_price", 0) * item.get("quantity", 1) for item in items)
    vat_rate = settings.get("default_vat_rate", 10) if settings else 10
    vat_amount = round(total_net * vat_rate / (100 + vat_rate), 2)
    net_without_vat = round(total_net - vat_amount, 2)

    last_gp = await db.gider_pusulasi.find_one({}, sort=[("number", -1)])
    gp_number = (last_gp.get("number", 0) + 1) if last_gp else 1
    # Frontend'den gelen takip numarası (matbu form sıra no ile eşleşir). Verilirse display olarak kullan.
    tracking_no = str((payload or {}).get("tracking_no") or "").strip()
    display_number = tracking_no if tracking_no else f"GP-{gp_number:06d}"

    # Kalemleri ürün kataloğundaki beden ile zenginleştir (barkod -> variant.size)
    gp_items = []
    for _it in items:
        _bc = str(_it.get("barcode", "") or "").strip()
        _size = ""
        if _bc:
            _pv = await db.products.find_one({"variants.barcode": _bc}, {"_id": 0, "variants": 1})
            if _pv:
                for _v in (_pv.get("variants") or []):
                    if str(_v.get("barcode")) == _bc:
                        _size = _v.get("size") or _v.get("beden") or ""
                        break
        gp_items.append({
            "name": _it.get("productName", ""),
            "barcode": _bc,
            "size": _size,
            "quantity": _it.get("quantity", 1),
            "unit_price": _it.get("unit_price", 0),
            "discount": _it.get("discount_amount", 0),
            "net_price": _it.get("price", 0),
            "reason": _it.get("reason", ""),
        })

    gider_pusulasi = {
        "number": gp_number,
        "display_number": display_number,
        "claim_id": claim_id,
        "order_number": claim.get("order_number", ""),
        "date": datetime.now(timezone.utc).isoformat(),
        "company": company,
        "customer": {
            "name": _cust_name,
            "address": _cust_addr,
            "district": _cust_district,
            "city": _cust_city,
            "country": _cust_country,
        },
        "sales_invoice_no": claim.get("invoice_number", ""),
        "cargo_company": claim.get("cargo_provider_name", ""),
        "sales_rep": "",
        "items": gp_items,
        "totals": {
            "gross": total_gross,
            "discount": total_discount,
            "net": total_net,
            "vat_rate": vat_rate,
            "vat_amount": vat_amount,
            "net_without_vat": net_without_vat,
        },
        "claim_type": claim.get("claim_type", ""),
        "claim_reason": claim.get("claim_reason", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.gider_pusulasi.update_one(
        {"claim_id": claim_id},
        {"$set": gider_pusulasi},
        upsert=True
    )

    await db.trendyol_claims.update_one(
        {"claim_id": claim_id},
        {"$set": {"has_gider_pusulasi": True, "gider_pusulasi_no": gider_pusulasi["display_number"]}}
    )

    return {"success": True, "gider_pusulasi": gider_pusulasi}


@router.post("/trendyol/claims/bulk-gider-pusulasi")
async def bulk_generate_gider_pusulasi(payload: dict, current_user: dict = Depends(require_admin)):
    """Generate expense receipts for multiple claims"""
    claim_ids = payload.get("claim_ids", [])
    if not claim_ids:
        raise HTTPException(status_code=400, detail="Claim ID listesi boş")

    start_no = str(payload.get("start_no") or "").strip()
    try:
        base = int(start_no) if start_no else None
    except ValueError:
        base = None

    results = []
    n = 0
    for cid in claim_ids:
        try:
            tno = f"{base + n:06d}" if base is not None else None
            result = await generate_gider_pusulasi(cid, {"tracking_no": tno} if tno else None, current_user)
            results.append(result.get("gider_pusulasi"))
            n += 1
        except Exception:
            pass

    next_no = f"{base + n:06d}" if base is not None else ""
    return {"success": True, "gider_pusulalari": results, "count": len(results), "next_no": next_no}



# ==================== TRENDYOL STOK & FİYAT GÜNCELLEME ====================

@router.post("/trendyol/products/{product_id}/update-stock-price")
async def update_trendyol_stock_price(
    product_id: str,
    current_user: dict = Depends(require_admin)
):
    """Tek bir ürünün stok ve fiyatını Trendyol'a gönderir."""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")

    from trendyol_client import TrendyolClient
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    # Varyantlı ürün mü?
    items = []
    variants = product.get("variants", [])
    trendyol_multiplier = float(config.get("default_markup", 0) or 0)
    base_price = product.get("price", 0)
    sale_price = base_price  # Trendyol: indirimsiz satis fiyati
    
    if trendyol_multiplier > 0:
        sale_price = sale_price * (1 + trendyol_multiplier / 100)
        base_price = base_price * (1 + trendyol_multiplier / 100)

    if variants:
        for v in variants:
            barcode = v.get("barcode", "")
            if not barcode:
                continue
            v_price = base_price + (v.get("price_diff", 0) or 0)
            v_sale = sale_price + (v.get("price_diff", 0) or 0)
            items.append({
                "barcode": barcode,
                "quantity": v.get("stock", 0),
                "salePrice": round(v_sale, 2),
                "listPrice": round(v_price, 2)
            })
    else:
        barcode = product.get("barcode", "")
        if barcode:
            items.append({
                "barcode": barcode,
                "quantity": product.get("stock", 0),
                "salePrice": round(sale_price, 2),
                "listPrice": round(base_price, 2)
            })

    if not items:
        raise HTTPException(status_code=400, detail="Ürünün barkodu bulunamadı")

    try:
        result = await client.update_price_and_inventory(items)
    except Exception as e:
        logger.error(f"Trendyol stock/price update error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

    batch_id = result.get("batchRequestId", "")
    await db.products.update_one(
        {"id": product_id},
        {"$set": {"trendyol_stock_price_batch": str(batch_id), "trendyol_stock_price_updated": datetime.now(timezone.utc).isoformat()}}
    )

    return {
        "success": True,
        "message": f"{len(items)} kalem stok/fiyat güncellendi",
        "batch_id": batch_id,
        "items_count": len(items)
    }

@router.post("/trendyol/categories/{category_id}/update-stock-price")
async def update_trendyol_category_stock_price(
    category_id: str,
    current_user: dict = Depends(require_admin)
):
    """Bir kategorideki tüm ürünlerin stok ve fiyatlarını Trendyol'a gönderir."""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    # Kategorideki tüm ürünleri bul
    category = await db.categories.find_one({"id": category_id}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Kategori bulunamadı")

    products = await db.products.find(
        {"category_name": category.get("name"), "is_active": True},
        {"_id": 0}
    ).to_list(500)

    if not products:
        # Fallback: try category id
        products = await db.products.find(
            {"category_id": category_id, "is_active": True},
            {"_id": 0}
        ).to_list(500)

    if not products:
        raise HTTPException(status_code=404, detail="Bu kategoride ürün bulunamadı")

    from trendyol_client import TrendyolClient
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    items = []
    for product in products:
        trendyol_multiplier = float(config.get("default_markup", 0) or 0)
        base_price = product.get("price", 0)
        sale_price = base_price  # Trendyol: indirimsiz satis fiyati
        
        if trendyol_multiplier > 0:
            sale_price = sale_price * (1 + trendyol_multiplier / 100)
            base_price = base_price * (1 + trendyol_multiplier / 100)

        variants = product.get("variants", [])
        if variants:
            for v in variants:
                barcode = v.get("barcode", "")
                if not barcode:
                    continue
                v_price = base_price + (v.get("price_diff", 0) or 0)
                v_sale = sale_price + (v.get("price_diff", 0) or 0)
                items.append({
                    "barcode": barcode,
                    "quantity": v.get("stock", 0),
                    "salePrice": round(v_sale, 2),
                    "listPrice": round(v_price, 2)
                })
        else:
            barcode = product.get("barcode", "")
            if barcode:
                items.append({
                    "barcode": barcode,
                    "quantity": product.get("stock", 0),
                    "salePrice": round(sale_price, 2),
                    "listPrice": round(base_price, 2)
                })

    if not items:
        raise HTTPException(status_code=400, detail="Bu kategorideki ürünlerin barkodu bulunamadı")

    # Trendyol max 1000 item per request (v2 client metodu — güncel endpoint)
    batch_ids = []
    for i in range(0, len(items), 1000):
        chunk = items[i:i+1000]
        try:
            result = await client.update_price_and_inventory(chunk)
            batch_ids.append(result.get("batchRequestId", ""))
        except Exception as e:
            logger.error(f"Trendyol category stock/price update error: {str(e)}")

    return {
        "success": True,
        "message": f"{category.get('name')} kategorisindeki {len(items)} kalem stok/fiyat güncellendi",
        "items_count": len(items),
        "batch_ids": batch_ids
    }

# ==================== TRENDYOL KARGO ETİKETİ ====================

@router.get("/trendyol/cargo/label/{shipment_package_id}")
async def get_trendyol_cargo_label_pkg(
    shipment_package_id: str,
    current_user: dict = Depends(require_admin)
):
    """Trendyol kargo etiketi PDF/ZPL verisini getirir."""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    from trendyol_client import TrendyolClient
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    try:
        url = f"{client.base_url}/shipment/sellers/{client.supplier_id}/shipment-packages/{shipment_package_id}/shipping-label"
        async with httpx.AsyncClient(timeout=30.0) as http_client:
            headers = client._get_headers()
            response = await http_client.get(url, headers=headers)
            response.raise_for_status()
            
            content_type = response.headers.get("content-type", "")
            
            if "application/pdf" in content_type:
                return Response(
                    content=response.content,
                    media_type="application/pdf",
                    headers={"Content-Disposition": f"inline; filename=label_{shipment_package_id}.pdf"}
                )
            else:
                # ZPL veya text format
                return Response(
                    content=response.content,
                    media_type=content_type or "application/octet-stream",
                    headers={"Content-Disposition": f"attachment; filename=label_{shipment_package_id}"}
                )
    except httpx.HTTPStatusError as e:
        logger.error(f"Cargo label error: {e.response.status_code} - {e.response.text}")
        raise HTTPException(status_code=e.response.status_code, detail=f"Etiket alınamadı: {e.response.text}")
    except Exception as e:
        logger.error(f"Cargo label error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== TRENDYOL CLAIMS APPROVE/ISSUE ====================

@router.post("/trendyol/claims/{claim_id}/approve")
async def approve_trendyol_claim(
    claim_id: str,
    payload: dict,
    current_user: dict = Depends(require_admin)
):
    """Approve a list of claim items in Trendyol"""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    claim_item_ids = payload.get("claim_item_ids", [])
    if not claim_item_ids:
        raise HTTPException(status_code=400, detail="Onaylanacak iade kalemleri belirtilmedi.")

    from trendyol_client import TrendyolClient
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    try:
        url = f"{client.base_url}/order/sellers/{client.supplier_id}/claims/{claim_id}/items/approve"
        async with httpx.AsyncClient(timeout=30.0) as http_client:
            headers = client._get_headers()
            req_data = {
                "claimLineItemIdList": claim_item_ids
            }
            response = await http_client.put(url, headers=headers, json=req_data)
            response.raise_for_status()
            
            await log_integration_event("trendyol", "claim_approve", current_user["email"], claim_id, "success", f"{len(claim_item_ids)} kalem onaylandı", req_data)
            
            # Update claim in DB with action status
            await db.trendyol_claims.update_one(
                {"claim_id": claim_id},
                {"$set": {
                    "panel_action": "approved",
                    "panel_action_date": datetime.now(timezone.utc).isoformat(),
                    "panel_action_by": current_user["email"],
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )

            # A7: İade onaylanınca stok otomatik geri iade
            try:
                claim_doc = await db.trendyol_claims.find_one({"claim_id": claim_id}, {"_id": 0, "items": 1, "order_number": 1, "stock_restored": 1})
                restocked_items = []
                for item in ([] if (claim_doc or {}).get("stock_restored") else ((claim_doc or {}).get("items") or [])):
                    if str(item.get("claim_item_id", "")) not in [str(x) for x in claim_item_ids]:
                        continue
                    barcode = item.get("barcode", "")
                    qty = int(item.get("quantity", 1) or 1)
                    if not barcode:
                        continue
                    # variant-level restock
                    prod = await db.products.find_one({"variants.barcode": barcode}, {"_id": 0, "id": 1, "variants": 1, "stock": 1})
                    if prod:
                        for v in (prod.get("variants") or []):
                            if v.get("barcode") == barcode:
                                v["stock"] = int(v.get("stock", 0) or 0) + qty
                                break
                        new_total_stock = sum(int(v.get("stock", 0) or 0) for v in prod.get("variants", []))
                        await db.products.update_one(
                            {"id": prod["id"]},
                            {"$set": {"variants": prod["variants"], "stock": new_total_stock, "updated_at": datetime.now(timezone.utc).isoformat()}}
                        )
                        restocked_items.append({"barcode": barcode, "qty": qty, "product_id": prod["id"]})
                    else:
                        # Fallback: try product.barcode
                        p2 = await db.products.find_one({"barcode": barcode}, {"_id": 0, "id": 1, "stock": 1})
                        if p2:
                            await db.products.update_one(
                                {"id": p2["id"]},
                                {"$inc": {"stock": qty}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
                            )
                            restocked_items.append({"barcode": barcode, "qty": qty, "product_id": p2["id"]})

                # Stok hareketi loglama
                if restocked_items:
                    await db.stock_movements.insert_one({
                        "id": str(uuid.uuid4()),
                        "type": "return_approved",
                        "claim_id": claim_id,
                        "order_number": claim_doc.get("order_number", ""),
                        "items": restocked_items,
                        "created_by": current_user["email"],
                        "created_at": datetime.now(timezone.utc).isoformat(),
                    })
                    await db.trendyol_claims.update_one(
                        {"claim_id": claim_id},
                        {"$set": {"stock_restored": True, "stock_restored_at": datetime.now(timezone.utc).isoformat(), "stock_restored_source": "approve"}}
                    )
            except Exception as restock_err:
                logger.error(f"Restock after claim approve failed: {restock_err}")
                # non-fatal

            return {"success": True, "message": "İade işlemi Trendyol tarafında onaylandı."}

    except httpx.HTTPStatusError as e:
        logger.error(f"Claim approve error: {e.response.status_code} - {e.response.text}")
        await log_integration_event("trendyol", "claim_approve", current_user["email"], claim_id, "error", f"API Hatası: {e.response.text}")
        raise HTTPException(status_code=e.response.status_code, detail=f"Onay işlemi başarısız: {e.response.text}")
    except Exception as e:
        logger.error(f"Claim approve error: {str(e)}")
        await log_integration_event("trendyol", "claim_approve", current_user["email"], claim_id, "error", str(e))
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/trendyol/claims/{claim_id}/issue")
async def issue_trendyol_claim(
    claim_id: str,
    payload: dict,
    current_user: dict = Depends(require_admin)
):
    """Reject/Issue a list of claim items in Trendyol"""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    claim_item_ids = payload.get("claim_item_ids", [])
    issue_reason_id = payload.get("issue_reason_id")
    description = payload.get("description", "")

    if not claim_item_ids or not issue_reason_id:
        raise HTTPException(status_code=400, detail="İtiraz edilecek kalemler veya itiraz sebebi belirtilmedi.")

    from trendyol_client import TrendyolClient
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    try:
        url = f"{client.base_url}/order/sellers/{client.supplier_id}/claims/{claim_id}/issue"
        async with httpx.AsyncClient(timeout=30.0) as http_client:
            headers = client._get_headers()
            req_data = {
                "claimIssueReasonId": int(issue_reason_id),
                "claimItemIdList": claim_item_ids,
                "description": description
            }
            response = await http_client.post(url, headers=headers, json=req_data)
            response.raise_for_status()
            
            await log_integration_event("trendyol", "claim_issue", current_user["email"], claim_id, "success", f"{len(claim_item_ids)} kalem için itiraz açıldı", req_data)
            
            # Update claim in DB with action status
            await db.trendyol_claims.update_one(
                {"claim_id": claim_id},
                {"$set": {
                    "panel_action": "issued",
                    "panel_action_date": datetime.now(timezone.utc).isoformat(),
                    "panel_action_by": current_user["email"],
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            return {"success": True, "message": "İade işlemi için Trendyol tarafında itiraz oluşturuldu."}

    except httpx.HTTPStatusError as e:
        logger.error(f"Claim issue error: {e.response.status_code} - {e.response.text}")
        await log_integration_event("trendyol", "claim_issue", current_user["email"], claim_id, "error", f"API Hatası: {e.response.text}")
        raise HTTPException(status_code=e.response.status_code, detail=f"İtiraz işlemi başarısız: {e.response.text}")
    except Exception as e:
        logger.error(f"Claim issue error: {str(e)}")
        await log_integration_event("trendyol", "claim_issue", current_user["email"], claim_id, "error", str(e))
        raise HTTPException(status_code=500, detail=str(e))


# ==================== TRENDYOL Q&A + REVIEWS ====================
# Iteration 37 refactor: Q&A (3 endpoint) + Reviews (2 endpoint) ayrı modüle
# taşındı: /app/backend/routes/integrations_trendyol_qna.py
# server.py'de ayrı router olarak include edilir.





# ==================== TRENDYOL INVOICE ====================

@router.post("/trendyol/invoices/{order_number}")
async def upload_invoice_to_trendyol(order_number: str, payload: dict, current_user: dict = Depends(require_admin)):
    """Upload invoice link to Trendyol for a given order number"""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapilandirilmamis")

    invoice_link = payload.get("invoice_link", "").strip()
    invoice_number = payload.get("invoice_number", "").strip()
    if not invoice_link:
        raise HTTPException(status_code=400, detail="Fatura linki bos olamaz")

    order = await db.orders.find_one({"order_number": order_number, "platform": "trendyol"})
    if not order:
        raise HTTPException(status_code=404, detail="Siparis bulunamadi")

    package_id = order.get("trendyol_package_id")
    if not package_id:
        raise HTTPException(status_code=400, detail="Trendyol paket ID bulunamadi")

    supplier_id = config["supplier_id"]
    headers = await get_trendyol_headers()
    base_url = config["base_url"]

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            url = f"{base_url}/sapigw/suppliers/{supplier_id}/shipment-packages/{package_id}/invoices"
            body = {
                "invoiceNumber": invoice_number or f"FAT-{order_number}",
                "invoiceLink": invoice_link
            }
            resp = await client.post(url, headers=headers, json=body)
            resp.raise_for_status()

        await db.orders.update_one(
            {"order_number": order_number},
            {"$set": {"invoice_link": invoice_link, "invoice_number": invoice_number, "invoice_uploaded_at": datetime.now(timezone.utc).isoformat()}}
        )

        await log_integration_event("trendyol", "upload_invoice", current_user["email"], order_number, "success", "Fatura yuklendi", body)
        return {"success": True, "message": "Fatura Trendyol'a basariyla yuklendi"}
    except httpx.HTTPStatusError as e:
        logger.error(f"Invoice upload error: {e.response.text}")
        await log_integration_event("trendyol", "upload_invoice", current_user["email"], order_number, "error", e.response.text)
        raise HTTPException(status_code=e.response.status_code, detail=f"Fatura yukleme hatasi: {e.response.text}")
    except Exception as e:
        logger.error(f"Invoice upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/trendyol/products/{product_id}/sync")
async def sync_product_to_trendyol(product_id: str, current_user: dict = Depends(require_admin)):
    """Full product synchronization to Trendyol"""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    from datetime import datetime, timezone
    started_at = datetime.now(timezone.utc).isoformat()
    product = None
    
    try:
        product = await db.products.find_one({"id": product_id}, {"_id": 0})
        if not product:
            raise Exception("Ürün bulunamadı")

        ty_cat_id = product.get("trendyol_category_id")
        if not ty_cat_id:
            # Try finding in the category mapped to this product
            cat = await db.categories.find_one({"name": product.get("category_name")})
            if not cat:
                cat = await db.categories.find_one({"id": product.get("category_id")})
            
            if cat and cat.get("trendyol_category_id"):
                ty_cat_id = cat.get("trendyol_category_id")
            else:
                raise Exception("Ürün için Trendyol kategorisi seçilmemiş")

        # Fetch mapping details from category
        mapping_cat = await db.categories.find_one({"trendyol_category_id": ty_cat_id})
        attr_mappings = mapping_cat.get("attribute_mappings", []) if mapping_cat else []
        val_mappings = mapping_cat.get("value_mappings", {}) if mapping_cat else {}
        default_mappings = mapping_cat.get("default_mappings", {}) if mapping_cat else {}

        from trendyol_client import TrendyolClient
        client = TrendyolClient(
            supplier_id=config["supplier_id"],
            api_key=config["api_key"],
            api_secret=config["api_secret"],
            mode=config["mode"]
        )

        # Calculate prices
        base_price = product.get("price", 0)
        list_price = calculate_trendyol_price(base_price, product, config)
        sale_price = calculate_trendyol_price(base_price, product, config)  # Trendyol: indirimsiz satis fiyati

        # Build items
        items = []
        variants = product.get("variants", [])
        
        # Common attributes for all variants
        common_attrs = []
        for am in attr_mappings:
            ty_attr_id = int(am["trendyol_attr_id"])
            local_name = am["local_attr"]
            # Find value in product attributes
            val = next((a["value"] for a in product.get("attributes", []) if a["type"] == local_name), None)
            # Try default if not found
            if not val:
                val = default_mappings.get(str(ty_attr_id))
                
            if val:
                mapping_key = f"{ty_attr_id}:{val}"
                ty_val_id = val_mappings.get(mapping_key)
                if ty_val_id:
                    common_attrs.append({"attributeId": ty_attr_id, "attributeValueId": int(ty_val_id)})
                else:
                    common_attrs.append({"attributeId": ty_attr_id, "customAttributeValue": val})

        if variants:
            for v in variants:
                v_attrs = common_attrs.copy()
                
                # Map Size (Beden) and Color (Renk)
                for am in attr_mappings:
                    ty_attr_id = str(am["trendyol_attr_id"])
                    local_name = am["local_attr"]
                    
                    # Check if it's Beden or Renk
                    if local_name.lower() == "beden":
                        sz = v.get("size")
                        if sz:
                            m_key = f"{ty_attr_id}:{sz}"
                            v_id = val_mappings.get(m_key)
                            if v_id:
                                v_attrs.append({"attributeId": int(ty_attr_id), "attributeValueId": int(v_id)})
                            else:
                                v_attrs.append({"attributeId": int(ty_attr_id), "customAttributeValue": sz})
                    
                    elif local_name.lower() == "renk":
                        clr = v.get("color")
                        if clr:
                            m_key = f"{ty_attr_id}:{clr}"
                            v_id = val_mappings.get(m_key)
                            if v_id:
                                v_attrs.append({"attributeId": int(ty_attr_id), "attributeValueId": int(v_id)})
                            else:
                                v_attrs.append({"attributeId": int(ty_attr_id), "customAttributeValue": clr})

                # Pricing with price_diff
                diff = float(v.get("price_diff", 0) or 0)
                v_list = round(list_price + diff, 2)
                v_sale = round(sale_price + diff, 2)

                item = {
                    "barcode": v.get("barcode") or product.get("barcode"),
                    "title": product.get("name"),
                    "productMainId": product.get("stock_code"),
                    "brandId": product.get("trendyol_brand_id") or 975755,
                    "categoryId": int(ty_cat_id),
                    "quantity": v.get("stock", 0),
                    "stockCode": v.get("stock_code") or product.get("stock_code"),
                    "dimensionalWeight": product.get("cargo_weight") or 1,
                    "description": product.get("description", ""),
                    "currencyType": "TRY",
                    "listPrice": v_list,
                    "salePrice": v_sale,
                    "vatRate": product.get("vat_rate", 20),
                    "cargoCompanyId": 10,
                    "images": [{"url": img} for img in product.get("images", [])],
                    "attributes": v_attrs
                }
                items.append(item)
        else:
            # Single product
            item = {
                "barcode": product.get("barcode"),
                "title": product.get("name"),
                "productMainId": product.get("stock_code"),
                "brandId": product.get("trendyol_brand_id") or 975755,
                "categoryId": int(ty_cat_id),
                "quantity": product.get("stock", 0),
                "stockCode": product.get("stock_code"),
                "dimensionalWeight": product.get("cargo_weight") or 1,
                "description": product.get("description", ""),
                "currencyType": "TRY",
                "listPrice": list_price,
                "salePrice": sale_price,
                "vatRate": product.get("vat_rate", 20),
                "cargoCompanyId": 10,
                "images": [{"url": img} for img in product.get("images", [])],
                "attributes": common_attrs
            }
            items.append(item)

        result = await client.create_products(items)
        batch_id = result.get("batchRequestId", "")
        
        # Log to the new sync logs screen
        log_doc = {
            "id": generate_id(),
            "started_at": started_at,
            "finished_at": datetime.now(timezone.utc).isoformat(),
            "status": "success",
            "products_attempted": 1,
            "products_sent": len(items),
            "batch_request_id": batch_id,
            "errors": [],
            "message": f"'{product.get('name')}' ürünü tekli olarak aktarıldı."
        }
        await db.trendyol_sync_logs.insert_one(log_doc)

        await log_integration_event(
            platform="trendyol",
            action="product_sync",
            entity_type="product",
            entity_id=product_id,
            status="success",
            message=f"Sync initiated. Batch ID: {batch_id}",
            details={"batch_id": batch_id, "items_count": len(items)}
        )
        
        await db.products.update_one(
            {"id": product_id},
            {"$set": {
                "trendyol_sync_batch": str(batch_id),
                "trendyol_sync_last": datetime.now(timezone.utc).isoformat(),
                "trendyol_status": "synced"
            }}
        )
        
        return {"success": True, "message": "Eşleştirme başlatıldı", "batch_id": batch_id}
    except Exception as e:
        logger.error(f"Trendyol sync error: {str(e)}")
        log_doc = {
            "id": generate_id(),
            "started_at": started_at if 'started_at' in locals() else datetime.now(timezone.utc).isoformat(),
            "status": "error",
            "products_attempted": 1,
            "products_sent": 0,
            "batch_request_id": None,
            "errors": [f"Hata: {str(e)}"],
            "message": f"'{product.get('name') if product else product_id}' aktarımı sırasında hata oluştu."
        }
        await db.trendyol_sync_logs.insert_one(log_doc)
        await log_integration_event("trendyol", "product_sync", "product", product_id, "error", str(e))
        raise HTTPException(status_code=400 if "Ürün" in str(e) or "kategori" in str(e).lower() else 500, detail=f"Trendyol senkronizasyon hatası: {str(e)}")



# Doğan e-Dönüşüm endpoint'leri ayrı modüle taşındı: integrations_dogan.py
# (Iteration 35 refactor — server.py'de ayrı router olarak include edilir)


# ==================== HEPSIBURADA & TEMU (Marketplace) ====================
# Generic marketplace settings + questions endpoints.
# Real API calls can be filled in later; UI scaffolding is ready.

ALLOWED_MARKETPLACES = {
    # customer message channels
    "hepsiburada", "temu", "whatsapp", "instagram", "messenger", "site",
    # cargo providers
    "mng", "aras", "yurtici", "ptt", "hepsijet", "trendyol_express", "surat", "ups", "dhl",
}

@router.get("/{marketplace}/settings")
async def get_marketplace_settings(marketplace: str, current_user: dict = Depends(require_admin)):
    """Get Hepsiburada / Temu settings (generic)."""
    if marketplace not in ALLOWED_MARKETPLACES:
        raise HTTPException(status_code=404, detail="Bilinmeyen pazaryeri")
    settings = await db.settings.find_one({"id": marketplace}, {"_id": 0})
    if not settings:
        return {
            "id": marketplace,
            "merchant_id": "",
            "username": "",
            "api_key": "",
            "api_secret": "",
            "mode": "sandbox",
            "is_active": False,
            "default_markup": 0
        }
    # Mask secret
    if settings.get("api_secret"):
        settings["api_secret"] = "********"
    if settings.get("password"):
        settings["password"] = "********"
    if settings.get("secret_key"):
        settings["secret_key"] = "********"
    return settings


@router.post("/{marketplace}/settings")
async def save_marketplace_settings(marketplace: str, payload: dict, current_user: dict = Depends(require_admin)):
    """Save Hepsiburada / Temu settings (generic)."""
    if marketplace not in ALLOWED_MARKETPLACES:
        raise HTTPException(status_code=404, detail="Bilinmeyen pazaryeri")

    # Required alan validasyonu — is_active=True ise pazaryeri bazlı zorunlu alanlar
    if payload.get("is_active"):
        existing = await db.settings.find_one({"id": marketplace}, {"_id": 0}) or {}
        fields_by_mp = {
            "hepsiburada": ["merchant_id", "secret_key", "dev_username"],
            "temu": ["api_key", "api_secret"],
        }
        required = fields_by_mp.get(marketplace, ["api_key", "api_secret"])
        missing = []
        for k in required:
            v = payload.get(k)
            if v in (None, "", "********"):
                v = existing.get(k)
            if not v:
                missing.append(k)
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"{marketplace.capitalize()} aktifleştirmek için zorunlu alanlar eksik: {', '.join(missing)}"
            )

    update_data = {
        "id": marketplace,
        "merchant_id": payload.get("merchant_id", ""),
        "username": payload.get("username", ""),
        "dev_username": payload.get("dev_username", ""),
        "api_key": payload.get("api_key", ""),
        "mode": payload.get("mode", "sandbox"),
        "is_active": payload.get("is_active", False),
        "default_markup": payload.get("default_markup", 0),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    # Only update secret if provided
    if payload.get("api_secret") and payload.get("api_secret") != "********":
        update_data["api_secret"] = payload.get("api_secret")
    if payload.get("password") and payload.get("password") != "********":
        update_data["password"] = payload.get("password")
    if payload.get("secret_key") and payload.get("secret_key") != "********":
        update_data["secret_key"] = payload.get("secret_key")

    await db.settings.update_one({"id": marketplace}, {"$set": update_data}, upsert=True)
    return {"success": True, "message": f"{marketplace.capitalize()} ayarları kaydedildi"}


@router.get("/{marketplace}/status")
async def get_marketplace_status(marketplace: str):
    """Get Hepsiburada / Temu integration status."""
    if marketplace not in ALLOWED_MARKETPLACES:
        raise HTTPException(status_code=404, detail="Bilinmeyen pazaryeri")
    settings = await db.settings.find_one({"id": marketplace}, {"_id": 0})
    if not settings:
        return {"configured": False, "mode": "sandbox"}
    mode_raw = (settings.get("mode") or "sandbox").strip().lower()
    is_live = mode_raw in ("live", "production", "prod", "canli", "canlı")
    if marketplace == "hepsiburada":
        configured = bool(settings.get("is_active") and settings.get("merchant_id")
                          and (settings.get("secret_key") or settings.get("password"))
                          and settings.get("dev_username"))
    else:
        configured = bool(settings.get("is_active") and (settings.get("api_key") or settings.get("merchant_id")))
    return {
        "configured": configured,
        "mode": "live" if is_live else "sandbox",
        "merchant_id": settings.get("merchant_id", "") if settings.get("is_active") else None
    }


@router.post("/{marketplace}/test-connection")
async def test_marketplace_connection(marketplace: str, current_user: dict = Depends(require_admin)):
    """REAL connection test — Hepsiburada/Temu use Basic Auth against a known probe endpoint.
    Returns success=False with a concrete error message when credentials fail or are missing."""
    import httpx
    import base64
    if marketplace not in ALLOWED_MARKETPLACES:
        raise HTTPException(status_code=404, detail="Bilinmeyen pazaryeri")
    settings = await db.settings.find_one({"id": marketplace}, {"_id": 0})
    if not settings:
        return {"success": False, "message": f"{marketplace.capitalize()} ayarları kaydedilmemiş"}

    try:
        if marketplace == "hepsiburada":
            merchant_id = (settings.get("merchant_id") or "").strip()
            # Secret Key (yeni model). Eski kayitlarda 'password' alaninda olabilir.
            secret_key = (settings.get("secret_key") or settings.get("password") or "").strip()
            dev_username = (settings.get("dev_username") or "").strip()
            if not (merchant_id and secret_key and dev_username):
                return {"success": False, "message": "Hepsiburada için Merchant ID, Secret Key ve Developer Username zorunlu"}
            mode = settings.get("mode", "sandbox")
            host = "https://mpop-sit.hepsiburada.com" if mode == "sandbox" else "https://mpop.hepsiburada.com"
            url = f"{host}/product/api/categories/get-all-categories?page=0&size=1"
            # Basic Auth: kullanici adi = Merchant ID, sifre = Secret Key. User-Agent = Developer Username.
            token = base64.b64encode(f"{merchant_id}:{secret_key}".encode()).decode()
            async with httpx.AsyncClient(timeout=20) as c:
                r = await c.get(url, headers={"Authorization": f"Basic {token}", "User-Agent": dev_username, "Accept": "application/json"})
            if r.status_code == 200:
                return {"success": True, "message": f"Hepsiburada bağlantısı başarılı ({mode})"}
            if r.status_code in (401, 403):
                return {"success": False, "message": f"Hepsiburada kimlik hatalı (HTTP {r.status_code}). Merchant ID / Secret Key / Developer Username kontrol edin."}
            # diger durumlarda body'den mesaj parse et
            try:
                err_body = r.json()
                err_msg = err_body.get("message") or err_body.get("errorMessage") or err_body.get("detail") or ""
                err_code = err_body.get("errorCode") or err_body.get("code") or ""
                detail_txt = f" — {err_code}: {err_msg}" if (err_code or err_msg) else ""
            except Exception:
                detail_txt = f" — {r.text[:120]}"
            return {"success": False, "message": f"Hepsiburada beklenmeyen yanıt: HTTP {r.status_code}{detail_txt}"}

        if marketplace == "temu":
            shop_id = (settings.get("merchant_id") or "").strip()
            api_key = (settings.get("api_key") or "").strip()
            app_secret = (settings.get("api_secret") or "").strip()
            if not (shop_id and api_key and app_secret):
                return {"success": False, "message": "Temu için Shop ID, App Key ve App Secret zorunlu"}
            if len(api_key) < 8 or len(app_secret) < 8:
                return {"success": False, "message": "Temu App Key/Secret çok kısa, doğru girdiğinize emin olun"}
            # Gerçek Temu Open Platform probe — bg.temu.com /api/v1/seller/info endpoint'i
            import time
            import json as _json
            mode_ = settings.get("mode", "sandbox")
            # Temu Open Platform hostları — sandbox vs live farklı
            host = (
                "https://openapi-b-us.temu.com"
                if mode_ == "live"
                else "https://openapi-b-global-stg.temu.com"
            )
            ts = str(int(time.time()))
            body = {
                "type": "bg.auth.access_token.info.get",
                "app_key": api_key,
                "timestamp": ts,
            }
            sign_base = app_secret + "".join(f"{k}{v}" for k, v in sorted(body.items())) + app_secret
            # NOT: Temu Open Platform imza algoritması olarak MD5 ZORUNLU kılar
            # (API kontratı gereği; güvenlik tercihi değil — değiştirmek imzayı bozar).
            body["sign"] = hashlib.md5(sign_base.encode()).hexdigest().upper()
            try:
                async with httpx.AsyncClient(timeout=15) as c:
                    r = await c.post(f"{host}/openapi/router", json=body)
                try:
                    data = r.json()
                except Exception:
                    data = {"raw": r.text[:200]}
                if data.get("success") or data.get("result"):
                    return {"success": True, "message": f"Temu bağlantısı başarılı ({mode_})"}
                return {"success": False, "message": f"Temu hata: {data.get('errorMsg') or data.get('errorCode') or str(data)[:200]}"}
            except Exception as e:
                return {"success": False, "message": f"Temu probe hatası: {str(e)[:150]}"}

        if marketplace in {"mng", "aras", "yurtici", "ptt", "hepsijet", "trendyol_express", "surat"}:
            user = (settings.get("username") or "").strip()
            pw = (settings.get("password") or "").strip()
            key = (settings.get("api_key") or "").strip()
            if not (user or key) or not (pw or settings.get("customer_code")):
                return {"success": False, "message": f"{marketplace.upper()} kimlik bilgileri eksik"}
            return {"success": True, "message": f"{marketplace.upper()} kimlik bilgileri kaydedildi (canlı test yapılmadı)"}

        # Default fallback for other marketplaces (whatsapp/instagram/site channels etc.)
        if not (settings.get("api_key") or settings.get("username")):
            return {"success": False, "message": f"{marketplace} kimlik bilgileri eksik"}
        return {"success": True, "message": f"{marketplace} ayarları geçerli"}
    except httpx.TimeoutException:
        return {"success": False, "message": f"{marketplace} API zaman aşımına uğradı (15s)"}
    except Exception as e:
        return {"success": False, "message": f"{marketplace} test hatası: {str(e)[:150]}"}


# -------- Unified Marketplace Questions (Trendyol + Hepsiburada + Temu) --------

QUESTIONS_COLLECTIONS = {
    "trendyol": "trendyol_questions",
    "hepsiburada": "hepsiburada_questions",
    "temu": "temu_questions",
    "whatsapp": "whatsapp_messages",
    "instagram": "instagram_messages",
    "messenger": "messenger_messages",
    "site": "site_messages",
}


@router.delete("/{marketplace}/questions/{question_id}")
async def delete_marketplace_question(marketplace: str, question_id: str, current_user: dict = Depends(require_admin)):
    """Delete a question/message (e.g., expired unanswered)."""
    if marketplace not in QUESTIONS_COLLECTIONS:
        raise HTTPException(status_code=404, detail="Bilinmeyen kanal")
    coll = db[QUESTIONS_COLLECTIONS[marketplace]]
    res = await coll.delete_one({"question_id": question_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Mesaj bulunamadı")
    return {"success": True}


@router.get("/marketplace/questions")
async def get_marketplace_questions(
    marketplace: Optional[str] = "all",
    status: Optional[str] = None,
    page: int = 0,
    size: int = 20,
    current_user: dict = Depends(require_admin)
):
    """Unified endpoint returning questions tagged with their marketplace."""
    query = {}
    if status:
        query["status"] = status

    skip = page * size
    sources = []
    if marketplace and marketplace != "all":
        if marketplace not in QUESTIONS_COLLECTIONS:
            raise HTTPException(status_code=400, detail="Bilinmeyen pazaryeri")
        sources = [marketplace]
    else:
        sources = list(QUESTIONS_COLLECTIONS.keys())

    all_items = []
    totals = {}
    for mp in sources:
        coll = db[QUESTIONS_COLLECTIONS[mp]]
        # Pull enough for merge+sort; cap per marketplace to avoid huge reads
        items = await coll.find(query, {"_id": 0}).sort("created_at", -1).limit(500).to_list(500)
        for it in items:
            it["marketplace"] = mp
        all_items.extend(items)
        totals[mp] = await coll.count_documents(query)

    # Sort merged by created_at desc
    def _key(x):
        return x.get("created_at") or ""
    all_items.sort(key=_key, reverse=True)
    total = sum(totals.values())
    paginated = all_items[skip: skip + size]

    return {
        "questions": paginated,
        "total": total,
        "totals": totals,
        "page": page,
        "size": size,
    }


@router.post("/{marketplace}/questions/sync")
async def sync_marketplace_questions_stub(marketplace: str, current_user: dict = Depends(require_admin)):
    """Stub sync for HB/Temu until real API integration is wired up."""
    if marketplace not in ALLOWED_MARKETPLACES:
        raise HTTPException(status_code=404, detail="Bilinmeyen pazaryeri")
    settings = await db.settings.find_one({"id": marketplace}, {"_id": 0})
    if not settings or not settings.get("is_active"):
        raise HTTPException(status_code=400, detail=f"{marketplace.capitalize()} entegrasyonu yapılandırılmamış")
    # Real implementation: call HB / Temu QNA API and upsert into respective collection
    return {"success": True, "synced": 0, "total_fetched": 0, "message": "API entegrasyonu yapılandırıldığında otomatik çalışacak (stub)"}


@router.post("/{marketplace}/questions/{question_id}/answer")
async def answer_marketplace_question_stub(
    marketplace: str,
    question_id: str,
    payload: dict,
    current_user: dict = Depends(require_admin)
):
    """Answer HB / Temu question – currently stores locally; real API wire-up to follow."""
    if marketplace not in ALLOWED_MARKETPLACES:
        raise HTTPException(status_code=404, detail="Bilinmeyen pazaryeri")
    answer_text = (payload or {}).get("answer", "").strip()
    if not answer_text:
        raise HTTPException(status_code=400, detail="Yanıt metni boş olamaz")
    coll = db[QUESTIONS_COLLECTIONS[marketplace]]
    await coll.update_one(
        {"question_id": question_id},
        {"$set": {
            "answer": answer_text,
            "status": "ANSWERED",
            "answered_at": datetime.now(timezone.utc).isoformat(),
        }}
    )
    return {"success": True, "message": "Cevap kaydedildi (yerel). API entegrasyonu tamamlandığında otomatik gönderilecek."}
