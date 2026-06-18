"""
Product routes - CRUD, search, filtering
"""
from fastapi import APIRouter, HTTPException, Query, Depends, Request
from typing import List, Optional
from datetime import datetime, timezone
import re

from .deps import db, logger, get_current_user, require_admin, generate_id, generate_short_id, generate_barcode_from_range, build_used_barcode_set, generate_urun_karti_id
from ticimax_schema import BOOL_COLS as TICIMAX_BOOL_COLS
from fastapi import Response, UploadFile, File
import pandas as pd
import io

router = APIRouter(prefix="/products", tags=["Products"])


# ---------------------------------------------------------------------------
# XML ürün feed'leri (Google Merchant / Facebook Katalog / Genel)
# Çoklu feed: her feed bir "target" ile public URL üretir → /products/feed/<slug>.xml
# google/generic = ürün-seviyesi (g:id = ürün id)
# facebook        = varyant-seviyesi (g:id = varyant id, item_group_id = ürün id,
#                    g:size/g:color) → Meta pixel content_id'leriyle birebir eşleşir.
# ---------------------------------------------------------------------------
import html as _html


def _build_merchant_xml(prods, site, shop, target="google", in_stock_only=False):
    """Ürün listesinden Google/Facebook uyumlu RSS 2.0 (g:) XML üretir."""
    target = (target or "google").lower()

    def esc(x):
        return _html.escape(str(x if x is not None else ""), quote=True)

    def first_image(pr):
        for im in (pr.get("images") or []):
            if isinstance(im, str) and im:
                return im
            if isinstance(im, dict):
                u = im.get("url") or im.get("src") or im.get("image")
                if u:
                    return u
        return pr.get("image") or ""

    def abs_url(u):
        if not u:
            return ""
        if str(u).startswith("http"):
            return u
        return site + ("" if str(u).startswith("/") else "/") + str(u)

    def fnum(x):
        try:
            return float(x)
        except Exception:
            return 0.0

    items = []
    for pr in prods:
        pid = pr.get("id") or pr.get("slug")
        if not pid:
            continue
        slug = pr.get("slug") or pid
        link = f"{site}/urun/{slug}"
        img = abs_url(first_image(pr))
        price = fnum(pr.get("price"))
        sale = pr.get("sale_price")
        sale = fnum(sale) if sale not in (None, "", 0) else None
        name = (pr.get("name") or "")[:150]
        desc = re.sub(r"<[^>]+>", " ", str(pr.get("description") or pr.get("short_description") or pr.get("name") or ""))
        desc = re.sub(r"\s+", " ", desc).strip()[:4500] or name
        brand = pr.get("brand") or shop
        cat = pr.get("category_name") or ""
        pcolor = pr.get("color") or ""
        variants = pr.get("variants") or []

        def price_rows(p_price, p_sale):
            r = []
            if p_sale and p_sale > 0 and p_price and p_sale < p_price:
                r.append(f"<g:price>{p_price:.2f} TRY</g:price>")
                r.append(f"<g:sale_price>{p_sale:.2f} TRY</g:sale_price>")
            else:
                eff = p_sale if (p_sale and p_sale > 0) else p_price
                r.append(f"<g:price>{eff:.2f} TRY</g:price>")
            return r

        def common_rows(item_id, avail, gtin, mpn, extra=None):
            r = [
                f"<g:id>{esc(item_id)}</g:id>",
                f"<g:title>{esc(name)}</g:title>",
                f"<g:description>{esc(desc)}</g:description>",
                f"<g:link>{esc(link)}</g:link>",
            ]
            if img:
                r.append(f"<g:image_link>{esc(img)}</g:image_link>")
            r.append(f"<g:availability>{avail}</g:availability>")
            r.append("<g:condition>new</g:condition>")
            if brand:
                r.append(f"<g:brand>{esc(brand)}</g:brand>")
            has_id = False
            if gtin and str(gtin).isdigit() and len(str(gtin)) in (8, 12, 13, 14):
                r.append(f"<g:gtin>{esc(gtin)}</g:gtin>")
                has_id = True
            if mpn:
                r.append(f"<g:mpn>{esc(mpn)}</g:mpn>")
                has_id = True
            if not has_id:
                r.append("<g:identifier_exists>no</g:identifier_exists>")
            if cat:
                r.append(f"<g:product_type>{esc(cat)}</g:product_type>")
            if extra:
                r.extend(extra)
            return r

        # ---- Facebook: varyant (beden) bazlı satırlar ----
        if target == "facebook" and variants:
            for v in variants:
                vid = v.get("id")
                vstock = 0
                try:
                    vstock = int(v.get("stock") or 0)
                except Exception:
                    vstock = 0
                if in_stock_only and vstock <= 0:
                    continue
                vsize = v.get("size") or ""
                vcolor = v.get("color") or pcolor
                vbarcode = (v.get("barcode") or "").strip()
                item_id = str(vid) if (vid is not None and str(vid) != "") else f"{pid}-{esc(vsize)}"
                vprice = fnum(v.get("price")) or price
                vsale = v.get("sale_price")
                vsale = fnum(vsale) if vsale not in (None, "", 0) else sale
                avail = "in stock" if vstock > 0 else "out of stock"
                extra = [f"<g:item_group_id>{esc(pid)}</g:item_group_id>"]
                if vsize:
                    extra.append(f"<g:size>{esc(vsize)}</g:size>")
                if vcolor:
                    extra.append(f"<g:color>{esc(vcolor)}</g:color>")
                rows = common_rows(item_id, avail, vbarcode, (pr.get("stock_code") or pr.get("sku") or "").strip(), extra)
                rows.extend(price_rows(vprice, vsale))
                items.append("<item>" + "".join(rows) + "</item>")
            continue

        # ---- Google / Generic: ürün-seviyesi tek satır ----
        stock = pr.get("stock") or 0
        if variants:
            try:
                vsum = sum(int(v.get("stock") or 0) for v in variants)
                if vsum:
                    stock = vsum
            except Exception:
                pass
        if in_stock_only and not (stock and stock > 0):
            continue
        avail = "in stock" if (stock and stock > 0) else "out of stock"
        gtin = (pr.get("barcode") or "").strip()
        mpn = (pr.get("stock_code") or pr.get("sku") or "").strip()
        rows = common_rows(pid, avail, gtin, mpn)
        rows.extend(price_rows(price, sale))
        items.append("<item>" + "".join(rows) + "</item>")

    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<rss xmlns:g="http://base.google.com/ns/1.0" version="2.0"><channel>'
        f"<title>{esc(shop)}</title><link>{esc(site)}</link>"
        f"<description>{esc(shop)} urun feed</description>"
        + "".join(items) +
        "</channel></rss>"
    )


async def _feed_site_shop_prods():
    main = await db.settings.find_one({"id": "main"}, {"_id": 0}) or {}
    site = (main.get("site_url") or "https://facette.com.tr").rstrip("/")
    shop = main.get("site_name") or "FACETTE"
    prods = await db.products.find({"is_active": True, "is_deleted": {"$ne": True}}, {"_id": 0}).to_list(50000)
    return site, shop, prods


@router.get("/google-merchant-feed.xml")
async def google_merchant_feed():
    """Geriye-uyumlu varsayilan feed (tum aktif urunler, urun-seviyesi)."""
    site, shop, prods = await _feed_site_shop_prods()
    xml = _build_merchant_xml(prods, site, shop, "google", False)
    return Response(content=xml, media_type="application/xml; charset=utf-8")


@router.get("/feed/{slug}.xml")
async def dynamic_feed(slug: str):
    """Yapilandirilmis XML feed (slug ile). target'a gore urun/varyant seviyesi."""
    feed = await db.xml_feeds.find_one({"slug": slug}, {"_id": 0})
    if not feed or not feed.get("enabled", True):
        return Response(content='<?xml version="1.0"?><error>feed not found</error>',
                        media_type="application/xml", status_code=404)
    site, shop, prods = await _feed_site_shop_prods()
    xml = _build_merchant_xml(prods, site, shop, feed.get("target", "google"), bool(feed.get("in_stock_only")))
    return Response(content=xml, media_type="application/xml; charset=utf-8")


@router.get("/feeds", dependencies=[Depends(require_admin)])
async def list_feeds():
    return await db.xml_feeds.find({}, {"_id": 0}).sort("created_at", 1).to_list(200)


@router.post("/feeds", dependencies=[Depends(require_admin)])
async def create_feed(payload: dict):
    name = (payload.get("name") or "").strip() or "Yeni Feed"
    target = (payload.get("target") or "google").strip().lower()
    if target not in ("google", "facebook", "generic"):
        target = "google"
    base = generate_slug(name) or "feed"
    s = base
    i = 2
    while await db.xml_feeds.find_one({"slug": s}):
        s = f"{base}-{i}"
        i += 1
    doc = {
        "id": generate_id(), "name": name, "slug": s, "target": target,
        "enabled": bool(payload.get("enabled", True)),
        "in_stock_only": bool(payload.get("in_stock_only", False)),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.xml_feeds.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.put("/feeds/{fid}", dependencies=[Depends(require_admin)])
async def update_feed(fid: str, payload: dict):
    upd = {}
    for k in ("name", "target", "enabled", "in_stock_only"):
        if k in payload:
            upd[k] = payload[k]
    if upd.get("target") and upd["target"] not in ("google", "facebook", "generic"):
        upd["target"] = "google"
    if upd:
        await db.xml_feeds.update_one({"id": fid}, {"$set": upd})
    f = await db.xml_feeds.find_one({"id": fid}, {"_id": 0})
    return f or {"ok": True}


@router.delete("/feeds/{fid}", dependencies=[Depends(require_admin)])
async def delete_feed(fid: str):
    await db.xml_feeds.delete_one({"id": fid})
    return {"ok": True}


def generate_slug(name: str) -> str:
    """Generate URL-friendly slug from name"""
    slug = name.lower()
    # Turkish character replacements
    tr_map = {'ı': 'i', 'ğ': 'g', 'ü': 'u', 'ş': 's', 'ö': 'o', 'ç': 'c', 'İ': 'i', 'Ğ': 'g', 'Ü': 'u', 'Ş': 's', 'Ö': 'o', 'Ç': 'c'}
    for tr, en in tr_map.items():
        slug = slug.replace(tr, en)
    slug = re.sub(r'[^a-z0-9\s-]', '', slug)
    slug = re.sub(r'[\s_]+', '-', slug)
    slug = re.sub(r'-+', '-', slug).strip('-')
    return slug


def slug_with_card_id(name: str, card_id) -> str:
    """Ürün slug'ını her zaman `{urun-adi}-{kart_id}` biçiminde üretir; böylece
    tüm ürün linkleri TEK formatta olur. card_id yoksa çağıran iç id'yi geçer."""
    base = generate_slug(name or "") or "urun"
    cid = str(card_id).strip() if card_id not in (None, "") else ""
    return f"{base}-{cid}" if cid else base


def _slug_to_diacritic_regex(slug: str) -> str:
    """Türkçe karakter DUYARSIZ regex deseni üretir.

    Storefront menüsü Türkçe karaktersiz slug gönderir (örn. 'takim', 'tisort',
    'giyim') ama DB'de kategori adları/breadcrumb Türkçe karakterlidir
    ('Takım', 'Tişört', 'GİYİM'). Düz regex (case-insensitive) 'i'≠'ı', 's'≠'ş'
    olduğu için eşleşmez. Bu helper her belirsiz latin harfi, Türkçe varyantları
    da kapsayan bir karakter sınıfına çevirir → 'giyim' deseni 'GİYİM'i de yakalar.
    """
    char_map = {
        'i': '[iıİI]', 'o': '[oöÖO]', 'u': '[uüÜU]',
        's': '[sşŞS]', 'c': '[cçÇC]', 'g': '[gğĞG]',
    }
    parts = []
    for ch in slug:
        if ch in '-_ ':
            parts.append(r'[\s_>-]*')
        elif ch in char_map:
            parts.append(char_map[ch])
        else:
            parts.append(re.escape(ch))
    return ''.join(parts)


def _search_tr_regex(s: str) -> str:
    """Serbest metin arama için Türkçe duyarsız regex.
    MongoDB $options:'i' Türkçe İ↔i / ı↔I eşlemesini yapmadığından her Türkçe
    harf ailesini kapsayan bir karakter sınıfına çeviririz → 'büstiyer',
    'Büstiyer' ve 'BÜSTİYER' hepsi aynı sonucu verir."""
    cls = {
        'i': '[iıİI]', 'ı': '[iıİI]', 'İ': '[iıİI]', 'I': '[iıİI]',
        'o': '[oöÖO]', 'ö': '[oöÖO]', 'O': '[oöÖO]', 'Ö': '[oöÖO]',
        'u': '[uüÜU]', 'ü': '[uüÜU]', 'U': '[uüÜU]', 'Ü': '[uüÜU]',
        's': '[sşŞS]', 'ş': '[sşŞS]', 'S': '[sşŞS]', 'Ş': '[sşŞS]',
        'c': '[cçÇC]', 'ç': '[cçÇC]', 'C': '[cçÇC]', 'Ç': '[cçÇC]',
        'g': '[gğĞG]', 'ğ': '[gğĞG]', 'G': '[gğĞG]', 'Ğ': '[gğĞG]',
    }
    return ''.join(cls.get(ch, re.escape(ch)) for ch in (s or '').strip())

@router.get("")
async def get_products(
    request: Request,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=500),
    category: Optional[str] = None,
    category_id: Optional[str] = None,
    search: Optional[str] = None,
    sort: str = Query("created_at"),
    order: str = Query("desc"),
    is_featured: Optional[bool] = None,
    is_new: Optional[bool] = None,
    min_price: Optional[float] = None,
    max_price: Optional[float] = None,
    status: Optional[str] = None,
    admin_view: Optional[str] = None,
    brand: Optional[str] = None,
    min_stock: Optional[int] = None,
    max_stock: Optional[int] = None,
    is_showcase: Optional[bool] = None,
    is_opportunity: Optional[bool] = None,
    is_free_shipping: Optional[bool] = None,
    stock_code: Optional[str] = None,
    barcode: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    # --- Gelişmiş Ticimax tarzı filtreler ---
    urun_karti_id: Optional[str] = None,
    varyasyon_id: Optional[str] = None,
    name: Optional[str] = None,
    gtip: Optional[str] = None,
    breadcrumb: Optional[str] = None,
    supplier: Optional[str] = None,
    tag: Optional[str] = None,
    has_image: Optional[str] = None,
    has_variants: Optional[str] = None,
    has_video: Optional[str] = None,
    multi_barcode: Optional[str] = None,
    discounted: Optional[str] = None,
    attr_key: Optional[str] = None,
    attr_value: Optional[str] = None,
    pub_date_from: Optional[str] = None,
    pub_date_to: Optional[str] = None,
):
    """Get products with filtering and pagination.

    Gelişmiş filtreler (Ticimax paneli ile birebir): yukarıdaki açık parametrelerin
    yanı sıra `ticimax_fields` ham verisi üzerinde çalışan dinamik parametreler de
    kabul edilir:
      - tf_<KOLON>=deger        → ticimax_fields.<KOLON> (BOOL ise eşitlik, değilse regex)
      - tf_<KOLON>=__nonempty__ → alan dolu (Var)
      - tf_<KOLON>=__empty__    → alan boş (Yok)
      - tfmin_<KOLON> / tfmax_<KOLON> → sayısal aralık (ticimax_fields.<KOLON>)
    Veri henüz yoksa bile yapı hazırdır; senkron aktifleşince otomatik sorgulanır.
    """
    skip = (page - 1) * limit
    query = {}
    # `and_clauses` — yeni gelişmiş filtreler birbirini ezmeden eklenir.
    and_clauses: list = []
    # Çöp kutusundaki (soft-deleted) ürünler normal listelerde/storefront'ta görünmez
    query["is_deleted"] = {"$ne": True}

    # Admin (token'lı istek): status verilmemişse pasif ürünler de görünsün ki
    # arama/listede hiçbir ürün "kaybolmasın". Storefront token göndermez →
    # aktif default korunur, vitrin etkilenmez.
    _is_admin = False
    _auth = request.headers.get("authorization") or ""
    if _auth.lower().startswith("bearer "):
        try:
            from .deps import _decode_jwt_strict
            if _decode_jwt_strict(_auth.split(" ", 1)[1]).get("is_admin"):
                _is_admin = True
        except Exception:
            pass
    # KRİTİK: AuthContext token'ı GLOBAL eklediği için (axios.defaults) storefront
    # istekleri de admin token taşır → _is_admin TEK BAŞINA storefront/admin ayrımı
    # YAPAMAZ. Bu yüzden pasif/görselsiz ürünleri yalnızca admin panelin AÇIK
    # `admin_view=1` bayrağıyla gösteririz. Storefront bu bayrağı göndermez → token
    # olsa bile MÜŞTERİ görünümü alır (pasifler gizli).
    _admin_view = _is_admin and str(admin_view or "").strip().lower() in ("1", "true", "yes", "evet")
    if status is None and _admin_view:
        status = "all"

    # Status default to active for storefront compatibility unless specified differently
    if status == "all":
        pass
    elif status == "passive":
        query["is_active"] = False
    else:
        query["is_active"] = True

    # ===================================================================
    # STOREFRONT (müşteri) GÖRÜNÜRLÜK KURALI — admin DEĞİLSE zorunlu:
    #   • Yalnızca aktif ürünler (pasife alınanlar müşteriye görünmez).
    #   • En az bir görseli olan ürünler — görselsiz/placeholder ürünler
    #     (çoğu import'tan gelen yarım kayıt) vitrinde gizlenir.
    # Bu, status parametresi ne gelirse gelsin storefront'u korur.
    # ===================================================================
    if not _admin_view:
        query["is_active"] = True
        query["images.0"] = {"$exists": True}

    if brand:
        query["brand"] = {"$regex": brand, "$options": "i"}

    if min_stock is not None:
        query["stock"] = {"$gte": min_stock}

    if max_stock is not None:
        if "stock" in query:
            query["stock"]["$lte"] = max_stock
        else:
            query["stock"] = {"$lte": max_stock}
            
    if is_showcase is not None:
        query["is_showcase"] = is_showcase
        
    if is_opportunity is not None:
        query["is_opportunity"] = is_opportunity
        
    if is_free_shipping is not None:
        query["is_free_shipping"] = is_free_shipping

    if stock_code:
        query["stock_code"] = {"$regex": stock_code, "$options": "i"}

    if barcode:
        import re as _re2
        bce = _re2.escape(barcode.strip())
        query.setdefault("$and", []).append({"$or": [
            {"barcode": {"$regex": bce, "$options": "i"}},
            {"variants.barcode": {"$regex": bce, "$options": "i"}},
        ]})
    
    if date_from or date_to:
        date_q = {}
        try:
            from datetime import datetime, timezone
            if date_from:
                date_q["$gte"] = datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=timezone.utc).isoformat()
            if date_to:
                # end of day
                date_q["$lte"] = datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc).isoformat()
            query["created_at"] = date_q
        except Exception:
            pass
    
    if category:
        # =====================================================================
        # KATEGORİ FİLTRESİ — Türkçe karakter duyarsız + breadcrumb ağaç eşleşmesi
        # Storefront 'takim'/'tisort'/'giyim' gibi diakritiksiz slug yollar; DB
        # 'Takım'/'Tişört'/'GİYİM' tutar. _slug_to_diacritic_regex ile eşleştirilir.
        # =====================================================================
        cat_slug = category.strip().lower()
        SHOW_ALL = {"tum-urunler", "tumu", "all"}
        EN_YENILER = {"en-yeniler", "en-yeni", "yeniler", "yeni", "yeni-urunler"}
        if cat_slug in SHOW_ALL:
            # "Tümü" — gerçekten tüm aktif ürünler, default sıralama (created_at desc)
            pass
        elif cat_slug in EN_YENILER:
            # "En Yeniler" GERÇEK kategori üyeliğiyle filtrelenir (tüm ürünler DEĞİL).
            # Yerel "En Yeniler" kategorisinin id'sini bulup category_ids ile eşleştir.
            # (Ticimax En Yeniler eşitlemesi bu kategoriye üyelik yazar.)
            _en_id = None
            async for _c in db.categories.find({}, {"_id": 0, "id": 1, "name": 1, "slug": 1}):
                _sl = (_c.get("slug") or "").strip().lower()
                _nm = (_c.get("name") or "").strip()
                if _sl in EN_YENILER or generate_slug(_nm) in EN_YENILER:
                    _en_id = _c.get("id")
                    break
            if _en_id:
                and_clauses.append({"category_ids": _en_id})
            else:
                # Yerel "En Yeniler" kategorisi yoksa: yanlışlıkla tüm ürünleri
                # göstermek yerine boş döndür (eşleşmeyen sentinel).
                and_clauses.append({"id": "__en_yeniler_category_missing__"})
        elif cat_slug == "sale":
            and_clauses.append({"$or": [
                {"sale_price": {"$gt": 0}},
                {"discount_price": {"$gt": 0}},
                {"is_on_sale": True},
                {"sale_active": True},
            ]})
        else:
            dia = _slug_to_diacritic_regex(cat_slug)
            # Slug'tan tam kategori adını çöz (generate_slug ile ters eşleme)
            distinct_names = await db.products.distinct("category_name", {"is_deleted": {"$ne": True}})
            matched_names = [n for n in distinct_names if n and generate_slug(n) == cat_slug]
            cat_or = []
            if matched_names:
                cat_or.append({"category_name": {"$in": matched_names}})
            # Kategori adı tam eşleşme (diakritik duyarsız)
            cat_or.append({"category_name": {"$regex": f"^{dia}$", "$options": "i"}})
            # Breadcrumb segment eşleşmesi: üst kategori (GİYİM>...) ya da yaprak (...>Şort)
            cat_or.append({"breadcrumb": {"$regex": f"(?:^|>){dia}(?:>|$)", "$options": "i"}})
            cat_or.append({"category_slug": cat_slug})
            # ÇOKLU KATEGORİ: slug'a karşılık gelen kategori id'lerini category_ids içinde de ara.
            # Ürün ana kategoride olmasa bile (ör. ana "Pantolon") admin'de ekstra eklendiği
            # "Şort" kategorisinde de müşteriye görünür.
            _slug_cat_ids = []
            async for _c in db.categories.find({}, {"_id": 0, "id": 1, "name": 1, "slug": 1}):
                _csl = (_c.get("slug") or "").strip().lower()
                _cnm = (_c.get("name") or "").strip()
                if _csl == cat_slug or (_cnm and generate_slug(_cnm) == cat_slug):
                    if _c.get("id"):
                        _slug_cat_ids.append(_c["id"])
            if _slug_cat_ids:
                cat_or.append({"category_ids": {"$in": _slug_cat_ids}})
            and_clauses.append({"$or": cat_or})
    
    if search:
        import re as _re
        esc = _search_tr_regex(search)  # Türkçe duyarsız (İ/ı/ş/ç/ğ/ö/ü dahil)
        query["$or"] = [
            {"name": {"$regex": esc, "$options": "i"}},
            {"description": {"$regex": esc, "$options": "i"}},
            {"keywords": {"$regex": esc, "$options": "i"}},
            {"stock_code": {"$regex": esc, "$options": "i"}},
            {"sku": {"$regex": esc, "$options": "i"}},
            {"barcode": {"$regex": esc, "$options": "i"}},
            {"variants.barcode": {"$regex": esc, "$options": "i"}},
            {"variants.sku": {"$regex": esc, "$options": "i"}},
            {"variants.stock_code": {"$regex": esc, "$options": "i"}},
            {"urun_karti_id": {"$regex": esc, "$options": "i"}},
            {"variants.urun_id": {"$regex": esc, "$options": "i"}},
        ]
    
    if is_featured is not None:
        query["is_featured"] = is_featured
    
    if is_new is not None:
        query["is_new"] = is_new
    
    if min_price is not None:
        query["price"] = {"$gte": min_price}
    
    if max_price is not None:
        if "price" in query:
            query["price"]["$lte"] = max_price
        else:
            query["price"] = {"$lte": max_price}

    # ============================================================
    # GELİŞMİŞ FİLTRELER (Ticimax paneli)
    # ============================================================
    # --- Metin (regex) filtreleri ---
    if category_id:
        # category_ids atalar dahil tutulduğu için üst kategori seçilince
        # tüm alt kategori ürünleri de eşleşir (ağaç filtreleme).
        and_clauses.append({"category_ids": category_id})
    if urun_karti_id:
        and_clauses.append({"urun_karti_id": {"$regex": re.escape(urun_karti_id.strip()), "$options": "i"}})
    if varyasyon_id:
        and_clauses.append({"variants.urun_id": {"$regex": re.escape(varyasyon_id.strip()), "$options": "i"}})
    if name:
        and_clauses.append({"name": {"$regex": re.escape(name.strip()), "$options": "i"}})
    if gtip:
        ge = re.escape(gtip.strip())
        and_clauses.append({"$or": [
            {"gtip_code": {"$regex": ge, "$options": "i"}},
            {"ticimax_fields.GTIPKODU": {"$regex": ge, "$options": "i"}},
        ]})
    if breadcrumb:
        be = re.escape(breadcrumb.strip())
        and_clauses.append({"$or": [
            {"breadcrumb": {"$regex": be, "$options": "i"}},
            {"breadcrumb_category": {"$regex": be, "$options": "i"}},
            {"ticimax_fields.BREADCRUMBKAT": {"$regex": be, "$options": "i"}},
        ]})
    if supplier:
        se = re.escape(supplier.strip())
        and_clauses.append({"$or": [
            {"supplier": {"$regex": se, "$options": "i"}},
            {"ticimax_fields.TEDARIKCI": {"$regex": se, "$options": "i"}},
        ]})
    if tag:
        te = re.escape(tag.strip())
        and_clauses.append({"$or": [
            {"keywords": {"$regex": te, "$options": "i"}},
            {"tags": {"$regex": te, "$options": "i"}},
            {"ticimax_fields.ANAHTARKELIME": {"$regex": te, "$options": "i"}},
        ]})

    # --- Varlık (Var/Yok) filtreleri ---
    def _bool_flag(v):
        return str(v).strip() in ("1", "true", "evet", "yes")

    if has_image is not None and has_image != "":
        and_clauses.append({"images.0": {"$exists": _bool_flag(has_image)}})
    if has_variants is not None and has_variants != "":
        and_clauses.append({"variants.0": {"$exists": _bool_flag(has_variants)}})
    if has_video is not None and has_video != "":
        if _bool_flag(has_video):
            and_clauses.append({"video_url": {"$nin": [None, ""]}})
        else:
            and_clauses.append({"$or": [{"video_url": {"$in": [None, ""]}}, {"video_url": {"$exists": False}}]})
    if multi_barcode is not None and multi_barcode != "":
        if _bool_flag(multi_barcode):
            and_clauses.append({"variants.1": {"$exists": True}})
        else:
            and_clauses.append({"variants.1": {"$exists": False}})
    if discounted is not None and discounted != "":
        if _bool_flag(discounted):
            and_clauses.append({"$or": [
                {"sale_price": {"$gt": 0}},
                {"ticimax_fields.INDIRIMLIFIYAT": {"$gt": 0}},
            ]})

    # --- Teknik detay (attributes) filtresi ---
    if attr_key:
        if attr_value:
            and_clauses.append({f"attributes.{attr_key}.value": {"$regex": re.escape(attr_value.strip()), "$options": "i"}})
        else:
            and_clauses.append({f"attributes.{attr_key}": {"$exists": True}})

    # --- Yayın tarihi aralığı (ticimax_fields.YAYINTARIHI, metinsel ISO karşılaştırma) ---
    if pub_date_from or pub_date_to:
        pub_q = {}
        if pub_date_from:
            pub_q["$gte"] = pub_date_from
        if pub_date_to:
            pub_q["$lte"] = pub_date_to + "T23:59:59"
        and_clauses.append({"ticimax_fields.YAYINTARIHI": pub_q})

    # --- Dinamik ticimax_fields parametreleri (tf_ / tfmin_ / tfmax_) ---
    range_acc: dict = {}
    for pkey, pval in request.query_params.items():
        if pval is None or pval == "":
            continue
        if pkey.startswith("tfmin_"):
            col = pkey[6:]
            try:
                range_acc.setdefault(col, {})["$gte"] = float(pval)
            except ValueError:
                pass
        elif pkey.startswith("tfmax_"):
            col = pkey[6:]
            try:
                range_acc.setdefault(col, {})["$lte"] = float(pval)
            except ValueError:
                pass
        elif pkey.startswith("tf_"):
            col = pkey[3:]
            field = f"ticimax_fields.{col}"
            if pval == "__nonempty__":
                and_clauses.append({field: {"$nin": ["", None, 0]}})
            elif pval == "__empty__":
                and_clauses.append({"$or": [{field: {"$in": ["", None, 0]}}, {field: {"$exists": False}}]})
            elif col in TICIMAX_BOOL_COLS:
                try:
                    and_clauses.append({field: int(float(pval))})
                except ValueError:
                    pass
            else:
                and_clauses.append({field: {"$regex": re.escape(str(pval).strip()), "$options": "i"}})
    for col, rng in range_acc.items():
        and_clauses.append({f"ticimax_fields.{col}": rng})

    if and_clauses:
        query.setdefault("$and", []).extend(and_clauses)

    sort_order = -1 if order == "desc" else 1
    # Kategori sayfasında kullanıcı özel sıralama seçmediyse (default created_at):
    # urun_karti_id (sayısal) DESC — yüksek kart id = en yeni ürün, kategoride en üstte.
    _cat_view = bool(category or category_id)
    _card_sort = _cat_view and sort == "created_at"

    if not _admin_view:
        # ===============================================================
        # STOREFRONT: tükenen ürünler (efektif stok 0) ilgili listenin/
        # kategorinin EN SONUNDA gösterilir. Efektif stok = varyant varsa
        # varyant stokları toplamı, yoksa ürün stoğu. Önce stoğu olanlar,
        # sonra istenen sıralama (created_at/price vb.).
        # ===============================================================
        pipeline = [
            {"$match": query},
            {"$addFields": {
                "_eff_stock": {
                    "$cond": [
                        {"$gt": [{"$size": {"$ifNull": ["$variants", []]}}, 0]},
                        {"$sum": {"$map": {
                            "input": {"$ifNull": ["$variants", []]},
                            "as": "v",
                            "in": {"$ifNull": ["$$v.stock", 0]},
                        }}},
                        {"$ifNull": ["$stock", 0]},
                    ]
                }
            }},
            {"$addFields": {"_in_stock": {"$cond": [{"$gt": ["$_eff_stock", 0]}, 1, 0]}}},
            {"$addFields": {"_card_num": {"$convert": {"input": "$urun_karti_id", "to": "long", "onError": 0, "onNull": 0}}}},
            ({"$sort": {"_in_stock": -1, "_card_num": -1, "_id": -1}} if _card_sort
             else {"$sort": {"_in_stock": -1, sort: sort_order, "_id": 1}}),
            {"$skip": skip},
            {"$limit": limit},
            {"$project": {"_id": 0, "_eff_stock": 0, "_in_stock": 0, "_card_num": 0}},
        ]
        products = await db.products.aggregate(pipeline, allowDiskUse=True).to_list(limit)
    else:
        if sort == "stock":
            # Admin stok sıralaması: gösterilen stok = varyant varsa varyant stokları
            # toplamı, yoksa ürün stoğu. DB 'stock' alanı varyantlı üründe güncel
            # olmayabildiği için efektif stoğu hesaplayıp ona göre sıralıyoruz —
            # böylece sıralama, listede görünen stok değeriyle birebir tutarlı olur.
            stock_pipeline = [
                {"$match": query},
                {"$addFields": {
                    "_eff_stock": {
                        "$cond": [
                            {"$gt": [{"$size": {"$ifNull": ["$variants", []]}}, 0]},
                            {"$sum": {"$map": {
                                "input": {"$ifNull": ["$variants", []]},
                                "as": "v",
                                "in": {"$ifNull": ["$$v.stock", 0]},
                            }}},
                            {"$ifNull": ["$stock", 0]},
                        ]
                    }
                }},
                {"$sort": {"_eff_stock": sort_order, "_id": 1}},
                {"$skip": skip},
                {"$limit": limit},
                {"$project": {"_id": 0, "_eff_stock": 0}},
            ]
            products = await db.products.aggregate(stock_pipeline, allowDiskUse=True).to_list(limit)
        else:
            products = await db.products.find(query, {"_id": 0}).sort(sort, sort_order).skip(skip).limit(limit).to_list(limit)
    total = await db.products.count_documents(query)
    
    return {
        "products": products,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }

@router.get("/meta/ticimax-schema")
async def get_ticimax_schema(current_user: dict = Depends(require_admin)):
    """Ürün kartında tüm Ticimax (113) alanını gruplu render etmek için şema."""
    from ticimax_schema import build_schema
    return {"groups": build_schema()}

@router.get("/meta/next-card-id")
async def get_next_card_id(current_user: dict = Depends(require_admin)):
    """Yeni urun icin onerilen Urun Kart ID (sistemdeki en buyuk + 1).
    Sayac TUKETMEZ; bir urun olusturmanin tum renk/varyantlarina ayni id verilebilsin
    diye frontend bunu bir kez alip hepsine atar (boylece kart id 'kendi icinde' artmaz)."""
    cid = await generate_urun_karti_id()
    return {"card_id": cid}

@router.get("/meta/filter-options")
async def get_filter_options(current_user: dict = Depends(require_admin)):
    """Gelişmiş filtre panelinin dropdown'larını besleyen dinamik veri.

    Markalar, tedarikçiler, para birimleri ve teknik detay (attribute) grupları
    canlı katalogdan distinct çekilir. Böylece sistem aktifleştikçe seçenekler
    otomatik büyür.
    """
    base = {"is_deleted": {"$ne": True}}
    brands = [b for b in await db.products.distinct("brand", base) if b]
    suppliers = sorted(set(
        [s for s in await db.products.distinct("supplier", base) if s]
        + [s for s in await db.products.distinct("ticimax_fields.TEDARIKCI", base) if s]
    ))
    currencies = sorted(set(
        [c for c in await db.products.distinct("currency", base) if c]
        + [c for c in await db.products.distinct("ticimax_fields.PARABIRIMI", base) if c]
    ))
    # Teknik detay grupları: attributes anahtarları + label'ları
    attr_groups: dict = {}
    async for d in db.products.find(base, {"attributes": 1, "_id": 0}):
        a = d.get("attributes")
        if isinstance(a, dict):
            for k, v in a.items():
                if k not in attr_groups:
                    label = v.get("label") if isinstance(v, dict) else None
                    attr_groups[k] = label or k
    attr_options = [{"key": k, "label": lbl} for k, lbl in sorted(attr_groups.items(), key=lambda x: str(x[1]))]
    # Kategori başına ürün sayısı (category_ids atalar dahil → üst kategori toplamı verir)
    category_counts: dict = {}
    pipeline = [
        {"$match": {"is_deleted": {"$ne": True}}},
        {"$unwind": "$category_ids"},
        {"$group": {"_id": "$category_ids", "count": {"$sum": 1}}},
    ]
    async for row in db.products.aggregate(pipeline):
        if row.get("_id"):
            category_counts[str(row["_id"])] = row["count"]
    return {
        "brands": sorted(brands),
        "suppliers": suppliers,
        "currencies": currencies,
        "attribute_groups": attr_options,
        "category_counts": category_counts,
    }

@router.get("/trash/list")
async def list_trashed_products(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    search: Optional[str] = None,
    current_user: dict = Depends(require_admin)
):
    """Çöp kutusundaki (soft-deleted) ürünleri listeler."""
    skip = (page - 1) * limit
    query = {"is_deleted": True}
    if search:
        esc = re.escape(search.strip())
        query["$or"] = [
            {"name": {"$regex": esc, "$options": "i"}},
            {"stock_code": {"$regex": esc, "$options": "i"}},
            {"urun_karti_id": {"$regex": esc, "$options": "i"}},
        ]
    products = await db.products.find(query, {"_id": 0}).sort("deleted_at", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.products.count_documents(query)
    return {"products": products, "total": total, "page": page, "pages": (total + limit - 1) // limit}

@router.get("/{product_id}")
async def get_product(product_id: str, request: Request):
    """Get single product by ID or slug"""
    product = await db.products.find_one(
        {"$or": [{"id": product_id}, {"slug": product_id}, {"slug_aliases": product_id}]},
        {"_id": 0}
    )
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    # PASİF/SİLİNMİŞ ürün direkt link ile açılamaz: yalnızca admin görebilir,
    # müşteri için "bulunamadı" döner (vitrin gizliliği detay sayfasında da geçerli).
    _is_admin = False
    _auth = request.headers.get("authorization") or ""
    if _auth.lower().startswith("bearer "):
        try:
            from .deps import _decode_jwt_strict
            if _decode_jwt_strict(_auth.split(" ", 1)[1]).get("is_admin"):
                _is_admin = True
        except Exception:
            pass
    if not _is_admin and (product.get("is_active") is not True or product.get("is_deleted")):
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    # Varyantları global Beden Havuzu (variant_options) sırasına göre diz —
    # böylece storefront'ta XS, S, M, L, XL... admin'in tanımladığı sırayla görünür.
    product["variants"] = await _sort_variants_by_pool(product.get("variants") or [])
    return product


async def _sort_variants_by_pool(variants: list) -> list:
    """Ürün varyantlarını `variant_options` (type=size) sort_order'ına göre sıralar.
    Havuzda olmayan bedenler (kombinasyonlar/numeric) orijinal sırada en sona eklenir
    (stable sort). Renk sıralaması beden eşitliğinde korunur."""
    if not variants:
        return variants
    size_order = {}
    async for vo in db.variant_options.find({"type": "size"}, {"_id": 0, "value": 1, "sort_order": 1}):
        key = str(vo.get("value", "")).strip().lower()
        if key:
            size_order[key] = vo.get("sort_order", 9999)
    if not size_order:
        return variants
    def _k(v):
        s = str(v.get("size", "")).strip().lower()
        return size_order.get(s, 10000)
    return sorted(variants, key=_k)


@router.get("/{product_id}/color-siblings")
async def get_color_siblings(product_id: str):
    """Aynı modelin (csv_card_id) farklı renk varyantlarını getir.
    Ürün detay sayfasında "Diğer Renkler" swatch listesi için kullanılır.
    """
    p = await db.products.find_one(
        {"$or": [{"id": product_id}, {"slug": product_id}]},
        {"_id": 0, "id": 1, "csv_card_id": 1}
    )
    if not p:
        return {"siblings": []}
    card_id = p.get("csv_card_id")
    if not card_id:
        return {"siblings": []}
    siblings = []
    cursor = db.products.find(
        {"csv_card_id": card_id, "id": {"$ne": p["id"]}, "is_active": True},
        {"_id": 0, "id": 1, "slug": 1, "name": 1, "thumbnail": 1, "images": 1,
         "variants": 1, "attributes": 1}
    ).limit(20)
    async for s in cursor:
        # Renk: önce variants[0].color, yoksa attributes Web Color
        color = ""
        if s.get("variants"):
            for v in s["variants"]:
                if v.get("color"):
                    color = v["color"]
                    break
        if not color and isinstance(s.get("attributes"), list):
            for a in s["attributes"]:
                if (a.get("name") or "").strip().lower() in ("web color", "renk", "color"):
                    color = a.get("value") or ""
                    break
        siblings.append({
            "id": s["id"],
            "slug": s.get("slug") or s["id"],
            "name": s.get("name") or "",
            "color": color,
            "image": (s.get("images") or [s.get("thumbnail")] or [None])[0],
        })
    return {"siblings": siblings}

async def _expand_category_ids(selected_ids):
    """Seçilen kategori id'lerini atalarıyla birlikte düzleştirir (vitrin category_ids için)."""
    sel = [str(c) for c in (selected_ids or []) if c]
    if not sel:
        return []
    cats = await db.categories.find({}, {"_id": 0, "id": 1, "parent_id": 1}).to_list(5000)
    parent = {c.get("id"): c.get("parent_id") for c in cats if c.get("id")}
    result, seen = [], set()
    for cid in sel:
        cur, guard = cid, 0
        while cur and cur not in seen and guard < 50:
            seen.add(cur); result.append(cur)
            cur = parent.get(cur); guard += 1
    return result


def _distinct_variant_colors(variants):
    """Varyant listesindeki BENZERSIZ renkleri (ilk görülme sırasıyla) döndürür."""
    seen = []
    for v in (variants or []):
        c = (v.get("color") or "").strip()
        if c and c.lower() not in [s.lower() for s in seen]:
            seen.append(c)
    return seen


def _variants_for_color(variants, color):
    """Verilen renge ait varyantları (beden vb.) döndürür."""
    cl = (color or "").strip().lower()
    return [v for v in (variants or []) if (v.get("color") or "").strip().lower() == cl]


@router.post("")
async def create_product(
    product_data: dict,
    current_user: dict = Depends(require_admin)
):
    """Create new product (admin only)"""
    variants = product_data.get("variants", [])
    
    # Fetch settings for default VAT
    settings = await db.settings.find_one({"id": "main"})
    default_vat = settings.get("default_vat_rate", 20) if settings else 20

    # Auto-generate barcodes efficiently
    used_barcodes_set = await build_used_barcode_set()
    
    # Auto-generate barcodes for variants if missing
    for v in variants:
        if not v.get("barcode"):
            barcode = await generate_barcode_from_range(used_barcodes_set)
            if barcode:
                v["barcode"] = barcode
    
    # Also generate for main product if no variants and barcode is empty
    if not variants and not product_data.get("barcode"):
        barcode = await generate_barcode_from_range(used_barcodes_set)
        if barcode:
            product_data["barcode"] = barcode

    _pid = await generate_short_id("products")
    # Urun Kart ID: form "Kimlik & Kodlar > Urun Kart ID" (ticimax_fields.URUNKARTIID)
    # veya ust-seviye urun_karti_id; ikisi de bossa SON kart id + 1 otomatik atanir.
    _tf_in = product_data.get("ticimax_fields") or {}
    urun_karti_id = str(_tf_in.get("URUNKARTIID") or product_data.get("urun_karti_id") or product_data.get("csv_card_id") or "").strip()
    if not urun_karti_id:
        urun_karti_id = await generate_urun_karti_id()
    # liste/etiket (urun_karti_id) ile form (ticimax_fields.URUNKARTIID) senkron
    product_data["ticimax_fields"] = {**_tf_in, "URUNKARTIID": urun_karti_id}
    _card = urun_karti_id
    _sel_cats = product_data.get("categories")
    if not _sel_cats and product_data.get("category_id"):
        _sel_cats = [product_data.get("category_id")]
    _sel_cats = [str(x) for x in (_sel_cats or []) if x]
    _cat_ids_all = await _expand_category_ids(_sel_cats)
    _primary_cat = _sel_cats[0] if _sel_cats else ""
    product = {
        "id": _pid,
        "urun_karti_id": urun_karti_id,
        "ticimax_fields": product_data.get("ticimax_fields", {}),
        "name": product_data.get("name", ""),
        "slug": product_data.get("slug") or slug_with_card_id(product_data.get("name", ""), _card),
        "description": product_data.get("description", ""),
        "short_description": product_data.get("short_description", ""),
        "price": float(product_data.get("price", 0)),
        "sale_price": product_data.get("sale_price"),
        "category_name": product_data.get("category_name", ""),
        "category_id": _primary_cat,
        "category_ids": _cat_ids_all,
        "categories": _sel_cats,
        "brand": product_data.get("brand", "FACETTE"),
        "images": product_data.get("images", []),
        "variants": product_data.get("variants", []),
        "attributes": product_data.get("attributes", []),
        "stock": int(product_data.get("stock", 0)),
        "stock_code": product_data.get("stock_code", ""),
        "barcode": product_data.get("barcode", ""),
        "sku": product_data.get("sku", ""),
        "supplier": product_data.get("supplier", ""),
        "manufacturer": product_data.get("manufacturer", "FACETTE"),
        # FAZ 7 — İmalat modülü için ek alanlar
        "collection": product_data.get("collection", ""),   # ör. "2026 İlkbahar/Yaz"
        "purchase_price": float(product_data.get("purchase_price", 0) or 0),  # Alış fiyatı
        "color": product_data.get("color", ""),  # Renk (varyant dışı global)
        "is_active": product_data.get("is_active", True),
        "is_featured": product_data.get("is_featured", False),
        "is_new": product_data.get("is_new", False),
        "is_showcase": product_data.get("is_showcase", False),
        "is_opportunity": product_data.get("is_opportunity", False),
        "is_free_shipping": product_data.get("is_free_shipping", False),
        "vat_rate": product_data.get("vat_rate", default_vat),
        "use_default_markup": product_data.get("use_default_markup", True),
        "markup_rate": float(product_data.get("markup_rate", 0)),
        "trendyol_attributes": product_data.get("trendyol_attributes", {}),
        "hepsiburada_attributes": product_data.get("hepsiburada_attributes", {}),
        "temu_attributes": product_data.get("temu_attributes", {}),
        "hepsiburada_category_id": product_data.get("hepsiburada_category_id", ""),
        "hepsiburada_category_name": product_data.get("hepsiburada_category_name", ""),
        "temu_category_id": product_data.get("temu_category_id", ""),
        "temu_category_name": product_data.get("temu_category_name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # FARKLI RENK = AYRI ÜRÜN kuralı:
    # Varyantlarda birden fazla renk varsa her renk AYRI ürün olarak açılır.
    # Hepsi aynı csv_card_id + urun_karti_id'yi paylaşır → storefront "Diğer Renkler"
    # swatch'ında renk kardeşi olarak bağlanır. Bedenler her renk altında varyant kalır.
    _all_variants = product.get("variants") or []
    _colors = _distinct_variant_colors(_all_variants)
    # Renk-kardeşi gruplama anahtarı HER ZAMAN yazılır (manuel ürünler de bağlansın)
    product["csv_card_id"] = urun_karti_id

    if len(_colors) <= 1:
        if _colors and not product.get("color"):
            product["color"] = _colors[0]
        await db.products.insert_one(product)
        logger.info(f"Product created: {product['id']}")
        return {"id": product["id"], "message": "Ürün oluşturuldu"}

    # Çok renkli → her renk ayrı ürün (ilk renk ana üründe, diğerleri yeni id)
    _base_name = product.get("name") or ""
    _created_ids = []
    for _idx, _col in enumerate(_colors):
        _doc = {k: v for k, v in product.items()}
        _doc["id"] = product["id"] if _idx == 0 else await generate_short_id("products")
        _doc["csv_card_id"] = urun_karti_id
        _doc["urun_karti_id"] = urun_karti_id
        _doc["color"] = _col
        _doc["variants"] = _variants_for_color(_all_variants, _col)
        _doc["slug"] = slug_with_card_id(f"{_base_name} {_col}", urun_karti_id)
        _now = datetime.now(timezone.utc).isoformat()
        _doc["created_at"] = _doc.get("created_at") or _now
        _doc["updated_at"] = _now
        await db.products.insert_one(_doc)
        _created_ids.append(_doc["id"])
    logger.info(f"Product created with color split: {_created_ids} (card {urun_karti_id})")
    return {
        "id": _created_ids[0],
        "split": True,
        "color_count": len(_colors),
        "product_ids": _created_ids,
        "message": f"{len(_colors)} renk ayrı ürün olarak oluşturuldu",
    }


@router.post("/{product_id}/split-by-color", dependencies=[Depends(require_admin)])
async def split_product_by_color(product_id: str):
    """Mevcut bir ürünün farklı RENK varyantlarını AYRI ürünlere böler.
    İlk renk ana üründe kalır; diğer renkler yeni ürün olur. Hepsi aynı
    csv_card_id + urun_karti_id'yi paylaşır → "Diğer Renkler" swatch'ında bağlı kalır.
    Bedenler her renk ürününün altında varyant olarak kalır.
    """
    p = await db.products.find_one({"id": product_id})
    if not p:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    variants = p.get("variants") or []
    colors = _distinct_variant_colors(variants)
    if len(colors) <= 1:
        return {
            "success": False,
            "color_count": len(colors),
            "message": "Bu üründe birden fazla renk yok; ayırma gerekmedi.",
        }
    card = str(p.get("csv_card_id") or p.get("urun_karti_id") or p.get("id"))
    base_name = p.get("name") or ""
    now = datetime.now(timezone.utc).isoformat()
    created_ids = [product_id]
    # İlk renk → ana üründe kalır
    first = colors[0]
    await db.products.update_one({"id": product_id}, {"$set": {
        "variants": _variants_for_color(variants, first),
        "color": first,
        "csv_card_id": card,
        "urun_karti_id": p.get("urun_karti_id") or card,
        "slug": slug_with_card_id(f"{base_name} {first}", card),
        "updated_at": now,
    }})
    # Diğer renkler → yeni ürün
    for col in colors[1:]:
        nid = await generate_short_id("products")
        clone = {k: v for k, v in p.items() if k not in ("_id", "id", "slug")}
        clone["id"] = nid
        clone["csv_card_id"] = card
        clone["urun_karti_id"] = p.get("urun_karti_id") or card
        clone["color"] = col
        clone["variants"] = _variants_for_color(variants, col)
        clone["slug"] = slug_with_card_id(f"{base_name} {col}", card)
        clone["created_at"] = now
        clone["updated_at"] = now
        await db.products.insert_one(clone)
        created_ids.append(nid)
    logger.info(f"Product split by color: {product_id} -> {created_ids} (card {card})")
    return {
        "success": True,
        "color_count": len(colors),
        "product_ids": created_ids,
        "message": f"{len(colors)} renk ayrı ürüne bölündü.",
    }

@router.post("/{product_id}/duplicate", dependencies=[Depends(require_admin)])
async def duplicate_product(product_id: str):
    """Bir ürünü BAĞIMSIZ kopya olarak çoğaltır.

    - Her çoğaltmada YENİ ve benzersiz Ürün Kart ID atanır (paylaşılmaz; orijinalle
      aynı kart id'yi ALMAZ). Böylece "farklı ürün = farklı kart id" sağlanır.
    - Tüm varyantlara aralıktan YENİ benzersiz barkod üretilir → orijinalin
      barkodlarıyla çakışmaz (eski 'duplicate'in patlama sebebi buydu).
    - Varyant id'leri yenilenir; Ticimax varyant id'si (urun_id) temizlenir.
    - stock_code AYNI bırakılır (aynı modelin başka rengini açmak için pratik;
      gerekiyorsa kopyada elle değiştirilir).
    Kopya orijinalin renk-kardeşi DEĞİLDİR (csv_card_id yeni kart id'ye eşitlenir).
    """
    p = await db.products.find_one({"id": product_id})
    if not p:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")

    new_card = await generate_urun_karti_id()
    new_id = await generate_short_id("products")
    used = await build_used_barcode_set()

    clone = {k: v for k, v in p.items() if k not in ("_id", "id", "slug", "slug_aliases")}
    clone["id"] = new_id
    clone["urun_karti_id"] = new_card
    clone["csv_card_id"] = new_card          # bağımsız kart (renk-kardeşi değil)
    clone["urun_id"] = ""
    _tf = dict(clone.get("ticimax_fields") or {})
    _tf["URUNKARTIID"] = new_card
    _tf["URUNID"] = ""
    clone["ticimax_fields"] = _tf

    base_name = p.get("name") or ""
    clone["name"] = f"{base_name} (Kopya)"
    clone["slug"] = slug_with_card_id(clone["name"], new_card)

    # Varyantlar: yeni id + yeni barkod, Ticimax varyant id'si temizlenir
    new_vars = []
    for v in (p.get("variants") or []):
        nv = dict(v)
        nv["id"] = generate_id()
        nv["urun_id"] = None
        nv["barcode"] = (await generate_barcode_from_range(used)) or ""
        new_vars.append(nv)
    clone["variants"] = new_vars
    if new_vars:
        clone["barcode"] = ""
    else:
        clone["barcode"] = (await generate_barcode_from_range(used)) or ""

    now = datetime.now(timezone.utc).isoformat()
    clone["created_at"] = now
    clone["updated_at"] = now

    await db.products.insert_one(clone)
    logger.info(f"Product duplicated: {product_id} -> {new_id} (card {new_card})")
    return {"id": new_id, "urun_karti_id": new_card, "message": "Ürün kopyalandı"}


@router.post("/assign-variant-ids", dependencies=[Depends(require_admin)])
async def assign_variant_ids(payload: dict):
    """Varyantlara BEDEN bazında Ticimax varyant id'si (urun_id) atar.

    Beden id'si eksik ürünler için tek seferlik düzeltme (ör. siyah bermuda şort).
    Body:
      {"product_id": "...", "map": {"S":"8618","XS":"8620",...}}
      veya
      {"name": "bermuda", "color": "siyah", "map": {...}}
    Yalnızca urun_id'si BOŞ olan varyantlara yazar (dolu olanı ezmez).
    """
    size_map = {
        str(k).strip().upper(): str(v).strip()
        for k, v in (payload.get("map") or {}).items()
        if str(v).strip()
    }
    if not size_map:
        raise HTTPException(status_code=400, detail="map (beden->id) zorunlu")

    pid = str(payload.get("product_id") or "").strip()
    if pid:
        query = {"id": pid}
    else:
        name = str(payload.get("name") or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="product_id veya name gerekli")
        query = {"name": {"$regex": re.escape(name), "$options": "i"}}

    color = str(payload.get("color") or "").strip().lower()
    updated_products = 0
    updated_variants = 0
    touched = []
    async for p in db.products.find(query):
        variants = p.get("variants") or []
        if color and not any((v.get("color") or "").strip().lower() == color for v in variants):
            continue
        changed = False
        for v in variants:
            sz = str(v.get("size") or "").strip().upper()
            if sz in size_map and not str(v.get("urun_id") or "").strip():
                v["urun_id"] = size_map[sz]
                updated_variants += 1
                changed = True
        if changed:
            await db.products.update_one(
                {"id": p["id"]},
                {"$set": {"variants": variants, "updated_at": datetime.now(timezone.utc).isoformat()}},
            )
            updated_products += 1
            touched.append({"id": p["id"], "name": p.get("name")})
    return {
        "updated_products": updated_products,
        "updated_variants": updated_variants,
        "products": touched,
        "map": size_map,
    }


@router.put("/{product_id}")
async def update_product(
    product_id: str,
    product_data: dict,
    current_user: dict = Depends(require_admin)
):
    """Update product (admin only)"""
    existing = await db.products.find_one({"id": product_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    
    # Update slug if name changed — format: isim-kartid (tüm linkler tek biçim).
    # Eski slug, kırılan link/SEO olmaması için slug_aliases'a eklenir.
    if product_data.get("name") and product_data.get("name") != existing.get("name"):
        _card = existing.get("csv_card_id") or existing.get("urun_karti_id") or product_id
        new_slug = slug_with_card_id(product_data["name"], _card)
        old_slug = existing.get("slug")
        product_data["slug"] = new_slug
        if old_slug and old_slug != new_slug:
            aliases = list(existing.get("slug_aliases") or [])
            if old_slug not in aliases:
                aliases.append(old_slug)
            product_data["slug_aliases"] = aliases
    
    
    # Auto-generate barcodes for variants if missing
    variants = product_data.get("variants", [])
    used_barcodes_set = await build_used_barcode_set()
    for v in variants:
        if not v.get("barcode") or v.get("barcode") == "":
            barcode = await generate_barcode_from_range(used_barcodes_set)
            if barcode:
                v["barcode"] = barcode
    
    if ("categories" in product_data) or ("category_id" in product_data):
        _sel = product_data.get("categories")
        if _sel is None and product_data.get("category_id"):
            _sel = [product_data.get("category_id")]
        _sel = [str(x) for x in (_sel or []) if x]
        product_data["categories"] = _sel
        product_data["category_ids"] = await _expand_category_ids(_sel)
        if _sel:
            product_data["category_id"] = _sel[0]
            _pc = await db.categories.find_one({"id": _sel[0]}, {"_id": 0, "name": 1})
            if _pc and _pc.get("name"):
                product_data["category_name"] = _pc["name"]

    # Form "Urun Kart ID" alani ticimax_fields.URUNKARTIID'e yazar -> ust-seviye ile senkronla
    _tf_u = product_data.get("ticimax_fields") or {}
    _kid = str(_tf_u.get("URUNKARTIID") or product_data.get("urun_karti_id") or "").strip()
    if _kid:
        product_data["urun_karti_id"] = _kid
    # Bos urun_karti_id gonderilirse mevcut degeri ezme (sadece dolu ise guncelle)
    if "urun_karti_id" in product_data and not str(product_data.get("urun_karti_id") or "").strip():
        product_data.pop("urun_karti_id", None)

    product_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.products.update_one({"id": product_id}, {"$set": product_data})
    
    return {"message": "Ürün güncellendi"}

@router.delete("/{product_id}")
async def delete_product(
    product_id: str,
    current_user: dict = Depends(require_admin)
):
    """Ürünü çöp kutusuna taşır (soft delete). Kalıcı silme için /permanent kullanın."""
    product = await db.products.find_one({"id": product_id}, {"_id": 0, "is_active": 1})
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    await db.products.update_one(
        {"id": product_id},
        {"$set": {
            "is_deleted": True,
            "is_active": False,
            "prev_active": product.get("is_active", True),
            "deleted_at": datetime.now(timezone.utc).isoformat(),
        }}
    )
    return {"message": "Ürün çöp kutusuna taşındı"}

@router.post("/{product_id}/restore")
async def restore_product(
    product_id: str,
    current_user: dict = Depends(require_admin)
):
    """Çöp kutusundaki ürünü geri yükler."""
    product = await db.products.find_one({"id": product_id}, {"_id": 0, "prev_active": 1})
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    await db.products.update_one(
        {"id": product_id},
        {"$set": {"is_deleted": False, "is_active": product.get("prev_active", True),
                  "updated_at": datetime.now(timezone.utc).isoformat()},
         "$unset": {"deleted_at": "", "prev_active": ""}}
    )
    return {"message": "Ürün geri yüklendi"}

@router.delete("/{product_id}/permanent")
async def permanent_delete_product(
    product_id: str,
    current_user: dict = Depends(require_admin)
):
    """Ürünü veritabanından KALICI olarak siler (geri alınamaz)."""
    result = await db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    return {"message": "Ürün kalıcı olarak silindi"}

@router.post("/{product_id}/toggle-active")
async def toggle_product_active(
    product_id: str,
    current_user: dict = Depends(require_admin)
):
    """Toggle product active status"""
    product = await db.products.find_one({"id": product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    
    new_status = not product.get("is_active", True)
    await db.products.update_one(
        {"id": product_id},
        {"$set": {"is_active": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"is_active": new_status}

@router.get("/search/popular")
async def get_popular_searches():
    """Get popular search terms"""
    # In production, track and return actual popular searches
    return [
        {"term": "elbise", "count": 150},
        {"term": "bluz", "count": 120},
        {"term": "pantolon", "count": 100},
        {"term": "jean", "count": 90},
        {"term": "kazak", "count": 80},
    ]


# ==================== PRODUCT ATTRIBUTE IMPORT ====================

@router.post("/attributes/import-xlsx")
async def import_attributes_from_xlsx(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_admin)
):
    """
    Parse an XLSX file and extract product attributes per stock_code.
    Columns: one column for stock_code, rest are attribute types with values in cells.
    Returns: list of {stock_code, attributes: [{type, value}], matched_product_id}
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl yuklenmemis. pip install openpyxl")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Bos dosya")

    headers = [str(h).strip() if h else "" for h in rows[0]]

    # Find stock code column - look for "stok kodu", "stock_code", "barkod", "kod" etc.
    stock_col_idx = None
    stock_col_keywords = ["stok kodu", "stock code", "stock_code", "barkod", "barcode", "kod", "urun kodu", "urun_kodu"]
    for i, h in enumerate(headers):
        if any(kw in h.lower().replace(" ", " ") for kw in stock_col_keywords):
            stock_col_idx = i
            break
    if stock_col_idx is None:
        stock_col_idx = 0  # Default to first column

    # Attribute columns = all other columns
    attr_headers = [(i, h) for i, h in enumerate(headers) if i != stock_col_idx and h]

    results = []
    stock_codes = []

    for row in rows[1:]:
        stock_code = str(row[stock_col_idx]).strip() if row[stock_col_idx] else None
        if not stock_code or stock_code.lower() in ("none", "null", ""):
            continue

        attributes = []
        for col_idx, attr_type in attr_headers:
            value = row[col_idx] if col_idx < len(row) else None
            if value is not None and str(value).strip() not in ("", "None", "null"):
                attributes.append({
                    "type": attr_type,
                    "value": str(value).strip()
                })

        if attributes:
            results.append({
                "stock_code": stock_code,
                "attributes": attributes,
                "matched_product_id": None,
                "matched_product_name": None
            })
            stock_codes.append(stock_code)

    # Match with products by stock_code
    if stock_codes:
        products = await db.products.find(
            {"stock_code": {"$in": stock_codes}},
            {"_id": 0, "id": 1, "name": 1, "stock_code": 1}
        ).to_list(1000)
        product_map = {p["stock_code"]: p for p in products}

        for r in results:
            p = product_map.get(r["stock_code"])
            if p:
                r["matched_product_id"] = p["id"]
                r["matched_product_name"] = p["name"]

    return {
        "total_rows": len(results),
        "matched": sum(1 for r in results if r["matched_product_id"]),
        "unmatched": sum(1 for r in results if not r["matched_product_id"]),
        "attribute_types": [h for _, h in attr_headers],
        "results": results
    }


@router.post("/attributes/save-bulk")
async def save_attributes_bulk(payload: dict, current_user: dict = Depends(require_admin)):
    """
    Save attributes to multiple products.
    Payload: { updates: [{product_id, attributes: [{type, value, trendyol_attr_id, trendyol_attr_value_id}]}] }
    """
    updates = payload.get("updates", [])
    if not updates:
        raise HTTPException(status_code=400, detail="Guncellenecek urun yok")

    updated = 0
    for update in updates:
        product_id = update.get("product_id")
        attributes = update.get("attributes", [])
        if not product_id or not attributes:
            continue

        await db.products.update_one(
            {"id": product_id},
            {"$set": {
                "attributes": attributes,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        updated += 1

    return {"success": True, "updated": updated}


@router.get("/{product_id}/combine-products")
async def get_combine_products(product_id: str):
    """Bu ürünle birlikte gösterilecek kombin ürünlerin LİSTESİNİ döner.
    Public endpoint — sepet/ürün detay sayfası kullanır."""
    product = await db.products.find_one(
        {"id": product_id},
        {"_id": 0, "combine_products": 1, "category_id": 1, "categories": 1}
    )
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")

    combine_ids = product.get("combine_products") or []
    items = []
    if combine_ids:
        async for p in db.products.find(
            {"id": {"$in": combine_ids}, "is_active": {"$ne": False}},
            {"_id": 0, "id": 1, "name": 1, "slug": 1, "price": 1, "discount_price": 1,
             "images": 1, "image": 1, "stock": 1, "category_id": 1}
        ):
            items.append(p)
    return {"items": items, "source": "combine"}


@router.put("/{product_id}/combine-products")
async def update_combine_products(
    product_id: str,
    payload: dict,
    current_user: dict = Depends(require_admin)
):
    """Bu ürün için kombin ürün ID listesini günceller (admin)."""
    combine_ids = payload.get("combine_products") or []
    if not isinstance(combine_ids, list):
        raise HTTPException(status_code=400, detail="combine_products bir liste olmalıdır")
    # Self-reference temizliği
    combine_ids = [str(cid) for cid in combine_ids if str(cid) != product_id]
    # En fazla 12 kombin ürün
    combine_ids = combine_ids[:12]

    # Var olan ürün ID'lerini doğrula — fake/stale id'leri filtrele
    if combine_ids:
        existing = await db.products.distinct("id", {"id": {"$in": combine_ids}})
        existing_set = set(existing)
        combine_ids = [cid for cid in combine_ids if cid in existing_set]

    result = await db.products.update_one(
        {"id": product_id},
        {"$set": {
            "combine_products": combine_ids,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    return {"success": True, "count": len(combine_ids)}


@router.post("/{product_id}/auto-combine")
async def auto_assign_combine_products(
    product_id: str,
    payload: dict = None,
    current_user: dict = Depends(require_admin),
):
    """Geçmiş siparişlerdeki co-occurrence verisinden bu ürünle en sık birlikte
    satılan top-N ürünü otomatik kombin olarak atar."""
    payload = payload or {}
    max_n = min(int(payload.get("max", 8)), 12)
    dry_run = bool(payload.get("dry_run", False))
    replace = bool(payload.get("replace", True))

    base = await db.products.find_one({"id": product_id}, {"_id": 0, "combine_products": 1})
    if not base:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")

    co_count = {}
    async for order in db.orders.find(
        {"items.product_id": product_id}, {"_id": 0, "items.product_id": 1}
    ).limit(2000):
        ids_in_order = {it.get("product_id") for it in (order.get("items") or []) if it.get("product_id")}
        if product_id not in ids_in_order:
            continue
        for pid in ids_in_order:
            if pid and pid != product_id:
                co_count[pid] = co_count.get(pid, 0) + 1

    sorted_ids = sorted(co_count.items(), key=lambda kv: kv[1], reverse=True)
    if not sorted_ids:
        return {"success": False, "message": "Bu ürün için yeterli sipariş geçmişi yok", "candidates": []}

    candidate_ids = [pid for pid, _ in sorted_ids[:max_n * 3]]
    existing_ids = set(await db.products.distinct(
        "id", {"id": {"$in": candidate_ids}, "is_active": {"$ne": False}}
    ))

    candidates = []
    for pid, cnt in sorted_ids:
        if len(candidates) >= max_n:
            break
        if pid not in existing_ids:
            continue
        prod = await db.products.find_one(
            {"id": pid}, {"_id": 0, "id": 1, "name": 1, "price": 1, "images": 1, "image": 1}
        )
        if prod:
            candidates.append({**prod, "_co_count": cnt})

    selected_ids = [c["id"] for c in candidates]
    if dry_run:
        return {"success": True, "candidates": candidates, "would_assign": selected_ids, "dry_run": True}

    new_ids = selected_ids if replace else list(dict.fromkeys((base.get("combine_products") or []) + selected_ids))[:12]
    await db.products.update_one(
        {"id": product_id},
        {"$set": {
            "combine_products": new_ids,
            "combine_auto_generated_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }}
    )
    return {
        "success": True, "assigned": new_ids, "candidates": candidates,
        "count": len(new_ids),
        "message": f"{len(new_ids)} kombin ürün atandı (geçmiş siparişlerden)",
    }


@router.post("/auto-combine-all")
async def auto_assign_combine_all(
    payload: dict = None,
    current_user: dict = Depends(require_admin),
):
    """Tüm aktif ürünler için tek tıkla otomatik kombin atama (admin)."""
    payload = payload or {}
    max_n = min(int(payload.get("max", 8)), 12)
    only_empty = bool(payload.get("only_empty", True))

    query = {"is_active": {"$ne": False}}
    if only_empty:
        query["$or"] = [
            {"combine_products": {"$exists": False}},
            {"combine_products": {"$size": 0}},
        ]

    processed = 0
    assigned_total = 0
    skipped_no_data = 0

    cursor = db.products.find(query, {"_id": 0, "id": 1})
    async for p in cursor:
        pid = p["id"]
        co_count = {}
        async for order in db.orders.find(
            {"items.product_id": pid}, {"_id": 0, "items.product_id": 1}
        ).limit(500):
            ids_in_order = {it.get("product_id") for it in (order.get("items") or []) if it.get("product_id")}
            for cid in ids_in_order:
                if cid and cid != pid:
                    co_count[cid] = co_count.get(cid, 0) + 1
        sorted_ids = sorted(co_count.items(), key=lambda kv: kv[1], reverse=True)
        if not sorted_ids:
            skipped_no_data += 1
            processed += 1
            continue
        candidate_ids = [cid for cid, _ in sorted_ids[:max_n * 2]]
        existing_ids = set(await db.products.distinct(
            "id", {"id": {"$in": candidate_ids}, "is_active": {"$ne": False}}
        ))
        selected = [cid for cid, _ in sorted_ids if cid in existing_ids][:max_n]
        if selected:
            await db.products.update_one(
                {"id": pid},
                {"$set": {
                    "combine_products": selected,
                    "combine_auto_generated_at": datetime.now(timezone.utc).isoformat(),
                }}
            )
            assigned_total += 1
        processed += 1

    return {
        "success": True, "processed": processed,
        "products_with_combine_assigned": assigned_total,
        "skipped_no_order_history": skipped_no_data,
        "message": f"{assigned_total}/{processed} ürüne kombin atandı",
    }


@router.post("/cart-suggestions")
async def get_cart_suggestions(payload: dict):
    """Sepetteki ürünlere göre öneriler döner (public).
    
    Öncelik:
      1) Sepetteki ürünlerin combine_products listesi (cross-sell, manuel atama)
      2) Sale/indirim kategorisindeki aktif ürünler (fallback)
    """
    cart_product_ids = payload.get("product_ids") or []
    limit = int(payload.get("limit", 8))

    suggestions = []
    seen = set(cart_product_ids)

    # 1) Sepetteki her ürünün combine_products'ını topla
    if cart_product_ids:
        cart_products = []
        async for p in db.products.find(
            {"id": {"$in": cart_product_ids}},
            {"_id": 0, "combine_products": 1}
        ):
            cart_products.append(p)
        combine_ids = []
        for cp in cart_products:
            for cid in (cp.get("combine_products") or []):
                if cid not in seen:
                    combine_ids.append(cid)
                    seen.add(cid)
        if combine_ids:
            async for p in db.products.find(
                {"id": {"$in": combine_ids[:limit]}, "is_active": {"$ne": False}},
                {"_id": 0, "id": 1, "name": 1, "slug": 1, "price": 1, "discount_price": 1,
                 "images": 1, "image": 1, "stock": 1, "category_id": 1}
            ):
                suggestions.append({**p, "_source": "combine"})

    # 2) Yetersizse → sale/discount kategorisindeki aktif ürünlerle doldur
    needed = max(0, limit - len(suggestions))
    if needed > 0:
        sale_query = {
            "is_active": {"$ne": False},
            "id": {"$nin": list(seen)},
            "$or": [
                {"discount_price": {"$gt": 0}},
                {"is_on_sale": True},
                {"sale_active": True},
            ],
        }
        async for p in db.products.find(
            sale_query,
            {"_id": 0, "id": 1, "name": 1, "slug": 1, "price": 1, "discount_price": 1,
             "images": 1, "image": 1, "stock": 1, "category_id": 1}
        ).limit(needed):
            suggestions.append({**p, "_source": "sale"})
            seen.add(p["id"])

    # 3) Hala yetersizse → en son eklenmiş aktif ürünler
    needed = max(0, limit - len(suggestions))
    if needed > 0:
        async for p in db.products.find(
            {"is_active": {"$ne": False}, "id": {"$nin": list(seen)}},
            {"_id": 0, "id": 1, "name": 1, "slug": 1, "price": 1, "discount_price": 1,
             "images": 1, "image": 1, "stock": 1, "category_id": 1}
        ).sort("created_at", -1).limit(needed):
            suggestions.append({**p, "_source": "new"})

    return {"items": suggestions[:limit], "total": len(suggestions[:limit])}


@router.post("/checkout-deals")
async def get_checkout_deals(payload: dict):
    """Sepet sayfasındaki "Kasa Önü Fırsatları" — yalnızca indirimdeki aktif ürünler.

    Sepetteki ürünleri hariç tutar, indirimli olanları rastgele döner.
    """
    cart_product_ids = payload.get("product_ids") or []
    limit = int(payload.get("limit", 8))

    sale_query = {
        "is_active": {"$ne": False},
        "id": {"$nin": cart_product_ids},
        "$or": [
            {"discount_price": {"$gt": 0}},
            {"is_on_sale": True},
            {"sale_active": True},
        ],
    }

    deals = []
    async for p in db.products.find(
        sale_query,
        {"_id": 0, "id": 1, "name": 1, "slug": 1, "price": 1, "discount_price": 1,
         "images": 1, "image": 1, "stock": 1, "category_id": 1}
    ).limit(limit * 2):
        # Yalnızca gerçekten indirimi olanları kabul et
        if (p.get("discount_price") or 0) > 0 and p["discount_price"] < (p.get("price") or 0):
            deals.append(p)
        elif p.get("is_on_sale") or p.get("sale_active"):
            deals.append(p)
        if len(deals) >= limit:
            break

    return {"items": deals[:limit], "total": len(deals[:limit])}


@router.get("/{product_id}/attributes")
async def get_product_attributes(product_id: str, current_user: dict = Depends(require_admin)):
    """Get attributes for a single product"""
    product = await db.products.find_one({"id": product_id}, {"_id": 0, "attributes": 1, "name": 1, "stock_code": 1})
    if not product:
        raise HTTPException(status_code=404, detail="Urun bulunamadi")
    return {"attributes": product.get("attributes", []), "name": product.get("name"), "stock_code": product.get("stock_code")}


@router.put("/{product_id}/attributes")
async def update_product_attributes(product_id: str, payload: dict, current_user: dict = Depends(require_admin)):
    """Update attributes for a single product"""
    attributes = payload.get("attributes", [])
    await db.products.update_one(
        {"id": product_id},
        {"$set": {"attributes": attributes, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"success": True, "message": "Ozellikler guncellendi"}

@router.post("/bulk-update-vat")
async def bulk_update_vat(
    payload: dict,
    current_user: dict = Depends(require_admin)
):
    """Bulk update VAT for all products"""
    vat_rate = payload.get("vat_rate")
    if vat_rate is None:
        raise HTTPException(status_code=400, detail="VAT rate is required")
    
    result = await db.products.update_many({}, {"$set": {"vat_rate": vat_rate}})
    return {"message": f"{result.modified_count} ürünün KDV oranı %{vat_rate} olarak güncellendi."}


@router.post("/bulk/add-to-category-after")
async def bulk_add_to_category_after(
    after_card_id: int = Query(..., description="Bu Ürün Kart ID'den BÜYÜK sayısal kartlı ürünler kategoriye eklenir"),
    category_slug: str = Query("en-yeniler", description="Hedef kategori slug'ı (varsayılan: en-yeniler)"),
    set_is_new: bool = Query(True, description="Eşleşen ürünlerde is_new=True yapılsın mı"),
    dry_run: bool = Query(True, description="True → sadece önizleme (YAZMAZ). False → uygular."),
    current_user: dict = Depends(require_admin),
):
    """urun_karti_id (sayısal) > after_card_id olan TÜM ürünleri verilen kategoriye ekler.

    - Üyelik çoklu kategori dizilerine (category_ids + categories) addToSet ile yazılır;
      ürünün mevcut/ana kategorisini BOZMAZ (category_id/category_name'e dokunulmaz).
    - İdempotent: tekrar çalıştırmak güvenli (zaten ekliyse değişmez).
    - dry_run=True (VARSAYILAN) → yalnız eşleşen sayıyı + örnek döner, hiçbir şey yazmaz.
      Önce dry-run ile sayıyı gör, doğruysa dry_run=false ile uygula.
    """
    target_slug = (category_slug or "").strip().lower()
    cat_id = None
    cat_name = None
    async for c in db.categories.find({}, {"_id": 0, "id": 1, "name": 1, "slug": 1}):
        sl = (c.get("slug") or "").strip().lower()
        nm = (c.get("name") or "").strip()
        if sl == target_slug or generate_slug(nm) == target_slug:
            cat_id = c.get("id")
            cat_name = nm
            break
    if not cat_id:
        raise HTTPException(status_code=404, detail=f"'{category_slug}' kategorisi bulunamadı.")

    cat_ids_all = await _expand_category_ids([cat_id])  # ataları dahil (En Yeniler kökse sadece kendisi)

    matched = 0
    updated = 0
    sample = []
    cursor = db.products.find(
        {"urun_karti_id": {"$nin": [None, ""]}},
        {"_id": 0, "id": 1, "urun_karti_id": 1, "name": 1},
    )
    async for p in cursor:
        v = str(p.get("urun_karti_id") or "").strip()
        if not v.isdigit() or int(v) <= after_card_id:
            continue
        matched += 1
        if len(sample) < 15:
            sample.append({"urun_karti_id": v, "name": (p.get("name") or "")[:40]})
        if not dry_run:
            ops = {
                "$addToSet": {"category_ids": {"$each": cat_ids_all}, "categories": cat_id},
                "$set": {"updated_at": datetime.now(timezone.utc).isoformat()},
            }
            if set_is_new:
                ops["$set"]["is_new"] = True
            await db.products.update_one({"id": p["id"]}, ops)
            updated += 1

    return {
        "category": {"id": cat_id, "name": cat_name, "expanded_ids": cat_ids_all},
        "after_card_id": after_card_id,
        "matched": matched,
        "updated": (0 if dry_run else updated),
        "dry_run": dry_run,
        "sample": sample,
        "message": (
            f"DRY-RUN: Kart ID > {after_card_id} olan {matched} ürün '{cat_name}' kategorisine eklenecek (henüz YAZILMADI). "
            f"Uygulamak için aynı isteği dry_run=false ile çağırın."
            if dry_run else
            f"{updated} ürün '{cat_name}' kategorisine eklendi (Kart ID > {after_card_id})."
        ),
    }


@router.get("/export/excel")
async def export_products_excel(current_user: dict = Depends(require_admin)):
    """Export all products to an Excel file (variants as rows with dynamic attributes)"""
    try:
        products = await db.products.find({}, {"_id": 0}).to_list(None)
        
        # Collect all unique attribute names
        all_attr_names = set()
        for p in products:
            for attr in p.get("attributes", []):
                attr_name = attr.get("name") or attr.get("type")
                if attr_name:
                    all_attr_names.add(attr_name)
        
        rows = []
        for p in products:
            variants = p.get("variants", [])
            if not variants:
                variants = [{
                    "barcode": p.get("barcode", ""),
                    "stock_code": p.get("stock_code", ""),
                    "price": p.get("price", 0),
                    "sale_price": p.get("sale_price"),
                    "stock": p.get("stock", 0),
                    "size": "",
                    "color": ""
                }]
            
            for v in variants:
                row = {
                    "ID": p.get("id"),
                    "Ürün Adı": p.get("name"),
                    "Kategori": p.get("category_name"),
                    "Marka": p.get("brand"),
                    "Stok Kodu": v.get("stock_code") or p.get("stock_code"),
                    "Barkod": v.get("barcode") or p.get("barcode"),
                    "Beden": v.get("size", ""),
                    "Renk": v.get("color", ""),
                    "Piyasa Fiyatı": v.get("price") or p.get("price", 0),
                    "Satış Fiyatı": v.get("sale_price") or p.get("sale_price") or p.get("price", 0),
                    "Stok": v.get("stock", 0),
                    "Açıklama": p.get("description", ""),
                    "Aktif": "Evet" if p.get("is_active") else "Hayır"
                }
                
                # pre-fill attributes with empty string
                for attr_name in all_attr_names:
                    row[f"Özellik: {attr_name}"] = ""
                    
                # apply product attributes
                for attr in p.get("attributes", []):
                    attr_name = attr.get("name") or attr.get("type")
                    if attr_name and attr.get("value"):
                        row[f"Özellik: {attr_name}"] = attr["value"]
                        
                rows.append(row)
        
        df = pd.DataFrame(rows)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Ürünler')
        
        headers = {
            'Content-Disposition': 'attachment; filename="urunler.xlsx"',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        return Response(content=output.getvalue(), headers=headers)
        
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        raise HTTPException(status_code=500, detail=f"Dışa aktarma hatası: {str(e)}")

@router.post("/import/excel")
async def import_products_excel(file: UploadFile = File(...), current_user: dict = Depends(require_admin)):
    """Import or update products from an Excel file"""
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        # Validation
        required = ["Ürün Adı", "Kategori", "Satış Fiyatı"]
        for col in required:
            if col not in df.columns:
                raise Exception(f"Eksik sütun: {col}")
        
        stats = {"created": 0, "updated": 0, "errors": 0}
        
        for _, row in df.iterrows():
            try:
                barcode = str(row.get("Barkod", "")).strip()
                if not barcode or barcode == "nan":
                    continue
                
                # Parse dynamic attributes from columns
                parsed_attrs = []
                import uuid
                for col in df.columns:
                    if str(col).startswith("Özellik: "):
                        attr_name = str(col).replace("Özellik: ", "").strip()
                        val = str(row.get(col, "")).strip()
                        
                        # Ensure attribute exists in global library
                        existing_global = await db.attributes.find_one({"name": attr_name})
                        if not existing_global:
                            await db.attributes.insert_one({
                                "id": f"attr_{uuid.uuid4().hex[:8]}",
                                "name": attr_name,
                                "values": []
                            })
                            
                        if val and val != "nan":
                            parsed_attrs.append({"type": attr_name, "name": attr_name, "value": val})
                
                # Try finding product by variant barcode
                existing = await db.products.find_one({"variants.barcode": barcode})
                
                if existing:
                    # Update variant in existing product
                    update_fields = {
                        "variants.$.stock": int(row.get("Stok", 0) if pd.notna(row.get("Stok")) else 0),
                        "variants.$.price": float(row.get("Piyasa Fiyatı", row.get("Satış Fiyatı", 0)) if pd.notna(row.get("Piyasa Fiyatı", row.get("Satış Fiyatı", 0))) else 0),
                        "variants.$.sale_price": float(row.get("Satış Fiyatı", 0) if pd.notna(row.get("Satış Fiyatı")) else 0),
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }
                    if parsed_attrs:
                        update_fields["attributes"] = parsed_attrs

                    await db.products.update_one(
                        {"id": existing["id"], "variants.barcode": barcode},
                        {"$set": update_fields}
                    )
                    stats["updated"] += 1
                else:
                    # Create new product or add as variant to existing product with same name
                    name = str(row.get("Ürün Adı"))
                    prod_by_name = await db.products.find_one({"name": name})
                    
                    variant = {
                        "barcode": barcode,
                        "stock_code": str(row.get("Stok Kodu", "")).replace("nan", ""),
                        "size": str(row.get("Beden", "")).replace("nan", ""),
                        "color": str(row.get("Renk", "")).replace("nan", ""),
                        "price": float(row.get("Piyasa Fiyatı", row.get("Satış Fiyatı", 0)) if pd.notna(row.get("Piyasa Fiyatı", row.get("Satış Fiyatı", 0))) else 0),
                        "sale_price": float(row.get("Satış Fiyatı", 0) if pd.notna(row.get("Satış Fiyatı")) else 0),
                        "stock": int(row.get("Stok", 0) if pd.notna(row.get("Stok")) else 0)
                    }
                    
                    if prod_by_name:
                        # Add as new variant
                        update_fields = {
                            "updated_at": datetime.now(timezone.utc).isoformat()
                        }
                        if parsed_attrs:
                            update_fields["attributes"] = parsed_attrs
                        
                        await db.products.update_one(
                            {"id": prod_by_name["id"]},
                            {"$push": {"variants": variant}, "$set": update_fields}
                        )
                        stats["updated"] += 1
                    else:
                        # Create full new product
                        new_id = await generate_short_id("products")
                        new_p = {
                            "id": new_id,
                            "name": name,
                            "slug": generate_slug(name),
                            "category_name": str(row.get("Kategori", "")).replace("nan", ""),
                            "brand": str(row.get("Marka", "")).replace("nan", ""),
                            "description": str(row.get("Açıklama", "")).replace("nan", ""),
                            "price": variant["price"],
                            "sale_price": variant["sale_price"],
                            "stock": variant["stock"],
                            "is_active": True,
                            "variants": [variant],
                            "images": [],
                            "attributes": parsed_attrs,
                            "created_at": datetime.now(timezone.utc).isoformat(),
                            "updated_at": datetime.now(timezone.utc).isoformat()
                        }
                        await db.products.insert_one(new_p)
                        stats["created"] += 1
            except Exception as row_err:
                logger.error(f"Import row error: {row_err}")
                stats["errors"] += 1
                
        return {"success": True, "stats": stats}
        
    except Exception as e:
        logger.error(f"Excel import error: {e}")
        raise HTTPException(status_code=500, detail=str(e))



@router.post("/attributes/import-technical-xlsx")
async def import_technical_details_xlsx(
    file: UploadFile = File(...),
    current_user: dict = Depends(require_admin)
):
    """
    Import technical details from Excel in format:
    UrunKartID | StokKodu | UrunAdi | Ozellik | Deger
    Groups by UrunAdi, fuzzy matches with existing products, returns preview.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl yüklenmemiş")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Dosya boş veya başlık satırı eksik")

    headers = [str(h).strip() if h else "" for h in rows[0]]

    # Find column indices
    name_col = None
    ozellik_col = None
    deger_col = None
    stok_kodu_col = None

    for i, h in enumerate(headers):
        hl = h.lower().replace("ı", "i").replace("ö", "o").replace("ü", "u")
        if "urunadi" in hl.replace(" ", "") or "ürün adı" in h.lower() or "urun adi" in h.lower():
            name_col = i
        elif "ozellik" in hl.replace(" ", "") or "özellik" in h.lower():
            ozellik_col = i
        elif "deger" in hl.replace(" ", "") or "değer" in h.lower():
            deger_col = i
        elif "stokkodu" in hl.replace(" ", "") or "stok kodu" in h.lower():
            stok_kodu_col = i

    if name_col is None:
        raise HTTPException(status_code=400, detail="UrunAdi sütunu bulunamadı")
    if deger_col is None:
        raise HTTPException(status_code=400, detail="Deger sütunu bulunamadı")

    # Group by product name - deduplicate attributes (last value wins for same type)
    product_groups = {}
    # Metadata column headers that should not be treated as attributes
    meta_headers_lower = {h.lower().strip() for h in headers if h}

    for row in rows[1:]:
        name = str(row[name_col]).strip() if row[name_col] else None
        if not name or name.lower() in ("none", "null", ""):
            continue

        ozellik = str(row[ozellik_col]).strip() if ozellik_col is not None and row[ozellik_col] else ""
        deger = str(row[deger_col]).strip() if row[deger_col] else ""
        stok_kodu = str(row[stok_kodu_col]).strip() if stok_kodu_col is not None and row[stok_kodu_col] else ""

        if not deger or deger.lower() in ("none", "null"):
            continue

        if name not in product_groups:
            product_groups[name] = {"stok_kodu": stok_kodu, "attributes": {}, "extra_colors": []}

        if ozellik and ozellik.lower() not in ("none", "null", "") and ozellik.lower() not in meta_headers_lower:
            # Use dict to deduplicate (last value wins)
            product_groups[name]["attributes"][ozellik] = deger
        elif not ozellik or ozellik.lower() in ("none", "null", ""):
            # Empty ozellik with a deger = extra color variant
            product_groups[name]["extra_colors"].append(deger)

    # Convert attribute dicts to list format
    for name, data in product_groups.items():
        data["attributes_list"] = [{"type": k, "value": v} for k, v in data["attributes"].items()]

    # Now match products by name - one Excel product can match MULTIPLE DB products
    all_products = await db.products.find({}, {"_id": 0, "id": 1, "name": 1, "stock_code": 1}).to_list(None)

    def normalize(s):
        return s.lower().replace("ı", "i").replace("ğ", "g").replace("ü", "u").replace("ş", "s").replace("ö", "o").replace("ç", "c").replace("İ", "i").strip()

    results = []
    used_product_ids = set()

    for excel_name, data in product_groups.items():
        excel_norm = normalize(excel_name)

        # Find ALL matching products (for color variants)
        matches = []
        for p in all_products:
            p_norm = normalize(p["name"])
            if p_norm == excel_norm:
                matches.append((p, 100))
            elif excel_norm in p_norm:
                # Excel name is a substring of DB name (e.g., "Basic Triko" in "Basic Triko Siyah")
                overlap = len(excel_norm.split()) / len(p_norm.split()) * 100
                if overlap >= 50:
                    matches.append((p, round(overlap, 1)))
            elif p_norm in excel_norm:
                overlap = len(p_norm.split()) / len(excel_norm.split()) * 100
                if overlap >= 50:
                    matches.append((p, round(overlap, 1)))

        if matches:
            matches.sort(key=lambda x: x[1], reverse=True)
            for match_p, match_score in matches:
                if match_p["id"] not in used_product_ids:
                    results.append({
                        "excel_name": excel_name,
                        "stok_kodu": data["stok_kodu"],
                        "attributes": data["attributes_list"],
                        "extra_colors": data["extra_colors"],
                        "matched_product_id": match_p["id"],
                        "matched_product_name": match_p["name"],
                        "match_score": match_score
                    })
                    used_product_ids.add(match_p["id"])
        else:
            results.append({
                "excel_name": excel_name,
                "stok_kodu": data["stok_kodu"],
                "attributes": data["attributes_list"],
                "extra_colors": data["extra_colors"],
                "matched_product_id": None,
                "matched_product_name": None,
                "match_score": 0
            })

    results.sort(key=lambda r: r["match_score"], reverse=True)

    return {
        "success": True,
        "total_excel_products": len(results),
        "matched": sum(1 for r in results if r["matched_product_id"]),
        "unmatched": sum(1 for r in results if not r["matched_product_id"]),
        "results": results
    }


@router.post("/attributes/apply-technical-xlsx")
async def apply_technical_details(payload: dict, current_user: dict = Depends(require_admin)):
    """
    Apply matched technical details to products.
    Payload: { updates: [{product_id, attributes: [{type, value}], extra_colors: []}] }
    """
    updates = payload.get("updates", [])
    if not updates:
        raise HTTPException(status_code=400, detail="Güncellenecek ürün yok")

    updated = 0
    attr_lib_updates = {}

    for update in updates:
        product_id = update.get("product_id")
        attributes = update.get("attributes", [])
        if not product_id or not attributes:
            continue

        # Replace attributes with Excel data (clean import)
        new_attrs = [{"type": a["type"], "name": a["type"], "value": a["value"]} for a in attributes]

        # Track for attribute library
        for new_attr in attributes:
            if new_attr["type"] not in attr_lib_updates:
                attr_lib_updates[new_attr["type"]] = set()
            attr_lib_updates[new_attr["type"]].add(new_attr["value"])

        await db.products.update_one(
            {"id": product_id},
            {"$set": {
                "attributes": new_attrs,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        updated += 1

    # Update global attribute library with new values
    for attr_name, val_set in attr_lib_updates.items():
        existing_lib = await db.attributes.find_one({"name": {"$regex": f"^{re.escape(attr_name)}$", "$options": "i"}})
        if existing_lib:
            current_vals = set(existing_lib.get("values", []))
            merged_vals = list(current_vals.union(val_set))
            if len(merged_vals) > len(current_vals):
                await db.attributes.update_one(
                    {"_id": existing_lib["_id"]},
                    {"$set": {"values": merged_vals, "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
        else:
            await db.attributes.insert_one({
                "id": generate_id(),
                "name": attr_name,
                "values": list(val_set),
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            })

    return {"success": True, "updated": updated, "message": f"{updated} ürünün özellikleri güncellendi"}
