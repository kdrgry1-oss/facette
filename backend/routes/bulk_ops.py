"""
=============================================================================
bulk_ops.py — Toplu Fiyat/Stok İşlemleri + Stok Uyarı
=============================================================================

AMAÇ:
  Giyim sektöründe günlük işleyen operasyonlar:
    1) Excel ile toplu fiyat/stok güncelleme (ideaSoft/T-Soft'ta standart)
    2) Kritik seviye altı ürünler (stok uyarı sistemi)
    3) Yeniden sipariş / reorder önerisi (stokta 0 olan ama satılmış ürünler)

ENDPOINT'LER:
  POST /api/bulk-ops/price-stock/template  → boş Excel şablonu
  POST /api/bulk-ops/price-stock/preview   → yüklenen Excel'i önizle (commit YOK)
  POST /api/bulk-ops/price-stock/apply     → Excel'i uygula (DB update)
  GET  /api/bulk-ops/stock-alerts          → kritik stok uyarıları
  GET  /api/bulk-ops/reorder-suggestions   → yeniden sipariş önerileri

KULLANAN FRONTEND:
  /app/frontend/src/pages/admin/BulkPriceStock.jsx
  /app/frontend/src/pages/admin/StockAlerts.jsx
=============================================================================
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
from io import BytesIO
import openpyxl

from .deps import db, require_admin

router = APIRouter(prefix="/bulk-ops", tags=["Bulk Operations"])


# ---------------------------------------------------------------------------
# Template (şablon)
# ---------------------------------------------------------------------------
@router.get("/price-stock/template")
async def get_template(current_user: dict = Depends(require_admin)):
    """
    İçi boş bir Excel şablonu döner. Kullanıcı bunu doldurup apply endpoint'e yükler.
    Kolonlar: stock_code, barcode, price, sale_price, stock, status
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fiyat-Stok"
    headers = ["stock_code", "barcode", "price", "sale_price", "stock", "status"]
    ws.append(headers)
    # Örnek veri satırı
    ws.append(["", "", "", "", "", "active"])
    # Genişlik ayarı
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[chr(64 + i)].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=fiyat-stok-sablon.xlsx"},
    )


# ---------------------------------------------------------------------------
# Yardımcı: Excel → dict rows
# ---------------------------------------------------------------------------
def _parse_xlsx(content: bytes):
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip().lower() for h in rows[0]]
    out = []
    for row in rows[1:]:
        if all(c is None or (isinstance(c, str) and c.strip() == "") for c in row):
            continue
        d = {}
        for i, h in enumerate(headers):
            if i < len(row):
                d[h] = row[i]
        out.append(d)
    return out


# ---------------------------------------------------------------------------
# Preview (dry-run)
# ---------------------------------------------------------------------------
@router.post("/price-stock/preview")
async def preview_upload(file: UploadFile = File(...),
                          current_user: dict = Depends(require_admin)):
    """
    Excel'i okur, hangi ürünlere hangi alanların güncelleneceğini gösterir.
    Hiçbir DB update YAPMAZ.
    """
    content = await file.read()
    try:
        rows = _parse_xlsx(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel okunamadı: {e}")

    preview = []
    match_stock = 0
    match_barcode = 0
    not_found = 0

    for idx, r in enumerate(rows):
        sc = str(r.get("stock_code") or "").strip()
        bc = str(r.get("barcode") or "").strip()
        query = None
        key_used = None
        if sc:
            query = {"$or": [{"stock_code": sc}, {"variants.stock_code": sc}]}
            key_used = "stock_code"
        elif bc:
            query = {"$or": [{"barcode": bc}, {"variants.barcode": bc}]}
            key_used = "barcode"
        if not query:
            preview.append({"row": idx + 2, "error": "stock_code ya da barcode zorunlu",
                            "data": r})
            continue
        product = await db.products.find_one(query, {"_id": 0, "id": 1, "name": 1, "variants": 1})
        if not product:
            not_found += 1
            preview.append({"row": idx + 2, "error": "Ürün bulunamadı", "key": key_used, "ref": sc or bc, "data": r})
            continue
        if key_used == "stock_code": match_stock += 1
        else: match_barcode += 1

        updates = {}
        if r.get("price") not in (None, ""): updates["price"] = r["price"]
        if r.get("sale_price") not in (None, ""): updates["sale_price"] = r["sale_price"]
        if r.get("stock") not in (None, ""): updates["stock"] = r["stock"]
        if r.get("status") not in (None, ""): updates["status"] = r["status"]

        preview.append({
            "row": idx + 2,
            "product_id": product.get("id"),
            "product_name": product.get("name", ""),
            "updates": updates,
            "ref": sc or bc,
        })

    return {
        "total_rows": len(rows),
        "matched_by_stock_code": match_stock,
        "matched_by_barcode": match_barcode,
        "not_found": not_found,
        "preview": preview[:500],  # ilk 500 satır gösterilir
    }


# ---------------------------------------------------------------------------
# Apply
# ---------------------------------------------------------------------------
@router.post("/price-stock/apply")
async def apply_upload(file: UploadFile = File(...),
                        current_user: dict = Depends(require_admin)):
    """Yüklenen Excel'i uygular — eşleşen her ürün için updateOne çağırır."""
    content = await file.read()
    try:
        rows = _parse_xlsx(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel okunamadı: {e}")

    ok = 0; fail = 0; skipped = 0
    now = datetime.now(timezone.utc).isoformat()
    for r in rows:
        sc = str(r.get("stock_code") or "").strip()
        bc = str(r.get("barcode") or "").strip()
        query = None; key_used = None
        if sc:
            query = {"$or": [{"stock_code": sc}, {"variants.stock_code": sc}]}
            key_used = "stock_code"; ref = sc
        elif bc:
            query = {"$or": [{"barcode": bc}, {"variants.barcode": bc}]}
            key_used = "barcode"; ref = bc
        else:
            skipped += 1; continue
        product = await db.products.find_one(query, {"_id": 0, "id": 1, "variants": 1})
        if not product:
            fail += 1; continue

        root_updates = {}
        if r.get("price") not in (None, ""):
            try: root_updates["price"] = float(r["price"])
            except: pass
        if r.get("sale_price") not in (None, ""):
            try: root_updates["sale_price"] = float(r["sale_price"])
            except: pass
        if r.get("status") not in (None, ""):
            root_updates["status"] = str(r["status"]).strip()

        # Stok: varyant bazlı güncelle (doğru eşleşen varyant tek ise onun stock'u)
        variant_updated = False
        if r.get("stock") not in (None, ""):
            try:
                new_stock = int(float(r["stock"]))
            except:
                new_stock = None
            if new_stock is not None:
                variants = product.get("variants") or []
                target_idx = None
                for i, v in enumerate(variants):
                    if key_used == "stock_code" and v.get("stock_code") == ref:
                        target_idx = i; break
                    if key_used == "barcode" and v.get("barcode") == ref:
                        target_idx = i; break
                if target_idx is not None:
                    await db.products.update_one(
                        {"id": product["id"]},
                        {"$set": {f"variants.{target_idx}.stock": new_stock, "updated_at": now}}
                    )
                    variant_updated = True
                else:
                    # Ürün seviyesinde stock tutuluyorsa root'a yaz
                    root_updates["stock"] = new_stock

        if root_updates:
            root_updates["updated_at"] = now
            await db.products.update_one({"id": product["id"]}, {"$set": root_updates})

        if root_updates or variant_updated: ok += 1
        else: skipped += 1

    return {"success": True, "applied": ok, "failed": fail, "skipped": skipped,
            "message": f"{ok} ürün güncellendi, {fail} bulunamadı, {skipped} atlandı"}


# ---------------------------------------------------------------------------
# Stock Alerts
# ---------------------------------------------------------------------------
@router.get("/stock-alerts")
async def stock_alerts(threshold: int = 3,
                        current_user: dict = Depends(require_admin)):
    """
    Stok seviyesi threshold'un altındaki ürün-varyant kombinasyonlarını döner.
    Varyantsız ürünlerde kök product.stock (varsa) ile kontrol edilir.
    """
    alerts = []
    async for p in db.products.find(
        {"status": {"$ne": "archived"}},
        {"_id": 0, "id": 1, "name": 1, "stock_code": 1, "images": 1, "variants": 1, "stock": 1, "price": 1}
    ):
        variants = p.get("variants") or []
        if variants:
            for v in variants:
                s = v.get("stock")
                if s is not None and s <= threshold:
                    alerts.append({
                        "product_id": p.get("id"),
                        "product_name": p.get("name", ""),
                        "image": (p.get("images") or [None])[0],
                        "variant": f"{v.get('size','')} {v.get('color','')}".strip(),
                        "stock_code": v.get("stock_code") or p.get("stock_code"),
                        "barcode": v.get("barcode"),
                        "stock": s,
                        "price": p.get("price"),
                    })
        else:
            s = p.get("stock")
            if s is not None and s <= threshold:
                alerts.append({
                    "product_id": p.get("id"),
                    "product_name": p.get("name", ""),
                    "image": (p.get("images") or [None])[0],
                    "variant": "—",
                    "stock_code": p.get("stock_code"),
                    "barcode": None,
                    "stock": s,
                    "price": p.get("price"),
                })
    alerts.sort(key=lambda x: x["stock"])
    return {"threshold": threshold, "total": len(alerts), "items": alerts[:500]}


# ---------------------------------------------------------------------------
# Reorder suggestions (stokta 0 ama geçmişte satılmış)
# ---------------------------------------------------------------------------
@router.get("/reorder-suggestions")
async def reorder_suggestions(current_user: dict = Depends(require_admin)):
    """
    Son 60 gündeki sipariş kalemlerinden satılan ürünler arasında, şu anda
    stok == 0 olanları "yeniden sipariş gerekli" olarak listeler.
    """
    from datetime import timedelta
    cutoff = (datetime.now(timezone.utc) - timedelta(days=60)).isoformat()
    pipeline = [
        {"$match": {"created_at": {"$gte": cutoff}}},
        {"$unwind": "$items"},
        {"$group": {
            "_id": "$items.product_id",
            "sold": {"$sum": "$items.quantity"},
            "last_sold": {"$max": "$created_at"},
        }},
        {"$sort": {"sold": -1}},
        {"$limit": 500},
    ]
    agg = await db.orders.aggregate(pipeline).to_list(length=500)
    out = []
    for a in agg:
        pid = a["_id"]
        if not pid: continue
        p = await db.products.find_one({"id": pid},
            {"_id": 0, "id": 1, "name": 1, "stock": 1, "variants": 1, "images": 1, "price": 1, "stock_code": 1})
        if not p: continue
        total_stock = p.get("stock") or 0
        variants = p.get("variants") or []
        if variants:
            total_stock = sum((v.get("stock") or 0) for v in variants)
        if total_stock <= 0:
            out.append({
                "product_id": p["id"],
                "product_name": p.get("name", ""),
                "image": (p.get("images") or [None])[0],
                "stock_code": p.get("stock_code"),
                "sold_60_days": a["sold"],
                "last_sold": a["last_sold"],
                "current_stock": total_stock,
            })
    return {"total": len(out), "items": out}


# ---------------------------------------------------------------------------
# Stokta biten ürünleri pazaryerlerinde pasife alma
# ---------------------------------------------------------------------------
@router.post("/stock-alerts/deactivate-on-marketplaces")
async def deactivate_out_of_stock_on_marketplaces(
    payload: dict = None,
    current_user: dict = Depends(require_admin),
):
    """
    Stokta `threshold` (vars.: 0) olan seçili ürünleri ya da tüm out-of-stock
    ürünleri, aktif pazaryerlerindeki envanter update endpoint'ini (qty=0
    göndererek) çağırır. Canlı API yoksa sadece `integration_logs` düşer.
    """
    from .marketplace_hub import log_integration_event
    payload = payload or {}
    product_ids = payload.get("product_ids") or []
    threshold = int(payload.get("threshold", 0))

    # Hedef ürünleri topla
    q = {"is_active": True}
    if product_ids:
        q["id"] = {"$in": product_ids}
    targets = []
    async for p in db.products.find(q, {"_id": 0}):
        variants = p.get("variants") or []
        total_stock = sum((v.get("stock") or 0) for v in variants) if variants else (p.get("stock") or 0)
        if total_stock <= threshold:
            targets.append(p)

    if not targets:
        return {"success": True, "message": "Pasife alınacak stoksuz ürün yok", "processed": 0}

    # Aktif pazaryerleri
    accounts = await db.marketplace_accounts.find(
        {"enabled": True}, {"_id": 0, "key": 1}
    ).to_list(100)
    active_mps = [a["key"] for a in accounts]

    results = {"processed": 0, "marketplaces": {}}

    for mp in active_mps:
        mp_ok = 0
        mp_fail = 0
        # Trendyol için gerçek push
        if mp == "trendyol":
            try:
                from .integrations import _sync_inventory_to_trendyol, get_trendyol_config
                cfg = await get_trendyol_config()
                if cfg.get("is_active"):
                    # qty=0 göndermek için product listesini kopyala ve stock=0 zorla
                    patched = []
                    for p in targets:
                        pc = {**p, "stock": 0, "variants": [{**v, "stock": 0} for v in (p.get("variants") or [])]}
                        patched.append(pc)
                    res = await _sync_inventory_to_trendyol(patched)
                    mp_ok = len(targets) if res.get("success") else 0
                    mp_fail = 0 if res.get("success") else len(targets)
                    await log_integration_event(
                        marketplace="trendyol", action="stock_update",
                        status=("success" if res.get("success") else "failed"),
                        direction="outbound",
                        message=f"Stoksuz {len(targets)} ürün Trendyol'da pasifleştirildi (qty=0)"
                    )
                else:
                    await log_integration_event(
                        marketplace="trendyol", action="stock_update", status="failed",
                        direction="outbound",
                        message="Trendyol konfigürasyonu yok — pasifleme atlandı"
                    )
                    mp_fail = len(targets)
            except Exception as e:
                mp_fail = len(targets)
                await log_integration_event(
                    marketplace="trendyol", action="stock_update", status="failed",
                    direction="outbound", message=f"Pasifleme hatası: {e}"
                )
        else:
            # Diğer pazaryerleri (hepsiburada/temu/..): şu an log ile kaydedilir
            await log_integration_event(
                marketplace=mp, action="stock_update", status="queued",
                direction="outbound",
                message=f"{len(targets)} ürün için pazaryerinde pasifleme kuyruğa alındı (canlı API bağlandığında uygulanır)"
            )
            mp_ok = len(targets)
        results["marketplaces"][mp] = {"success": mp_ok, "failed": mp_fail}

    results["processed"] = len(targets)
    # Genel başarı: en az bir pazaryerinde tam başarı ve hiç kritik hata yok
    total_fail = sum(v.get("failed", 0) for v in results["marketplaces"].values())
    overall_success = total_fail == 0
    msg = f"{len(targets)} ürün için pasifleme işlemi tetiklendi"
    if total_fail:
        msg += f" — {total_fail} pazaryeri güncelleme başarısız (log detayı)"
    return {"success": overall_success, "result": results, "message": msg}
