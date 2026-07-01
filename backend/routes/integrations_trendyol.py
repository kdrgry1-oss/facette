"""
integrations_trendyol.py — Trendyol pazaryeri entegrasyonu (ürün/stok/sipariş/iade/soru-cevap).
2026-07-01 refactor: integrations.py'den ayrıştırıldı (bkz. integrations_common.py başlığı).
"""
from fastapi import APIRouter, HTTPException, Query, Depends, Response, BackgroundTasks, Request, Body, UploadFile, File
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone, timedelta
import os
import base64
import uuid
import re
import xml.etree.ElementTree as ET
import httpx
import hashlib

from .deps import db, logger, get_current_user, require_admin, generate_id, generate_short_id
from facette_defaults import facette_fixed_value_for  # tüm-pazaryeri sabit varsayılan (gap-fill)

router = APIRouter(tags=["Integrations-Trendyol"])

from .integrations_common import (
    _BAD_COMPOSITION_VALUES,
    _ORDER_STATUS_TR,
    _RETURN_STATUS_KEYS,
    _build_product_query_from_payload,
    _claim_bucket,
    _decrement_stock_for_imported_order,
    _dedupe_products_by_stock_code,
    _derive_claim_status,
    _first_seen_stamps,
    _mp_base_price,
    _ms_to_iso,
    _norm_val,
    _normalize_attr_key,
    _order_payment_type,
    _resolve_stock_code,
    _resolve_value_id,
    _search_tr_regex,
    log_integration_event,
    restock_claim_once,
)

# ---- Request/response modelleri ----
class TrendyolOrderPreviewReq(BaseModel):
    order_number: Optional[str] = None
    start_date_ms: Optional[int] = None
    end_date_ms: Optional[int] = None

class TrendyolOrderImportReq(BaseModel):
    orders: List[dict]

class CategoryMappingReq(BaseModel):
    local_category_id: str
    local_name: str
    trendyol_category_id: int
    trendyol_category_name: str

class AttributeMapping(BaseModel):
    local_attr: str
    trendyol_attr_id: int

class AttributeMappingReq(BaseModel):
    attribute_mappings: List[AttributeMapping]
    default_mappings: Optional[dict] = {}


async def get_trendyol_config():
    """Get Trendyol configuration from DB or env.
    NOTE: default_markup için Ana Ayarlar > trendyol_markup ÖNCELİKLİ — kullanıcı UI'da
    en son nereye girerse oradan okunur."""
    settings = await db.settings.find_one({"id": "trendyol"})
    # Main settings'ten markup override
    main_settings = await db.settings.find_one({"id": "main"}) or {}
    main_markup = main_settings.get("trendyol_markup")
    try:
        main_markup_f = float(main_markup) if (main_markup is not None and main_markup != "") else None
    except Exception:
        main_markup_f = None

    if settings:
        mode = settings.get("mode", "sandbox")
        local_markup = settings.get("default_markup", 0) or 0
        effective_markup = main_markup_f if main_markup_f is not None else local_markup
        return {
            "api_key": settings.get("api_key", ""),
            "api_secret": settings.get("api_secret", ""),
            "supplier_id": settings.get("supplier_id", ""),
            "is_active": settings.get("is_active", False),
            "mode": mode,
            "default_markup": effective_markup,
            "default_brand_id": settings.get("default_brand_id"),
            "default_cargo_company_id": settings.get("default_cargo_company_id"),
            "default_vat_rate": settings.get("default_vat_rate"),
            "base_url": 'https://api.trendyol.com' if mode == 'live' else 'https://stageapigw.trendyol.com'
        }
    
    # Fallback to env
    mode = os.environ.get('TRENDYOL_MODE', 'sandbox')
    return {
        "api_key": os.environ.get('TRENDYOL_API_KEY', ''),
        "api_secret": os.environ.get('TRENDYOL_API_SECRET', ''),
        "supplier_id": os.environ.get('TRENDYOL_SUPPLIER_ID', ''),
        "is_active": bool(os.environ.get('TRENDYOL_API_KEY')),
        "mode": mode,
        "base_url": 'https://api.trendyol.com' if mode == 'live' else 'https://stageapigw.trendyol.com'
    }
async def get_trendyol_headers():
    config = await get_trendyol_config()
    if not config["api_key"] or not config["api_secret"]:
        return None
    credentials = f'{config["api_key"]}:{config["api_secret"]}'
    encoded = base64.b64encode(credentials.encode()).decode()
    return {
        "Authorization": f"Basic {encoded}",
        "User-Agent": f'{config["supplier_id"]} - FacetteIntegration',
        "Content-Type": "application/json"
    }
def calculate_trendyol_price(base_price: float, product_data: dict, trendyol_config: dict) -> float:
    """Calculate price with markup logic"""
    # SSOT: her zaman global default_markup uygulanir; urun bazli override yok sayilir
    # (kullanici talebi: belirlenen oran disinda fiyat guncellenmez)
    markup = float(trendyol_config.get("default_markup", 0) or 0)

    final_price = base_price * (1 + markup / 100)
    return round(final_price, 2)
@router.get("/trendyol/settings")
async def get_trendyol_settings(current_user: dict = Depends(require_admin)):
    """Get Trendyol settings"""
    config = await get_trendyol_config()

    # Single source of truth: Ana Ayarlar sayfasındaki `trendyol_markup`
    # main settings'ten yazıldıysa ÖNCELİKLİ olarak onu kullan; aksi halde
    # Trendyol Integration kartındaki default_markup'a düş.
    main_settings = await db.settings.find_one({"id": "main"})
    main_markup = (main_settings or {}).get("trendyol_markup")
    if main_markup is not None and main_markup != "":
        try:
            default_markup = float(main_markup)
        except Exception:
            default_markup = config.get("default_markup", 0) or 0
    else:
        default_markup = config.get("default_markup", 0) or 0

    # Mask secrets
    return {
        "supplier_id": config.get("supplier_id", ""),
        "api_key": config.get("api_key", ""),
        "api_secret": "********" if config.get("api_secret") else "",
        "mode": config.get("mode", "sandbox"),
        "is_active": config.get("is_active", False),
        "default_markup": default_markup,
        "default_brand_id": config.get("default_brand_id"),
        "default_cargo_company_id": config.get("default_cargo_company_id"),
        "default_vat_rate": config.get("default_vat_rate")
    }
@router.post("/trendyol/settings")
async def save_trendyol_settings(
    settings: dict,
    current_user: dict = Depends(require_admin)
):
    """Save Trendyol settings"""
    from datetime import datetime, timezone

    # Required alan validasyonu — is_active=True ise supplier_id/api_key/api_secret zorunlu
    if settings.get("is_active"):
        existing = await db.settings.find_one({"id": "trendyol"}, {"_id": 0}) or {}
        supplier_id = settings.get("supplier_id") or existing.get("supplier_id")
        api_key = settings.get("api_key") or existing.get("api_key")
        api_secret = settings.get("api_secret")
        if api_secret in (None, "", "********"):
            api_secret = existing.get("api_secret")
        missing = [k for k, v in {"supplier_id": supplier_id, "api_key": api_key, "api_secret": api_secret}.items() if not v]
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"Trendyol aktifleştirmek için zorunlu alanlar eksik: {', '.join(missing)}"
            )

    update_data = {
        "supplier_id": settings.get("supplier_id", ""),
        "api_key": settings.get("api_key", ""),
        "mode": settings.get("mode", "sandbox"),
        "is_active": settings.get("is_active", False),
        "default_markup": settings.get("default_markup", 0),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }

    if settings.get("api_secret") and settings.get("api_secret") != "********":
        update_data["api_secret"] = settings.get("api_secret")

    # Faz T3 (white-label): kanal varsayilanlari — yalniz gonderildiyse yaz (yoksa mevcut korunur).
    for _k in ("default_brand_id", "default_cargo_company_id", "default_vat_rate"):
        if _k in settings:
            _v = settings.get(_k)
            if _v in (None, ""):
                update_data[_k] = None
            else:
                try:
                    update_data[_k] = int(_v)
                except (TypeError, ValueError):
                    update_data[_k] = None

    await db.settings.update_one(
        {"id": "trendyol"},
        {"$set": update_data},
        upsert=True
    )

    # SSOT: trendyol_markup'i main settings'e de yansıt (UI'da ana ayarlar sayfası bu key'i okur)
    try:
        await db.settings.update_one(
            {"id": "main"},
            {"$set": {"trendyol_markup": update_data["default_markup"]}},
            upsert=True,
        )
    except Exception:
        pass

    return {"success": True, "message": "Trendyol ayarları kaydedildi"}
@router.post("/trendyol/test-connection")
async def test_trendyol_connection(current_user: dict = Depends(require_admin)):
    """Trendyol gerçek bağlantı testi — brands endpoint'i üzerinden."""
    cfg = await get_trendyol_config()
    if not cfg.get("api_key") or not cfg.get("api_secret") or not cfg.get("supplier_id"):
        return {"success": False, "message": "Trendyol API bilgileri eksik (api_key / api_secret / supplier_id zorunlu)"}
    try:
        from trendyol_client import TrendyolClient
        client = TrendyolClient(
            supplier_id=cfg["supplier_id"],
            api_key=cfg["api_key"],
            api_secret=cfg["api_secret"],
            mode=cfg["mode"],
        )
        # Hafif bir probe — ilk 1 marka yeter
        data = await client.get_brands(size=1, page=0)
        if isinstance(data, dict) and ("brands" in data or "content" in data or "totalElements" in data):
            return {"success": True, "message": "Trendyol bağlantısı başarılı", "mode": cfg["mode"]}
        return {"success": False, "message": f"Beklenmeyen yanıt: {str(data)[:200]}"}
    except Exception as e:
        msg = str(e)[:300]
        status_hint = "401/403 kimlik hatası" if ("401" in msg or "403" in msg or "Unauthorized" in msg) else "HTTP hatası"
        return {"success": False, "message": f"{status_hint}: {msg}"}
@router.get("/trendyol/status")
async def get_trendyol_status():
    """Get Trendyol integration status"""
    config = await get_trendyol_config()
    return {
        "configured": config["is_active"],
        "mode": config["mode"],
        "supplier_id": config["supplier_id"] if config["is_active"] else None
    }
@router.get("/trendyol/debug")
async def debug_trendyol_orders():
    config = await get_trendyol_config()
    import sys
    import os
    import time
    from datetime import datetime, timezone
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )
    
    now = datetime.now()
    start = now - timedelta(days=14)
    end_date_ms = int(now.timestamp() * 1000)
    start_date_ms = int(start.timestamp() * 1000)
    
    try:
        resp = await client.get_orders(start_date_ms=start_date_ms, end_date_ms=end_date_ms, size=50)
        return resp
    except Exception as e:
        return {"error": str(e)}
@router.post("/trendyol/categories/sync")
async def sync_trendyol_categories(current_user: dict = Depends(require_admin)):
    """Sync and save category tree from Trendyol API to local DB"""
    config = await get_trendyol_config()
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    
    try:
        client = TrendyolClient(
            supplier_id=config["supplier_id"],
            api_key=config["api_key"],
            api_secret=config["api_secret"],
            mode=config["mode"]
        )
        categories = await client.get_categories()
        
        # Save to DB (Drop and re-insert for clean sync)
        if categories:
            await db.trendyol_categories.delete_many({})
            await db.trendyol_categories.insert_many(categories)
            
        return {"success": True, "message": f"{len(categories)} kategori senkronize edildi."}
    except Exception as e:
        logger.error(f"Error fetching trendyol categories: {str(e)}")
        raise HTTPException(status_code=500, detail="Trendyol kategorileri alınamadı")
@router.get("/trendyol/categories/{category_id}/attributes")
async def get_trendyol_category_attributes(category_id: int, refresh: bool = Query(False, description="True ise cache atlanır, Trendyol'dan taze ve eksiksiz çekilir"), current_user: dict = Depends(require_admin)):
    """Get attributes for a specific category (From DB or Trendyol API directly)"""
    config = await get_trendyol_config()
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    
    # Check if already in local DB (refresh=true ise cache atla — eksik/eski cache'i yenile)
    if not refresh:
        existing = await db.trendyol_attributes.find_one({"category_id": category_id})
        if existing and existing.get("attributes"):
            return {"success": True, "attributes": existing.get("attributes", []), "cached": True}
        
    try:
        client = TrendyolClient(
            supplier_id=config["supplier_id"],
            api_key=config["api_key"],
            api_secret=config["api_secret"],
            mode=config["mode"]
        )
        attributes = await client.get_category_attributes(category_id)
        
        # Save to local DB for future use.
        # T1 (TEK KAYNAK): editör "özellikleri yenile" dediğinde, GÖNDERİMİN (push)
        # okuduğu kanonik cache'i (`trendyol_category_attributes`) de aynı veriyle tazele
        # ki iki cache ayrışmasın ve push eksik şemayla göndermesin. Eklemeli + geri alınabilir.
        if attributes:
            from datetime import datetime, timezone
            _now = datetime.now(timezone.utc).isoformat()
            await db.trendyol_attributes.update_one(
                {"category_id": category_id},
                {"$set": {"category_id": category_id, "attributes": attributes, "updated_at": _now}},
                upsert=True
            )
            await db.trendyol_category_attributes.update_one(
                {"category_id": category_id},
                {"$set": {"category_id": category_id, "attributes": attributes, "updated_at": _now}},
                upsert=True
            )

        _val_count = sum(len(a.get("attributeValues") or []) for a in (attributes or []))
        return {"success": True, "attributes": attributes,
                "count": len(attributes or []), "value_count": _val_count}
    except Exception as e:
        logger.error(f"Error fetching trendyol attributes for category {category_id}: {str(e)}")
        raise HTTPException(status_code=500, detail="Özellikler (attributes) alınamadı")
@router.post("/trendyol/brands/sync")
async def sync_trendyol_brands(current_user: dict = Depends(require_admin)):
    """Sync brands from Trendyol API to local DB"""
    config = await get_trendyol_config()
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    
    try:
        client = TrendyolClient(
            supplier_id=config["supplier_id"],
            api_key=config["api_key"],
            api_secret=config["api_secret"],
            mode=config["mode"]
        )
        # Assuming maximum 500 per page, just fetching first page for demonstration. In prod, paginate.
        brands_data = await client.get_brands(size=5000)
        brands = brands_data.get("brands", [])
        
        if brands:
            await db.trendyol_brands.delete_many({})
            await db.trendyol_brands.insert_many(brands)
            
        return {"success": True, "message": f"{len(brands)} marka senkronize edildi."}
    except Exception as e:
        logger.error(f"Error fetching trendyol brands: {str(e)}")
        raise HTTPException(status_code=500, detail="Trendyol markaları alınamadı")
_TRENDYOL_ATTR_SYNONYMS = {
    "materyal bileşeni": ["urun icerik bilgisi", "kumas bilgisi", "kumas icerigi", "urun icerigi", "icerik bilgisi"],
}
def _bridge_trendyol_attr_synonyms(local_vals: dict) -> dict:
    """Trendyol özellik adlarıyla (ör. 'Materyal Bileşeni') lokal özellik adları
    (ör. 'Ürün İçerik Bilgisi') farklı olabilir. Eksik Trendyol anahtarını uygun
    lokal kaynaktan köprüler. local_vals: {lower(label): value}."""
    if not local_vals:
        return local_vals
    norm_index: dict = {}
    for k, v in local_vals.items():
        norm_index.setdefault(_normalize_attr_key(k), v)
    for target, sources in _TRENDYOL_ATTR_SYNONYMS.items():
        if target in local_vals and local_vals.get(target):
            continue
        for src in sources:
            val = norm_index.get(src)
            if val and str(val).strip().casefold() not in _BAD_COMPOSITION_VALUES:
                local_vals[target] = val
                break
    return local_vals
@router.post("/trendyol/products/validate")
async def validate_products_for_trendyol(
    request: Request,
    current_user: dict = Depends(require_admin),
):
    """
    Aktarım öncesi DOĞRULAMA paneli — ürün(ler)i Trendyol'a göndermeden,
    eksik zorunlu alanları (kategori mapping, barkod, görsel, zorunlu attribute)
    listeleyip raporlar. Body sync ile aynı.
    """
    payload = await request.json()
    query = await _build_product_query_from_payload(payload)
    products = await db.products.find(query, {"_id": 0}).to_list(length=None)
    # 🛡️ Aynı stock_code ile dublike doküman varsa, görseli/source'u en iyi olanı seç
    products = _dedupe_products_by_stock_code(products)

    # Category mappings (category_id -> mapping doc) — Trendyol için
    cm_list = await db.category_mappings.find(
        {"marketplace": "trendyol"}, {"_id": 0}
    ).to_list(length=3000)
    cm_by_local = {str(c.get("category_id")): c for c in cm_list}
    # Fallback: kategori adına göre
    all_cats = await db.categories.find({}, {"_id": 0}).to_list(length=5000)
    cat_by_id = {str(c.get("id")): c for c in all_cats}
    cat_by_name = {(c.get("name") or "").strip(): c for c in all_cats}
    cm_by_name = {}
    for c in all_cats:
        nm = (c.get("name") or "").strip()
        if nm and str(c.get("id")) in cm_by_local:
            cm_by_name[nm] = cm_by_local[str(c.get("id"))]

    # Trendyol mp_cat -> required attribute listesini cache'le
    attr_cache: dict = {}

    async def get_required_attrs(mp_cat_id):
        if not mp_cat_id:
            return []
        key = str(mp_cat_id)
        if key in attr_cache:
            return attr_cache[key]
        try:
            cached = await db.trendyol_category_attributes.find_one(
                {"category_id": int(mp_cat_id) if str(mp_cat_id).isdigit() else str(mp_cat_id)},
                {"_id": 0},
            )
            attrs = (cached or {}).get("attributes", []) or []
        except Exception:
            attrs = []
        required = [a for a in attrs if a.get("required")]
        attr_cache[key] = required
        return required

    results = []
    valid_count = 0
    invalid_count = 0

    for p in products:
        errors: list[str] = []
        warnings: list[str] = []
        missing_required_attrs: list[dict] = []
        unmatched_values: list[dict] = []

        cat_id = p.get("category_id")
        cat_name = p.get("category_name") or ""
        # Sync ile aynı sıra: category_id → category_name → categories.trendyol_category_id
        cm = cm_by_local.get(str(cat_id)) if cat_id else None
        if not cm and cat_name:
            cm = cm_by_name.get(cat_name.strip())
        cat_doc = cat_by_id.get(str(cat_id)) or cat_by_name.get(cat_name.strip())

        mp_cat_id = None
        if cm and cm.get("marketplace_category_id"):
            mp_cat_id = cm.get("marketplace_category_id")
        elif cat_doc and cat_doc.get("trendyol_category_id"):
            mp_cat_id = cat_doc.get("trendyol_category_id")

        if not mp_cat_id:
            errors.append("Trendyol kategori eşleştirmesi yok")

        # Görsel kontrolü
        if not (p.get("images") or []):
            errors.append("En az 1 ürün görseli yok")

        # Barkod kontrolü (varyantlı ürünlerde tüm varyantlar)
        # ⚠️ barcode_uncertain=True ise barkod yok kabul edilir (stock_code'tan kopyalanmış olabilir)
        variants = p.get("variants") or []
        if not variants:
            if not p.get("barcode") or p.get("barcode_uncertain"):
                errors.append("Barkod yok / belirsiz (Ticimax'tan doğrulayın)")
            if not p.get("stock_code"):
                warnings.append("Stok kodu yok")
        else:
            missing_v_barcode = sum(
                1 for v in variants if not v.get("barcode") or v.get("barcode_uncertain")
            )
            if missing_v_barcode:
                errors.append(f"{missing_v_barcode} varyantın barkodu eksik/belirsiz")

        # Fiyat
        try:
            if float(p.get("price", 0) or 0) <= 0:
                errors.append("Fiyat 0 veya boş")
        except Exception:
            errors.append("Fiyat geçersiz")

        # Stok
        total_stock = int(p.get("stock", 0) or 0) + sum(int(v.get("stock", 0) or 0) for v in variants)
        if total_stock <= 0:
            warnings.append("Toplam stok 0")

        # Açıklama
        if not (p.get("description") or p.get("short_description")):
            warnings.append("Açıklama boş")

        # Zorunlu attribute kontrolü (category_mappings → attribute_mappings + default_mappings)
        if mp_cat_id:
            req_attrs = await get_required_attrs(mp_cat_id)
            attr_mappings = (cm or {}).get("attribute_mappings", []) or []
            default_mappings = (cm or {}).get("default_mappings", {}) or {}
            # local_attr → mp_attr_id map'ını da kur
            mp_id_to_local = {}
            for am in attr_mappings:
                mid = str(am.get("mp_attr_id") or am.get("trendyol_attr_id") or "")
                if mid:
                    mp_id_to_local[mid] = am.get("local_attr")

            # Ürünün attribute'larından lokal isim → değer (DICT veya LIST formatını destekler)
            local_vals: dict = {}

            def _add_lv(nm, vv):
                if not nm or vv in (None, ""):
                    return
                local_vals.setdefault(str(nm).lower().strip(), str(vv))

            def _walk(attrs):
                if isinstance(attrs, dict):
                    items = sorted(
                        attrs.items(),
                        key=lambda kv: 1 if str(kv[0]).lower().startswith("ticimax_") else 0,
                    )
                    for k, v in items:
                        if isinstance(v, dict):
                            nm = v.get("label") or v.get("name") or k
                            vv = v.get("value") or v.get("attribute_value")
                            _add_lv(nm, vv)
                        elif v is not None:
                            _add_lv(k, v)
                elif isinstance(attrs, list):
                    for a in attrs:
                        if isinstance(a, dict):
                            nm = a.get("label") or a.get("name") or a.get("type") or a.get("attribute_name")
                            vv = a.get("value") or a.get("attribute_value")
                            _add_lv(nm, vv)

            _walk(p.get("attributes"))
            for v in variants:
                _walk(v.get("attributes"))
                if v.get("color"):
                    _add_lv("Renk", v["color"])
                    _add_lv("Web Color", v["color"])
                if v.get("size"):
                    _add_lv("Beden", v["size"])
            _bridge_trendyol_attr_synonyms(local_vals)

            for ra in req_attrs:
                ra_id = str(ra.get("id") or ra.get("attribute", {}).get("id") or "")
                ra_name = ra.get("name") or ra.get("attribute", {}).get("name") or "(?)"
                if not ra_id:
                    continue
                # default mapping var mı?
                default_val = default_mappings.get(ra_id) or default_mappings.get(str(ra_id))
                if default_val:
                    continue
                # local attribute mapping var mı + üründe değer var mı?
                local_attr = mp_id_to_local.get(ra_id)
                has_val = False
                if local_attr:
                    has_val = bool(local_vals.get(local_attr.lower()))
                if not has_val:
                    missing_required_attrs.append({
                        "id": ra_id,
                        "name": ra_name,
                        "mapped_local": local_attr,
                    })
            if missing_required_attrs:
                errors.append(f"{len(missing_required_attrs)} zorunlu özellik eksik")

            # 🔎 Listeli (enum) değerlerin Trendyol karşılığı var mı? Yoksa aktarım engellenir
            # (allowCustom alanlar serbest metindir, her zaman gönderilir → kontrol edilmez).
            val_mappings_v = (cm or {}).get("value_mappings", {}) or {}
            try:
                _cache_attrs = await db.trendyol_category_attributes.find_one(
                    {"category_id": int(mp_cat_id) if str(mp_cat_id).isdigit() else str(mp_cat_id)},
                    {"_id": 0},
                )
            except Exception:
                _cache_attrs = None
            local_norm_index_v = {}
            for lk, lv in local_vals.items():
                local_norm_index_v.setdefault(_normalize_attr_key(lk), lv)
            for a in (_cache_attrs or {}).get("attributes", []) or []:
                aid = a.get("id") or a.get("attribute", {}).get("id")
                if aid is None:
                    continue
                if bool(a.get("allowCustom") or a.get("attribute", {}).get("allowCustom")):
                    continue
                aname = a.get("name") or a.get("attribute", {}).get("name") or ""
                lval = local_norm_index_v.get(_normalize_attr_key(aname))
                if not lval:
                    la = mp_id_to_local.get(str(aid))
                    if la:
                        lval = local_vals.get(la.lower())
                if not lval:
                    continue
                if val_mappings_v.get(f"{aid}|{lval}"):
                    continue
                vname_map = {
                    _norm_val(v.get("name")): str(v.get("id"))
                    for v in (a.get("attributeValues") or [])
                    if v.get("id") is not None and v.get("name")
                }
                if _resolve_value_id(vname_map, lval):
                    continue
                unmatched_values.append({
                    "mp_attr_id": int(aid),
                    "attr_name": aname,
                    "local_value": lval,
                    "required": bool(a.get("required")),
                })
            if unmatched_values:
                errors.append(f"{len(unmatched_values)} değerin Trendyol karşılığı yok (eşleştirme gerekli)")

        is_valid = len(errors) == 0
        if is_valid:
            valid_count += 1
        else:
            invalid_count += 1

        results.append({
            "id": p.get("id"),
            "name": p.get("name"),
            "stock_code": _resolve_stock_code(p) or p.get("barcode") or "",
            "barcode": p.get("barcode"),
            "category_name": cat_name,
            "marketplace_category_id": mp_cat_id,
            "is_valid": is_valid,
            "errors": errors,
            "warnings": warnings,
            "missing_required_attrs": missing_required_attrs,
            "unmatched_values": unmatched_values,
        })

    # Eksik attribute istatistikleri (en sık eksik olanları üstte göster)
    attr_freq: dict = {}
    for r in results:
        for m in r["missing_required_attrs"]:
            k = m["name"]
            attr_freq[k] = attr_freq.get(k, 0) + 1
    top_missing = sorted(attr_freq.items(), key=lambda x: -x[1])[:10]

    return {
        "success": True,
        "total": len(products),
        "valid_count": valid_count,
        "invalid_count": invalid_count,
        "results": results,
        "top_missing_attrs": [{"name": k, "count": v} for k, v in top_missing],
    }
@router.get("/trendyol/batch/{batch_id}")
async def get_trendyol_batch_status(
    batch_id: str,
    current_user: dict = Depends(require_admin),
):
    """Trendyol batch işleminin (ürün oluşturma vb.) gerçek durumunu döndürür.
    UI'da kullanıcı 'Detayları Gör' butonuna basınca her item'ın SUCCESS/FAILED
    durumu ve failureReasons listesini görür."""
    config = await get_trendyol_config()
    if not config or not config.get("is_active"):
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu aktif değil")
    from trendyol_client import TrendyolClient
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config.get("mode", "live"),
    )
    try:
        data = await client.get_batch_request_result(batch_id)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Trendyol batch detayı alınamadı: {e}")

    items = data.get("items") or []
    success_count = sum(1 for it in items if str(it.get("status")).upper() == "SUCCESS")
    failed_count = sum(1 for it in items if str(it.get("status")).upper() == "FAILED")
    # Hata özetlerini topla
    fail_freq: dict = {}
    for it in items:
        for fr in (it.get("failureReasons") or []):
            key = str(fr).split(".")[0][:120]
            fail_freq[key] = fail_freq.get(key, 0) + 1
    return {
        "batch_id": batch_id,
        "status": data.get("status"),
        "source_type": data.get("sourceType"),
        "item_count": data.get("itemCount") or len(items),
        "success_count": success_count,
        "failed_count": failed_count,
        "top_failures": [{"reason": k, "count": v} for k, v in
                         sorted(fail_freq.items(), key=lambda x: -x[1])[:10]],
        "items": items,
        "raw": data,
    }
@router.get("/trendyol/barcode-duplicates")
async def trendyol_barcode_duplicates(current_user: dict = Depends(require_admin)):
    """DB içinde aynı barkoda atanmış birden fazla varyantı tespit eder.
    Bu Trendyol'a aktarımı bloklayan ana sebep oluyor. Manuel düzeltme için liste döner.
    """
    pipeline = [
        {"$match": {"variants.barcode": {"$nin": [None, ""]}}},
        {"$unwind": "$variants"},
        {"$match": {"variants.barcode": {"$nin": [None, ""]}}},
        {"$group": {
            "_id": "$variants.barcode",
            "count": {"$sum": 1},
            "products": {"$push": {
                "product_id": "$id",
                "stock_code": "$stock_code",
                "name": "$name",
                "is_active": "$is_active",
                "variant_size": "$variants.size",
                "variant_color": "$variants.color",
                "variant_stock": "$variants.stock",
            }}
        }},
        {"$match": {"count": {"$gt": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 500},
    ]
    rows = await db.products.aggregate(pipeline).to_list(length=500)
    out = []
    for r in rows:
        out.append({
            "barcode": r["_id"],
            "count": r["count"],
            "assignments": r["products"],
        })
    return {"total": len(out), "duplicates": out}
@router.post("/trendyol/ghost-scanner")
async def trendyol_ghost_scanner(
    payload: dict = Body(default={}),
    current_user: dict = Depends(require_admin),
):
    """Trendyol panelindeki tüm ürünleri (max 5000) sayfalı çekip,
    DB'de KARŞILIK BULAMAYAN ya da DB'de archive edilmiş olanları "hayalet" olarak listeler.
    Bu hayaletler genelde eski yanlış barkod kayıtlarıdır ve duplicate çakışmalara sebep olur.
    """
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    import sys
    import os as _os
    sys.path.insert(0, _os.path.dirname(_os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient

    cli = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    only_unmatched = bool(payload.get("only_unmatched", True))
    include_archived = bool(payload.get("include_archived", False))
    page_limit = int(payload.get("page_limit", 50))  # 50 sayfa x 200 = 10K ürün max

    # DB barkodlarını VE stock_code'larını topla
    db_barcodes = set()
    db_stock_codes = set()
    async for p in db.products.find({}, {"_id": 0, "barcode": 1, "stock_code": 1, "variants.barcode": 1, "variants.stock_code": 1}):
        if p.get("barcode"):
            db_barcodes.add(str(p["barcode"]))
        if p.get("stock_code"):
            db_stock_codes.add(str(p["stock_code"]))
        for v in (p.get("variants") or []):
            if v.get("barcode"):
                db_barcodes.add(str(v["barcode"]))
            if v.get("stock_code"):
                db_stock_codes.add(str(v["stock_code"]))

    ghosts = []
    matched = 0
    total_scanned = 0
    page = 0
    while page < page_limit:
        try:
            res = await cli.get_filtered_products(
                page=page, size=200,
                archived=None if include_archived else False,
            )
        except Exception as e:
            return {"error": str(e), "scanned": total_scanned, "ghosts": ghosts}
        content = res.get("content") or []
        total_pages = res.get("totalPages") or 0
        if not content:
            break
        for row in content:
            total_scanned += 1
            bc = str(row.get("barcode") or "")
            sc = str(row.get("stockCode") or "")
            pmi = str(row.get("productMainId") or "")
            if not bc:
                continue
            # Match: barkod DB'de varsa VEYA stockCode/productMainId DB'de varsa "matched"
            if bc in db_barcodes or sc in db_barcodes or sc in db_stock_codes or pmi in db_stock_codes:
                matched += 1
                if only_unmatched:
                    continue
            else:
                ghosts.append({
                    "barcode": bc,
                    "stockCode": row.get("stockCode"),
                    "productMainId": row.get("productMainId"),
                    "title": row.get("title"),
                    "brand": row.get("brand"),
                    "approved": row.get("approved"),
                    "archived": row.get("archived"),
                    "onSale": row.get("onSale"),
                    "salePrice": row.get("salePrice"),
                    "quantity": row.get("quantity"),
                    "rejectReasonDetails": row.get("rejectReasonDetails"),
                })
        page += 1
        if page >= total_pages:
            break

    return {
        "scanned": total_scanned,
        "matched_in_db": matched,
        "ghosts_count": len(ghosts),
        "ghosts": ghosts,
    }
@router.post("/trendyol/archive-barcodes")
async def trendyol_archive_barcodes(
    payload: dict = Body(...),
    current_user: dict = Depends(require_admin),
):
    """Verilen barkodları Trendyol'da arşivler (panelden gizler, slot iadesi sağlar).
    Hayalet ürünlerin temizliği için kullanılır.
    """
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    barcodes = [str(b).strip() for b in (payload.get("barcodes") or []) if str(b).strip()]
    if not barcodes:
        raise HTTPException(status_code=400, detail="Arşivlenecek barkod listesi boş.")
    if len(barcodes) > 1000:
        raise HTTPException(status_code=400, detail="Tek seferde max 1000 barkod arşivlenebilir.")

    import sys
    import os as _os
    sys.path.insert(0, _os.path.dirname(_os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient

    cli = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    try:
        resp = await cli.archive_products(barcodes)
        batch_id = (resp or {}).get("batchRequestId")
        from datetime import datetime, timezone
        log_doc = {
            "id": generate_id(),
            "started_at": datetime.now(timezone.utc).isoformat(),
            "status": "archive",
            "products_attempted": len(barcodes),
            "batch_request_id": batch_id,
            "archived_barcodes": barcodes,
            "trendyol_response": resp,
            "message": f"{len(barcodes)} barkod için arşivleme batch'i Trendyol'a gönderildi (batch: {batch_id}).",
        }
        await db.trendyol_sync_logs.insert_one(log_doc)
        log_doc.pop("_id", None)
        return {
            "success": bool(batch_id),
            "batchRequestId": batch_id,
            "barcodes_count": len(barcodes),
            "response": resp,
        }
    except Exception as e:
        logger.error(f"Archive barcodes error: {e}")
        raise HTTPException(status_code=500, detail=f"Trendyol arşiv hatası: {str(e)}")
@router.post("/trendyol/products/sync")
async def sync_products_to_trendyol(
    request: Request,
    current_user: dict = Depends(require_admin)
):
    """Sync products to Trendyol via Batch Request"""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")
    
    payload = await request.json()
    product_ids = payload.get("product_ids", [])
    category_filters = payload.get("category_filters", [])
    # Yeni: Barkod/stok kodu ile filtre (kullanıcı UI'da yazıp aktarabilsin)
    barcodes_raw = payload.get("barcodes", [])
    stock_codes_raw = payload.get("stock_codes", [])
    card_ids_raw = payload.get("card_ids", [])
    barcodes = [str(b).strip() for b in (barcodes_raw or []) if str(b).strip()]
    stock_codes = [str(s).strip() for s in (stock_codes_raw or []) if str(s).strip()]
    card_ids = [str(c).strip() for c in (card_ids_raw or []) if str(c).strip()]
    # Yeni: Tarih aralığı (created_at — ürün eklenme tarihi)
    date_from = payload.get("date_from")  # ISO format "2026-01-01"
    date_to = payload.get("date_to")
    
    query = {}
    if product_ids:
        query = {"id": {"$in": product_ids}}
    elif barcodes or stock_codes or card_ids:
        # Hem ürünün kendi barcode/stock_code'una hem de variants[] içine bak
        or_conditions = []
        if barcodes:
            or_conditions.append({"barcode": {"$in": barcodes}})
            or_conditions.append({"variants.barcode": {"$in": barcodes}})
        if stock_codes:
            or_conditions.append({"stock_code": {"$in": stock_codes}})
            or_conditions.append({"sku": {"$in": stock_codes}})
            or_conditions.append({"variants.stock_code": {"$in": stock_codes}})
        if card_ids:
            # Ürün Kart ID ile aktarım — urun_karti_id (asıl) + csv_card_id (yedek)
            or_conditions.append({"urun_karti_id": {"$in": card_ids}})
            or_conditions.append({"csv_card_id": {"$in": card_ids}})
        query = {"$or": or_conditions}
    elif category_filters:
        # Build an $or query for each category + its filters
        or_conditions = []
        for cf in category_filters:
            cat_id = cf.get("category_id")
            filters = cf.get("filters", {})
            try:
                from bson.objectid import ObjectId
                cat = await db.categories.find_one({"_id": ObjectId(cat_id)})
            except Exception:
                cat = await db.categories.find_one({"id": cat_id})

            if not cat:
                continue
            
            cat_q = {"category_name": cat.get("name")}
            if filters.get("stock_code"):
                cat_q["stock_code"] = {"$regex": filters["stock_code"], "$options": "i"}
            if filters.get("date_range"):
                try:
                    from datetime import datetime, timezone
                    # Format: YYYY-MM-DD
                    date_obj = datetime.strptime(filters["date_range"], "%Y-%m-%d").replace(tzinfo=timezone.utc)
                    cat_q["created_at"] = {"$gte": date_obj.isoformat()}
                except Exception:
                    pass
            or_conditions.append(cat_q)
            
        if or_conditions:
            query = {"$or": or_conditions}
        else:
            query = {"is_active": True}
    else:
        # Sync only active products if no specific IDs are provided
        query = {"is_active": True}

    # Tarih aralığı filtresi (created_at) — diğer filtrelerle AND ile birleşir
    if date_from or date_to:
        date_q: dict = {}
        if date_from:
            date_q["$gte"] = date_from
        if date_to:
            date_q["$lte"] = date_to + "T23:59:59" if "T" not in str(date_to) else date_to
        query = {"$and": [query, {"created_at": date_q}]} if query else {"created_at": date_q}

    products = await db.products.find(query).to_list(length=None)
    # 🛡️ Aynı stock_code ile dublike doküman varsa, görseli/source'u en iyi olanı seç
    products = _dedupe_products_by_stock_code(products)

    # 🔍 Hangi stok kodları / barkodlar DB'de bulundu? Bulunmayanları topla.
    not_found_codes = []
    if barcodes or stock_codes:
        found_barcodes = set()
        found_stock_codes = set()
        for p in products:
            if p.get("barcode"):
                found_barcodes.add(str(p["barcode"]))
            if p.get("stock_code"):
                found_stock_codes.add(str(p["stock_code"]))
            if p.get("sku"):
                found_stock_codes.add(str(p["sku"]))
            for v in (p.get("variants") or []):
                if v.get("barcode"):
                    found_barcodes.add(str(v["barcode"]))
                if v.get("stock_code"):
                    found_stock_codes.add(str(v["stock_code"]))
        requested = set([str(c) for c in (barcodes or [])] + [str(c) for c in (stock_codes or [])])
        not_found_codes = sorted([c for c in requested if c not in found_barcodes and c not in found_stock_codes])

    # ❗ Hiç ürün bulunamadıysa anında net mesajla dön — kullanıcı "DB'de yok mu, validasyon mu hatalı?" diye anlamayabiliyor
    if not products:
        from datetime import datetime, timezone
        msg = (
            f"DB'de bu kod(lar)la ürün bulunamadı: {', '.join(not_found_codes[:10])}"
            + (f" (+{len(not_found_codes)-10} daha)" if len(not_found_codes) > 10 else "")
        ) if not_found_codes else "Sorgu kriterlerine uyan ürün bulunamadı."
        await db.trendyol_sync_logs.insert_one({
            "id": generate_id(),
            "started_at": datetime.now(timezone.utc).isoformat(),
            "status": "failed",
            "products_attempted": 0,
            "products_sent": 0,
            "batch_request_id": None,
            "errors": [msg],
            "not_found_codes": not_found_codes,
            "message": msg,
        })
        return {
            "success": False,
            "message": msg,
            "total": 0,
            "successful": 0,
            "failed": 0,
            "not_found_codes": not_found_codes,
            "errors": [msg],
        }
    
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )
    
    items_to_send = []
    errors = []

    # Trendyol kategori özellik cache'lerini bu sync için yükle (attribute meta + geçerli value_id'ler)
    _attr_meta_cache: dict = {}

    async def _get_attr_meta(mp_cat_id):
        if mp_cat_id in _attr_meta_cache:
            return _attr_meta_cache[mp_cat_id]
        meta: dict = {}
        try:
            cache = await db.trendyol_category_attributes.find_one(
                {"category_id": int(mp_cat_id) if str(mp_cat_id).isdigit() else str(mp_cat_id)},
                {"_id": 0},
            )
        except Exception:
            cache = None
        # T1 EMNİYET: kanonik cache boşsa (ör. kategori yalnız ürün editöründen yenilenmiş
        # ya da kategori-eşleme tazelemesi hata vermişse), editörün yazdığı
        # `trendyol_attributes` koleksiyonuna düş. Böylece push hiçbir zaman boş/eksik
        # şemayla kalmaz; aynı ham `categoryAttributes` listesi okunur (şekil birebir uyumlu).
        if not (cache or {}).get("attributes"):
            try:
                _legacy = await db.trendyol_attributes.find_one(
                    {"category_id": int(mp_cat_id) if str(mp_cat_id).isdigit() else str(mp_cat_id)},
                    {"_id": 0},
                )
                if _legacy and _legacy.get("attributes"):
                    cache = _legacy
            except Exception:
                pass
        for a in (cache or {}).get("attributes", []) or []:
            aid = a.get("id") or a.get("attribute", {}).get("id")
            if aid is None:
                continue
            valid_value_ids = {str(v.get("id")) for v in (a.get("attributeValues") or []) if v.get("id") is not None}
            # Normalize edilmiş "değer adı → value_id" haritası (isimle otomatik eşleştirme için)
            value_name_to_id = {}
            for v in (a.get("attributeValues") or []):
                if v.get("id") is None or not v.get("name"):
                    continue
                value_name_to_id[_norm_val(v["name"])] = str(v["id"])
            meta[int(aid)] = {
                "allow_custom": bool(a.get("allowCustom") or a.get("attribute", {}).get("allowCustom")),
                "required": bool(a.get("required")),
                "valid_value_ids": valid_value_ids,
                "value_name_to_id": value_name_to_id,
                "name": a.get("name") or a.get("attribute", {}).get("name") or "",
            }
        _attr_meta_cache[mp_cat_id] = meta
        return meta

    def _collect_local_values(product, variant):
        """Ürünün ve varyantın attribute değerlerini lokalName(lower) → value şeklinde toplar.
        attributes hem LIST hem DICT formatını destekler."""
        out = {}
        def _put(nm, vv):
            if not nm or vv in (None, ""):
                return
            out.setdefault(str(nm).lower().strip(), str(vv))

        def _walk(attrs):
            if isinstance(attrs, dict):
                # Çakışan mükerrer özelliklerde TEMİZ (güncel, elle/teknik) anahtarlar
                # önce işlenir; `ticimax_*` (eski/ham 113-kolon) anahtarları yalnızca
                # fallback olur. _put setdefault olduğundan ilk gelen kazanır.
                items = sorted(
                    attrs.items(),
                    key=lambda kv: 1 if str(kv[0]).lower().startswith("ticimax_") else 0,
                )
                for k, v in items:
                    if isinstance(v, dict):
                        nm = v.get("label") or v.get("name") or k
                        vv = v.get("value") or v.get("attribute_value")
                        _put(nm, vv)
                    elif v is not None:
                        _put(k, v)
            elif isinstance(attrs, list):
                for a in attrs:
                    if isinstance(a, dict):
                        nm = a.get("label") or a.get("name") or a.get("type") or a.get("attribute_name")
                        vv = a.get("value") or a.get("attribute_value")
                        _put(nm, vv)
        _walk(product.get("attributes"))
        _walk((product.get("trendyol_attributes_labels") or {}))  # opsiyonel
        if variant:
            _walk(variant.get("attributes"))
            if variant.get("color"):
                _put("Renk", variant["color"])
                _put("Web Color", variant["color"])
            if variant.get("size"):
                _put("Beden", variant["size"])
        _bridge_trendyol_attr_synonyms(out)
        return out

    def resolve_attributes(base_attrs, product, variant, category, meta):
        """meta = _get_attr_meta sonucu: {attr_id: {allow_custom, required, valid_value_ids, name}}
        Geçersiz value_id ya da allow_custom=False olan custom değerleri SESSİZCE atlar."""
        item_attrs = list(base_attrs)
        processed = {int(a["attributeId"]) for a in item_attrs if "attributeId" in a}
        
        attr_mappings = category.get("attribute_mappings", []) or []
        val_mappings = category.get("value_mappings", {}) or {}
        default_mappings = category.get("default_mappings", {}) or {}

        local_vals = _collect_local_values(product, variant)

        def _push(ty_id: int, value_id=None, custom=None):
            """Cache'e karşı doğrulayarak append."""
            am = meta.get(ty_id) or {}
            # Dosya linki / sertifika gerektiren attribute'ları skip (custom text kabul etmez)
            am_name = (am.get("name") or "").lower()
            if any(p in am_name for p in ["analiz testi", "test raporu", "sertifika dosya", "dosya linki"]):
                return False
            if value_id is not None:
                vid = str(value_id)
                # Cache'de yoksa custom'a düşür
                if am.get("valid_value_ids") and vid not in am["valid_value_ids"]:
                    if am.get("allow_custom") and custom:
                        item_attrs.append({"attributeId": ty_id, "customAttributeValue": str(custom)})
                        processed.add(ty_id)
                        return True
                    return False
                item_attrs.append({"attributeId": ty_id, "attributeValueId": int(vid)})
                processed.add(ty_id)
                return True
            if custom is not None:
                if not am.get("allow_custom"):
                    return False
                item_attrs.append({"attributeId": ty_id, "customAttributeValue": str(custom)})
                processed.add(ty_id)
                return True
            return False
        
        for mapping in attr_mappings:
            # Yeni format: mp_attr_id, eski format: trendyol_attr_id
            ty_id = mapping.get("mp_attr_id") or mapping.get("trendyol_attr_id")
            if not ty_id:
                continue
            try:
                ty_id = int(ty_id)
            except (ValueError, TypeError):
                continue
            if ty_id in processed:
                continue
                
            local_attr_name = str(mapping.get("local_attr") or "").strip()
            local_key = local_attr_name.lower()
            local_val = None

            if variant:
                if local_key in ["renk", "color", "web color"]:
                    local_val = variant.get("color")
                elif local_key in ["beden", "size"]:
                    local_val = variant.get("size") or local_vals.get(local_key)
                    
            if not local_val:
                local_val = local_vals.get(local_key)
            if not local_val and local_attr_name:
                local_val = local_vals.get(local_attr_name.lower())
                        
            if local_val:
                str_ty_id = str(ty_id)
                # value_mapping format: "343|Erkek" → "value_id"
                mapped_val = val_mappings.get(f"{str_ty_id}|{local_val}")
                # eski yapı: {"343": {"Erkek": "value_id"}}
                if not mapped_val and str_ty_id in val_mappings and isinstance(val_mappings[str_ty_id], dict):
                    mapped_val = val_mappings[str_ty_id].get(str(local_val))
                if mapped_val:
                    if str(mapped_val).isdigit():
                        if _push(ty_id, value_id=mapped_val, custom=local_val):
                            continue
                    else:
                        if _push(ty_id, custom=mapped_val):
                            continue
                # Kaydedilmiş eşleştirme yok → listeli (enum) değeri ADIYLA otomatik eşle.
                # Örn. local "Dokuma"/"Örme"/"Kısa" → Trendyol value_id (Türkçe/boşluk duyarsız).
                am_meta = meta.get(ty_id) or {}
                name_map = am_meta.get("value_name_to_id") or {}
                if name_map:
                    auto_vid = _resolve_value_id(name_map, local_val)
                    if auto_vid and _push(ty_id, value_id=auto_vid, custom=local_val):
                        continue
                # Mapping yok ama allow_custom varsa local_val'i custom olarak yolla
                if _push(ty_id, custom=local_val):
                    continue
                
            # Default mapping
            str_ty_id = str(ty_id)
            if str_ty_id in default_mappings and default_mappings[str_ty_id]:
                def_val = default_mappings[str_ty_id]
                if str(def_val).isdigit():
                    _push(ty_id, value_id=def_val, custom=local_val)
                else:
                    _push(ty_id, custom=def_val)

        # Default mapping'de olup attribute_mappings'de olmayanları da ekle
        for ty_str, def_val in default_mappings.items():
            if not def_val:
                continue
            try:
                ty_id = int(ty_str)
            except (ValueError, TypeError):
                continue
            if ty_id not in processed:
                if str(def_val).isdigit():
                    _push(ty_id, value_id=def_val)
                else:
                    _push(ty_id, custom=def_val)

        # 🎯 GARANTİ: Trendyol "Materyal Bileşeni" (serbest metin / allowCustom) alanı,
        # mapping yapılmamış olsa bile HER kategoride ürünün "Ürün İçerik Bilgisi"
        # değerinden otomatik gönderilir. local_vals["materyal bileşeni"] köprü ile dolar.
        icerik_val = local_vals.get("materyal bileşeni")
        if icerik_val:
            for m_ty_id, m_meta in meta.items():
                if m_ty_id in processed:
                    continue
                mname = (m_meta.get("name") or "").lower()
                if "materyal bileşeni" in mname and m_meta.get("allow_custom"):
                    _push(m_ty_id, custom=icerik_val)

        # 🎯 GENEL GARANTİ: Trendyol özellik adı ile lokal özellik adı eşleşen ama
        # attribute_mappings'te tanımlanmamış alanları (Boy, Desen, Kumaş Tipi vb.)
        # otomatik gönder. Listeli alanlar değeri ADIYLA value_id'ye eşlenir,
        # serbest alanlar custom olarak yollanır.
        local_norm_index = {}
        for lk, lv in local_vals.items():
            local_norm_index.setdefault(_normalize_attr_key(lk), lv)
        for m_ty_id, m_meta in meta.items():
            if m_ty_id in processed:
                continue
            lval = local_norm_index.get(_normalize_attr_key(m_meta.get("name") or ""))
            if not lval:
                continue
            name_map = m_meta.get("value_name_to_id") or {}
            auto_vid = _resolve_value_id(name_map, lval)
            if auto_vid:
                _push(m_ty_id, value_id=auto_vid, custom=lval)
            else:
                _push(m_ty_id, custom=lval)

        # 🎯 FACETTE SABİT VARSAYILANLAR (gap-fill, EN SON):
        # Menşei=Türkiye · Cinsiyet=Kadın · Yaş Grubu=Yetişkin · Ortam/Koleksiyon=Casual/Günlük ·
        # Ek Özellik=Mevcut Değil · Kutu Durumu=Kutu Yok · Persona=Fashion Forward ·
        # Performans=Cool & Comfort · Üretici/İthalatçı (GPSR)=FACETTE bilgileri.
        # Yalnız ürünün DOLDURMADIĞI (processed olmayan) Trendyol özelliğine yazılır; listeli
        # alanlar ADIYLA value_id'ye çözülür (Türkiye→TR), serbest alanlar custom gider.
        for m_ty_id, m_meta in meta.items():
            if m_ty_id in processed:
                continue
            fv = facette_fixed_value_for(m_meta.get("name") or "")
            if not fv:
                continue
            name_map = m_meta.get("value_name_to_id") or {}
            auto_vid = _resolve_value_id(name_map, fv) if name_map else None
            if auto_vid:
                _push(m_ty_id, value_id=auto_vid, custom=fv)
            else:
                _push(m_ty_id, custom=fv)

        return item_attrs
    
    for product in products:
        try:
            # 1. Category Mapping check — önce category_id, sonra category_name üzerinden category_mappings
            trendyol_cat_id = None
            category = None
            cat_id = product.get("category_id")
            cat_name = (product.get("category_name") or "").strip()
            cm = None

            # 1a. Doğrudan category_id ile mapping arayalım
            if cat_id:
                cm = await db.category_mappings.find_one(
                    {"category_id": str(cat_id), "marketplace": "trendyol"}, {"_id": 0}
                )

            # 1b. Bulunamadıysa: kategori adından sistem kategorisini, oradan mapping'i bul
            if not cm and cat_name:
                sys_cat = await db.categories.find_one({"name": cat_name}, {"_id": 0, "id": 1, "trendyol_category_id": 1})
                if sys_cat and sys_cat.get("id"):
                    cm = await db.category_mappings.find_one(
                        {"category_id": str(sys_cat["id"]), "marketplace": "trendyol"}, {"_id": 0}
                    )

            if cm and cm.get("marketplace_category_id"):
                trendyol_cat_id = cm["marketplace_category_id"]
                category = cm  # category_mappings doc — attribute_mappings, default_mappings içerir

            # 1c. Eski legacy fallback: db.categories.trendyol_category_id
            if not trendyol_cat_id and cat_name:
                cat_doc = await db.categories.find_one({"name": cat_name})
                if cat_doc and cat_doc.get("trendyol_category_id"):
                    trendyol_cat_id = cat_doc["trendyol_category_id"]
                    category = cat_doc

            if not trendyol_cat_id:
                errors.append(f"{product.get('name')} - Trendyol kategori eşleştirmesi (Mapping) yok.")
                continue
                
            # 2. Attributes check and formatting
            raw_attrs = product.get("trendyol_attributes", {})
            attributes = []
            for attr_id, val_id in raw_attrs.items():
                if val_id:
                    # Trendyol expects attributeId and attributeValueId, or customAttributeValue
                    if str(val_id).isdigit():
                        attributes.append({
                            "attributeId": int(attr_id),
                            "attributeValueId": int(val_id)
                        })
                    else:
                        attributes.append({
                            "attributeId": int(attr_id),
                            "customAttributeValue": str(val_id)
                        })
            
            # Description (HTML → düz metin: br/li/p satır sonuna, entity decode, min 30 karakter)
            import re as _re
            import html as _html
            raw_desc = (product.get("description") or product.get("short_description")
                        or product.get("long_description") or "").strip()
            # 1) Blok/satır kıran tagları newline'a çevir (görsel paragraf yapısı korunur).
            _desc = raw_desc
            _desc = _re.sub(r"<\s*br\s*/?\s*>", "\n", _desc, flags=_re.IGNORECASE)
            _desc = _re.sub(r"</\s*(p|div|li|h[1-6]|tr)\s*>", "\n", _desc, flags=_re.IGNORECASE)
            _desc = _re.sub(r"<\s*li\b[^>]*>", "• ", _desc, flags=_re.IGNORECASE)
            # 2) Tüm kalan HTML tag'lerini at.
            _desc = _re.sub(r"<[^>]+>", " ", _desc)
            # 3) HTML entity'leri decode et (&nbsp;, &amp;, &uuml; vb.).
            _desc = _html.unescape(_desc)
            # 4) Yatay boşlukları sıkıştır ama satır sonlarını koru.
            _desc = _re.sub(r"[ \t]+", " ", _desc)
            _desc = _re.sub(r"\n[ \t]+", "\n", _desc)
            _desc = _re.sub(r"[ \t]+\n", "\n", _desc)
            _desc = _re.sub(r"\n{3,}", "\n\n", _desc)
            clean_desc = _desc.strip()
            if not clean_desc or len(clean_desc) < 30:
                # Description yoksa veya çok kısaysa ürün adından bir minimum açıklama üret
                fallback = (product.get("name") or "").strip()
                if fallback and len(fallback) >= 10:
                    clean_desc = f"{fallback}. Kaliteli kumaş, modern kesim, şık tasarım. Günlük ve özel kullanım için ideal."
                else:
                    errors.append(f"{product.get('name')} - Açıklama eksik (Trendyol min 30 karakter zorunlu).")
                    continue

            # 3. Base Product Details
            base_item = {
                "title": product.get("name"),
                "productMainId": _resolve_stock_code(product) or product.get("id"),
                "brandId": int(product.get("trendyol_brand_id") or config.get("default_brand_id") or 975755),
                "categoryId": int(trendyol_cat_id),
                "description": clean_desc,
                "currencyType": product.get("currency", "TRY"),
                "listPrice": calculate_trendyol_price(_mp_base_price(product), product, config),
                "salePrice": calculate_trendyol_price(_mp_base_price(product), product, config),
                "vatRate": int(product.get("vat_rate", config.get("default_vat_rate") or 20)),
                "cargoCompanyId": int(config.get("default_cargo_company_id") or 10), # Assuming 10 is MNG Kargo (Needs specific Cargo Provider ID)
                "dimensionalWeight": float(product.get("cargo_weight", 1)),
                "images": [{"url": img} for img in product.get("images", [])[:8]]
            }
            
            if not base_item["images"]:
                errors.append(f"{product.get('name')} - En az 1 görsel gerekli.")
                continue

            # 3b. Trendyol attribute meta (cache) — value_id ve allowCustom validasyonu için
            meta = await _get_attr_meta(trendyol_cat_id)
            
            # 4. Handle Variants or No-Variants
            variants = product.get("variants", [])
            if not variants:
                if not product.get("barcode") or product.get("barcode_uncertain"):
                    errors.append(f"{product.get('name')} - Barkod yok ya da belirsiz (Ticimax'tan doğrulayın).")
                    continue
                item = base_item.copy()
                item["barcode"] = product.get("barcode")
                item["stockCode"] = product.get("stock_code") or product.get("sku") or product.get("barcode")
                item["quantity"] = int(product.get("stock", 0))
                item["attributes"] = resolve_attributes(attributes, product, None, category, meta)
                items_to_send.append(item)
            else:
                # 🎯 Eğer kullanıcı barcodes parametresi ile spesifik barkodlar istediyse,
                # sadece o barkodları batch'e ekle. Aksi halde parent'ın tüm varyantları
                # gönderilir → Trendyol'da kardeş varyantlar zaten varsa duplicate hatası
                # oluşur ve yeni eklenmek istenen barkod da reject olur.
                # AMA: Kullanıcı stok kodu girip parent ürünü hedeflediyse, tüm varyantlar
                # gönderilmelidir (stockCode == "FCSSxxx" hiçbir varyant barkoduyla
                # eşleşmediği için aksi halde sıfır varyant gönderilir).
                product_match_by_stock = False
                if stock_codes:
                    sc_set = set(str(s) for s in stock_codes)
                    if (
                        str(product.get("stock_code") or "") in sc_set
                        or str(product.get("sku") or "") in sc_set
                        or any(str(v.get("stock_code") or "") in sc_set for v in variants)
                    ):
                        product_match_by_stock = True
                if product_match_by_stock:
                    requested_set = None  # tüm varyantları gönder
                else:
                    requested_set = set(barcodes) if barcodes else None
                for v in variants:
                    if not v.get("barcode") or v.get("barcode_uncertain"):
                        errors.append(f"{product.get('name')} - Varyant ({v.get('size') or v.get('color') or '?'}) barkodu yok / belirsiz.")
                        continue
                    if requested_set and str(v.get("barcode")) not in requested_set:
                        continue  # kullanıcı bu barkodu istemedi, atla
                    item = base_item.copy()
                    item["barcode"] = v.get("barcode")
                    # stockCode = parent stok kodu (FCSS.../FCFW...). Trendyol unique
                    # check'i barcode üzerinden yapar, stockCode aynı olabilir.
                    parent_stock = (product.get("stock_code") or product.get("sku") or "").strip()
                    v_stock = (v.get("stock_code") or v.get("sku") or "").strip()
                    # Varyantın kendi unique stock_code'u parent'tan farklıysa onu kullan,
                    # değilse parent stock_code'u gönder.
                    if v_stock and v_stock != parent_stock:
                        item["stockCode"] = v_stock
                    elif parent_stock:
                        item["stockCode"] = parent_stock
                    else:
                        item["stockCode"] = v.get("barcode")
                    item["quantity"] = int(v.get("stock", 0))
                    item["attributes"] = resolve_attributes(attributes, product, v, category, meta)
                    items_to_send.append(item)
                    
        except Exception as e:
            errors.append(f"{product.get('name')} - Hazırlama Hatası: {str(e)}")
            
    if not items_to_send:
        # Save failure log — products bulundu ama validasyondan geçemedi
        from datetime import datetime, timezone
        first_reason = errors[0] if errors else "Bilinmeyen validasyon hatası."
        msg = (
            f"{len(products)} ürün DB'de bulundu ama Trendyol'a gönderilemedi. "
            f"İlk hata: {first_reason}"
        )
        log_doc = {
            "id": generate_id(),
            "started_at": datetime.now(timezone.utc).isoformat(),
            "status": "failed",
            "products_attempted": len(products),
            "products_sent": 0,
            "batch_request_id": None,
            "errors": errors,
            "not_found_codes": not_found_codes,
            "message": msg,
        }
        await db.trendyol_sync_logs.insert_one(log_doc)
        return {
            "success": False,
            "message": msg,
            "total": len(products),
            "successful": 0,
            "failed": len(errors),
            "not_found_codes": not_found_codes,
            "errors": errors
        }
        
    try:
        from datetime import datetime, timezone
        import asyncio as _asyncio
        started_at = datetime.now(timezone.utc).isoformat()
        response = await client.create_products(items_to_send)
        batch_id = response.get("batchRequestId")
        # Trendyol returned an error response (no batchRequestId)?
        trendyol_error = None
        if not batch_id:
            # response may contain {"errors":[...]} or {"message":...}
            trendyol_error = (
                response.get("message")
                or (response.get("errors") if isinstance(response.get("errors"), list) else None)
                or str(response)[:500]
            )
            if isinstance(trendyol_error, list) and trendyol_error:
                trendyol_error = "; ".join([
                    (e.get("message") if isinstance(e, dict) else str(e)) for e in trendyol_error
                ])[:1000]

            # 🛟 FALLBACK: "Tekrarlı ürün oluşturma isteği atılamaz" (Trendyol anti-spam)
            # geldiyse, ürün zaten Trendyol'da yaşıyor demektir → create yerine
            # price-and-inventory POST (stok/fiyat) ile güncellemeyi dene. Bu endpoint
            # farklı bir kapı (inventory/...) ve recurring throttling'e takılmaz.
            # Hala başarısızsa update_products (PUT) ile son bir deneme yap.
            if trendyol_error and (
                "tekrarl" in trendyol_error.lower()
                or "duplicate request" in trendyol_error.lower()
                or "recurring" in trendyol_error.lower()
            ):
                # ⚠️ Ürün Trendyol'da zaten yaşıyor. İKİ İŞ birden yapılmalı:
                #   1) price-and-inventory  → stok/fiyat senkronu (ayrı, throttle'a takılmaz kapı)
                #   2) update_products PUT  → kategori/attribute/görsel senkronu
                # ÖNCEKİ HATA: price-and-inventory batch dönünce burada DURULUYORDU; PUT hiç
                # çalışmadığı için "tekrar aktardım ama özellikleri güncellenmedi" oluyordu.
                # Artık PUT her zaman çalışır ve raporlanan batch = PUT batch'i olur (attribute
                # sonucunu kullanıcı görsün). pi yalnızca stok/fiyat için kısa süre poll edilir.
                logger.info("Trendyol create reddetti (tekrarlı). price-and-inventory + update_products fallback.")
                try:
                    pi_items = []
                    for it in items_to_send:
                        if not it.get("barcode"):
                            continue
                        entry = {
                            "barcode": str(it.get("barcode")),
                            "quantity": int(it.get("quantity", 0)),
                        }
                        if it.get("salePrice") is not None:
                            try:
                                entry["salePrice"] = float(it["salePrice"])
                            except Exception:
                                pass
                        if it.get("listPrice") is not None:
                            try:
                                entry["listPrice"] = float(it["listPrice"])
                            except Exception:
                                pass
                        pi_items.append(entry)
                    # 1) Stok/fiyat senkronu (best-effort)
                    pi_batch = None
                    try:
                        pi_response = await client.update_price_and_inventory(pi_items) if pi_items else {}
                        pi_batch = (pi_response or {}).get("batchRequestId")
                        if pi_batch:
                            logger.info(f"Tekrarlı fallback price-and-inventory batch={pi_batch}")
                    except Exception as pi_err:
                        logger.warning(f"Tekrarlı fallback price-and-inventory exception: {pi_err}")
                    # 2) Kategori/attribute/görsel senkronu — HER ZAMAN çalışır
                    upd_response = await client.update_products(items_to_send)
                    upd_batch = (upd_response or {}).get("batchRequestId")
                    if upd_batch:
                        batch_id = upd_batch
                        response = upd_response
                        trendyol_error = None
                        logger.info(f"Tekrarlı fallback update_products (attribute) başarılı: batch={upd_batch}")
                    elif pi_batch:
                        # PUT batch dönmediyse en azından stok/fiyat batch'ini takip et
                        batch_id = pi_batch
                        response = pi_response
                        trendyol_error = None
                        logger.info(f"Tekrarlı fallback: PUT batch yok, price-and-inventory batch={pi_batch} izleniyor.")
                    else:
                        upd_err = (upd_response or {}).get("message") or str(upd_response)[:500]
                        logger.warning(f"Tekrarlı fallback update_products reddetti: {upd_err}")
                except Exception as upd_err:
                    logger.error(f"Update fallback exception: {upd_err}")

        # 🔍 Gerçek batch sonucunu sorgula: Trendyol asenkron işliyor; 5-15sn'de tamamlanır.
        # "aktarıldı diyor ama Trendyol'da yok" tuzağını önler.
        # Trendyol BUG: bazen status=COMPLETED dönüp items[]'in işlenmiş olduğunu söyler,
        # ama failedItemCount=0 olur halbuki birkaç sn sonra items[].status FAILED olur.
        # Bu yüzden polling birkaç tur daha sürdürüyoruz: items[].status sayısı itemCount'a ulaşmalı
        # ve "tüm item'lar terminal status'ta" (SUCCESS|FAILED|COMPLETED) olmalı.
        batch_failed_items = []
        batch_success_count = 0
        batch_final_status = "INPROGRESS"

        def _norm_status(s: str) -> str:
            """Trendyol bazen 'IN_PROGRESS' bazen 'INPROGRESS' döndürüyor — normalize."""
            return (s or "").upper().replace("_", "").replace(" ", "").replace("-", "")

        if batch_id:
            for attempt in range(16):  # 16 deneme x 2.5sn = max ~40sn (ingress 60sn limitine güvenli)
                await _asyncio.sleep(2.5)
                try:
                    br = await client.get_batch_request_result(batch_id)
                    batch_final_status = (br or {}).get("status") or "INPROGRESS"
                    items = (br or {}).get("items", []) or []
                    item_count = (br or {}).get("itemCount", 0) or len(items)
                    # Tüm item'ların terminal status'ta olmasını bekle
                    terminal_set = {"SUCCESS", "FAILED", "COMPLETED"}
                    statuses_present = [(it.get("status") or "").upper() for it in items]
                    all_terminal = (
                        len(items) >= item_count
                        and all(s in terminal_set for s in statuses_present)
                    )
                    if _norm_status(batch_final_status) in ("COMPLETED", "FAILED") and all_terminal:
                        batch_failed_items = []
                        for it in items:
                            if (it.get("status") or "").upper() == "FAILED":
                                req = it.get("requestItem", {}) or {}
                                prod = req.get("product", {}) if isinstance(req, dict) else {}
                                batch_failed_items.append({
                                    "stock_code": prod.get("stockCode") or req.get("barcode") or "",
                                    "barcode": prod.get("barcode") or req.get("barcode") or "",
                                    "title": prod.get("title") or "",
                                    "reasons": it.get("failureReasons") or [],
                                })
                        # ✔ Gerçek başarı = items listesinden say (failedItemCount değil!)
                        batch_success_count = sum(1 for s in statuses_present if s == "SUCCESS")
                        break
                except Exception as poll_err:
                    logger.warning(f"Batch poll {attempt+1}/12 failed: {poll_err}")

        # 🛟 Polling timeout durumunda da elde olan son durumu işle — fallback'ler
        # böylece her senaryoda tetiklenir (kritik: aksi halde duplicate'lar sessizce kalır).
        if batch_id and not batch_failed_items and batch_success_count == 0:
            try:
                br = await client.get_batch_request_result(batch_id)
                items = (br or {}).get("items", []) or []
                if items:
                    batch_final_status = (br or {}).get("status") or batch_final_status
                    for it in items:
                        if (it.get("status") or "").upper() == "FAILED":
                            req = it.get("requestItem", {}) or {}
                            prod = req.get("product", {}) if isinstance(req, dict) else {}
                            batch_failed_items.append({
                                "stock_code": prod.get("stockCode") or req.get("barcode") or "",
                                "barcode": prod.get("barcode") or req.get("barcode") or "",
                                "title": prod.get("title") or "",
                                "reasons": it.get("failureReasons") or [],
                            })
                    batch_success_count = sum(
                        1 for it in items if (it.get("status") or "").upper() == "SUCCESS"
                    )
            except Exception as final_poll_err:
                logger.warning(f"Final batch poll failed: {final_poll_err}")

        # 🔁 SMART CONFLICT RESOLUTION:
        # - "Self duplicate" (sent_bc == conflict_bc) → PUT (update_products)
        # - "Cross duplicate" (sent_bc != conflict_bc) → eski barkodu ARCHIVE + yeni barkodu CREATE
        upsert_attempted = 0
        upsert_succeeded = 0
        upsert_failed_items = []
        upsert_batch_id = None
        upsert_final_status = None
        archived_barcodes: list = []
        archive_batch_id = None
        retry_create_batch_id = None
        retry_create_succeeded = 0

        def _is_duplicate_error(reasons):
            """Trendyol'un duplicate (çakışma) hatalarını tespit eder."""
            if not reasons:
                return False
            text = " ".join([
                (r.get("message") if isinstance(r, dict) else str(r)) for r in reasons
            ]).lower()
            patterns = [
                "aynı barkodlu",
                "ayni barkodlu",
                "aynı barkod",
                "barkod zaten",
                "stockcode zaten",
                "stok kodu zaten",
                "duplicate barcode",
                "duplicate stockcode",
                "already exist",
                "zaten mevcut",
                "zaten kayıtlı",
                "zaten kayitli",
                "productmainid",
                "bulunduğundan",
                "bulundugundan",
            ]
            return any(p in text for p in patterns)

        def _parse_conflict_barcode(reasons):
            """Trendyol hata mesajından çakışan barkodu çıkar."""
            import re
            text = " ".join([
                (r.get("message") if isinstance(r, dict) else str(r)) for r in (reasons or [])
            ])
            m = re.search(r"Barkod[:\s]+(\d{6,})", text)
            return m.group(1) if m else None

        def _is_not_found_error(reasons):
            """Trendyol'un 'ürün bulunamadı' (price-and-inventory'de Trendyol'da olmayan
            barkod) hatasını tespit eder."""
            if not reasons:
                return False
            text = " ".join([
                (r.get("message") if isinstance(r, dict) else str(r)) for r in reasons
            ]).lower()
            patterns = [
                "ürün bulunamadı",
                "urun bulunamadi",
                "product not found",
                "tedarikçi id si",
                "tedarikci id si",
            ]
            return any(p in text for p in patterns)

        # Hangi item'lar duplicate yüzünden patladı?
        duplicate_failed = [f for f in batch_failed_items if _is_duplicate_error(f.get("reasons"))]
        if batch_id and duplicate_failed:
            # Cross conflict: eski barkod ARCHIVE + yeni barkod CREATE
            # Self conflict: PUT update_products
            cross_conflicts = []   # (sent_barcode, conflict_old_barcode, item_payload)
            self_conflicts = []    # item_payload

            # items_to_send → barcode lookup
            items_by_barcode = {str(it.get("barcode") or ""): it for it in items_to_send}

            for f in duplicate_failed:
                sent_bc = str(f.get("barcode") or "")
                conflict_bc = _parse_conflict_barcode(f.get("reasons")) or ""
                payload = items_by_barcode.get(sent_bc)
                if not payload:
                    continue
                if conflict_bc and conflict_bc != sent_bc:
                    cross_conflicts.append((sent_bc, conflict_bc, payload))
                else:
                    self_conflicts.append(payload)

            # 1) CROSS CONFLICTS: önce ARCHIVE + RETRY CREATE.
            # (Önceki versiyonlarda PUT update_products deneniyordu ama Trendyol
            # stockCode altında çoklu kayıt olduğunda hayalet SUCCESS dönüyor → kaldırıldı.)
            if cross_conflicts:
                put_cross_succeeded_barcodes = set()  # şimdilik boş (PUT yok)

                # Hâlâ başarısız olan cross item'lar için archive + retry create
                remaining_cross = [
                    c for c in cross_conflicts
                    if str(c[2].get("barcode") or "") not in put_cross_succeeded_barcodes
                ]
                if remaining_cross:
                    # 🧹 AGRESİF TEMİZLİK: cross-conflict varsa, conflict'in tek barkodunu
                    # arşivlemek yetmez. Trendyol cache'inde aynı stockCode altında onlarca
                    # eski varyant olabiliyor. DB'de OLMAYAN tüm Trendyol barkodlarını
                    # bul ve arşivle. Sonra retry create yapılabilir.
                    old_barcodes_to_archive: set = set()
                    db_barcodes_per_sc: dict = {}
                    # 1) cross_conflicts'tan stockCode'ları topla, her stockCode için DB barkodlarını çek
                    stock_codes_in_play: set = set()
                    for c in remaining_cross:
                        old_barcodes_to_archive.add(c[1])
                        sc = (c[2] or {}).get("productMainId") or (c[2] or {}).get("stockCode")
                        if sc:
                            stock_codes_in_play.add(str(sc))
                    for sc in stock_codes_in_play:
                        # DB'deki barkodlar (bu stockCode için)
                        if sc not in db_barcodes_per_sc:
                            db_prod = await db.products.find_one(
                                {"$or": [{"stock_code": sc}, {"sku": sc}, {"variants.stock_code": sc}]},
                                {"_id": 0, "barcode": 1, "variants.barcode": 1, "stock_code": 1}
                            )
                            db_bcs = set()
                            if db_prod:
                                if db_prod.get("barcode"):
                                    db_bcs.add(str(db_prod.get("barcode")))
                                for v in (db_prod.get("variants") or []):
                                    if v.get("barcode"):
                                        db_bcs.add(str(v.get("barcode")))
                            db_barcodes_per_sc[sc] = db_bcs
                        # 2) Trendyol'da bu stockCode altındaki tüm barkodları çek
                        try:
                            tr_resp = await client.get_filtered_products(stock_code=sc, archived=False, size=100)
                            for p in (tr_resp or {}).get("content", []):
                                tr_bc = p.get("barcode")
                                if tr_bc and str(tr_bc) not in db_barcodes_per_sc[sc]:
                                    old_barcodes_to_archive.add(str(tr_bc))
                        except Exception as fe:
                            logger.warning(f"get_filtered_products(stock_code={sc}) failed: {fe}")
                    old_barcodes_to_archive = list(old_barcodes_to_archive)
                    logger.info(f"Trendyol CROSS-CONFLICT (deep): {len(old_barcodes_to_archive)} eski barkod arşivlenecek: {old_barcodes_to_archive[:10]}…")
                    try:
                        arch_resp = await client.archive_products(old_barcodes_to_archive) if old_barcodes_to_archive else {}
                        archive_batch_id = (arch_resp or {}).get("batchRequestId")
                        archived_barcodes = old_barcodes_to_archive
                        # Archive batch poll
                        if archive_batch_id:
                            for attempt in range(5):
                                await _asyncio.sleep(1.5)
                                try:
                                    abr = await client.get_batch_request_result(archive_batch_id)
                                    if _norm_status((abr or {}).get("status", "")) in ("COMPLETED", "FAILED"):
                                        break
                                except Exception:
                                    pass
                    except Exception as ae:
                        logger.error(f"Archive products error: {ae}")

                    # Retry CREATE with the new barcodes
                    retry_items = [c[2] for c in remaining_cross]
                    if retry_items:
                        try:
                            retry_resp = await client.create_products(retry_items)
                            retry_create_batch_id = (retry_resp or {}).get("batchRequestId")
                            if retry_create_batch_id:
                                for attempt in range(5):
                                    await _asyncio.sleep(2.0)
                                    try:
                                        rbr = await client.get_batch_request_result(retry_create_batch_id)
                                        rstatus = (rbr or {}).get("status") or "INPROGRESS"
                                        if _norm_status(rstatus) in ("COMPLETED", "FAILED"):
                                            for rit in (rbr or {}).get("items", []) or []:
                                                if (rit.get("status") or "").upper() == "FAILED":
                                                    req = rit.get("requestItem", {}) or {}
                                                    prod = req.get("product", {}) if isinstance(req, dict) else {}
                                                    upsert_failed_items.append({
                                                        "stock_code": prod.get("stockCode") or req.get("barcode") or "",
                                                        "barcode": prod.get("barcode") or req.get("barcode") or "",
                                                        "title": prod.get("title") or "",
                                                        "reasons": rit.get("failureReasons") or [],
                                                        "phase": "retry_create",
                                                    })
                                            retry_create_succeeded = max(
                                                0,
                                                (rbr or {}).get("itemCount", 0) - (rbr or {}).get("failedItemCount", 0),
                                            )
                                            upsert_succeeded += retry_create_succeeded
                                            break
                                    except Exception as rp:
                                        logger.warning(f"Retry-create poll {attempt+1}/5 failed: {rp}")
                        except Exception as re_err:
                            logger.error(f"Retry create after archive error: {re_err}")

            # 2) SELF CONFLICTS: PUT update_products + price-and-inventory POST
            # PUT mevcut ürünün kategori/attribute/image alanlarını günceller; ancak
            # stok/fiyat değerleri için Trendyol'un ayrı bir endpoint'i daha güvenilir
            # (`/inventory/.../price-and-inventory`). Stok güncellemesinin Trendyol
            # panelinde garantili yansıması için ikisini birden çalıştırıyoruz.
            if self_conflicts:
                # 2a) Önce price-and-inventory POST (stok/fiyat senkronu)
                try:
                    pi_items = []
                    for it in self_conflicts:
                        if not it.get("barcode"):
                            continue
                        entry = {
                            "barcode": str(it.get("barcode")),
                            "quantity": int(it.get("quantity", 0)),
                        }
                        if it.get("salePrice") is not None:
                            try:
                                entry["salePrice"] = float(it["salePrice"])
                            except Exception:
                                pass
                        if it.get("listPrice") is not None:
                            try:
                                entry["listPrice"] = float(it["listPrice"])
                            except Exception:
                                pass
                        pi_items.append(entry)
                    if pi_items:
                        pi_resp = await client.update_price_and_inventory(pi_items)
                        pi_batch = (pi_resp or {}).get("batchRequestId")
                        if pi_batch:
                            logger.info(f"Self-conflict price-and-inventory batch: {pi_batch}")
                            # Polling — kısa süre içinde tamamlanır
                            for attempt in range(5):
                                await _asyncio.sleep(1.5)
                                try:
                                    pbr = await client.get_batch_request_result(pi_batch)
                                    if _norm_status((pbr or {}).get("status", "")) in ("COMPLETED", "FAILED"):
                                        break
                                except Exception:
                                    pass
                except Exception as pi_err:
                    logger.warning(f"Self-conflict price-and-inventory exception: {pi_err}")

                # 2b) Sonra update_products PUT (kategori/attribute/image senkronu)
                upsert_attempted += len(self_conflicts)
                try:
                    upd_resp = await client.update_products(self_conflicts)
                    upsert_batch_id = (upd_resp or {}).get("batchRequestId")
                    if upsert_batch_id:
                        for attempt in range(8):
                            await _asyncio.sleep(2.5)
                            try:
                                ubr = await client.get_batch_request_result(upsert_batch_id)
                                upsert_final_status = (ubr or {}).get("status") or "INPROGRESS"
                                if _norm_status(upsert_final_status) in ("COMPLETED", "FAILED"):
                                    put_succeeded = max(
                                        0,
                                        (ubr or {}).get("itemCount", 0) - (ubr or {}).get("failedItemCount", 0),
                                    )
                                    for uit in (ubr or {}).get("items", []) or []:
                                        if (uit.get("status") or "").upper() == "FAILED":
                                            req = uit.get("requestItem", {}) or {}
                                            prod = req.get("product", {}) if isinstance(req, dict) else {}
                                            upsert_failed_items.append({
                                                "stock_code": prod.get("stockCode") or req.get("barcode") or "",
                                                "barcode": prod.get("barcode") or req.get("barcode") or "",
                                                "title": prod.get("title") or "",
                                                "reasons": uit.get("failureReasons") or [],
                                                "phase": "put_update",
                                            })
                                    upsert_succeeded += put_succeeded
                                    break
                            except Exception as up_poll_err:
                                logger.warning(f"Upsert PUT poll {attempt+1}/8 failed: {up_poll_err}")
                    else:
                        upd_err = (upd_resp or {}).get("message") or str(upd_resp)[:500]
                        logger.warning(f"Trendyol PUT update_products reddetti: {upd_err}")
                except Exception as up_err:
                    logger.error(f"Trendyol PUT update_products exception: {up_err}")

        # 🔁 3. FALLBACK: Price-and-inventory "ürün bulunamadı" hatası verirse
        # → Bu varyant Trendyol'da hiç yok. Asıl payload ile create_products'a gönder.
        # (Recurring fallback'te tüm itemlar price-and-inventory'ye yönlendirilince,
        # Trendyol'da olmayan varyantlar "ürün bulunamadı" der; onları create'e geri gönderelim.)
        not_found_failed = [f for f in batch_failed_items if _is_not_found_error(f.get("reasons"))]
        if not_found_failed:
            items_by_barcode = {str(it.get("barcode") or ""): it for it in items_to_send}
            nf_create_items = []
            for f in not_found_failed:
                bc = str(f.get("barcode") or "")
                payload = items_by_barcode.get(bc)
                if payload:
                    nf_create_items.append(payload)
            if nf_create_items:
                upsert_attempted += len(nf_create_items)
                try:
                    nf_resp = await client.create_products(nf_create_items)
                    nf_batch = (nf_resp or {}).get("batchRequestId")
                    if nf_batch:
                        logger.info(f"Not-found → create_products fallback batch: {nf_batch}")
                        if not retry_create_batch_id:
                            retry_create_batch_id = nf_batch
                        for attempt in range(5):
                            await _asyncio.sleep(2.0)
                            try:
                                nbr = await client.get_batch_request_result(nf_batch)
                                if _norm_status((nbr or {}).get("status", "")) in ("COMPLETED", "FAILED"):
                                    nf_succeeded = 0
                                    nf_failed_barcodes = set()
                                    for nit in (nbr or {}).get("items", []) or []:
                                        if (nit.get("status") or "").upper() == "SUCCESS":
                                            nf_succeeded += 1
                                        else:
                                            req = nit.get("requestItem", {}) or {}
                                            prod = req.get("product", {}) if isinstance(req, dict) else {}
                                            bc_n = prod.get("barcode") or req.get("barcode") or ""
                                            nf_failed_barcodes.add(str(bc_n))
                                            upsert_failed_items.append({
                                                "stock_code": prod.get("stockCode") or bc_n,
                                                "barcode": bc_n,
                                                "title": prod.get("title") or "",
                                                "reasons": nit.get("failureReasons") or [],
                                                "phase": "not_found_create",
                                            })
                                    upsert_succeeded += nf_succeeded
                                    # not-found-success olanları batch_failed_items'tan düş
                                    succeeded_barcodes = {
                                        str(f.get("barcode"))
                                        for f in not_found_failed
                                        if str(f.get("barcode")) not in nf_failed_barcodes
                                    }
                                    if succeeded_barcodes:
                                        batch_failed_items = [
                                            f for f in batch_failed_items
                                            if str(f.get("barcode")) not in succeeded_barcodes
                                        ]
                                        batch_success_count += nf_succeeded
                                    break
                            except Exception as nf_poll_err:
                                logger.warning(f"Not-found-create poll {attempt+1}/5 failed: {nf_poll_err}")
                except Exception as nf_err:
                    logger.error(f"Not-found → create_products exception: {nf_err}")

        # Upsert ile düzelen item'ları batch_failed_items'tan düş
        if upsert_succeeded > 0 and duplicate_failed:
            still_failed_keys = set()
            for f in upsert_failed_items:
                if f.get("barcode"):
                    still_failed_keys.add(str(f["barcode"]))

            def _still_failed(f):
                if not _is_duplicate_error(f.get("reasons")):
                    return True
                return str(f.get("barcode") or "") in still_failed_keys

            batch_failed_items = [f for f in batch_failed_items if _still_failed(f)]
            batch_success_count = batch_success_count + upsert_succeeded

        # Local "errors" + Trendyol API hataları + Batch failure'ları birleştir
        all_errors = list(errors)
        if trendyol_error:
            all_errors.append(f"Trendyol API: {trendyol_error}")
        for f in batch_failed_items:
            reason = "; ".join([
                (r.get("message") if isinstance(r, dict) else str(r)) for r in (f.get("reasons") or [])
            ])
            all_errors.append(f"{f.get('title') or f.get('stock_code')} [{f.get('stock_code')}]: {reason}")
        for f in upsert_failed_items:
            reason = "; ".join([
                (r.get("message") if isinstance(r, dict) else str(r)) for r in (f.get("reasons") or [])
            ])
            all_errors.append(f"[UPDATE] {f.get('title') or f.get('stock_code')} [{f.get('stock_code')}]: {reason}")

        # Final status hesapla
        if not batch_id:
            final_status = "failed"
        elif batch_failed_items and batch_success_count == 0:
            final_status = "failed"
        elif batch_failed_items:
            final_status = "partial"
        elif _norm_status(batch_final_status) == "INPROGRESS":
            final_status = "pending"  # 60sn'de tamamlanmadı, kullanıcıya logdan takibi öneriliyor
        else:
            final_status = "success" if not errors else "partial"

        log_doc = {
            "id": generate_id(),
            "started_at": started_at,
            "finished_at": datetime.now(timezone.utc).isoformat(),
            "status": final_status,
            "products_attempted": len(products),
            "products_sent": batch_success_count if batch_id else 0,
            "products_failed": len(batch_failed_items),
            "batch_request_id": batch_id,
            "batch_final_status": batch_final_status,
            "upsert_attempted": upsert_attempted,
            "upsert_succeeded": upsert_succeeded,
            "upsert_batch_id": upsert_batch_id,
            "upsert_final_status": upsert_final_status,
            "upsert_failed_items": upsert_failed_items,
            "archived_barcodes": archived_barcodes,
            "archive_batch_id": archive_batch_id,
            "retry_create_batch_id": retry_create_batch_id,
            "retry_create_succeeded": retry_create_succeeded,
            "errors": all_errors,
            "failed_items": batch_failed_items,
            "trendyol_response": response,
            "message": (
                f"{batch_success_count} ürün başarıyla aktarıldı." + (f" ({upsert_succeeded} adet UPDATE ile)." if upsert_succeeded else "") if batch_id and _norm_status(batch_final_status) == "COMPLETED" and not batch_failed_items
                else f"{batch_success_count} başarılı, {len(batch_failed_items)} HATA — detaylar loglarda." if batch_failed_items
                else f"Batch alındı, Trendyol işliyor (durum: {batch_final_status}). Loglardan takip edin." if batch_id
                else f"Trendyol kabul etmedi: {trendyol_error}"
            ),
        }
        await db.trendyol_sync_logs.insert_one(log_doc)
        # _id ekleniyor MongoDB tarafından — response'a koyma
        log_doc.pop("_id", None)

        # 🤖 OTOMATİK RETRY KUYRUĞU: "Aynı barkod var" / "ürün bulunamadı" / "Trendyol cache"
        # tipi sıkışmalarda barkodları otomatik kuyruğa ekle (saatte bir retry edilir).
        # Header'da X-Internal-Retry varsa atla (loop önlemi).
        try:
            from fastapi import Request as _Req  # noqa
            # request header'ına direkt erişim yok; bunun yerine Request inject etmiyoruz.
            # Bunun yerine sadece "stuck" tipi hataları kuyruğa al.
            stuck_barcodes = []
            for f in batch_failed_items:
                reasons_text = " ".join([
                    (r.get("message") if isinstance(r, dict) else str(r))
                    for r in (f.get("reasons") or [])
                ]).lower()
                if any(p in reasons_text for p in ["aynı barkodlu", "ayni barkodlu", "ürün bulunamadı", "urun bulunamadi", "tedarikçi id si", "tedarikci id si"]):
                    if f.get("barcode"):
                        stuck_barcodes.append(str(f["barcode"]))
            for f in upsert_failed_items:
                reasons_text = " ".join([
                    (r.get("message") if isinstance(r, dict) else str(r))
                    for r in (f.get("reasons") or [])
                ]).lower()
                if any(p in reasons_text for p in ["aynı barkodlu", "ayni barkodlu", "ürün bulunamadı", "urun bulunamadi"]):
                    if f.get("barcode"):
                        stuck_barcodes.append(str(f["barcode"]))
            if stuck_barcodes:
                from routes.trendyol_retry_queue import add_to_queue as _add_to_queue
                added = await _add_to_queue(db, stuck_barcodes, reason="Trendyol cache conflict — auto-queued")
                logger.info(f"Trendyol auto-queue: {added} barkod retry kuyruğuna eklendi")
        except Exception as q_err:
            logger.warning(f"Auto-queue failed: {q_err}")

        return {
            "success": bool(batch_id) and not batch_failed_items and _norm_status(batch_final_status) != "FAILED",
            "message": log_doc["message"],
            "total": len(products),
            "successful": batch_success_count if batch_id else 0,
            "failed": len(batch_failed_items) + (0 if batch_id else len(items_to_send)),
            "batchRequestId": batch_id,
            "batch_final_status": batch_final_status,
            "failed_items": batch_failed_items,
            "upsert_attempted": upsert_attempted,
            "upsert_succeeded": upsert_succeeded,
            "upsert_batch_id": upsert_batch_id,
            "upsert_failed_items": upsert_failed_items,
            "archived_barcodes": archived_barcodes,
            "archive_batch_id": archive_batch_id,
            "retry_create_batch_id": retry_create_batch_id,
            "errors": all_errors,
            "trendyol_response": response,
        }
    except Exception as e:
        logger.error(f"Error syncing products to Trendyol: {str(e)}")
        from datetime import datetime, timezone
        log_doc = {
            "id": generate_id(),
            "started_at": datetime.now(timezone.utc).isoformat(),
            "status": "error",
            "products_attempted": len(products),
            "products_sent": 0,
            "batch_request_id": None,
            "errors": errors + [f"API Hatası: {str(e)}"],
            "message": "Trendyol API hatası oluştu."
        }
        await db.trendyol_sync_logs.insert_one(log_doc)
        if hasattr(e, "response") and hasattr(e.response, "json"):
            raise HTTPException(status_code=500, detail=e.response.json())
        raise HTTPException(status_code=500, detail="Trendyol API ürün gönderimi başarısız oldu.")
@router.get("/trendyol/sync-logs")
async def get_trendyol_sync_logs(
    page: int = 1,
    limit: int = 20,
    current_user: dict = Depends(require_admin)
):
    """Get paginated Trendyol sync logs"""
    skip = (page - 1) * limit
    logs = await db.trendyol_sync_logs.find({}, {"_id": 0}).sort("started_at", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.trendyol_sync_logs.count_documents({})
    return {"logs": logs, "total": total, "page": page}
@router.post("/trendyol/products/inventory-sync")
async def sync_trendyol_inventory(payload: dict = Body(default={}), current_user: dict = Depends(require_admin)):
    """Bulk sync stock and prices to Trendyol.

    payload BOŞSA  -> tüm aktif ürünler (eski davranış, byte-aynı).
    payload barcodes/stock_codes/product_ids/category_filters İÇERİYORSA
                   -> yalnız hedeflenen ürünler (kod-bazlı stok/fiyat güncelleme, Faz T2).
    Aktarmadaki _build_product_query_from_payload aynen yeniden kullanılır.
    """
    has_filter = any(payload.get(k) for k in ("barcodes", "stock_codes", "product_ids", "category_filters"))
    if has_filter:
        query = await _build_product_query_from_payload(payload)
        # query boş kalırsa (örn. eşleşmeyen kategori) HER ŞEYİ değil, hiçbir şeyi güncelle -> güvenli.
        products = await db.products.find(query).to_list(length=None) if query else []
    else:
        products = await db.products.find({"is_active": True}).to_list(length=None)
    return await _sync_inventory_to_trendyol(products)
@router.post("/trendyol/products/{product_id}/sync-inventory")
async def sync_single_product_inventory(
    product_id: str,
    current_user: dict = Depends(require_admin)
):
    """Sync stock and prices for a single product to Trendyol"""
    product = await db.products.find_one({"id": product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")
    return await _sync_inventory_to_trendyol([product])
async def _sync_inventory_to_trendyol(products: list):
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")
        
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )
    
    items_to_send = []
    # Global markup (Ana Ayarlar/Trendyol default_markup) uygula — cron ile manuel push
    # AYNI mantıkta çalışsın (aksi halde otomatik senkron zammı geri alıyordu).
    default_markup = float(config.get("default_markup", 0) or 0)

    for product in products:
        _mult = product.get("trendyol_multiplier")
        markup = float(_mult) if (_mult is not None and float(_mult) > 0) else default_markup
        factor = 1 + markup / 100.0
        base_price = _mp_base_price(product) * factor  # #4: üye fiyatı baz
        sale_price = _mp_base_price(product) * factor  # Trendyol: indirimsiz satis fiyati
        variants = product.get("variants", [])
        if not variants:
            if product.get("barcode"):
                items_to_send.append({
                    "barcode": product["barcode"],
                    "quantity": int(product.get("stock", 0)),
                    "salePrice": round(sale_price, 2),
                    "listPrice": round(base_price, 2)
                })
        else:
            for v in variants:
                if v.get("barcode"):
                    diff = float(v.get("price_diff") or 0)
                    items_to_send.append({
                        "barcode": v["barcode"],
                        "quantity": int(v.get("stock", 0)),
                        "salePrice": round(sale_price + diff, 2),
                        "listPrice": round(base_price + diff, 2)
                    })
    
    from datetime import datetime, timezone
    started_at = datetime.now(timezone.utc).isoformat()
    
    if not items_to_send:
        # Log failure to sync screen
        log_doc = {
            "id": generate_id(),
            "started_at": started_at,
            "status": "failed",
            "products_attempted": len(products),
            "products_sent": 0,
            "batch_request_id": None,
            "errors": ["Gönderilecek stok/fiyat bilgisi bulunamadı (barkodlar eksik olabilir)."],
            "message": "Envanter güncellemesi başarısız (geçerli barkod yok)."
        }
        await db.trendyol_sync_logs.insert_one(log_doc)
        return {"success": False, "message": "Gönderilecek stok/fiyat bilgisi bulunamadı (barkod eksik?)"}
        
    try:
        res = await client.update_price_and_inventory(items_to_send)
        batch_id = res.get("batchRequestId", "")
        
        # Log to the new sync logs screen
        log_doc = {
            "id": generate_id(),
            "started_at": started_at,
            "finished_at": datetime.now(timezone.utc).isoformat(),
            "status": "success",
            "products_attempted": len(products),
            "products_sent": len(items_to_send),
            "batch_request_id": batch_id,
            "errors": [],
            "message": "Stok ve fiyat güncellemesi başarıyla gönderildi."
        }
        await db.trendyol_sync_logs.insert_one(log_doc)
        
        return {"success": True, "message": f"{len(items_to_send)} kalem ürünün stok/fiyat bilgisi Trendyol'a gönderildi.", "batch_id": batch_id}
    except Exception as e:
        logger.error(f"Trendyol inventory sync error: {str(e)}")
        log_doc = {
            "id": generate_id(),
            "started_at": started_at,
            "status": "error",
            "products_attempted": len(products),
            "products_sent": 0,
            "batch_request_id": None,
            "errors": [f"API Hatası: {str(e)}"],
            "message": "Stok/fiyat güncellemesi sırasında hata oluştu."
        }
        await db.trendyol_sync_logs.insert_one(log_doc)
        raise HTTPException(status_code=500, detail=f"Trendyol stok/fiyat güncelleme hatası: {str(e)}")
@router.get("/trendyol/products/batch-status/{batch_id}")
async def get_trendyol_batch_status_v2(batch_id: str, current_user: dict = Depends(require_admin)):
    """Check the status of a batch request"""
    config = await get_trendyol_config()
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    
    try:
        client = TrendyolClient(
            supplier_id=config["supplier_id"],
            api_key=config["api_key"],
            api_secret=config["api_secret"],
            mode=config["mode"]
        )
        status = await client.get_batch_request_result(batch_id)
        return {"success": True, "status": status}
    except Exception as e:
        logger.error(f"Error fetching batch status: {str(e)}")
        raise HTTPException(status_code=500, detail="Batch durumu alınamadı.")
async def _sync_trendyol_status_passes(client, start_date_ms, end_date_ms, widen_cancel=True, statuses=("Cancelled", "Returned", "UnDelivered")):
    """İptal/iade/teslim edilemedi gibi durum değişikliklerini ayrı status
    sorgularıyla yakalar; yalnızca MEVCUT siparişlerin status alanını günceller
    (kaydı baştan ezmez). Trendyol orders ucu ~2 haftalık pencereyle sınırlıdır;
    daha eski iadeler İadeler (claims) modülünden takip edilir."""
    from datetime import datetime, timezone, timedelta
    updated = 0
    _logged_sample = False
    for st in statuses:
        # İptaller daha GENİŞ pencerede taranır: Trendyol startDate/endDate sipariş
        # TARİHİNE göre filtreler → 14 günden eski bir sipariş sonradan iptal olursa dar
        # pencereye girmez ve İptaller'e hiç düşmezdi ("bazıları düşmemiş" kök nedeni).
        # Cancelled için pencereyi 45 güne kadar geriye çek.
        _start = start_date_ms
        if st == "Cancelled" and widen_cancel:
            _wide = int((datetime.now(timezone.utc) - timedelta(days=45)).timestamp() * 1000)
            _start = min(start_date_ms, _wide) if start_date_ms else _wide
        try:
            page = 0
            while page < 25:
                resp = await client.get_orders(
                    start_date_ms=_start, end_date_ms=end_date_ms,
                    status=st, size=200, page=page,
                )
                chunk = resp.get("content", []) or []
                for t_order in chunk:
                    onum = str(t_order.get("orderNumber"))
                    mapped = map_trendyol_order(t_order)
                    _raw_status = t_order.get("status")
                    # GERÇEK SEBEP ARAYIŞI — ground truth için: iptal kaydının HAM yapısını
                    # her taramada BİR KEZ logla; Trendyol'un sebebi hangi alanda verdiğini
                    # (varsa) buradan kesin görürüz.
                    if st == "Cancelled" and not _logged_sample:
                        try:
                            import json as _json
                            _lines0 = (t_order.get("lines") or [{}])[0]
                            logger.info(
                                "[trendyol cancel sample %s] top_keys=%s line_keys=%s "
                                "cancellationReason=%r cancelReason=%r packageHistories=%r "
                                "line.cancellationReason=%r line.orderLineItemStatusName=%r",
                                onum, sorted([str(k) for k in t_order.keys()]),
                                sorted([str(k) for k in _lines0.keys()]),
                                t_order.get("cancellationReason"), t_order.get("cancelReason"),
                                str(t_order.get("packageHistories"))[:300],
                                _lines0.get("cancellationReason"), _lines0.get("orderLineItemStatusName"),
                            )
                            _logged_sample = True
                        except Exception:
                            _logged_sample = True
                    # Sebebi orders ucundaki TÜM olası alanlardan dene (sipariş + line seviyesi).
                    _l0 = (t_order.get("lines") or [{}])[0] if t_order.get("lines") else {}
                    _reason = (
                        t_order.get("cancellationReason") or t_order.get("cancelReason")
                        or t_order.get("cancellationReasonText") or t_order.get("customerCancellationReason")
                        or _l0.get("cancellationReason") or _l0.get("cancelReason") or ""
                    )
                    if isinstance(_reason, dict):
                        _reason = _reason.get("name") or _reason.get("text") or _reason.get("reason") or ""
                    _reason = str(_reason or "").strip()
                    if not _reason and st == "Cancelled":
                        # Yedek: müşteri iade/iptal claim'i varsa oradaki sebep.
                        try:
                            _cl = await db.trendyol_claims.find_one(
                                {"order_number": onum, "claim_reason": {"$nin": ["", None]}},
                                {"_id": 0, "claim_reason": 1},
                                sort=[("updated_at", -1)],
                            )
                            if _cl and _cl.get("claim_reason"):
                                _reason = _cl["claim_reason"]
                        except Exception:
                            pass
                    if not _reason:
                        _reason = ("Trendyol iptali" if st == "Cancelled"
                                   else ("Teslim edilemedi (Trendyol)" if st == "UnDelivered"
                                         else "Trendyol iadesi"))
                    _set = {
                        "status": mapped.get("status"),
                        "trendyol_status_raw": _raw_status,
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    }
                    if st == "Cancelled":
                        _set["cancel_reason"] = _reason
                        _set["cancel_source"] = "trendyol"
                    res = await db.orders.update_one(
                        {"order_number": onum, "platform": "trendyol"},
                        {"$set": _set},
                    )
                    if res.modified_count:
                        updated += 1
                    # ÖNEMLİ ("bi düşüyor bi düşmüyor" kök nedeni): sipariş bizde YOKSA
                    # (eski / hiç senkronlanmamış) eskiden update_one sessizce düşüyordu.
                    # Artık tam kaydı iptal/iade durumuyla EKLERİZ → İptaller/İadeler
                    # sayfasına garanti düşer. created_at = gerçek sipariş tarihi (orderDate)
                    # ki tarih sıralaması doğru olsun.
                    _was_new = (res.matched_count == 0)
                    if _was_new:
                        try:
                            mapped["id"] = generate_id()
                            mapped["created_at"] = _ms_to_iso(t_order.get("orderDate")) or datetime.now(timezone.utc).isoformat()
                            mapped["trendyol_status_raw"] = _raw_status
                            if st == "Cancelled":
                                mapped["cancel_reason"] = _reason
                                mapped["cancel_source"] = "trendyol"
                            await db.orders.insert_one(mapped)
                            updated += 1
                        except Exception as _ie:
                            logger.error(f"[trendyol status upsert {onum}] {_ie}")
                    # İptal senkronu → stoğu BİR KEZ geri ekle (idempotent; manuel iptalle aynı
                    # order_cancelled guard). YENİ eklenen kayıtta stok geri EKLENMEZ: bu sipariş
                    # bizde hiç olmadığı için stok daha önce DÜŞÜLMEDİ → +1 yanlış olurdu.
                    if st == "Cancelled" and not _was_new:
                        try:
                            _o = await db.orders.find_one({"order_number": onum, "platform": "trendyol"}, {"_id": 0, "id": 1, "items": 1})
                            if _o:
                                _already = await db.stock_movements.find_one({"order_id": _o.get("id"), "type": "order_cancelled"}, {"_id": 1})
                                if not _already:
                                    from routes.orders import _stock_delta_for_order
                                    _moves = await _stock_delta_for_order(_o, +1)
                                    await db.stock_movements.insert_one({
                                        "id": str(uuid.uuid4()), "type": "order_cancelled",
                                        "order_id": _o.get("id"), "order_number": onum,
                                        "items": _moves, "source": "trendyol_cancel_sync",
                                        "created_at": datetime.now(timezone.utc).isoformat(),
                                    })
                        except Exception as _re:
                            logger.error(f"[trendyol cancel restock {onum}] {_re}")
                total_pages = resp.get("totalPages") or 0
                page += 1
                if not chunk or page >= total_pages:
                    break
        except Exception as e:
            logger.error(f"[trendyol status pass {st}] {e}")
    return updated
def map_trendyol_order(t_order: dict) -> dict:
    from datetime import datetime, timezone
    order_number = t_order.get("orderNumber")
    
    items = []
    total_price = t_order.get("totalPrice", 0)
    gross_amount = t_order.get("grossAmount", 0)
    total_discount = t_order.get("totalDiscount", 0)
    
    for line in t_order.get("lines", []):
        qty = max(line.get("quantity", 1), 1)
        line_gross = line.get("lineGrossAmount", line.get("amount", 0))
        unit_price = line_gross / qty # İndirimsiz birim fiyat
        
        # Trendyol 'price' field is actually the net discounted price
        net_price = line.get("price", line.get("lineUnitPrice", 0))
        
        discount = line.get("discount", 0)
        discount_per_item = discount / qty if discount else 0

        items.append({
            "product_id": line.get("productCode"),
            "product_name": line.get("productName"),
            "quantity": qty,
            "unit_price": unit_price, # İndirimsiz birim fiyat
            "discount_amount": discount_per_item, # Birim başına indirim
            "price": net_price, # Net ödenen birim fiyat (Faturalandırılan)
            "size": line.get("productSize", ""),
            "color": line.get("productColor", ""),
            "barcode": line.get("barcode", ""),
            "currency": line.get("currencyCode", "TRY")
        })

    shipment_address = t_order.get("shipmentAddress", {})
    invoice_address = t_order.get("invoiceAddress", {})

    # --- Kurumsal fatura alanları (VKN / vergi dairesi / ünvan) ---
    # Trendyol bu alanları invoiceAddress İÇİNDE de, sipariş ÜST SEVİYESİNDE de
    # (taxNumber / taxOffice) verebilir. Eskiden yalnızca invoiceAddress.* okunduğu
    # için müşteri kurumsal fatura talebinde VKN girse bile (üst seviye taxNumber)
    # billing'e düşmüyor → e-Arşiv VKN'siz/vergi-dairesiz kalıp Doğan reddediyordu.
    # Artık ikisi de okunur, dolu olan kullanılır.
    _cust_tax_number = (str(invoice_address.get("taxNumber") or "").strip()
                        or str(t_order.get("taxNumber") or "").strip())
    _cust_tax_office = (str(invoice_address.get("taxOffice") or "").strip()
                        or str(t_order.get("taxOffice") or "").strip())
    _cust_company = (str(invoice_address.get("company") or "").strip()
                     or str(invoice_address.get("companyName") or "").strip()
                     or str(invoice_address.get("companyTitle") or "").strip())
    _is_corporate = (len(_cust_tax_number) == 10) or bool(_cust_company)

    # --- Mikro İhracat Tespiti ---
    # KESİN gösterge: Trendyol paketindeki `micro` bayrağı (ve ETGB no/tarihi).
    # Adres ülkesine güvenilmez: mikro ihracatta kargo Türkiye içi aktarım
    # noktasına teslim edildiği için shipmentAddress.country = "Türkiye" görünür;
    # yabancı ülke invoiceAddress (alıcı) tarafındadır. Bu yüzden yedek ülke
    # kontrolü invoiceAddress üzerinden yapılır.
    delivery_type = (t_order.get("deliveryType") or "").lower()
    _buyer_country = (invoice_address.get("country") or "").strip().lower()
    _is_intl_buyer = bool(_buyer_country) and _buyer_country not in ("turkey", "türkiye", "tr")
    is_micro_export = bool(
        t_order.get("micro")                 # birincil/kesin bayrak
        or t_order.get("etgbNo")             # ETGB numarası => mikro ihracat
        or t_order.get("etgbDate")
        or "micro" in delivery_type
        or "international" in delivery_type
        or _is_intl_buyer                    # yedek: ALICI ülkesi TR dışı
    )
    
    # Pazaryeri siparisi panele DAIMA "confirmed" (Onaylandi) duser; Trendyol'un is
    # akisi durumu (Picking/Invoiced/Shipped/Delivered) bizim operasyonel durumumuzu
    # EZMEZ. Yalnizca iptal/iade/teslim-edilemedi terminal durumlari yansir.
    status_map = {
        "Cancelled": "cancelled",
        "Returned": "returned",
        "UnDelivered": "returned",
    }
    
    order_doc = {
        "order_number": str(order_number),
        "platform": "trendyol",
        "trendyol_package_id": t_order.get("id"),
        "user_id": None,
        "items": items,
        "shipping_address": {
            "first_name": shipment_address.get("firstName", "Trendyol"),
            "last_name": shipment_address.get("lastName", "Müşterisi"),
            "phone": shipment_address.get("phone", ""),
            "email": t_order.get("customerEmail", ""),
            "address": shipment_address.get("fullAddress", ""),
            "city": shipment_address.get("city", ""),
            "district": shipment_address.get("district", ""),
            "country": shipment_address.get("country", "")
        },
        "billing_address": {
            "first_name": invoice_address.get("firstName", "Trendyol"),
            "last_name": invoice_address.get("lastName", "Müşterisi"),
            "phone": invoice_address.get("phone", ""),
            "address": invoice_address.get("fullAddress", ""),
            "city": invoice_address.get("city", ""),
            "district": invoice_address.get("district", ""),
            "country": invoice_address.get("country", ""),
            "company_name": _cust_company,
            "tax_number": _cust_tax_number,
            "tax_office": _cust_tax_office,
            "is_corporate": _is_corporate
        },
        "billing_info": {
            "is_corporate": _is_corporate,
            "company_name": _cust_company,
            "tax_number": _cust_tax_number,
            "tax_office": _cust_tax_office,
            # Müşteri self-deklarasyonu burada yok; e-Fatura mükellefiyeti kesimde
            # Doğan CheckUser ile sorgulanır → mükellefse e-Fatura, değilse e-Arşiv.
            "e_invoice_user": False,
        },
        "subtotal": gross_amount if gross_amount else total_price,
        "shipping_cost": 0,
        "discount_amount": total_discount,
        "total": total_price,
        "payment_method": "marketplace",
        "payment_status": "paid",
        "status": status_map.get(t_order.get("status"), "confirmed"),
        "cargo_tracking_number": t_order.get("cargoTrackingNumber", ""),
        "cargo_tracking_link": t_order.get("cargoTrackingLink", ""),
        "cargo_provider_name": t_order.get("cargoProviderName", ""),
        "invoice_link": t_order.get("invoiceLink", ""),
        # Mikro ihracat bilgileri — ayrı faturalama/ETGB akışı için
        "is_micro_export": is_micro_export,
        "shipment_country": shipment_address.get("country", ""),
        "delivery_type": t_order.get("deliveryType", ""),
        "trendyol_customer_id": str(t_order.get("customerId") or ""),
        "trendyol_identity_number": str(t_order.get("identityNumber") or ""),
        # --- FAZ 1b: Pazaryeri / SLA bilgileri (ham .get; alan yoksa boş) ---
        "marketplace_status": t_order.get("status", ""),
        "marketplace_agreed_delivery_date": _ms_to_iso(t_order.get("agreedDeliveryDate")),
        "marketplace_estimated_delivery_start": _ms_to_iso(t_order.get("estimatedDeliveryStartDate")),
        "marketplace_estimated_delivery_end": _ms_to_iso(t_order.get("estimatedDeliveryEndDate")),
        "marketplace_last_modified": _ms_to_iso(t_order.get("lastModifiedDate")),
        "cargo_sender_number": str(t_order.get("cargoSenderNumber") or ""),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    return order_doc
@router.post("/trendyol/orders/preview")
async def preview_trendyol_orders(req: TrendyolOrderPreviewReq, current_user: dict = Depends(require_admin)):
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    client = TrendyolClient(
        supplier_id=config["supplier_id"], api_key=config["api_key"],
        api_secret=config["api_secret"], mode=config["mode"]
    )
    try:
        resp = await client.get_orders(
            start_date_ms=req.start_date_ms, 
            end_date_ms=req.end_date_ms, 
            order_number=req.order_number, 
            size=100
        )
        content = resp.get("content", [])
        return {"success": True, "orders": content}
    except Exception as e:
        logger.error(f"Error previewing Trendyol orders: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
@router.post("/trendyol/orders/import-selected")
async def import_selected_trendyol_orders(req: TrendyolOrderImportReq, current_user: dict = Depends(require_admin)):
    try:
        from datetime import datetime, timezone
        from .deps import generate_id
        imported_count = 0
        updated_count = 0
        errors = []
        for t_order in req.orders:
            order_number = str(t_order.get("orderNumber"))
            try:
                order_data = map_trendyol_order(t_order)
                existing = await db.orders.find_one({"order_number": order_number, "platform": "trendyol"})
                if existing:
                    await db.orders.update_one({"_id": existing["_id"]}, {"$set": {k: v for k, v in order_data.items() if k != "status"}})
                    updated_count += 1
                else:
                    order_data["id"] = generate_id()
                    order_data["created_at"] = _ms_to_iso(t_order.get("orderDate")) or datetime.now(timezone.utc).isoformat()
                    await db.orders.insert_one(order_data)
                    imported_count += 1
                    await _decrement_stock_for_imported_order(order_data, "trendyol")
                await log_integration_event("trendyol", "import_order", "order", order_number, "success", "Sipariş başarıyla aktarıldı.")
            except Exception as e:
                err_msg = str(e)
                errors.append({"orderNumber": order_number, "error": err_msg})
                await log_integration_event("trendyol", "import_order", "order", order_number, "error", f"Aktarım hatası: {err_msg}", {"raw": t_order})
                
        return {"success": True, "imported": imported_count, "updated": updated_count, "errors": errors}
    except Exception as e:
        logger.error(f"Error importing selected orders: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
@router.post("/trendyol/orders/import")
async def import_trendyol_orders(current_user: dict = Depends(require_admin)):
    """Import orders from Trendyol (Last 15 days) auto job"""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")
    
    import sys
    import os
    import time
    from datetime import datetime, timezone
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    from .deps import generate_id
    
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )
    
    # Trendyol siparis ucu en fazla ~2 hafta (14 gun) araliga izin verir.
    now = datetime.now()
    start = now - timedelta(days=14)
    end_date_ms = int(now.timestamp() * 1000)
    start_date_ms = int(start.timestamp() * 1000)
    
    imported_count = 0
    updated_count = 0
    
    try:
        # TUM sayfalari dolas - daha once tek sayfa (200) cekilip fazlasi atlaniyordu.
        content = []
        page = 0
        MAX_PAGES = 50
        while page < MAX_PAGES:
            resp = await client.get_orders(
                start_date_ms=start_date_ms, end_date_ms=end_date_ms, size=200, page=page
            )
            chunk = resp.get("content", []) or []
            content.extend(chunk)
            total_pages = resp.get("totalPages") or 0
            page += 1
            if not chunk or page >= total_pages:
                break
        
        for t_order in content:
            order_number = t_order.get("orderNumber")
            existing_order = await db.orders.find_one({"order_number": str(order_number), "platform": "trendyol"})
            try:
                order_data = map_trendyol_order(t_order)
                if existing_order:
                    await db.orders.update_one(
                        {"_id": existing_order["_id"]},
                        {"$set": {k: v for k, v in order_data.items() if k != "status"}}
                    )
                    updated_count += 1
                else:
                    order_data["id"] = generate_id()
                    order_data["created_at"] = _ms_to_iso(t_order.get("orderDate")) or datetime.now(timezone.utc).isoformat()
                    await db.orders.insert_one(order_data)
                    imported_count += 1
                    await _decrement_stock_for_imported_order(order_data, "trendyol")
            except Exception as e:
                logger.error(f"Error mapping/saving order {order_number}: {e}")
                await log_integration_event("trendyol", "auto_import", "order", str(order_number), "error", f"Otomatik aktarım hatası: {str(e)}", {"raw": t_order})
        
        try:
            await _sync_trendyol_status_passes(client, start_date_ms, end_date_ms)
        except Exception as _e:
            logger.error(f"[trendyol status passes] {_e}")

        return {
            "success": True, 
            "message": f"Trendyol'dan {imported_count} yeni sipariş aktarıldı, {updated_count} sipariş güncellendi.",
            "imported": imported_count,
            "updated": updated_count
        }
    except Exception as e:
        logger.error(f"Error importing Trendyol orders: {str(e)}")
        await log_integration_event("trendyol", "auto_import_job", "system", "-", "error", f"Toplu aktarım hatası: {str(e)}")
        raise HTTPException(status_code=500, detail="Sipariş aktarımı sırasında bir hata oluştu.")
@router.get("/trendyol/category-mappings")
async def get_trendyol_category_mappings(current_user: dict = Depends(require_admin)):
    """Get all local categories with their Trendyol mappings, excluding hidden ones"""
    try:
        categories = await db.categories.find({"trendyol_hidden": {"$ne": True}}).to_list(1000)
        mappings = []
        for c in categories:
            mappings.append({
                "id": str(c["_id"]) if "_id" in c else c.get("id"),
                "local_name": c.get("name"),
                "trendyol_category_id": c.get("trendyol_category_id"),
                "trendyol_category_name": c.get("trendyol_category_name"),
                "attribute_mappings": c.get("attribute_mappings", []),
                "value_mappings": c.get("value_mappings", {}),
                "default_mappings": c.get("default_mappings", {}),
                "has_children": c.get("has_children", c.get("children_count", 0) > 0),
                "is_matched": bool(c.get("trendyol_category_id"))
            })
        return {"success": True, "mappings": mappings}
    except Exception as e:
        logger.error(f"Error fetching category mappings: {e}")
        raise HTTPException(status_code=500, detail="Kategori eşleştirmeleri alınamadı.")
@router.post("/trendyol/category-mappings")
async def save_trendyol_category_mapping(req: CategoryMappingReq, current_user: dict = Depends(require_admin)):
    """Save a Trendyol category mapping to a local category"""
    from bson.objectid import ObjectId
    try:
        # Support both string UUID and ObjectId
        filter_q = {"id": req.local_category_id} if len(req.local_category_id) > 24 else {"$or": [{"id": req.local_category_id}, {"_id": ObjectId(req.local_category_id)}]}
        
        await db.categories.update_one(
            filter_q,
            {"$set": {
                "trendyol_category_id": req.trendyol_category_id,
                "trendyol_category_name": req.trendyol_category_name
            }}
        )
        return {"success": True, "message": "Eşleştirme kaydedildi"}
    except Exception as e:
        logger.error(f"Error saving category mapping: {e}")
        raise HTTPException(status_code=500, detail="Eşleştirme kaydedilemedi.")
@router.post("/trendyol/category-mappings/{local_category_id}/value-mappings")
async def save_trendyol_category_value_mappings(local_category_id: str, req: Request, current_user: dict = Depends(require_admin)):
    payload = await req.json()
    value_mappings = payload.get("value_mappings", {})
    from bson.objectid import ObjectId
    filter_q = {"id": local_category_id} if len(local_category_id) > 24 else {"$or": [{"id": local_category_id}, {"_id": ObjectId(local_category_id)}]}
    
    await db.categories.update_one(
        filter_q,
        {"$set": {"value_mappings": value_mappings}}
    )
    return {"success": True}
@router.get("/trendyol/category-values/{local_category_id}")
async def get_local_category_values(local_category_id: str, current_user: dict = Depends(require_admin)):
    from bson.objectid import ObjectId
    import re
    
    filter_q = {"id": local_category_id} if len(local_category_id) > 24 else {"$or": [{"id": local_category_id}, {"_id": ObjectId(local_category_id)}]}
    category = await db.categories.find_one(filter_q)
    
    if not category:
         raise HTTPException(status_code=404, detail="Kategori bulunamadı")
         
    # Case-insensitive category name search
    name = category.get("name")
    cat_id = category.get("id") or str(category.get("_id"))
    
    query = {
        "$or": [
            {"category_name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}},
            {"category_id": cat_id}
        ]
    }
    
    products = await db.products.find(query).to_list(None)
    
    val_map = {
        "Renk": set(),
        "Beden": set(),
        "Boy": set()
    }
    
    for p in products:
        # Pull from variants (standard for clothing)
        for v in p.get("variants", []):
            if v.get("color"):
                c = str(v["color"]).strip()
                if c and c.lower() != "none":
                    val_map["Renk"].add(c)
            if v.get("size"): 
                s = str(v["size"]).strip()
                if s and s.lower() != "none":
                    val_map["Beden"].add(s)
        
        # Pull from attributes array (from CSV imports or manual entry)
        for a in p.get("attributes", []):
            t = str(a.get("type", "")).strip()
            val = str(a.get("value", "")).strip()
            if t and val and val.lower() != "none":
                if t not in val_map:
                    val_map[t] = set()
                val_map[t].add(val)
                
    result = []
    for k, v in val_map.items():
        if v:
            # Sort values naturally if possible
            sorted_vals = sorted(list(v))
            result.append({"attribute_name": k, "values": sorted_vals})
            
    return {"success": True, "local_values": result}
@router.delete("/trendyol/category-mappings/{local_category_id}")
async def delete_trendyol_category_mapping(local_category_id: str, current_user: dict = Depends(require_admin)):
    """Hide a category from the Trendyol mappings list and clear its mapping"""
    from bson.objectid import ObjectId
    try:
        filter_q = {"id": local_category_id} if len(local_category_id) > 24 else {"$or": [{"id": local_category_id}, {"_id": ObjectId(local_category_id)}]}
        
        await db.categories.update_one(
            filter_q,
            {
                "$unset": {
                    "trendyol_category_id": "",
                    "trendyol_category_name": "",
                    "attribute_mappings": ""
                },
                "$set": {
                    "trendyol_hidden": True
                }
            }
        )
        return {"success": True, "message": "Kategori listeden kaldırıldı ve eşleştirmesi silindi"}
    except Exception as e:
        logger.error(f"Error hiding category mapping: {e}")
        raise HTTPException(status_code=500, detail="Eşleştirme silinemedi.")
@router.post("/trendyol/category-mappings/bulk-delete")
async def bulk_delete_trendyol_category_mappings(req: Request, current_user: dict = Depends(require_admin)):
    payload = await req.json()
    category_ids = payload.get("category_ids", [])
    if not category_ids:
        return {"success": True}
    from bson.objectid import ObjectId
    try:
        flat_filters = []
        for cid in category_ids:
            cid_str = str(cid)
            flat_filters.append({"id": cid_str})
            if len(cid_str) <= 24:
                try:
                    flat_filters.append({"_id": ObjectId(cid_str)})
                except Exception:
                    pass

        if not flat_filters:
            return {"success": True}

        await db.categories.update_many(
            {"$or": flat_filters},
            {
                "$unset": {
                    "trendyol_category_id": "",
                    "trendyol_category_name": "",
                    "attribute_mappings": ""
                },
                "$set": {
                    "trendyol_hidden": True
                }
            }
        )
        return {"success": True, "message": "Seçili kategoriler kaldırıldı"}
    except Exception as e:
        logger.error(f"Error bulk hiding categories: {e}")
        raise HTTPException(status_code=500, detail="Kategoriler silinemedi.")
@router.post("/trendyol/category-mappings/{local_category_id}/attributes")
async def save_trendyol_attribute_mapping(local_category_id: str, req: AttributeMappingReq, current_user: dict = Depends(require_admin)):
    from bson.objectid import ObjectId
    try:
        filter_q = {"id": local_category_id} if len(local_category_id) > 24 else {"$or": [{"id": local_category_id}, {"_id": ObjectId(local_category_id)}]}
        
        mappings = [{"local_attr": m.local_attr, "trendyol_attr_id": m.trendyol_attr_id} for m in req.attribute_mappings]
        
        await db.categories.update_one(
            filter_q,
            {"$set": {
                "attribute_mappings": mappings,
                "default_mappings": req.default_mappings
            }}
        )
        return {"success": True, "message": "Özellik eşleştirmeleri kaydedildi"}
    except Exception as e:
        logger.error(f"Error saving attribute mapping: {e}")
        raise HTTPException(status_code=500, detail="Özellik eşleştirmeleri kaydedilemedi.")
@router.get("/trendyol/categories")
async def get_local_trendyol_categories(current_user: dict = Depends(require_admin)):
    """Fetch previously downloaded Trendyol categories for UI lists"""
    try:
        categories = await db.trendyol_categories.find({}, {"_id": 0, "id": 1, "name": 1, "subCategories": 1}).to_list(1000)
        # Flatten simple list for datalist mapping
        def flatten(cats, parent_name=""):
            result = []
            for c in cats:
                full_name = f"{parent_name} > {c['name']}" if parent_name else c["name"]
                result.append({"id": c["id"], "name": full_name})
                if c.get("subCategories"):
                    result.extend(flatten(c["subCategories"], full_name))
            return result
        flat_list = flatten(categories)
        return {"success": True, "categories": flat_list}
    except Exception as e:
        logger.error(f"Error fetching trendyol categories from db: {e}")
        return {"success": False, "categories": []}
@router.get("/trendyol/orders/label/{cargo_tracking_number}")
async def get_trendyol_cargo_label(cargo_tracking_number: str, current_user: dict = Depends(require_admin)):
    """Fetch PDF Cargo label from Trendyol"""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")
        
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
    from trendyol_client import TrendyolClient
    from fastapi.responses import Response
    
    try:
        client = TrendyolClient(
            supplier_id=config["supplier_id"],
            api_key=config["api_key"],
            api_secret=config["api_secret"],
            mode=config["mode"]
        )
        pdf_bytes = await client.get_cargo_label(cargo_tracking_number)
        return Response(content=pdf_bytes, media_type="application/pdf")
    except Exception as e:
        logger.error(f"Error fetching Trendyol cargo label: {str(e)}")
        raise HTTPException(status_code=500, detail="Kargo etiketi alınırken hata oluştu.")
async def _order_derived_trendyol_returns(search: str = "", claim_type: str = "",
                                          exclude_order_numbers=None):
    """İade durumundaki Trendyol siparişlerinden — senkron claim'i OLMAYANLARI —
    iade satırına çevirir. Tamamen read-side; satırlar `manual=True` işaretlenir."""
    if claim_type == "CANCEL":
        return []  # manuel iade satırları RETURN tipidir; iptal sekmesine girmez
    exclude = exclude_order_numbers or set()
    q = {"platform": "trendyol", "status": {"$in": _RETURN_STATUS_KEYS}}
    proj = {
        "_id": 0, "id": 1, "order_number": 1, "status": 1, "items": 1,
        "shipping_address": 1, "billing_address": 1, "customer_name": 1, "full_name": 1,
        "total": 1, "subtotal": 1, "invoice_number": 1, "created_at": 1, "updated_at": 1,
        "payment_method": 1, "return_request": 1, "cargo_tracking_number": 1,
        "cargo_provider_name": 1,
    }
    _rx = re.compile(_search_tr_regex(search.strip()), re.IGNORECASE) if search else None
    out = []
    async for o in db.orders.find(q, proj).sort("updated_at", -1):
        onum = str(o.get("order_number") or "")
        if onum and onum in exclude:
            continue  # Trendyol senkronundan gerçek claim zaten geldi → tekrarlama
        addr = o.get("shipping_address") or {}
        bill = o.get("billing_address") or {}
        name = (" ".join([addr.get("first_name") or "", addr.get("last_name") or ""]).strip()
                or addr.get("full_name") or o.get("customer_name") or o.get("full_name")
                or " ".join([bill.get("first_name") or "", bill.get("last_name") or ""]).strip()
                or "—")
        st = o.get("status") or ""
        rr = o.get("return_request") or {}
        items = []
        for it in (o.get("items") or []):
            up = float(it.get("unit_price") or it.get("list_price") or it.get("price") or 0)
            pr = float(it.get("price") or up or 0)
            dc = float(it.get("discount_amount") or it.get("discount") or 0)
            _nm = it.get("name") or it.get("product_name") or "Ürün"
            items.append({
                "claim_item_id": "", "productName": _nm, "product_name": _nm,
                "barcode": it.get("barcode") or it.get("product_id") or it.get("sku") or "",
                "size": it.get("size", ""), "color": it.get("color", ""),
                "quantity": int(it.get("quantity", 1) or 1),
                "unit_price": up, "price": pr, "discount_amount": dc, "reason": "",
            })
        net = float(o.get("total") or 0) or sum(i["price"] for i in items)
        row = {
            "claim_id": "ord:" + str(o.get("id") or onum),
            "order_id": o.get("id"), "order_number": onum,
            "claim_type": "RETURN", "claim_reason": rr.get("reason") or "",
            "claim_status": st, "order_status": st,
            "manual": True, "source": "order_status",
            "customer_name": name,
            "created_date": o.get("updated_at") or o.get("created_at") or "",
            "items": items, "refund_amount": net,
            "invoice_number": str(o.get("invoice_number") or ""),
            "cargo_tracking_number": str(o.get("cargo_tracking_number") or ""),
            "cargo_provider_name": o.get("cargo_provider_name") or "",
            "payment_type": _order_payment_type(o.get("payment_method")),
        }
        if _rx is not None:
            hay = " ".join([onum, name, row["invoice_number"], row["cargo_tracking_number"],
                            " ".join(i["productName"] for i in items)])
            if not _rx.search(hay):
                continue
        out.append(row)
    # GP bilgisini ekle: bu siparişlere ait customer_returns köprü kaydı varsa
    # (gider pusulası kesildiyse) numarayı/işaretini satıra taşı (tek toplu sorgu).
    _oids = [r["order_id"] for r in out if r.get("order_id")]
    if _oids:
        _gp = {}
        async for cr in db.customer_returns.find(
            {"order_id": {"$in": _oids}},
            {"_id": 0, "order_id": 1, "id": 1, "has_gider_pusulasi": 1, "gider_pusulasi_no": 1}
        ).sort("created_at", -1):
            _gp.setdefault(cr.get("order_id"), cr)
        for r in out:
            _c = _gp.get(r.get("order_id"))
            if _c:
                r["return_id"] = _c.get("id")
                r["has_gider_pusulasi"] = bool(_c.get("has_gider_pusulasi"))
                r["gider_pusulasi_no"] = _c.get("gider_pusulasi_no") or ""
    return out
async def _sync_trendyol_claims_core(days_back: int = 1095):
    """Trendyol claims çekirdek senkron — endpoint + scheduler ortak kullanır.

    days_back: geçmiş tarama penceresi (ilk backfill 1095=3yıl; scheduler kısa pencere).
    """
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    from trendyol_client import TrendyolClient

    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    end_date = datetime.now(timezone.utc)
    total_synced = 0
    order_cache = {}  # order_number -> order_data cache

    # Trendyol max 15 günlük aralık destekliyor, parçalıyoruz
    chunk_days = 15
    current_end = end_date

    while True:
        current_start = current_end - timedelta(days=chunk_days)
        days_elapsed = (end_date - current_start).days

        if days_elapsed > days_back:
            current_start = end_date - timedelta(days=days_back)

        start_ts = int(current_start.timestamp() * 1000)
        end_ts = int(current_end.timestamp() * 1000)

        current_page = 0
        page_size = 200

        while True:
            try:
                url = f"{client.base_url}/order/sellers/{client.supplier_id}/claims"
                params = {
                    "page": current_page,
                    "size": page_size,
                    "startDate": start_ts,
                    "endDate": end_ts,
                }
                async with httpx.AsyncClient(timeout=30.0) as http_client:
                    headers = client._get_headers()
                    response = await http_client.get(url, headers=headers, params=params)
                    response.raise_for_status()
                    result = response.json()
            except Exception as e:
                logger.error(f"Claims sync error: {str(e)}")
                break

            data = result if isinstance(result, dict) else {}
            content = data.get("content", [])
            total_pages = data.get("totalPages", 0)

            if not content:
                break

            for claim in content:
                claim_id = str(claim.get("claimId", claim.get("id", "")))
                if not claim_id:
                    continue

                # Mevcut claim: pahalı order API'sini TEKRAR ÇEKME (iskonto zaten kayıtlı),
                # ama DURUM/kargo/tarih alanlarını canlı tazele — durum geçişleri (Created→
                # Accepted/Rejected) yansısın. Trendyol kayıtları lastModifiedDate sırasıyla gelir.
                existing_claim = await db.trendyol_claims.find_one(
                    {"claim_id": claim_id},
                    {"_id": 1, "return_approved_at": 1, "return_rejected_at": 1, "manual_locked": 1}
                )
                if existing_claim:
                    _new_status = _derive_claim_status(claim)
                    _lm = claim.get("lastModifiedDate")
                    _lm_iso = ""
                    if _lm:
                        try:
                            _lm_iso = datetime.fromtimestamp(_lm / 1000, tz=timezone.utc).isoformat()
                        except Exception:
                            _lm_iso = ""
                    _now_iso = datetime.now(timezone.utc).isoformat()
                    _set = {
                        "cargo_tracking_number": str(claim.get("cargoTrackingNumber", "")),
                        "cargo_provider_name": claim.get("cargoProviderName", ""),
                        "raw_data": claim,
                        "updated_at": _now_iso,
                    }
                    # MANUEL KİLİT: admin durumu elle ilerlettiyse (manual_locked) Trendyol senkronu
                    # claim_status'u EZMEZ — yalnız kargo/raw bilgisi tazelenir, durum manuel kalır.
                    if not existing_claim.get("manual_locked"):
                        _set["claim_status"] = _new_status
                        # Durum geçiş tarih damgaları (idempotent — yalnız ilk geçişte yazılır)
                        if _new_status == "Accepted" and not existing_claim.get("return_approved_at"):
                            _set["return_approved_at"] = _lm_iso or _now_iso
                        if _new_status in ("Rejected", "Cancelled") and not existing_claim.get("return_rejected_at"):
                            _set["return_rejected_at"] = _lm_iso or _now_iso
                    await db.trendyol_claims.update_one({"claim_id": claim_id}, {"$set": _set})
                    total_synced += 1
                    continue

                # Claim items'dan tip ve sebep çıkar
                claim_items = []
                claim_type = ""
                claim_reason = ""
                refund_amount = 0
                
                # Claims API'sinde iskonto bilgisi yok. Sipariş API'sinden çek.
                order_number = str(claim.get("orderNumber", ""))
                order_discount_map = {}  # barcode -> {discount, gross_price, net_price}
                if order_number:
                    # Cache kontrolü: aynı sipariş numarasını tekrar çekme
                    cache_key = f"order_{order_number}"
                    if cache_key not in order_cache:
                        try:
                            order_data = await client.get_orders(order_number=order_number)
                            order_cache[cache_key] = order_data
                        except Exception as e:
                            logger.warning(f"Could not fetch order {order_number} for discount: {e}")
                            order_cache[cache_key] = {}
                    
                    cached = order_cache.get(cache_key, {})
                    for pkg in cached.get("content", []):
                        for line in pkg.get("lines", []):
                            bc = line.get("barcode", "")
                            line_gross = line.get("lineGrossAmount", line.get("amount", 0))
                            line_net = line.get("price", 0)
                            line_disc = line.get("discount", 0)
                            qty = max(line.get("quantity", 1), 1)
                            if bc:
                                order_discount_map[bc] = {
                                    "gross": line_gross / qty if line_gross else 0,
                                    "net": line_net / qty if line_net else 0,
                                    "discount": line_disc / qty if line_disc else 0,
                                }

                for item in claim.get("items", []):
                    order_line = item.get("orderLine", {})
                    for ci in item.get("claimItems", []):
                        reason_info = ci.get("customerClaimItemReason", {})
                        if not claim_type:
                            code = reason_info.get("code", "").upper()
                            if code in ["ABANDON", "UNDELIVERED", "NOTDELIVERED"]:
                                claim_type = "CANCEL"
                            else:
                                claim_type = "RETURN"
                        if not claim_reason:
                            claim_reason = reason_info.get("name", "")

                        barcode = order_line.get("barcode", "")
                        claim_price = order_line.get("price", 0)
                        
                        # İskontoyu sipariş verisinden al
                        order_info = order_discount_map.get(barcode, {})
                        if order_info:
                            gross_price = order_info.get("gross", claim_price)
                            net_price = order_info.get("net", claim_price)
                            discount = order_info.get("discount", 0)
                        else:
                            # Fallback: Claims API verisini kullan (iskonto yok)
                            gross_price = claim_price
                            net_price = claim_price
                            discount = 0
                        
                        claim_items.append({
                            "claim_item_id": str(ci.get("id", "")),
                            "productName": order_line.get("productName", ""),
                            "barcode": barcode,
                            "unit_price": gross_price,
                            "discount_amount": discount,
                            "price": net_price,
                            "quantity": 1,
                            "reason": reason_info.get("name", "")
                        })
                        refund_amount += net_price

                # Tarih formatı
                claim_date = claim.get("claimDate")
                created_date_str = ""
                if claim_date:
                    try:
                        created_date_str = datetime.fromtimestamp(claim_date / 1000, tz=timezone.utc).isoformat()
                    except Exception:
                        created_date_str = str(claim_date)

                # Fatura numarasını çıkar: sipariş verisinden veya claim'den
                invoice_number = ""
                for item in claim.get("items", []):
                    ol = item.get("orderLine", {})
                    inv = ol.get("invoiceNumber", "") or item.get("invoiceNumber", "")
                    if inv:
                        invoice_number = str(inv)
                        break
                if not invoice_number:
                    invoice_number = str(claim.get("invoiceNumber", "") or "")
                # Sipariş verisinden fatura no çek
                if not invoice_number and order_discount_map:
                    try:
                        _order_data = await client.get_orders(order_number=order_number)
                        for pkg in _order_data.get("content", []):
                            inv_no = pkg.get("invoiceNumber", "")
                            if inv_no:
                                invoice_number = str(inv_no)
                                break
                    except Exception:
                        pass

                claim_doc = {
                    "claim_id": claim_id,
                    "order_number": order_number,
                    "claim_type": claim_type,
                    "claim_reason": claim_reason,
                    "claim_status": _derive_claim_status(claim),
                    **_first_seen_stamps(claim),
                    "customer_name": f"{claim.get('customerFirstName', '')} {claim.get('customerLastName', '')}".strip(),
                    "created_date": created_date_str,
                    "items": claim_items,
                    "refund_amount": refund_amount,
                    "invoice_number": invoice_number,
                    "invoice_link": claim.get("invoiceLink", ""), # Yeni eklendi
                    "cargo_tracking_number": str(claim.get("cargoTrackingNumber", "")),
                    "cargo_provider_name": claim.get("cargoProviderName", ""),
                    "raw_data": claim,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }

                await db.trendyol_claims.update_one(
                    {"claim_id": claim_id},
                    {"$set": claim_doc, "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": datetime.now(timezone.utc).isoformat()}},
                    upsert=True
                )
                total_synced += 1

                # Claim sebebini eşleşen İPTAL siparişine bağla. claim_type sınıflandırması
                # (CANCEL/RETURN) Trendyol sebep koduna göre dar kalabildiğinden ONA GÜVENMEYİZ:
                # sipariş zaten status=cancelled ise (iade post-teslimat olduğundan iptalle
                # karışmaz) o claim'in sebebi = iptal sebebidir. Böylece "Stok tükendi",
                # "Adreste bulunmayacağım" gibi gerçek sebepler İptaller'e yazılır.
                if claim_reason and order_number:
                    try:
                        await db.orders.update_one(
                            {"order_number": str(order_number), "platform": "trendyol", "status": "cancelled"},
                            {"$set": {"cancel_reason": claim_reason, "cancel_source": "trendyol",
                                      "updated_at": datetime.now(timezone.utc).isoformat()}},
                        )
                    except Exception as _ce:
                        logger.error(f"[trendyol claim->order reason {order_number}] {_ce}")

            current_page += 1
            if current_page >= total_pages:
                break

        current_end = current_start
        if days_elapsed >= days_back:
            break

    return {
        "message": f"Son {days_back} gündeki toplam {total_synced} iade/iptal kaydı senkronize edildi",
        "total_synced": total_synced,
        "days_back": days_back
    }
@router.get("/trendyol/claims/sync")
async def sync_trendyol_claims(
    days_back: int = 1095,
    current_user: dict = Depends(require_admin)
):
    """Trendyol'dan iade/iptal (claim) kayıtlarını çeker ve MongoDB'ye kaydeder.

    days_back varsayılan 1095 (3 yıl) — geçmiş backfill için. UI bu ucu tetikler;
    durum geçişleri her çağrıda canlı tazelenir.
    """
    return await _sync_trendyol_claims_core(days_back)
@router.post("/trendyol/claims/fix-discounts")
async def fix_claim_discounts(current_user: dict = Depends(require_admin)):
    """Fix discount data for existing claims by fetching from order API"""
    settings = await db.settings.find_one({"id": "trendyol"}, {"_id": 0})
    if not settings or not settings.get("api_key"):
        raise HTTPException(status_code=400, detail="Trendyol API ayarları eksik")
    
    from trendyol_client import TrendyolClient
    client = TrendyolClient(settings["supplier_id"], settings["api_key"], settings["api_secret"])
    
    # Get claims that need discount fix (where items have 0 discount)
    claims = await db.trendyol_claims.find({}, {"_id": 0, "claim_id": 1, "order_number": 1, "items": 1}).to_list(None)
    
    order_cache = {}
    fixed = 0
    
    for claim in claims:
        order_number = claim.get("order_number", "")
        if not order_number:
            continue
        
        items = claim.get("items", [])
        needs_fix = any(item.get("discount_amount", 0) == 0 and item.get("unit_price", 0) == item.get("price", 0) for item in items)
        if not needs_fix:
            continue
        
        # Get order data (cached)
        if order_number not in order_cache:
            try:
                order_cache[order_number] = await client.get_orders(order_number=order_number)
            except Exception:
                order_cache[order_number] = {}
        
        cached = order_cache.get(order_number, {})
        discount_map = {}
        invoice_number = ""
        for pkg in cached.get("content", []):
            if not invoice_number:
                invoice_number = pkg.get("invoiceNumber", "")
            for line in pkg.get("lines", []):
                bc = line.get("barcode", "")
                qty = max(line.get("quantity", 1), 1)
                if bc:
                    discount_map[bc] = {
                        "gross": (line.get("lineGrossAmount", line.get("amount", 0)) or 0) / qty,
                        "net": (line.get("price", 0) or 0) / qty,
                        "discount": (line.get("discount", 0) or 0) / qty,
                    }
        
        updated_items = []
        refund_amount = 0
        for item in items:
            bc = item.get("barcode", "")
            if bc in discount_map:
                item["unit_price"] = discount_map[bc]["gross"]
                item["discount_amount"] = discount_map[bc]["discount"]
                item["price"] = discount_map[bc]["net"]
            refund_amount += item.get("price", 0)
            updated_items.append(item)
        
        update_set = {"items": updated_items, "refund_amount": refund_amount}
        if invoice_number:
            update_set["invoice_number"] = invoice_number
        
        await db.trendyol_claims.update_one(
            {"claim_id": claim["claim_id"]},
            {"$set": update_set}
        )
        fixed += 1
    
    return {"success": True, "fixed": fixed, "message": f"{fixed} iadenin iskonto bilgisi güncellendi"}
@router.get("/trendyol/claims")
async def get_trendyol_claims(
    page: int = 1,
    limit: int = 20,
    claim_type: str = "",
    search: str = "",
    status: str = "",
    platform: str = "trendyol",
    current_user: dict = Depends(require_admin)
):
    """Yerel veritabanındaki iade kayıtlarını listele.

    Tüm claim'ler İADE'dir (İptal ayrı menüde, sipariş durumu üzerinden).
    `status` = durum sekmesi anahtarı: all / talep_olusturulan / kargoya_verilen /
    aksiyon_bekleyen / onaylanan / reddedilen.
    `platform` = 'facette' (web/site iadeleri) veya bir pazaryeri anahtarı
    (trendyol, hepsiburada, ...). Ayrım `_claim_is_site_order` kuralıyla yapılır.
    """
    # Sekme kovası eşlemesi artık _claim_bucket(c) helper'ında (status + kargo takip durumu).
    _VALID_TABS = {"talep_olusturulan", "kargoya_verilen", "acik_iade", "aksiyon_bekleyen", "onaylanan", "reddedilen"}

    base_query = {}
    if claim_type:
        base_query["claim_type"] = claim_type
    if search:
        _s = search.strip()
        _rx = {"$regex": _search_tr_regex(_s), "$options": "i"}
        _ors = [
            {"order_number": _rx},
            {"customer_name": _rx},
            {"claim_id": _rx},
            {"invoice_number": _rx},
            {"cargo_tracking_number": _rx},
            {"cargo_provider_name": _rx},
            {"claim_reason": _rx},
            {"items.productName": _rx},
            {"items.product_name": _rx},
            {"items.barcode": _rx},
            {"items.merchantSku": _rx},
        ]
        # Telefon: sadece rakam → son 10 hane (kayıtta varsa)
        _digits = re.sub(r"\D", "", _s)
        if len(_digits) >= 7:
            _ors.append({"customer_phone": {"$regex": re.escape(_digits[-10:])}})
        # Çok kelimeli ad: tüm kelimeler customer_name içinde geçsin (sıra önemsiz)
        _words = [w for w in _s.split() if len(w) >= 2]
        if len(_words) >= 2:
            _ors.append({"$and": [
                {"customer_name": {"$regex": _search_tr_regex(w), "$options": "i"}}
                for w in _words
            ]})
        base_query["$or"] = _ors

    want_tab = status if (status and status != "all" and status in _VALID_TABS) else None

    # base_query (arama/tip filtreli, AMA status filtresiz) ile TÜM kayıtları çek.
    # Status filtresi bellekte uygulanır ki tab_counts tüm kovaları doğru sayabilsin.
    raw = await db.trendyol_claims.find(
        base_query, {"_id": 0, "raw_data": 0}
    ).sort("created_date", -1).to_list(None)

    # (a) claim_id'ye göre tekilleştir — aynı claim birden fazla belge olarak yazılmışsa
    # en güncel (created_date'e göre zaten sıralı) ilk görüleni tut.
    seen = set()
    deduped = []
    for c in raw:
        cid = c.get("claim_id") or c.get("order_number")
        if cid in seen:
            continue
        seen.add(cid)
        deduped.append(c)

    # Sipariş-durumu köprüsü: Trendyol'da claim'i OLMAYAN ama elle iade durumuna
    # alınmış Trendyol siparişlerini de ekle (Web Sitesi/Rooftr deseninin aynası).
    _seen_orders = {c.get("order_number") for c in deduped if c.get("order_number")}
    _manual_rows = await _order_derived_trendyol_returns(
        search=search, claim_type=claim_type, exclude_order_numbers=_seen_orders)
    deduped = deduped + _manual_rows

    # (b) trendyol_claims koleksiyonu TAMAMEN pazaryeri (Trendyol) kayitlaridir.
    # Site iadeleri ayri koleksiyonda (customer_returns) ve ayri sekmede (Web Sitesi)
    # gosterilir. Bu yuzden burada site/pazaryeri ayrimi YAPILMAZ; mikro ihracat dahil
    # TUM claim'ler dondurulur. (`platform` parametresi geriye-donuk uyumluluk icin
    # kabul edilir ama bu endpoint'te artik filtreleme yapmaz.)
    platform_scoped = deduped

    # İptal (Cancelled iade statüsü) bu iade ekranından TAMAMEN dışlanır; iptaller
    # ayrı bir alandan yönetilir. Böylece "Tüm İadeler" sekmesi ve "Toplam İade"
    # kartı aynı evreni (iptal-hariç tekil iade) sayar.
    iade_scoped = [c for c in platform_scoped if _claim_bucket(c) != "iptal"]
    iade_scoped.sort(key=lambda c: (c.get("created_date") or ""), reverse=True)

    # (c) status sekmesi filtresi (bellekte) — _claim_bucket ile
    if want_tab == "acik_iade":
        filtered = [c for c in iade_scoped if _claim_bucket(c) in ("talep_olusturulan", "kargoya_verilen")]
    elif want_tab is not None:
        filtered = [c for c in iade_scoped if _claim_bucket(c) == want_tab]
    else:
        filtered = iade_scoped

    total = len(filtered)
    skip = (page - 1) * limit
    claims = filtered[skip: skip + limit]

    # Her claim'e kova + Türkçe durum etiketi ekle (frontend durum rozeti için).
    # talep/kargoda tek "Açık İade" altında birleşir; geçmiş kovalar korunur.
    _BUCKET_LABEL = {
        "talep_olusturulan": "Açık İade",
        "kargoya_verilen": "Açık İade",
        "aksiyon_bekleyen": "Aksiyon Bekleyen",
        "onaylanan": "Onaylandı",
        "reddedilen": "Reddedildi",
        "iptal": "İptal",
    }
    for c in claims:
        _b = _claim_bucket(c)
        c["bucket"] = _b
        c["bucket_label"] = _BUCKET_LABEL.get(_b, "—")
        if c.get("manual") and c.get("order_status"):
            c["bucket_label"] = _ORDER_STATUS_TR.get(c.get("order_status"), c["bucket_label"])

    # Sekme adetleri — iade_scoped (iptal hariç) üzerinden, _claim_bucket ile.
    _bcount = {"talep_olusturulan": 0, "kargoya_verilen": 0, "aksiyon_bekleyen": 0, "onaylanan": 0, "reddedilen": 0}
    for c in iade_scoped:
        _b = _claim_bucket(c)
        if _b in _bcount:
            _bcount[_b] += 1
    tab_counts = {"all": len(iade_scoped), **_bcount, "acik_iade": _bcount["talep_olusturulan"] + _bcount["kargoya_verilen"]}

    # İstatistikler — "Toplam İade" kartı = "Tüm İadeler" sekmesi (iptal hariç tekil iade).
    total_returns = len(iade_scoped)
    total_cancels = await db.trendyol_claims.count_documents({"claim_type": "CANCEL"})
    total_refund = sum((c.get("refund_amount") or 0) for c in iade_scoped)

    return {
        "claims": claims,
        "total": total,
        "page": page,
        "limit": limit,
        "tab_counts": tab_counts,
        "stats": {
            "total_returns": total_returns,
            "total_cancels": total_cancels,
            "total_refund": total_refund
        }
    }
@router.get("/trendyol/claims/export")
async def export_trendyol_claims(
    status: str = "",
    search: str = "",
    current_user: dict = Depends(require_admin),
):
    """Trendyol iadelerini bulunulan sekme (status) + arama filtresiyle Excel'e aktarır.
    Liste endpoint'iyle (get_trendyol_claims) AYNI dedup + kova mantığını kullanır;
    böylece hangi sekmedeyse o sekmenin kayıtları döner.
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    _VALID_TABS = {"talep_olusturulan", "kargoya_verilen", "acik_iade", "aksiyon_bekleyen", "onaylanan", "reddedilen"}
    base_query = {}
    if search:
        _rx = {"$regex": _search_tr_regex(search.strip()), "$options": "i"}
        base_query["$or"] = [
            {"order_number": _rx},
            {"customer_name": _rx},
            {"claim_id": _rx},
            {"invoice_number": _rx},
            {"cargo_tracking_number": _rx},
            {"cargo_provider_name": _rx},
            {"items.productName": _rx},
            {"items.product_name": _rx},
            {"items.barcode": _rx},
        ]
    want_tab = status if (status and status != "all" and status in _VALID_TABS) else None

    raw = await db.trendyol_claims.find(base_query, {"_id": 0, "raw_data": 0}).sort("created_date", -1).to_list(None)
    seen = set(); deduped = []
    for c in raw:
        cid = c.get("claim_id") or c.get("order_number")
        if cid in seen:
            continue
        seen.add(cid); deduped.append(c)
    _seen_orders = {c.get("order_number") for c in deduped if c.get("order_number")}
    _manual_rows = await _order_derived_trendyol_returns(search=search, exclude_order_numbers=_seen_orders)
    deduped = deduped + _manual_rows
    iade_scoped = [c for c in deduped if _claim_bucket(c) != "iptal"]
    iade_scoped.sort(key=lambda c: (c.get("created_date") or ""), reverse=True)
    if want_tab == "acik_iade":
        rows = [c for c in iade_scoped if _claim_bucket(c) in ("talep_olusturulan", "kargoya_verilen")]
    elif want_tab is not None:
        rows = [c for c in iade_scoped if _claim_bucket(c) == want_tab]
    else:
        rows = iade_scoped

    _BUCKET_LABEL = {
        "talep_olusturulan": "Açık İade", "kargoya_verilen": "Açık İade",
        "aksiyon_bekleyen": "Aksiyon Bekleyen", "onaylanan": "Onaylandı",
        "reddedilen": "Reddedildi", "iptal": "İptal",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Trendyol İadeleri"
    headers = ["Sipariş No", "Müşteri", "Ürün", "Tutar", "Tarih", "Durum", "Gider Pusulası No"]
    ws.append(headers)
    hfill = PatternFill("solid", fgColor="FCE4B6")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for c in rows:
        items = c.get("items") or []
        urun = ", ".join([
            (i.get("product_name") or i.get("name") or "").strip()
            for i in items if (i.get("product_name") or i.get("name"))
        ])
        b = _claim_bucket(c)
        _dlabel = _ORDER_STATUS_TR.get(c.get("order_status")) if c.get("manual") else None
        ws.append([
            c.get("order_number") or "",
            c.get("customer_name") or "",
            urun,
            float(c.get("refund_amount") or 0),
            str(c.get("created_date") or "")[:10],
            _dlabel or _BUCKET_LABEL.get(b, "—"),
            c.get("gider_pusulasi_no") or "",
        ])

    for col in ws.columns:
        ml = max((len(str(cc.value)) for cc in col if cc.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 2, 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"trendyol-iadeleri-{status or 'tum'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
@router.get("/trendyol/claims/diagnostics")
async def trendyol_claims_diagnostics(current_user: dict = Depends(require_admin)):
    """Kova kalibrasyonu teşhisi: ham status dağılımı + kova eşlemesi + kargo kırılımı.

    Panel'deki Trendyol sekme sayılarını (talep/kargoda/aksiyon/onay/ret) gerçek veriyle
    karşılaştırıp _claim_bucket eşlemesini doğrulamak/ayarlamak için kullanılır.
    Backfill bitince çağır; by_bucket panel sayılarıyla tutmazsa by_status_raw'a bakıp
    _claim_bucket tek noktadan düzeltilir.
    """
    rows = await db.trendyol_claims.find(
        {}, {"_id": 0, "claim_status": 1, "cargo_tracking_number": 1, "claim_id": 1}
    ).to_list(None)
    seen = set()
    uniq = []
    for c in rows:
        cid = c.get("claim_id")
        if cid in seen:
            continue
        seen.add(cid)
        uniq.append(c)

    by_status = {}
    by_bucket = {"talep_olusturulan": 0, "kargoya_verilen": 0, "aksiyon_bekleyen": 0, "onaylanan": 0, "reddedilen": 0, "iptal": 0}
    created_with_cargo = 0
    created_without_cargo = 0
    for c in uniq:
        st = (c.get("claim_status") or "").strip() or "(boş)"
        by_status[st] = by_status.get(st, 0) + 1
        b = _claim_bucket(c)
        if b in by_bucket:
            by_bucket[b] += 1
        if (c.get("claim_status") or "").strip() == "Created":
            if str(c.get("cargo_tracking_number") or "").strip():
                created_with_cargo += 1
            else:
                created_without_cargo += 1

    return {
        "total_unique_claims": len(uniq),
        "by_status_raw": by_status,
        "by_bucket": by_bucket,
        "created_with_cargo": created_with_cargo,
        "created_without_cargo": created_without_cargo,
        "expected_from_panel": {
            "talep_olusturulan": 34, "kargoya_verilen": 54,
            "aksiyon_bekleyen": 16, "onaylanan": 3583, "reddedilen": 49,
        },
        "note": "by_bucket panel ile tutmazsa _claim_bucket eşlemesi by_status_raw'a göre ayarlanır.",
    }
@router.post("/trendyol/claims/repair-status")
async def repair_trendyol_claim_status(current_user: dict = Depends(require_admin)):
    """Mevcut TÜM claim kayıtlarının statüsünü KAYITLI raw_data'dan yeniden türetir.

    Eski sync, claim-level olmayan `status` alanını okuduğu için kayıtlar statüsüz
    yazılmıştı. Bu uç Trendyol API'sine YENİ istek atmadan raw_data'daki gerçek item
    statülerinden (claimItemStatus.name) claim_status'ü, kargo alanlarını ve onay/ret
    tarih damgalarını yeniden hesaplar. İdempotent — tekrar çağrılabilir.
    """
    cursor = db.trendyol_claims.find(
        {}, {"_id": 0, "claim_id": 1, "raw_data": 1, "return_approved_at": 1, "return_rejected_at": 1}
    )
    scanned = 0
    fixed = 0
    bucket_after = {"talep_olusturulan": 0, "kargoya_verilen": 0, "aksiyon_bekleyen": 0, "onaylanan": 0, "reddedilen": 0, "iptal": 0}
    async for c in cursor:
        scanned += 1
        raw = c.get("raw_data") or {}
        if not raw:
            continue
        new_status = _derive_claim_status(raw)
        cargo_no = str(raw.get("cargoTrackingNumber", "") or "")
        _set = {
            "claim_status": new_status,
            "cargo_tracking_number": cargo_no,
            "cargo_provider_name": raw.get("cargoProviderName", "") or "",
        }
        stamps = _first_seen_stamps(raw)
        if stamps.get("return_approved_at") and not c.get("return_approved_at"):
            _set["return_approved_at"] = stamps["return_approved_at"]
        if stamps.get("return_rejected_at") and not c.get("return_rejected_at"):
            _set["return_rejected_at"] = stamps["return_rejected_at"]
        await db.trendyol_claims.update_one({"claim_id": c["claim_id"]}, {"$set": _set})
        fixed += 1
        b = _claim_bucket({"claim_status": new_status, "cargo_tracking_number": cargo_no})
        if b in bucket_after:
            bucket_after[b] += 1
    return {"scanned": scanned, "fixed": fixed, "bucket_after": bucket_after}
@router.post("/trendyol/claims/dedupe")
async def dedupe_trendyol_claims(current_user: dict = Depends(require_admin)):
    """Aynı claim_id'ye sahip MÜKERRER belgeleri temizler (her claim_id için en güncel
    updated_at olanı tutar), sonra claim_id üzerinde unique index kurarak gelecekte
    mükerrerlenmeyi engeller. İdempotent — tekrar çağrılabilir.
    """
    pipeline = [
        {"$group": {"_id": "$claim_id", "count": {"$sum": 1}}},
        {"$match": {"count": {"$gt": 1}}},
    ]
    groups = await db.trendyol_claims.aggregate(pipeline).to_list(None)
    removed = 0
    affected = 0
    for g in groups:
        cid = g["_id"]
        docs = await db.trendyol_claims.find(
            {"claim_id": cid}, {"_id": 1, "updated_at": 1, "created_date": 1, "created_at": 1}
        ).to_list(None)
        if len(docs) <= 1:
            continue
        docs.sort(key=lambda d: (str(d.get("updated_at") or ""), str(d.get("created_date") or ""), str(d.get("created_at") or "")), reverse=True)
        to_delete = [d["_id"] for d in docs[1:]]
        if to_delete:
            res = await db.trendyol_claims.delete_many({"_id": {"$in": to_delete}})
            removed += res.deleted_count
            affected += 1
    index_created = False
    index_error = ""
    try:
        await db.trendyol_claims.create_index("claim_id", unique=True, name="uniq_claim_id")
        index_created = True
    except Exception as e:
        index_error = str(e)[:200]
    remaining = await db.trendyol_claims.count_documents({})
    return {
        "duplicate_groups": len(groups),
        "claims_affected": affected,
        "removed": removed,
        "index_created": index_created,
        "index_error": index_error,
        "remaining_docs": remaining,
    }
@router.post("/trendyol/claims/{claim_id}/set-status")
async def set_trendyol_claim_status(claim_id: str, payload: dict, current_user: dict = Depends(require_admin)):
    """Trendyol iade durumunu MANUEL ilerletir ve KİLİTLER.

    manual_locked=True olunca sonraki Trendyol senkronları bu claim'in claim_status'unu
    EZMEZ (yalnız kargo/raw bilgisi tazelenir). Kilidi kaldırmak için /unlock çağrılır.
    """
    new_status = (payload.get("status") or "").strip()
    valid = {"Created", "WaitingInAction", "InAnalysis", "Accepted", "Rejected", "Unresolved", "Cancelled"}
    if new_status not in valid:
        raise HTTPException(status_code=400, detail=f"Geçersiz durum. Geçerli: {', '.join(sorted(valid))}")
    now = datetime.now(timezone.utc).isoformat()
    _set = {"claim_status": new_status, "manual_locked": True, "manual_status_at": now, "updated_at": now}
    if new_status == "Accepted":
        _set["return_approved_at"] = now
    if new_status in ("Rejected", "Cancelled"):
        _set["return_rejected_at"] = now
    res = await db.trendyol_claims.update_one({"claim_id": claim_id}, {"$set": _set})
    if not res.matched_count:
        raise HTTPException(status_code=404, detail="İade (claim) bulunamadı")
    return {"success": True, "claim_id": claim_id, "status": new_status, "manual_locked": True}
@router.post("/trendyol/claims/{claim_id}/unlock")
async def unlock_trendyol_claim(claim_id: str, current_user: dict = Depends(require_admin)):
    """Manuel kilidi kaldırır → durum tekrar Trendyol senkronundan güncellenmeye başlar."""
    res = await db.trendyol_claims.update_one(
        {"claim_id": claim_id},
        {"$set": {"manual_locked": False, "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    if not res.matched_count:
        raise HTTPException(status_code=404, detail="İade (claim) bulunamadı")
    return {"success": True, "claim_id": claim_id, "manual_locked": False}
@router.get("/trendyol/claims/shipment-probe")
async def trendyol_shipment_probe(order_number: str = "", current_user: dict = Depends(require_admin)):
    """GEÇİCİ ARAŞTIRMA: Trendyol sipariş paketi servisini (getShipmentPackages) bir iadenin
    orderNumber'ı ile sorgular ve paket durum/satır alanlarını döndürür. Amaç: talep vs
    kargoya-verilen ayrımı için kullanılabilir bir 'shipped/returned' sinyali var mı görmek.
    """
    headers = await get_trendyol_headers()
    if not headers:
        raise HTTPException(status_code=400, detail="Trendyol kimliği yapılandırılmamış")
    config = await get_trendyol_config()
    base = config["base_url"]
    sid = config["supplier_id"]
    probed = []
    # order_number verilmezse birkaç açık (Created) claim üzerinde dene
    targets = []
    if order_number:
        targets = [order_number]
    else:
        async for c in db.trendyol_claims.find({"claim_status": "Created"}, {"_id": 0, "order_number": 1}).limit(3):
            on = c.get("order_number")
            if on:
                targets.append(on)
    for on in targets:
        url = f"{base}/sapigw/suppliers/{sid}/orders?orderNumber={on}"
        try:
            async with httpx.AsyncClient(timeout=30) as cx:
                r = await cx.get(url, headers=headers)
            try:
                data = r.json()
            except Exception:
                data = {"_text": r.text[:500]}
            pkgs = []
            for pkg in (data.get("content") or [])[:5]:
                pkgs.append({
                    "id": pkg.get("id"),
                    "status": pkg.get("status"),
                    "cargoTrackingNumber": pkg.get("cargoTrackingNumber"),
                    "lines_status": [l.get("orderLineItemStatusName") for l in (pkg.get("lines") or [])[:4]],
                    "top_keys": list(pkg.keys()),
                })
            probed.append({"order_number": on, "http": r.status_code, "package_count": len(data.get("content") or []), "packages": pkgs})
        except Exception as e:
            probed.append({"order_number": on, "error": str(e)[:200]})
    return {"base": base, "supplier_id": sid, "probed": probed}
@router.get("/trendyol/claims/issue-reasons")
async def get_trendyol_issue_reasons(current_user: dict = Depends(require_admin)):
    """Fetch claim issue reasons from Trendyol"""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    from trendyol_client import TrendyolClient
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    try:
        url = f"{client.base_url}/order/sellers/{client.supplier_id}/claim-issue-reasons"
        async with httpx.AsyncClient(timeout=30.0) as http_client:
            headers = client._get_headers()
            response = await http_client.get(url, headers=headers)
            response.raise_for_status()
            return response.json()
    except Exception as e:
        logger.error(f"Failed to fetch issue reasons: {str(e)}")
        # Return common fallback reasons
        return [
            {"id": 1, "name": "Kullanım Hatası / Tüketici Kaynaklı Hasar"},
            {"id": 2, "name": "Ürün Orijinal Kutusunda / Ambalajında Değil"},
            {"id": 4, "name": "Eksik Aksesuar / Parça"},
            {"id": 6, "name": "İade Süresi Geçmiş"},
            {"id": 21, "name": "Ürün Kullanılmış / Etiketi Koparılmış"}
        ]
@router.get("/trendyol/claims/{claim_id}")
async def get_trendyol_claim_detail(claim_id: str, current_user: dict = Depends(require_admin)):
    """Tek bir iade/iptal kaydının detayını getir."""
    claim = await db.trendyol_claims.find_one({"claim_id": claim_id}, {"_id": 0})
    if not claim:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    return claim
@router.post("/trendyol/claims/{claim_id}/gider-pusulasi")
async def generate_gider_pusulasi(claim_id: str, payload: Optional[dict] = Body(default=None), current_user: dict = Depends(require_admin)):
    """Generate expense receipt (gider pusulası) data for a return claim"""
    claim = await db.trendyol_claims.find_one({"claim_id": claim_id}, {"_id": 0})
    if not claim:
        raise HTTPException(status_code=404, detail="İade kaydı bulunamadı")

    # Kurumsal/e-Fatura siparişinde gider pusulası DÜZENLENEMEZ — iade faturası gerekir.
    _onum = claim.get("order_number", "")
    if _onum:
        _ord = await db.orders.find_one({"order_number": _onum}, {"_id": 0, "invoice_type": 1, "billing_info": 1})
        if _ord and ((_ord.get("invoice_type") == "e-fatura") or bool((_ord.get("billing_info") or {}).get("is_corporate"))):
            raise HTTPException(status_code=400, detail="Bu sipariş kurumsal/e-Fatura siparişi — gider pusulası düzenlenemez. Müşteriden iade faturası gerekir (Doğan'dan panele düşecek, onaylayınca stok +1).")

    # İade onayımız = gider pusulası oluşturmak. Stoğu BİR KEZ geri ekle (idempotent).
    await restock_claim_once(claim_id, "gider_pusulasi", current_user.get("email", ""))

    settings = await db.settings.find_one({"id": "main"}, {"_id": 0})
    company = settings.get("company_info", {}) if settings else {}
    # Alıcı adresi claim'de tutulmuyor; sipariş kaydından çek
    order = await db.orders.find_one({"order_number": claim.get("order_number", "")}, {"_id": 0}) or {}
    _ship = order.get("shipping_address", {}) or {}
    _cust_name = claim.get("customer_name", "") or (f"{_ship.get('first_name','')} {_ship.get('last_name','')}".strip())
    _cust_addr = _ship.get("address", "") or (claim.get("shipping_address", "") if isinstance(claim.get("shipping_address"), str) else "")
    _cust_district = _ship.get("district", "")
    _cust_city = _ship.get("city", "") or claim.get("shipping_city", "")
    _cust_country = _ship.get("country", "") or "Türkiye"

    items = claim.get("items", [])
    # Kısmi gider pusulası: yalnızca seçili kalemler (item_indexes verilirse SADECE onlar hesaplanır)
    _sel_idx = (payload or {}).get("item_indexes")
    if isinstance(_sel_idx, list) and _sel_idx:
        _filtered = []
        for _i in _sel_idx:
            try:
                _filtered.append(items[int(_i)])
            except Exception:
                continue
        if _filtered:
            items = _filtered
    total_net = sum(item.get("price", 0) * item.get("quantity", 1) for item in items)
    total_discount = sum(item.get("discount_amount", 0) * item.get("quantity", 1) for item in items)
    total_gross = sum(item.get("unit_price", 0) * item.get("quantity", 1) for item in items)
    vat_rate = settings.get("default_vat_rate", 10) if settings else 10
    vat_amount = round(total_net * vat_rate / (100 + vat_rate), 2)
    net_without_vat = round(total_net - vat_amount, 2)

    last_gp = await db.gider_pusulasi.find_one({}, sort=[("number", -1)])
    gp_number = (last_gp.get("number", 0) + 1) if last_gp else 1
    # Frontend'den gelen takip numarası (matbu form sıra no ile eşleşir). Verilirse display olarak kullan.
    tracking_no = str((payload or {}).get("tracking_no") or "").strip()
    display_number = tracking_no if tracking_no else f"GP-{gp_number:06d}"

    # Kalemleri ürün kataloğundaki beden ile zenginleştir (barkod -> variant.size)
    gp_items = []
    for _it in items:
        _bc = str(_it.get("barcode", "") or "").strip()
        _size = ""
        if _bc:
            _pv = await db.products.find_one({"variants.barcode": _bc}, {"_id": 0, "variants": 1})
            if _pv:
                for _v in (_pv.get("variants") or []):
                    if str(_v.get("barcode")) == _bc:
                        _size = _v.get("size") or _v.get("beden") or ""
                        break
        gp_items.append({
            "name": _it.get("productName", ""),
            "barcode": _bc,
            "size": _size,
            "quantity": _it.get("quantity", 1),
            "unit_price": _it.get("unit_price", 0),
            "discount": _it.get("discount_amount", 0),
            "net_price": _it.get("price", 0),
            "reason": _it.get("reason", ""),
        })

    gider_pusulasi = {
        "number": gp_number,
        "display_number": display_number,
        "claim_id": claim_id,
        "order_number": claim.get("order_number", ""),
        "date": datetime.now(timezone.utc).isoformat(),
        "company": company,
        "customer": {
            "name": _cust_name,
            "address": _cust_addr,
            "district": _cust_district,
            "city": _cust_city,
            "country": _cust_country,
        },
        "sales_invoice_no": claim.get("invoice_number", ""),
        "cargo_company": claim.get("cargo_provider_name", ""),
        "sales_rep": "",
        "items": gp_items,
        "totals": {
            "gross": total_gross,
            "discount": total_discount,
            "net": total_net,
            "vat_rate": vat_rate,
            "vat_amount": vat_amount,
            "net_without_vat": net_without_vat,
        },
        "claim_type": claim.get("claim_type", ""),
        "claim_reason": claim.get("claim_reason", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.gider_pusulasi.update_one(
        {"claim_id": claim_id},
        {"$set": gider_pusulasi},
        upsert=True
    )

    await db.trendyol_claims.update_one(
        {"claim_id": claim_id},
        {"$set": {"has_gider_pusulasi": True, "gider_pusulasi_no": gider_pusulasi["display_number"]}}
    )

    return {"success": True, "gider_pusulasi": gider_pusulasi}
@router.post("/trendyol/claims/bulk-gider-pusulasi")
async def bulk_generate_gider_pusulasi(payload: dict, current_user: dict = Depends(require_admin)):
    """Generate expense receipts for multiple claims"""
    claim_ids = payload.get("claim_ids", [])
    if not claim_ids:
        raise HTTPException(status_code=400, detail="Claim ID listesi boş")

    start_no = str(payload.get("start_no") or "").strip()
    try:
        base = int(start_no) if start_no else None
    except ValueError:
        base = None

    results = []
    n = 0
    for cid in claim_ids:
        try:
            tno = f"{base + n:06d}" if base is not None else None
            result = await generate_gider_pusulasi(cid, {"tracking_no": tno} if tno else None, current_user)
            results.append(result.get("gider_pusulasi"))
            n += 1
        except Exception:
            pass

    next_no = f"{base + n:06d}" if base is not None else ""
    return {"success": True, "gider_pusulalari": results, "count": len(results), "next_no": next_no}
@router.post("/trendyol/products/{product_id}/update-stock-price")
async def update_trendyol_stock_price(
    product_id: str,
    current_user: dict = Depends(require_admin)
):
    """Tek bir ürünün stok ve fiyatını Trendyol'a gönderir."""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Ürün bulunamadı")

    from trendyol_client import TrendyolClient
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    # Varyantlı ürün mü?
    items = []
    variants = product.get("variants", [])
    trendyol_multiplier = float(config.get("default_markup", 0) or 0)
    base_price = product.get("price", 0)
    sale_price = base_price  # Trendyol: indirimsiz satis fiyati
    
    if trendyol_multiplier > 0:
        sale_price = sale_price * (1 + trendyol_multiplier / 100)
        base_price = base_price * (1 + trendyol_multiplier / 100)

    if variants:
        for v in variants:
            barcode = v.get("barcode", "")
            if not barcode:
                continue
            v_price = base_price + (v.get("price_diff", 0) or 0)
            v_sale = sale_price + (v.get("price_diff", 0) or 0)
            items.append({
                "barcode": barcode,
                "quantity": v.get("stock", 0),
                "salePrice": round(v_sale, 2),
                "listPrice": round(v_price, 2)
            })
    else:
        barcode = product.get("barcode", "")
        if barcode:
            items.append({
                "barcode": barcode,
                "quantity": product.get("stock", 0),
                "salePrice": round(sale_price, 2),
                "listPrice": round(base_price, 2)
            })

    if not items:
        raise HTTPException(status_code=400, detail="Ürünün barkodu bulunamadı")

    try:
        result = await client.update_price_and_inventory(items)
    except Exception as e:
        logger.error(f"Trendyol stock/price update error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

    batch_id = result.get("batchRequestId", "")
    await db.products.update_one(
        {"id": product_id},
        {"$set": {"trendyol_stock_price_batch": str(batch_id), "trendyol_stock_price_updated": datetime.now(timezone.utc).isoformat()}}
    )

    return {
        "success": True,
        "message": f"{len(items)} kalem stok/fiyat güncellendi",
        "batch_id": batch_id,
        "items_count": len(items)
    }
@router.post("/trendyol/categories/{category_id}/update-stock-price")
async def update_trendyol_category_stock_price(
    category_id: str,
    current_user: dict = Depends(require_admin)
):
    """Bir kategorideki tüm ürünlerin stok ve fiyatlarını Trendyol'a gönderir."""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    # Kategorideki tüm ürünleri bul
    category = await db.categories.find_one({"id": category_id}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Kategori bulunamadı")

    products = await db.products.find(
        {"category_name": category.get("name"), "is_active": True},
        {"_id": 0}
    ).to_list(500)

    if not products:
        # Fallback: try category id
        products = await db.products.find(
            {"category_id": category_id, "is_active": True},
            {"_id": 0}
        ).to_list(500)

    if not products:
        raise HTTPException(status_code=404, detail="Bu kategoride ürün bulunamadı")

    from trendyol_client import TrendyolClient
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    items = []
    for product in products:
        trendyol_multiplier = float(config.get("default_markup", 0) or 0)
        base_price = product.get("price", 0)
        sale_price = base_price  # Trendyol: indirimsiz satis fiyati
        
        if trendyol_multiplier > 0:
            sale_price = sale_price * (1 + trendyol_multiplier / 100)
            base_price = base_price * (1 + trendyol_multiplier / 100)

        variants = product.get("variants", [])
        if variants:
            for v in variants:
                barcode = v.get("barcode", "")
                if not barcode:
                    continue
                v_price = base_price + (v.get("price_diff", 0) or 0)
                v_sale = sale_price + (v.get("price_diff", 0) or 0)
                items.append({
                    "barcode": barcode,
                    "quantity": v.get("stock", 0),
                    "salePrice": round(v_sale, 2),
                    "listPrice": round(v_price, 2)
                })
        else:
            barcode = product.get("barcode", "")
            if barcode:
                items.append({
                    "barcode": barcode,
                    "quantity": product.get("stock", 0),
                    "salePrice": round(sale_price, 2),
                    "listPrice": round(base_price, 2)
                })

    if not items:
        raise HTTPException(status_code=400, detail="Bu kategorideki ürünlerin barkodu bulunamadı")

    # Trendyol max 1000 item per request (v2 client metodu — güncel endpoint)
    batch_ids = []
    for i in range(0, len(items), 1000):
        chunk = items[i:i+1000]
        try:
            result = await client.update_price_and_inventory(chunk)
            batch_ids.append(result.get("batchRequestId", ""))
        except Exception as e:
            logger.error(f"Trendyol category stock/price update error: {str(e)}")

    return {
        "success": True,
        "message": f"{category.get('name')} kategorisindeki {len(items)} kalem stok/fiyat güncellendi",
        "items_count": len(items),
        "batch_ids": batch_ids
    }
@router.get("/trendyol/cargo/label/{shipment_package_id}")
async def get_trendyol_cargo_label_pkg(
    shipment_package_id: str,
    current_user: dict = Depends(require_admin)
):
    """Trendyol kargo etiketi PDF/ZPL verisini getirir."""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    from trendyol_client import TrendyolClient
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    try:
        url = f"{client.base_url}/shipment/sellers/{client.supplier_id}/shipment-packages/{shipment_package_id}/shipping-label"
        async with httpx.AsyncClient(timeout=30.0) as http_client:
            headers = client._get_headers()
            response = await http_client.get(url, headers=headers)
            response.raise_for_status()
            
            content_type = response.headers.get("content-type", "")
            
            if "application/pdf" in content_type:
                return Response(
                    content=response.content,
                    media_type="application/pdf",
                    headers={"Content-Disposition": f"inline; filename=label_{shipment_package_id}.pdf"}
                )
            else:
                # ZPL veya text format
                return Response(
                    content=response.content,
                    media_type=content_type or "application/octet-stream",
                    headers={"Content-Disposition": f"attachment; filename=label_{shipment_package_id}"}
                )
    except httpx.HTTPStatusError as e:
        logger.error(f"Cargo label error: {e.response.status_code} - {e.response.text}")
        raise HTTPException(status_code=e.response.status_code, detail=f"Etiket alınamadı: {e.response.text}")
    except Exception as e:
        logger.error(f"Cargo label error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
@router.post("/trendyol/claims/{claim_id}/approve")
async def approve_trendyol_claim(
    claim_id: str,
    payload: dict,
    current_user: dict = Depends(require_admin)
):
    """Approve a list of claim items in Trendyol"""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    claim_item_ids = payload.get("claim_item_ids", [])
    if not claim_item_ids:
        raise HTTPException(status_code=400, detail="Onaylanacak iade kalemleri belirtilmedi.")

    from trendyol_client import TrendyolClient
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    try:
        url = f"{client.base_url}/order/sellers/{client.supplier_id}/claims/{claim_id}/items/approve"
        async with httpx.AsyncClient(timeout=30.0) as http_client:
            headers = client._get_headers()
            req_data = {
                "claimLineItemIdList": claim_item_ids
            }
            response = await http_client.put(url, headers=headers, json=req_data)
            response.raise_for_status()
            
            await log_integration_event("trendyol", "claim_approve", current_user["email"], claim_id, "success", f"{len(claim_item_ids)} kalem onaylandı", req_data)
            
            # Update claim in DB with action status
            await db.trendyol_claims.update_one(
                {"claim_id": claim_id},
                {"$set": {
                    "panel_action": "approved",
                    "panel_action_date": datetime.now(timezone.utc).isoformat(),
                    "panel_action_by": current_user["email"],
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )

            # A7: İade onaylanınca stok otomatik geri iade
            try:
                claim_doc = await db.trendyol_claims.find_one({"claim_id": claim_id}, {"_id": 0, "items": 1, "order_number": 1, "stock_restored": 1})
                restocked_items = []
                for item in ([] if (claim_doc or {}).get("stock_restored") else ((claim_doc or {}).get("items") or [])):
                    if str(item.get("claim_item_id", "")) not in [str(x) for x in claim_item_ids]:
                        continue
                    barcode = item.get("barcode", "")
                    qty = int(item.get("quantity", 1) or 1)
                    if not barcode:
                        continue
                    # variant-level restock
                    prod = await db.products.find_one({"variants.barcode": barcode}, {"_id": 0, "id": 1, "variants": 1, "stock": 1})
                    if prod:
                        for v in (prod.get("variants") or []):
                            if v.get("barcode") == barcode:
                                v["stock"] = int(v.get("stock", 0) or 0) + qty
                                break
                        new_total_stock = sum(int(v.get("stock", 0) or 0) for v in prod.get("variants", []))
                        await db.products.update_one(
                            {"id": prod["id"]},
                            {"$set": {"variants": prod["variants"], "stock": new_total_stock, "updated_at": datetime.now(timezone.utc).isoformat()}}
                        )
                        restocked_items.append({"barcode": barcode, "qty": qty, "product_id": prod["id"]})
                    else:
                        # Fallback: try product.barcode
                        p2 = await db.products.find_one({"barcode": barcode}, {"_id": 0, "id": 1, "stock": 1})
                        if p2:
                            await db.products.update_one(
                                {"id": p2["id"]},
                                {"$inc": {"stock": qty}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
                            )
                            restocked_items.append({"barcode": barcode, "qty": qty, "product_id": p2["id"]})

                # Stok hareketi loglama
                if restocked_items:
                    await db.stock_movements.insert_one({
                        "id": str(uuid.uuid4()),
                        "type": "return_approved",
                        "claim_id": claim_id,
                        "order_number": claim_doc.get("order_number", ""),
                        "items": restocked_items,
                        "created_by": current_user["email"],
                        "created_at": datetime.now(timezone.utc).isoformat(),
                    })
                    await db.trendyol_claims.update_one(
                        {"claim_id": claim_id},
                        {"$set": {"stock_restored": True, "stock_restored_at": datetime.now(timezone.utc).isoformat(), "stock_restored_source": "approve"}}
                    )
            except Exception as restock_err:
                logger.error(f"Restock after claim approve failed: {restock_err}")
                # non-fatal

            return {"success": True, "message": "İade işlemi Trendyol tarafında onaylandı."}

    except httpx.HTTPStatusError as e:
        logger.error(f"Claim approve error: {e.response.status_code} - {e.response.text}")
        await log_integration_event("trendyol", "claim_approve", current_user["email"], claim_id, "error", f"API Hatası: {e.response.text}")
        raise HTTPException(status_code=e.response.status_code, detail=f"Onay işlemi başarısız: {e.response.text}")
    except Exception as e:
        logger.error(f"Claim approve error: {str(e)}")
        await log_integration_event("trendyol", "claim_approve", current_user["email"], claim_id, "error", str(e))
        raise HTTPException(status_code=500, detail=str(e))
@router.post("/trendyol/claims/{claim_id}/issue")
async def issue_trendyol_claim(
    claim_id: str,
    payload: dict,
    current_user: dict = Depends(require_admin)
):
    """Reject/Issue a list of claim items in Trendyol"""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    claim_item_ids = payload.get("claim_item_ids", [])
    issue_reason_id = payload.get("issue_reason_id")
    description = payload.get("description", "")

    if not claim_item_ids or not issue_reason_id:
        raise HTTPException(status_code=400, detail="İtiraz edilecek kalemler veya itiraz sebebi belirtilmedi.")

    from trendyol_client import TrendyolClient
    client = TrendyolClient(
        supplier_id=config["supplier_id"],
        api_key=config["api_key"],
        api_secret=config["api_secret"],
        mode=config["mode"]
    )

    try:
        url = f"{client.base_url}/order/sellers/{client.supplier_id}/claims/{claim_id}/issue"
        async with httpx.AsyncClient(timeout=30.0) as http_client:
            headers = client._get_headers()
            req_data = {
                "claimIssueReasonId": int(issue_reason_id),
                "claimItemIdList": claim_item_ids,
                "description": description
            }
            response = await http_client.post(url, headers=headers, json=req_data)
            response.raise_for_status()
            
            await log_integration_event("trendyol", "claim_issue", current_user["email"], claim_id, "success", f"{len(claim_item_ids)} kalem için itiraz açıldı", req_data)
            
            # Update claim in DB with action status
            await db.trendyol_claims.update_one(
                {"claim_id": claim_id},
                {"$set": {
                    "panel_action": "issued",
                    "panel_action_date": datetime.now(timezone.utc).isoformat(),
                    "panel_action_by": current_user["email"],
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            return {"success": True, "message": "İade işlemi için Trendyol tarafında itiraz oluşturuldu."}

    except httpx.HTTPStatusError as e:
        logger.error(f"Claim issue error: {e.response.status_code} - {e.response.text}")
        await log_integration_event("trendyol", "claim_issue", current_user["email"], claim_id, "error", f"API Hatası: {e.response.text}")
        raise HTTPException(status_code=e.response.status_code, detail=f"İtiraz işlemi başarısız: {e.response.text}")
    except Exception as e:
        logger.error(f"Claim issue error: {str(e)}")
        await log_integration_event("trendyol", "claim_issue", current_user["email"], claim_id, "error", str(e))
        raise HTTPException(status_code=500, detail=str(e))
@router.post("/trendyol/invoices/{order_number}")
async def upload_invoice_to_trendyol(order_number: str, payload: dict, current_user: dict = Depends(require_admin)):
    """Upload invoice link to Trendyol for a given order number"""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapilandirilmamis")

    invoice_link = payload.get("invoice_link", "").strip()
    invoice_number = payload.get("invoice_number", "").strip()
    if not invoice_link:
        raise HTTPException(status_code=400, detail="Fatura linki bos olamaz")

    order = await db.orders.find_one({"order_number": order_number, "platform": "trendyol"})
    if not order:
        raise HTTPException(status_code=404, detail="Siparis bulunamadi")

    package_id = order.get("trendyol_package_id")
    if not package_id:
        raise HTTPException(status_code=400, detail="Trendyol paket ID bulunamadi")

    supplier_id = config["supplier_id"]
    headers = await get_trendyol_headers()
    base_url = config["base_url"]

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            url = f"{base_url}/sapigw/suppliers/{supplier_id}/shipment-packages/{package_id}/invoices"
            body = {
                "invoiceNumber": invoice_number or f"FAT-{order_number}",
                "invoiceLink": invoice_link
            }
            resp = await client.post(url, headers=headers, json=body)
            resp.raise_for_status()

        await db.orders.update_one(
            {"order_number": order_number},
            {"$set": {"invoice_link": invoice_link, "invoice_number": invoice_number, "invoice_uploaded_at": datetime.now(timezone.utc).isoformat()}}
        )

        await log_integration_event("trendyol", "upload_invoice", current_user["email"], order_number, "success", "Fatura yuklendi", body)
        return {"success": True, "message": "Fatura Trendyol'a basariyla yuklendi"}
    except httpx.HTTPStatusError as e:
        logger.error(f"Invoice upload error: {e.response.text}")
        await log_integration_event("trendyol", "upload_invoice", current_user["email"], order_number, "error", e.response.text)
        raise HTTPException(status_code=e.response.status_code, detail=f"Fatura yukleme hatasi: {e.response.text}")
    except Exception as e:
        logger.error(f"Invoice upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
@router.post("/trendyol/products/{product_id}/sync")
async def sync_product_to_trendyol(product_id: str, current_user: dict = Depends(require_admin)):
    """Full product synchronization to Trendyol"""
    config = await get_trendyol_config()
    if not config["is_active"]:
        raise HTTPException(status_code=400, detail="Trendyol entegrasyonu yapılandırılmamış")

    from datetime import datetime, timezone
    started_at = datetime.now(timezone.utc).isoformat()
    product = None
    
    try:
        product = await db.products.find_one({"id": product_id}, {"_id": 0})
        if not product:
            raise Exception("Ürün bulunamadı")

        ty_cat_id = product.get("trendyol_category_id")
        if not ty_cat_id:
            # Try finding in the category mapped to this product
            cat = await db.categories.find_one({"name": product.get("category_name")})
            if not cat:
                cat = await db.categories.find_one({"id": product.get("category_id")})
            
            if cat and cat.get("trendyol_category_id"):
                ty_cat_id = cat.get("trendyol_category_id")
            else:
                raise Exception("Ürün için Trendyol kategorisi seçilmemiş")

        # Fetch mapping details from category
        mapping_cat = await db.categories.find_one({"trendyol_category_id": ty_cat_id})
        attr_mappings = mapping_cat.get("attribute_mappings", []) if mapping_cat else []
        val_mappings = mapping_cat.get("value_mappings", {}) if mapping_cat else {}
        default_mappings = mapping_cat.get("default_mappings", {}) if mapping_cat else {}

        from trendyol_client import TrendyolClient
        client = TrendyolClient(
            supplier_id=config["supplier_id"],
            api_key=config["api_key"],
            api_secret=config["api_secret"],
            mode=config["mode"]
        )

        # Calculate prices
        base_price = product.get("price", 0)
        list_price = calculate_trendyol_price(base_price, product, config)
        sale_price = calculate_trendyol_price(base_price, product, config)  # Trendyol: indirimsiz satis fiyati

        # Build items
        items = []
        variants = product.get("variants", [])
        
        # Common attributes for all variants
        common_attrs = []
        for am in attr_mappings:
            ty_attr_id = int(am["trendyol_attr_id"])
            local_name = am["local_attr"]
            # Find value in product attributes
            val = next((a["value"] for a in product.get("attributes", []) if a["type"] == local_name), None)
            # Try default if not found
            if not val:
                val = default_mappings.get(str(ty_attr_id))
                
            if val:
                mapping_key = f"{ty_attr_id}:{val}"
                ty_val_id = val_mappings.get(mapping_key)
                if ty_val_id:
                    common_attrs.append({"attributeId": ty_attr_id, "attributeValueId": int(ty_val_id)})
                else:
                    common_attrs.append({"attributeId": ty_attr_id, "customAttributeValue": val})

        if variants:
            for v in variants:
                v_attrs = common_attrs.copy()
                
                # Map Size (Beden) and Color (Renk)
                for am in attr_mappings:
                    ty_attr_id = str(am["trendyol_attr_id"])
                    local_name = am["local_attr"]
                    
                    # Check if it's Beden or Renk
                    if local_name.lower() == "beden":
                        sz = v.get("size")
                        if sz:
                            m_key = f"{ty_attr_id}:{sz}"
                            v_id = val_mappings.get(m_key)
                            if v_id:
                                v_attrs.append({"attributeId": int(ty_attr_id), "attributeValueId": int(v_id)})
                            else:
                                v_attrs.append({"attributeId": int(ty_attr_id), "customAttributeValue": sz})
                    
                    elif local_name.lower() == "renk":
                        clr = v.get("color")
                        if clr:
                            m_key = f"{ty_attr_id}:{clr}"
                            v_id = val_mappings.get(m_key)
                            if v_id:
                                v_attrs.append({"attributeId": int(ty_attr_id), "attributeValueId": int(v_id)})
                            else:
                                v_attrs.append({"attributeId": int(ty_attr_id), "customAttributeValue": clr})

                # Pricing with price_diff
                diff = float(v.get("price_diff", 0) or 0)
                v_list = round(list_price + diff, 2)
                v_sale = round(sale_price + diff, 2)

                item = {
                    "barcode": v.get("barcode") or product.get("barcode"),
                    "title": product.get("name"),
                    "productMainId": product.get("stock_code"),
                    "brandId": product.get("trendyol_brand_id") or config.get("default_brand_id") or 975755,
                    "categoryId": int(ty_cat_id),
                    "quantity": v.get("stock", 0),
                    "stockCode": v.get("stock_code") or product.get("stock_code"),
                    "dimensionalWeight": product.get("cargo_weight") or 1,
                    "description": product.get("description", ""),
                    "currencyType": "TRY",
                    "listPrice": v_list,
                    "salePrice": v_sale,
                    "vatRate": product.get("vat_rate", config.get("default_vat_rate") or 20),
                    "cargoCompanyId": int(config.get("default_cargo_company_id") or 10),
                    "images": [{"url": img} for img in product.get("images", [])],
                    "attributes": v_attrs
                }
                items.append(item)
        else:
            # Single product
            item = {
                "barcode": product.get("barcode"),
                "title": product.get("name"),
                "productMainId": product.get("stock_code"),
                "brandId": product.get("trendyol_brand_id") or config.get("default_brand_id") or 975755,
                "categoryId": int(ty_cat_id),
                "quantity": product.get("stock", 0),
                "stockCode": product.get("stock_code"),
                "dimensionalWeight": product.get("cargo_weight") or 1,
                "description": product.get("description", ""),
                "currencyType": "TRY",
                "listPrice": list_price,
                "salePrice": sale_price,
                "vatRate": product.get("vat_rate", config.get("default_vat_rate") or 20),
                "cargoCompanyId": int(config.get("default_cargo_company_id") or 10),
                "images": [{"url": img} for img in product.get("images", [])],
                "attributes": common_attrs
            }
            items.append(item)

        result = await client.create_products(items)
        batch_id = result.get("batchRequestId", "")
        
        # Log to the new sync logs screen
        log_doc = {
            "id": generate_id(),
            "started_at": started_at,
            "finished_at": datetime.now(timezone.utc).isoformat(),
            "status": "success",
            "products_attempted": 1,
            "products_sent": len(items),
            "batch_request_id": batch_id,
            "errors": [],
            "message": f"'{product.get('name')}' ürünü tekli olarak aktarıldı."
        }
        await db.trendyol_sync_logs.insert_one(log_doc)

        await log_integration_event(
            platform="trendyol",
            action="product_sync",
            entity_type="product",
            entity_id=product_id,
            status="success",
            message=f"Sync initiated. Batch ID: {batch_id}",
            details={"batch_id": batch_id, "items_count": len(items)}
        )
        
        await db.products.update_one(
            {"id": product_id},
            {"$set": {
                "trendyol_sync_batch": str(batch_id),
                "trendyol_sync_last": datetime.now(timezone.utc).isoformat(),
                "trendyol_status": "synced"
            }}
        )
        
        return {"success": True, "message": "Eşleştirme başlatıldı", "batch_id": batch_id}
    except Exception as e:
        logger.error(f"Trendyol sync error: {str(e)}")
        log_doc = {
            "id": generate_id(),
            "started_at": started_at if 'started_at' in locals() else datetime.now(timezone.utc).isoformat(),
            "status": "error",
            "products_attempted": 1,
            "products_sent": 0,
            "batch_request_id": None,
            "errors": [f"Hata: {str(e)}"],
            "message": f"'{product.get('name') if product else product_id}' aktarımı sırasında hata oluştu."
        }
        await db.trendyol_sync_logs.insert_one(log_doc)
        await log_integration_event("trendyol", "product_sync", "product", product_id, "error", str(e))
        raise HTTPException(status_code=400 if "Ürün" in str(e) or "kategori" in str(e).lower() else 500, detail=f"Trendyol senkronizasyon hatası: {str(e)}")
